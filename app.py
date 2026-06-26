from flask import Flask, render_template, request, jsonify, Response, redirect, url_for, session, send_from_directory, send_file
import sqlite3
import csv
import io
import zipfile
import os
import os.path
import shutil
import calendar
import threading
import json
import re
import secrets
import string
import time
import uuid
import base64
from urllib.parse import quote, unquote
from urllib.request import Request as UrlRequest, urlopen
from urllib.error import HTTPError, URLError
from datetime import datetime, timedelta
from functools import wraps
import smtplib
from email.message import EmailMessage
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix

from app_modules.pagination import get_page_args, like_filter, pagination_meta
from app_modules.jobs import job_manager
from app_modules.db_compat import connect_database, is_postgres_enabled, table_exists as compat_table_exists, table_columns as compat_table_columns

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow, Flow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
app.secret_key = os.environ.get('SECRET_KEY', 'change-me-for-render-testing')

# --- Security: session inactivity timeout ---
DESKTOP_SESSION_IDLE_TIMEOUT_SECONDS = int(os.environ.get('DESKTOP_SESSION_IDLE_TIMEOUT_SECONDS', '900'))
SESSION_IDLE_TIMEOUT_SECONDS = DESKTOP_SESSION_IDLE_TIMEOUT_SECONDS
SESSION_TIMEOUT_LOGIN_URL = '/login?timeout=1'
app.config['SESSION_REFRESH_EACH_REQUEST'] = True


# --- Render / Cloud-ready configuration ---
def get_database_path():
    return os.environ.get('DATABASE_PATH') or os.environ.get('DB_PATH') or 'database.db'

def ensure_parent_dir(path):
    parent = os.path.dirname(os.path.abspath(path))
    if parent:
        os.makedirs(parent, exist_ok=True)

# --- File Upload Configuration ---
def _default_upload_folder():
    # Render's normal app filesystem is rebuilt on deploy. Paid services with a
    # persistent disk should store uploads under /var/data so company logos,
    # documents and attachments survive redeploys/restarts.
    if os.environ.get('UPLOAD_FOLDER'):
        return os.environ.get('UPLOAD_FOLDER')
    if os.environ.get('RENDER') and os.path.isdir('/var/data'):
        return '/var/data/uploads'
    return 'uploads'

UPLOAD_FOLDER = _default_upload_folder()
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('MAX_CONTENT_LENGTH', 50 * 1024 * 1024))

if os.environ.get('FLASK_ENV') == 'production' or os.environ.get('RENDER'):
    app.config.update(
        SESSION_COOKIE_SECURE=True,
        SESSION_COOKIE_HTTPONLY=True,
        SESSION_COOKIE_SAMESITE='Lax'
    )

# --- Display formatting helpers ---
# General Ledger date style: YYYY-MM-DD. Trial Balance money style: en-ZA number format.
def format_display_date(value):
    if value in (None, ''):
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    text = str(value).strip()
    if not text:
        return ''
    # Preserve month-only values such as payroll months.
    if re.fullmatch(r'\d{4}-\d{2}', text):
        return text
    # ISO dates/datetimes used by the General Ledger and booking records.
    iso_match = re.match(r'^(\d{4})[-/](\d{2})[-/](\d{2})', text)
    if iso_match:
        return f"{iso_match.group(1)}-{iso_match.group(2)}-{iso_match.group(3)}"
    # Common South African display/input forms converted to the General Ledger style.
    sa_match = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})', text)
    if sa_match:
        day, month, year = [int(x) for x in sa_match.groups()]
        try:
            return datetime(year, month, day).strftime('%Y-%m-%d')
        except Exception:
            return text
    for fmt in ('%d %b %Y', '%d %B %Y', '%b %d, %Y', '%B %d, %Y'):
        try:
            return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return text


def format_display_money(value):
    try:
        amount = float(value or 0)
    except (TypeError, ValueError):
        amount = 0.0
    sign = '-' if amount < 0 else ''
    amount = abs(amount)
    base = f"{amount:,.2f}"
    # Match JavaScript Number.toLocaleString('en-ZA') used by the Trial Balance.
    base = base.replace(',', ' ').replace('.', ',')
    return sign + base.replace(' ', '\u00a0')

app.jinja_env.filters['fmt_date'] = format_display_date
app.jinja_env.filters['fmt_money'] = format_display_money
app.jinja_env.globals['fmt_date'] = format_display_date
app.jinja_env.globals['fmt_money'] = format_display_money

# --- Industry Template Defaults ---
INDUSTRY_TEMPLATES = {
    "Cleaning": {
        "labels": {
            "booking_name": "Booking", "client_name": "Client", "service_name": "Service", "employee_name": "Employee",
            "booking_app_title": "Booking & Ops", "finance_app_title": "Finance", "job_location_label": "Address"
        },
        "booking_fields": [
            ("bedrooms", "Bedrooms", "number", 0, 0), ("bathrooms", "Bathrooms", "number", 0, 0),
            ("keys_required", "Keys Required", "checkbox", 0, 0), ("pets_on_premises", "Pets on Premises", "checkbox", 0, 0),
            ("supplies_provided", "Cleaning Supplies Provided by Client", "checkbox", 0, 0),
            ("before_after_photos", "Before/After Photos Required", "checkbox", 0, 0)
        ],
        "finance_categories": ["Cleaning Supplies", "Transport / Fuel", "Staff Wages", "Uniforms", "Equipment", "Marketing", "Admin"],
        "evaluation_scorecard": [
            {"block_name": "Experience & Skills", "questions": [{"question_text": "Relevant Background", "max_score": 10}, {"question_text": "Technical Cleaning Skills", "max_score": 10}, {"question_text": "Product/Chemical Knowledge", "max_score": 10}]},
            {"block_name": "Professionalism & Attitude", "questions": [{"question_text": "Punctuality & Presentation", "max_score": 10}, {"question_text": "Communication Skills", "max_score": 10}, {"question_text": "Work Ethic / Motivation", "max_score": 10}]},
            {"block_name": "Reliability & Trust", "questions": [{"question_text": "Reference Checks / History", "max_score": 10}, {"question_text": "Honesty & Transparency", "max_score": 10}, {"question_text": "Problem Solving Scenario", "max_score": 10}]},
            {"block_name": "Logistics & Fit", "questions": [{"question_text": "Transport Independence", "max_score": 10}, {"question_text": "Schedule Flexibility", "max_score": 10}, {"question_text": "Overall Cultural Fit", "max_score": 10}]}
        ]
    },
    "Transportation": {
        "labels": {
            "booking_name": "Trip", "client_name": "Customer", "service_name": "Load Type", "employee_name": "Driver",
            "booking_app_title": "Trip & Dispatch", "finance_app_title": "Fleet Finance", "job_location_label": "Pickup / Drop-off"
        },
        "booking_fields": [
            ("pickup_location", "Pickup Location", "text", 1, 1), ("dropoff_location", "Drop-off Location", "text", 1, 1),
            ("truck", "Truck / Vehicle", "text", 1, 1), ("trailer", "Trailer", "text", 0, 1),
            ("distance_km", "Distance KM", "number", 0, 1), ("load_description", "Load Description", "textarea", 0, 1),
            ("fuel_estimate", "Fuel Estimate", "number", 0, 1), ("toll_fees", "Toll Fees", "number", 0, 1)
        ],
        "finance_categories": ["Fuel", "Truck Maintenance", "Tyres", "Tolls", "Licensing", "Insurance", "Driver Wages", "Subcontracted Transport"],
        "evaluation_scorecard": [
            {"block_name": "Driving & Route Experience", "questions": [{"question_text": "Long-distance driving experience", "max_score": 10}, {"question_text": "Route planning / navigation", "max_score": 10}, {"question_text": "Vehicle inspection knowledge", "max_score": 10}]},
            {"block_name": "Safety & Compliance", "questions": [{"question_text": "Defensive driving attitude", "max_score": 10}, {"question_text": "Load safety awareness", "max_score": 10}, {"question_text": "Licence / permit readiness", "max_score": 10}]},
            {"block_name": "Reliability & Conduct", "questions": [{"question_text": "Punctuality", "max_score": 10}, {"question_text": "Communication with dispatch", "max_score": 10}, {"question_text": "Incident honesty / reporting", "max_score": 10}]},
            {"block_name": "Shift Fit", "questions": [{"question_text": "Night / weekend availability", "max_score": 10}, {"question_text": "Long-trip stamina", "max_score": 10}, {"question_text": "Customer handling", "max_score": 10}]}
        ]
    },
    "Construction": {
        "labels": {
            "booking_name": "Project / Site Visit", "client_name": "Client", "service_name": "Work Type", "employee_name": "Worker",
            "booking_app_title": "Projects & Sites", "finance_app_title": "Project Finance", "job_location_label": "Site Address"
        },
        "booking_fields": [
            ("project_name", "Project Name", "text", 1, 1), ("site_address", "Site Address", "textarea", 1, 1),
            ("foreman", "Foreman", "text", 0, 1), ("work_type", "Work Type", "text", 0, 1),
            ("materials_used", "Materials Used", "textarea", 0, 1), ("equipment_used", "Equipment Used", "textarea", 0, 1),
            ("start_time", "Start Time", "time", 0, 1), ("end_time", "End Time", "time", 0, 1)
        ],
        "finance_categories": ["Materials", "Labour", "Equipment Hire", "Subcontractors", "Transport", "Tools", "Site Consumables", "Waste Removal", "Permits"],
        "evaluation_scorecard": [
            {"block_name": "Trade Skills", "questions": [{"question_text": "Relevant site experience", "max_score": 10}, {"question_text": "Tool / equipment handling", "max_score": 10}, {"question_text": "Quality of workmanship", "max_score": 10}]},
            {"block_name": "Site Safety", "questions": [{"question_text": "PPE and safety awareness", "max_score": 10}, {"question_text": "Working at heights / hazard awareness", "max_score": 10}, {"question_text": "Follows supervisor instructions", "max_score": 10}]},
            {"block_name": "Reliability & Productivity", "questions": [{"question_text": "Punctuality", "max_score": 10}, {"question_text": "Work pace", "max_score": 10}, {"question_text": "Teamwork", "max_score": 10}]},
            {"block_name": "Project Fit", "questions": [{"question_text": "Travel to site", "max_score": 10}, {"question_text": "Overtime availability", "max_score": 10}, {"question_text": "Foreman feedback", "max_score": 10}]}
        ]
    },
    "Finance Auditing": {
        "labels": {
            "booking_name": "Engagement / Site Visit", "client_name": "Client", "service_name": "Audit Service", "employee_name": "Consultant",
            "booking_app_title": "Engagements & Scheduling", "finance_app_title": "Practice Finance", "job_location_label": "Client Site / Remote"
        },
        "booking_fields": [
            ("engagement_type", "Engagement Type", "dropdown", 1, 1),
            ("audit_period", "Audit / Review Period", "text", 0, 1),
            ("client_contact", "Client Contact Person", "text", 0, 1),
            ("onsite_remote", "Onsite / Remote", "dropdown", 0, 1),
            ("document_request_status", "Document Request Status", "dropdown", 0, 1),
            ("risk_level", "Risk Level", "dropdown", 0, 1),
            ("working_papers_ref", "Working Papers Reference", "text", 0, 1),
            ("deliverable_due_date", "Deliverable Due Date", "date", 0, 1),
            ("review_notes", "Review / Audit Notes", "textarea", 0, 1)
        ],
        "finance_categories": ["Professional Staff Costs", "Software Subscriptions", "Travel to Client", "Printing & Stationery", "Training & CPD", "Professional Indemnity", "Subcontracted Auditors", "Admin", "Marketing", "Bank Charges"],
        "evaluation_scorecard": [
            {"block_name": "Technical Accounting & Audit", "questions": [{"question_text": "Accounting/audit experience", "max_score": 10}, {"question_text": "Understanding of audit/review procedures", "max_score": 10}, {"question_text": "Tax and compliance awareness", "max_score": 10}]},
            {"block_name": "Accuracy & Documentation", "questions": [{"question_text": "Attention to detail", "max_score": 10}, {"question_text": "Working paper discipline", "max_score": 10}, {"question_text": "Ability to follow checklists", "max_score": 10}]},
            {"block_name": "Client Communication", "questions": [{"question_text": "Professional communication", "max_score": 10}, {"question_text": "Deadline management", "max_score": 10}, {"question_text": "Confidentiality awareness", "max_score": 10}]},
            {"block_name": "Practice Fit", "questions": [{"question_text": "Ethical judgement", "max_score": 10}, {"question_text": "Team collaboration", "max_score": 10}, {"question_text": "Availability during deadline periods", "max_score": 10}]}
        ]
    },
    "Custom": {
        "labels": {
            "booking_name": "Booking", "client_name": "Client", "service_name": "Service", "employee_name": "Employee",
            "booking_app_title": "Booking & Ops", "finance_app_title": "Finance", "job_location_label": "Address"
        },
        "booking_fields": [],
        "finance_categories": [],
        "evaluation_scorecard": []
    }
}


def _normalise_template_config(template):
    labels = dict(template.get("labels") or {})
    default_labels = INDUSTRY_TEMPLATES["Cleaning"].get("labels", {})
    for key, value in default_labels.items():
        labels.setdefault(key, value)

    booking_fields = []
    for item in template.get("booking_fields") or []:
        if isinstance(item, dict):
            field_key = (item.get("field_key") or item.get("key") or item.get("field_label") or "").strip().lower().replace(" ", "_")
            field_label = item.get("field_label") or item.get("label") or field_key.replace("_", " ").title()
            field_type = item.get("field_type") or item.get("type") or "text"
            required = 1 if item.get("required") in (1, True, "1", "true", "True", "yes") else 0
            visible = 0 if item.get("visible") in (0, False, "0", "false", "False", "no") else 1
        else:
            row = list(item)
            field_key = str(row[0]) if len(row) > 0 else ""
            field_label = str(row[1]) if len(row) > 1 else field_key.replace("_", " ").title()
            field_type = str(row[2]) if len(row) > 2 else "text"
            required = int(row[3]) if len(row) > 3 else 0
            visible = int(row[4]) if len(row) > 4 else 1
        if not field_key:
            continue
        field_type = field_type if field_type in ["text", "number", "date", "time", "dropdown", "checkbox", "textarea"] else "text"
        booking_fields.append((field_key, field_label, field_type, required, visible))

    finance_categories = []
    for cat in template.get("finance_categories") or []:
        if isinstance(cat, dict):
            name = cat.get("category_name") or cat.get("name")
        else:
            name = cat
        if name:
            finance_categories.append(str(name))

    scorecard = template.get("evaluation_scorecard") or []
    if not isinstance(scorecard, list):
        scorecard = []

    return {
        "labels": labels,
        "booking_fields": booking_fields,
        "finance_categories": finance_categories,
        "evaluation_scorecard": scorecard
    }


def get_all_industry_templates(conn=None):
    templates = {name: _normalise_template_config(cfg) for name, cfg in INDUSTRY_TEMPLATES.items()}
    should_close = False
    if conn is None:
        try:
            conn = get_db_connection()
            should_close = True
        except Exception:
            conn = None
    if conn is not None:
        try:
            rows = conn.execute('SELECT template_name, template_json FROM industry_template_uploads WHERE active=1 ORDER BY template_name').fetchall()
            for row in rows:
                try:
                    templates[row['template_name']] = _normalise_template_config(json.loads(row['template_json'] or '{}'))
                except Exception:
                    continue
        except sqlite3.OperationalError:
            pass
        finally:
            if should_close:
                conn.close()
    return templates


def get_template_defaults(industry, conn=None):
    templates = get_all_industry_templates(conn)
    return templates.get(industry or "Cleaning", templates.get("Cleaning", _normalise_template_config(INDUSTRY_TEMPLATES["Cleaning"])))


def get_valid_industry_template_name(industry, conn=None):
    templates = get_all_industry_templates(conn)
    return industry if industry in templates else "Cleaning"


def get_safe_industry_templates(conn=None):
    safe_templates = {}
    for name, cfg in get_all_industry_templates(conn).items():
        safe_templates[name] = {
            "labels": cfg.get("labels", {}),
            "booking_fields": [
                {"field_key": f[0], "field_label": f[1], "field_type": f[2], "required": f[3], "visible": f[4]}
                for f in cfg.get("booking_fields", [])
            ],
            "finance_categories": cfg.get("finance_categories", []),
            "evaluation_scorecard": cfg.get("evaluation_scorecard", [])
        }
    return safe_templates


def ensure_tenant_template(conn, company_id, industry_template=None, force_reset=False):
    company = conn.execute('SELECT * FROM companies WHERE id=?', (company_id,)).fetchone()
    if not company:
        return
    industry = industry_template or dict(company).get('industry_template') or 'Cleaning'
    industry = get_valid_industry_template_name(industry, conn)
    defaults = get_template_defaults(industry, conn)
    conn.execute('UPDATE companies SET industry_template=? WHERE id=?', (industry, company_id))

    if force_reset:
        conn.execute('DELETE FROM tenant_labels WHERE company_id=?', (company_id,))
        conn.execute('DELETE FROM tenant_custom_fields WHERE company_id=? AND module_name="booking"', (company_id,))
        conn.execute('DELETE FROM finance_categories WHERE company_id=?', (company_id,))

    for key, value in defaults['labels'].items():
        exists = conn.execute('SELECT 1 FROM tenant_labels WHERE company_id=? AND label_key=?', (company_id, key)).fetchone()
        if not exists:
            conn.execute('INSERT INTO tenant_labels (company_id, module_name, label_key, label_value) VALUES (?, ?, ?, ?)',
                         (company_id, 'global', key, value))

    for idx, (field_key, field_label, field_type, required, visible) in enumerate(defaults['booking_fields'], start=1):
        exists = conn.execute('SELECT 1 FROM tenant_custom_fields WHERE company_id=? AND module_name="booking" AND field_key=?',
                              (company_id, field_key)).fetchone()
        if not exists:
            conn.execute("""INSERT INTO tenant_custom_fields
                            (company_id, module_name, field_key, field_label, field_type, required, visible, sort_order, options_json, industry_template)
                            VALUES (?, 'booking', ?, ?, ?, ?, ?, ?, ?, ?)""",
                         (company_id, field_key, field_label, field_type, required, visible, idx, '[]', industry))

    for idx, category in enumerate(defaults['finance_categories'], start=1):
        exists = conn.execute('SELECT 1 FROM finance_categories WHERE company_id=? AND category_name=?', (company_id, category)).fetchone()
        if not exists:
            conn.execute("""INSERT INTO finance_categories (company_id, category_name, category_type, active, industry_template, sort_order)
                            VALUES (?, ?, 'expense', 1, ?, ?)""", (company_id, category, industry, idx))

    if force_reset:
        block_rows = conn.execute('SELECT id FROM tenant_scorecard_blocks WHERE company_id=?', (company_id,)).fetchall()
        for br in block_rows:
            conn.execute('DELETE FROM tenant_scorecard_questions WHERE block_id=?', (br['id'],))
        conn.execute('DELETE FROM tenant_scorecard_blocks WHERE company_id=?', (company_id,))

    existing_blocks = conn.execute('SELECT COUNT(*) FROM tenant_scorecard_blocks WHERE company_id=?', (company_id,)).fetchone()[0]
    if existing_blocks == 0:
        seed_scorecard_template(conn, company_id, defaults.get('evaluation_scorecard') or [], industry)


def seed_scorecard_template(conn, company_id, blocks, industry):
    for block_idx, block in enumerate(blocks or [], start=1):
        block_name = (block.get('block_name') or '').strip()
        if not block_name:
            continue
        cur = conn.execute("""INSERT INTO tenant_scorecard_blocks (company_id, block_name, sort_order, active, industry_template)
                              VALUES (?, ?, ?, 1, ?)""", (company_id, block_name, block_idx, industry))
        block_id = cur.lastrowid
        for q_idx, q in enumerate(block.get('questions') or [], start=1):
            question_text = (q.get('question_text') or '').strip()
            if not question_text:
                continue
            try:
                max_score = int(q.get('max_score') or 10)
            except Exception:
                max_score = 10
            max_score = max(1, min(max_score, 100))
            conn.execute("""INSERT INTO tenant_scorecard_questions (company_id, block_id, question_key, question_text, max_score, sort_order, active)
                            VALUES (?, ?, ?, ?, ?, ?, 1)""", (company_id, block_id, f"q{block_idx}_{q_idx}", question_text, max_score, q_idx))


def get_tenant_labels(conn, company_id):
    rows = conn.execute("""SELECT label_key, label_value
                         FROM tenant_labels
                         WHERE company_id=?""", (company_id,)).fetchall()
    labels = {r['label_key']: r['label_value'] for r in rows}
    defaults = get_template_defaults('Cleaning')['labels']
    for key, value in defaults.items():
        labels.setdefault(key, value)
    return labels


def normalise_dropdown_option_value(label):
    return (label or '').strip()


def get_custom_field_options(conn, company_id, field_id, include_inactive=False):
    if include_inactive:
        rows = conn.execute("""SELECT id, option_label, option_value, sort_order, active
                             FROM tenant_custom_field_options
                             WHERE company_id=? AND field_id=?
                             ORDER BY sort_order, id""", (company_id, field_id)).fetchall()
    else:
        rows = conn.execute("""SELECT id, option_label, option_value, sort_order, active
                             FROM tenant_custom_field_options
                             WHERE company_id=? AND field_id=? AND active=1
                             ORDER BY sort_order, id""", (company_id, field_id)).fetchall()
    return [dict(r) for r in rows]


def parse_options_json(raw):
    try:
        parsed = json.loads(raw or '[]')
        if isinstance(parsed, list):
            return [str(x).strip() for x in parsed if str(x).strip()]
    except Exception:
        pass
    return [x.strip() for x in str(raw or '').split(',') if x.strip()]


def get_tenant_custom_fields(conn, company_id, module_name='booking', visible_only=True):
    if visible_only:
        rows = conn.execute("""SELECT id, field_key, field_label, field_type, required, visible, sort_order, options_json
                             FROM tenant_custom_fields
                             WHERE company_id=? AND module_name=? AND visible=1
                             ORDER BY sort_order, id""", (company_id, module_name)).fetchall()
    else:
        rows = conn.execute("""SELECT id, field_key, field_label, field_type, required, visible, sort_order, options_json
                             FROM tenant_custom_fields
                             WHERE company_id=? AND module_name=?
                             ORDER BY sort_order, id""", (company_id, module_name)).fetchall()
    result = []
    for row in rows:
        item = dict(row)
        item['required'] = bool(item.get('required'))
        item['visible'] = bool(item.get('visible'))
        if item.get('field_type') == 'dropdown':
            option_rows = get_custom_field_options(conn, company_id, item['id'], include_inactive=False)
            if option_rows:
                item['options'] = [o['option_label'] for o in option_rows if o.get('active')]
                item['option_records'] = option_rows
            else:
                item['options'] = parse_options_json(item.get('options_json'))
                item['option_records'] = []
        else:
            item['options'] = parse_options_json(item.get('options_json'))
            item['option_records'] = []
        result.append(item)
    return result


def save_custom_field_values(conn, company_id, module_name, record_id, values):
    values = values or {}
    fields = get_tenant_custom_fields(conn, company_id, module_name, visible_only=True)
    valid_keys = {f['field_key'] for f in fields}
    conn.execute('DELETE FROM custom_field_values WHERE company_id=? AND module_name=? AND record_id=?',
                 (company_id, module_name, record_id))
    for key, value in values.items():
        if key not in valid_keys:
            continue
        if isinstance(value, bool):
            clean_value = 'Yes' if value else ''
        elif value is None:
            clean_value = ''
        else:
            clean_value = str(value).strip()
        conn.execute("""INSERT INTO custom_field_values (company_id, module_name, record_id, field_key, field_value)
                        VALUES (?, ?, ?, ?, ?)""",
                     (company_id, module_name, record_id, key, clean_value))


def get_custom_field_values_for_record(conn, company_id, module_name, record_id):
    rows = conn.execute("""SELECT field_key, field_value
                         FROM custom_field_values
                         WHERE company_id=? AND module_name=? AND record_id=?""",
                       (company_id, module_name, record_id)).fetchall()
    return {r['field_key']: r['field_value'] for r in rows}





def table_exists(conn, table_name):
    return compat_table_exists(conn, table_name)

def table_columns(conn, table_name):
    return compat_table_columns(conn, table_name)

def rows_to_csv_bytes(rows):
    output = io.StringIO()
    writer = csv.writer(output)
    if rows:
        headers = list(rows[0].keys())
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row[h] for h in headers])
    else:
        writer.writerow([])
    return output.getvalue().encode('utf-8-sig')

def query_company_table(conn, table_name, company_id):
    cols = table_columns(conn, table_name)
    if not cols:
        return []
    if 'company_id' in cols:
        return [dict(r) for r in conn.execute(f"SELECT * FROM {table_name} WHERE company_id=?", (company_id,)).fetchall()]
    return []


def get_active_company_id_for_admin():
    """Return the company id a company-scoped admin action should use.

    Company admins are always restricted to their own tenant. Super admins may
    pass ?company_id=... or use the currently switched active company.
    """
    if session.get('is_superadmin'):
        requested = request.args.get('company_id') or request.form.get('company_id')
        return int(requested or session.get('company_id') or 0)
    return int(session.get('company_id') or 0)


def safe_table_name(name):
    if not name or not all(ch.isalnum() or ch == '_' for ch in name):
        raise ValueError('Invalid table name')
    return name


COMPANY_IMPORT_TABLES = {
    'clients': ['name', 'surname', 'company_name', 'registration_number', 'vat_number', 'building_number', 'street_name', 'suburb', 'postal_code', 'address', 'phone', 'email', 'client_type', 'discount_percent'],
    'employees': ['name', 'emp_number', 'id_passport', 'date_of_birth', 'job_title', 'status', 'emp_type', 'gross_salary', 'start_date', 'inactive_date', 'phone', 'email', 'address', 'emergency_contact', 'tax_number', 'paye_ref', 'bank_name', 'account_holder', 'account_number', 'branch_code', 'account_type', 'payment_reference', 'workday_hours', 'overtime_pay_treatment'],
    'services': ['name', 'client_price', 'company_cost'],
    'bookings': ['title', 'start', 'employee', 'booking_type', 'transport', 'booking_notes', 'overtime_hours', 'is_invoiced'],
    'expenses': ['date', 'category', 'supplier', 'description', 'amount', 'invoice_file'],
    'leave_records': ['employee_id', 'date_taken', 'days', 'leave_type', 'document_file']
}


def normalise_import_value(value):
    if value is None:
        return ''
    return str(value).strip()


def get_import_template_rows(import_type):
    fields = COMPANY_IMPORT_TABLES.get(import_type)
    if not fields:
        return None
    sample = {f: '' for f in fields}
    if import_type == 'clients':
        sample.update({'name': 'Example Client', 'phone': '0712345678', 'email': 'client@example.com'})
    elif import_type == 'employees':
        sample.update({'name': 'Example Employee', 'status': 'Active', 'emp_type': 'Contract >25 Hrs', 'gross_salary': '5000', 'start_date': '2026-01-01'})
    elif import_type == 'services':
        sample.update({'name': 'Example Service', 'client_price': '550', 'company_cost': '300'})
    elif import_type == 'bookings':
        sample.update({'title': 'Example Client', 'start': '2026-05-22T09:00:00', 'employee': 'Example Employee', 'booking_type': 'Example Service'})
    elif import_type == 'expenses':
        sample.update({'date': '2026-05-22', 'category': 'Supplies', 'description': 'Example expense', 'amount': '100'})
    elif import_type == 'leave_records':
        sample.update({'employee_id': '1', 'date_taken': '2026-05-22', 'days': '1', 'leave_type': 'Annual Leave'})
    return [sample]

def validate_custom_field_payload(custom_fields_config, values, require_missing=True):
    """Validate custom-field payloads for booking/custom module saves.

    Legacy bookings may have been created before tenant custom fields existed. The
    edit endpoint can pass require_missing=False to avoid blocking a small edit to
    an older booking when no custom-field payload was submitted by the legacy UI.
    """
    values = values or {}
    fields = custom_fields_config or []
    for field in fields:
        key = field.get('field_key') if isinstance(field, dict) else field['field_key']
        label = field.get('field_label') if isinstance(field, dict) else field['field_label']
        ftype = field.get('field_type') if isinstance(field, dict) else field['field_type']
        required = bool(field.get('required') if isinstance(field, dict) else field['required'])
        raw_value = values.get(key, '')
        value = '' if raw_value is None else str(raw_value).strip()

        if required and require_missing and ftype != 'checkbox' and not value:
            return False, f"{label} is required."

        if ftype == 'dropdown' and value:
            options = field.get('options') if isinstance(field, dict) else field['options']
            options = options or []
            # Allow legacy saved values even if the dropdown option was later removed.
            if options and value not in options:
                pass
    return True, ''



def get_tenant_scorecard_template(conn, company_id):
    blocks = conn.execute("""SELECT id, block_name, sort_order, active
                             FROM tenant_scorecard_blocks
                             WHERE company_id=? AND active=1
                             ORDER BY sort_order, id""", (company_id,)).fetchall()
    result = []
    for block in blocks:
        questions = conn.execute("""SELECT id, question_key, question_text, max_score, sort_order, active
                                    FROM tenant_scorecard_questions
                                    WHERE company_id=? AND block_id=? AND active=1
                                    ORDER BY sort_order, id""", (company_id, block['id'])).fetchall()
        result.append({
            "id": block['id'],
            "block_name": block['block_name'],
            "sort_order": block['sort_order'],
            "active": block['active'],
            "questions": [dict(q) for q in questions]
        })
    return result


def get_scorecard_max_score(blocks):
    total = 0
    for block in blocks or []:
        for q in block.get('questions') or []:
            try:
                total += int(q.get('max_score') or 0)
            except Exception:
                pass
    return total


@app.route('/uploads/<folder>/<filename>')
def uploaded_file(folder, filename):
    return send_from_directory(os.path.join(app.config['UPLOAD_FOLDER'], folder), filename)


def _safe_logo_url(filename):
    if not filename:
        return ''
    return url_for('uploaded_file', folder='logos', filename=os.path.basename(str(filename)))


def _clear_pdf_logo_cache():
    try:
        _PDF_IMAGE_CACHE.clear()
    except Exception:
        pass

# --- Google Calendar Configuration ---
SCOPES = ['https://www.googleapis.com/auth/calendar.events']
GOOGLE_CALENDAR_TIMEZONE = 'Africa/Johannesburg'


def _company_calendar_settings(company_id=None):
    cid = company_id if company_id is not None else session.get('company_id')
    settings = {}
    if not cid:
        return settings
    conn = get_db_connection()
    try:
        rows = conn.execute('SELECT key, value FROM settings WHERE company_id=?', (cid,)).fetchall()
        settings = {r['key']: r['value'] for r in rows}
    finally:
        conn.close()
    return settings


def _set_company_setting(conn, company_id, key, value):
    exists = conn.execute('SELECT 1 FROM settings WHERE key=? AND company_id=?', (key, company_id)).fetchone()
    if exists:
        conn.execute('UPDATE settings SET value=? WHERE key=? AND company_id=?', (value, key, company_id))
    else:
        conn.execute('INSERT INTO settings (company_id, key, value) VALUES (?, ?, ?)', (company_id, key, value))


def _google_calendar_storage_root():
    upload_folder = app.config.get('UPLOAD_FOLDER') or os.environ.get('UPLOAD_FOLDER') or '.'
    return os.path.join(upload_folder, 'google_calendar')


def _company_google_calendar_dir(company_id=None):
    cid = company_id if company_id is not None else session.get('company_id')
    if not cid:
        raise RuntimeError('Company context is required for Google Calendar settings.')
    return os.path.join(_google_calendar_storage_root(), f'company_{int(cid)}')


def _company_google_credentials_path(company_id=None):
    return os.path.join(_company_google_calendar_dir(company_id), 'credentials.json')


def _company_google_token_path(company_id=None):
    return os.path.join(_company_google_calendar_dir(company_id), 'token.json')


def _legacy_google_token_path_candidates():
    candidates = []
    env_path = os.environ.get('GOOGLE_TOKEN_FILE', '').strip()
    if env_path:
        candidates.append(env_path)
    upload_folder = app.config.get('UPLOAD_FOLDER') or ''
    if upload_folder:
        candidates.append(os.path.join(upload_folder, 'google_token.json'))
        candidates.append(os.path.join(upload_folder, 'token.json'))
    candidates.append('google_token.json')
    candidates.append('token.json')
    seen = set()
    return [p for p in candidates if p and not (p in seen or seen.add(p))]


def _legacy_google_credentials_file_candidates():
    candidates = []
    env_path = os.environ.get('GOOGLE_CREDENTIALS_FILE', '').strip()
    if env_path:
        candidates.append(env_path)
    upload_folder = app.config.get('UPLOAD_FOLDER') or ''
    if upload_folder:
        candidates.append(os.path.join(upload_folder, 'google_credentials.json'))
        candidates.append(os.path.join(upload_folder, 'credentials.json'))
    candidates.append('google_credentials.json')
    candidates.append('credentials.json')
    seen = set()
    return [p for p in candidates if p and not (p in seen or seen.add(p))]


def _existing_google_token_path(company_id=None):
    # Each company gets its own Calendar token. Legacy paths are only fallback
    # when no company context exists, to avoid one company accidentally using
    # another company's connected Google account.
    if company_id is not None or session.get('company_id'):
        path = _company_google_token_path(company_id)
        return path if os.path.exists(path) else None
    for path in _legacy_google_token_path_candidates():
        if path and os.path.exists(path):
            return path
    return None


def _preferred_google_token_path(company_id=None):
    if company_id is not None or session.get('company_id'):
        return _company_google_token_path(company_id)
    env_path = os.environ.get('GOOGLE_TOKEN_FILE', '').strip()
    if env_path:
        return env_path
    upload_folder = app.config.get('UPLOAD_FOLDER') or ''
    if upload_folder:
        return os.path.join(upload_folder, 'google_token.json')
    return 'token.json'


def _existing_google_credentials_file(company_id=None):
    # Prefer the company's uploaded OAuth client JSON.
    if company_id is not None or session.get('company_id'):
        path = _company_google_credentials_path(company_id)
        if os.path.exists(path):
            return path
    # Fallback for older deployments or environment-based OAuth client config.
    for path in _legacy_google_credentials_file_candidates():
        if path and os.path.exists(path):
            return path
    return None


def _google_client_config_from_env():
    client_id = os.environ.get('GOOGLE_CLIENT_ID', '').strip()
    client_secret = os.environ.get('GOOGLE_CLIENT_SECRET', '').strip()
    if not client_id or not client_secret:
        return None
    return {
        'web': {
            'client_id': client_id,
            'client_secret': client_secret,
            'auth_uri': 'https://accounts.google.com/o/oauth2/auth',
            'token_uri': 'https://oauth2.googleapis.com/token',
            'redirect_uris': []
        }
    }


def _google_redirect_uri():
    uri = url_for('google_calendar_oauth_callback', _external=True)
    if os.environ.get('RENDER') and uri.startswith('http://'):
        uri = 'https://' + uri[len('http://'):]
    return uri


def _build_google_oauth_flow(company_id=None, redirect_uri=None):
    redirect_uri = redirect_uri or _google_redirect_uri()
    credentials_file = _existing_google_credentials_file(company_id)
    if credentials_file:
        return Flow.from_client_secrets_file(credentials_file, scopes=SCOPES, redirect_uri=redirect_uri)
    client_config = _google_client_config_from_env()
    if client_config:
        client_config['web']['redirect_uris'] = [redirect_uri]
        return Flow.from_client_config(client_config, scopes=SCOPES, redirect_uri=redirect_uri)
    raise RuntimeError('Google Calendar OAuth client JSON is not configured for this company. Upload the Google OAuth Client JSON in Company Email & Calendar Settings, save, then click Connect Google Calendar.')


def _validate_google_credentials_json(raw_json):
    try:
        data = json.loads(raw_json)
    except Exception as exc:
        raise RuntimeError(f'Google OAuth Client JSON is not valid JSON: {exc}')
    if not isinstance(data, dict) or not (data.get('web') or data.get('installed')):
        raise RuntimeError('Google OAuth Client JSON must be the OAuth client JSON from Google Cloud and contain a web or installed client section.')
    client_section = data.get('web') or data.get('installed') or {}
    if not client_section.get('client_id') or not client_section.get('client_secret'):
        raise RuntimeError('Google OAuth Client JSON must include client_id and client_secret.')
    return data


def _save_company_google_credentials(company_id, raw_json):
    data = _validate_google_credentials_json(raw_json)
    company_dir = _company_google_calendar_dir(company_id)
    os.makedirs(company_dir, exist_ok=True)
    credentials_path = _company_google_credentials_path(company_id)
    with open(credentials_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)
    token_path = _company_google_token_path(company_id)
    if os.path.exists(token_path):
        os.remove(token_path)
    return credentials_path


def _read_uploaded_google_credentials_json(upload):
    if not upload or not getattr(upload, 'filename', ''):
        return ''
    filename = upload.filename or ''
    if not filename.lower().endswith('.json'):
        raise RuntimeError('Please upload the Google OAuth Client JSON file downloaded from Google Cloud. The file must end with .json.')
    raw_bytes = upload.read()
    if not raw_bytes:
        raise RuntimeError('The uploaded Google OAuth Client JSON file is empty.')
    if len(raw_bytes) > 1024 * 1024:
        raise RuntimeError('The uploaded Google OAuth Client JSON file is too large. Please upload the original JSON credentials file from Google Cloud.')
    try:
        return raw_bytes.decode('utf-8-sig')
    except UnicodeDecodeError:
        raise RuntimeError('The uploaded Google OAuth Client JSON file could not be read as text.')


def _clear_company_google_token(company_id, conn=None):
    token_path = _company_google_token_path(company_id)
    if os.path.exists(token_path):
        os.remove(token_path)
    if conn is not None:
        conn.execute("DELETE FROM settings WHERE company_id=? AND key IN ('gcal_token_id','gcal_token_saved_at')", (company_id,))


def _create_google_token_id():
    return secrets.token_urlsafe(12)


def get_google_service(company_id=None):
    cid = company_id if company_id is not None else session.get('company_id')
    token_path = _existing_google_token_path(cid)
    save_path = token_path or _preferred_google_token_path(cid)

    creds = None
    if token_path:
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            token_dir = os.path.dirname(save_path)
            if token_dir:
                os.makedirs(token_dir, exist_ok=True)
            with open(save_path, 'w', encoding='utf-8') as token:
                token.write(creds.to_json())
        else:
            if os.environ.get('RENDER'):
                raise RuntimeError('Google Calendar is not connected for this company. Open Company Email & Calendar Settings, upload/save the Google OAuth Client JSON, then click Connect Google Calendar.')
            flow = InstalledAppFlow.from_client_secrets_file(_existing_google_credentials_file(cid) or 'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
            token_dir = os.path.dirname(save_path)
            if token_dir:
                os.makedirs(token_dir, exist_ok=True)
            with open(save_path, 'w', encoding='utf-8') as token:
                token.write(creds.to_json())
    return build('calendar', 'v3', credentials=creds, cache_discovery=False)


def get_target_calendar(company_id=None):
    cal_id = get_setting('gcal_calendar_id', company_id)
    if cal_id and str(cal_id).strip():
        return str(cal_id).strip()
    return 'primary'


def _format_google_calendar_error(exc):
    message = str(exc)
    lower = message.lower()
    if 'oauth client json' in lower or 'not connected' in lower or 'not configured' in lower or 'credentials.json' in lower or 'company context' in lower:
        return message
    if 'redirect_uri_mismatch' in lower:
        return 'Google Calendar OAuth redirect URI mismatch. Add this redirect URI in Google Cloud for the OAuth Client: ' + _google_redirect_uri()
    if 'invalid_grant' in lower or 'token has been expired or revoked' in lower:
        return 'Google Calendar OAuth token has expired or was revoked. Open Company Email & Calendar Settings and reconnect Google Calendar.'
    if 'not found' in lower or '404' in lower:
        return 'Google Calendar not found. Check the Target Google Calendar ID, or leave it blank to use the connected Google account primary calendar.'
    if 'insufficient' in lower or 'forbidden' in lower or '403' in lower or 'permission denied' in lower:
        return 'Google Calendar permission denied. The connected Google account must have permission to create and update events on the selected calendar.'
    return message


def test_google_calendar_sync(company_id=None):
    cid = company_id if company_id is not None else session.get('company_id')
    target_calendar = get_target_calendar(cid)
    service = get_google_service(cid)
    service.events().list(calendarId=target_calendar, maxResults=1).execute()
    target_label = 'connected Google account primary calendar' if target_calendar == 'primary' else target_calendar
    token_id = get_setting('gcal_token_id', cid)
    token_suffix = f' Token ID: {token_id}.' if token_id else ''
    return {
        'status': 'success',
        'message': f'Google Calendar connection successful. Bookings will sync to {target_label}.{token_suffix}'
    }


def create_google_event(client_name, date_str, time_str, employees, booking_type, transport, company_name, company_id=None):
    service = get_google_service(company_id)
    start_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
    end_dt = start_dt + timedelta(hours=2)
    event = {
        'summary': f"[{company_name}] {client_name} ({employees})",
        'description': f"{company_name} Assignment\nService(s): {booking_type}\nStaff: {employees}\nTransport: {transport}",
        'start': {'dateTime': start_dt.isoformat() + "+02:00", 'timeZone': GOOGLE_CALENDAR_TIMEZONE},
        'end': {'dateTime': end_dt.isoformat() + "+02:00", 'timeZone': GOOGLE_CALENDAR_TIMEZONE}
    }
    created_event = service.events().insert(calendarId=get_target_calendar(company_id), body=event).execute()
    return created_event.get('id')


def update_google_event(event_id, client_name, date_str, time_str, employees, booking_type, transport, company_name, company_id=None):
    service = get_google_service(company_id)
    start_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
    end_dt = start_dt + timedelta(hours=2)
    target_cal = get_target_calendar(company_id)
    event = service.events().get(calendarId=target_cal, eventId=event_id).execute()
    event['summary'] = f"[{company_name}] {client_name} ({employees})"
    event['description'] = f"{company_name} Assignment\nService(s): {booking_type}\nStaff: {employees}\nTransport: {transport}"
    event['start']['dateTime'] = start_dt.isoformat() + "+02:00"
    event['end']['dateTime'] = end_dt.isoformat() + "+02:00"
    service.events().update(calendarId=target_cal, eventId=event_id, body=event).execute()
    return event_id


def delete_google_event(event_id, company_id=None):
    service = get_google_service(company_id)
    service.events().delete(calendarId=get_target_calendar(company_id), eventId=event_id).execute()

def _booking_google_payload_from_row(row):
    booking = dict(row)
    start_value = str(booking.get('start') or '')
    date_str = start_value[:10]
    time_str = '08:00'
    if 'T' in start_value and len(start_value.split('T', 1)[1]) >= 5:
        time_str = start_value.split('T', 1)[1][:5]
    elif len(start_value) >= 16:
        time_str = start_value[11:16]
    if not date_str:
        raise RuntimeError('Booking has no date saved.')
    return {
        'client_name': booking.get('title') or 'Booking',
        'date_str': date_str,
        'time_str': time_str,
        'employees': booking.get('employee') or '',
        'booking_type': booking.get('booking_type') or '',
        'transport': booking.get('transport') or '',
    }


def sync_existing_bookings_to_google_calendar(company_id, start_date, end_date):
    if not company_id:
        raise RuntimeError('Company context is required for Google Calendar sync.')
    try:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
    except Exception:
        raise RuntimeError('Please enter a valid From Date and To Date in YYYY-MM-DD format.')
    if end_dt < start_dt:
        raise RuntimeError('To Date cannot be before From Date.')
    if (end_dt - start_dt).days > 370:
        raise RuntimeError('Please sync a date range of 12 months or less at a time.')

    company_name = session.get('company_name') or 'Easy Admin'
    conn = get_db_connection()
    created = 0
    updated = 0
    failed = 0
    skipped = 0
    errors = []
    try:
        rows = conn.execute("""
            SELECT id, title, start, employee, booking_type, transport, google_event_id
            FROM bookings
            WHERE company_id=?
              AND substr(COALESCE(start,''), 1, 10) BETWEEN ? AND ?
            ORDER BY start ASC, id ASC
        """, (company_id, start_date, end_date)).fetchall()

        for row in rows:
            booking = dict(row)
            try:
                payload = _booking_google_payload_from_row(booking)
                existing_event_id = booking.get('google_event_id')
                if existing_event_id:
                    try:
                        new_event_id = update_google_event(
                            existing_event_id,
                            payload['client_name'],
                            payload['date_str'],
                            payload['time_str'],
                            payload['employees'],
                            payload['booking_type'],
                            payload['transport'],
                            company_name,
                            company_id,
                        )
                        updated += 1
                    except HttpError as http_exc:
                        if getattr(getattr(http_exc, 'resp', None), 'status', None) == 404:
                            new_event_id = create_google_event(
                                payload['client_name'],
                                payload['date_str'],
                                payload['time_str'],
                                payload['employees'],
                                payload['booking_type'],
                                payload['transport'],
                                company_name,
                                company_id,
                            )
                            created += 1
                        else:
                            raise
                else:
                    new_event_id = create_google_event(
                        payload['client_name'],
                        payload['date_str'],
                        payload['time_str'],
                        payload['employees'],
                        payload['booking_type'],
                        payload['transport'],
                        company_name,
                        company_id,
                    )
                    created += 1

                if new_event_id and new_event_id != existing_event_id:
                    conn.execute('UPDATE bookings SET google_event_id=? WHERE id=? AND company_id=?', (new_event_id, booking['id'], company_id))
                    conn.commit()
                elif not new_event_id:
                    skipped += 1
            except Exception as exc:
                failed += 1
                if len(errors) < 5:
                    errors.append(f"Booking #{booking.get('id')}: {_format_google_calendar_error(exc)}")

    finally:
        conn.close()

    total_processed = created + updated + skipped + failed
    return {
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'failed': failed,
        'total': total_processed,
        'errors': errors,
    }

# --- Database Setup & Multi-Tenant Migration ---
def get_db_connection():
    # Uses Supabase/PostgreSQL when DATABASE_URL is configured; otherwise SQLite.
    return connect_database(get_database_path())

def init_db():
    os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'cvs'), exist_ok=True)
    os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'ids'), exist_ok=True)
    os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'leave'), exist_ok=True)
    os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'logos'), exist_ok=True)
    os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'contracts'), exist_ok=True)
    os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'expenses'), exist_ok=True)
    os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'accounting'), exist_ok=True)

    conn = get_db_connection()
    
    conn.execute('''CREATE TABLE IF NOT EXISTS companies (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE)''')
    try: conn.execute('ALTER TABLE companies ADD COLUMN logo_file TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN transport_policy TEXT DEFAULT "standard"')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN transport_amount_per_lift REAL DEFAULT 25')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN can_booking INTEGER DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN can_finance INTEGER DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN can_payroll INTEGER DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN can_invoicing INTEGER DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN can_accounting INTEGER DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN google_calendar_sync INTEGER DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN address TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN registration_number TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN vat_number TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN industry_template TEXT DEFAULT "Cleaning"')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE companies ADD COLUMN setup_complete INTEGER DEFAULT 0')
    except sqlite3.OperationalError: pass

    conn.execute("""CREATE TABLE IF NOT EXISTS industry_template_uploads (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        template_name TEXT NOT NULL UNIQUE,
        template_json TEXT NOT NULL,
        active INTEGER DEFAULT 1,
        created_by TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )""")

    conn.execute("""CREATE TABLE IF NOT EXISTS tenant_labels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        module_name TEXT DEFAULT 'global',
        label_key TEXT NOT NULL,
        label_value TEXT NOT NULL,
        UNIQUE(company_id, label_key)
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS tenant_custom_fields (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        module_name TEXT NOT NULL,
        field_key TEXT NOT NULL,
        field_label TEXT NOT NULL,
        field_type TEXT DEFAULT 'text',
        required INTEGER DEFAULT 0,
        visible INTEGER DEFAULT 1,
        sort_order INTEGER DEFAULT 0,
        options_json TEXT DEFAULT '[]',
        industry_template TEXT,
        UNIQUE(company_id, module_name, field_key)
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS custom_field_values (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        module_name TEXT NOT NULL,
        record_id INTEGER NOT NULL,
        field_key TEXT NOT NULL,
        field_value TEXT,
        UNIQUE(company_id, module_name, record_id, field_key)
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS tenant_custom_field_options (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        field_id INTEGER NOT NULL,
        option_label TEXT NOT NULL,
        option_value TEXT NOT NULL,
        sort_order INTEGER DEFAULT 0,
        active INTEGER DEFAULT 1,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(company_id, field_id, option_value)
    )""")
    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_custom_field_options_company_field ON tenant_custom_field_options(company_id, field_id, sort_order)')
    except sqlite3.OperationalError: pass
    conn.execute("""CREATE TABLE IF NOT EXISTS finance_categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        category_name TEXT NOT NULL,
        category_type TEXT DEFAULT 'expense',
        active INTEGER DEFAULT 1,
        industry_template TEXT,
        sort_order INTEGER DEFAULT 0,
        UNIQUE(company_id, category_name)
    )""")

    conn.execute("""CREATE TABLE IF NOT EXISTS tenant_scorecard_blocks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        block_name TEXT NOT NULL,
        sort_order INTEGER DEFAULT 0,
        active INTEGER DEFAULT 1,
        industry_template TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS tenant_scorecard_questions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        block_id INTEGER NOT NULL,
        question_key TEXT NOT NULL,
        question_text TEXT NOT NULL,
        max_score INTEGER DEFAULT 10,
        sort_order INTEGER DEFAULT 0,
        active INTEGER DEFAULT 1
    )""")
    try: conn.execute('ALTER TABLE interviews ADD COLUMN interview_notes TEXT')
    except sqlite3.OperationalError: pass
    conn.execute('CREATE INDEX IF NOT EXISTS idx_scorecard_blocks_company ON tenant_scorecard_blocks(company_id, sort_order)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_scorecard_questions_company_block ON tenant_scorecard_questions(company_id, block_id, sort_order)')

    if conn.execute('SELECT COUNT(*) FROM companies').fetchone()[0] == 0:
        conn.execute('INSERT INTO companies (name, can_booking, can_finance, can_payroll, can_invoicing, can_accounting, google_calendar_sync) VALUES ("Marvellous Maids", 1, 1, 1, 1, 1, 0)')
    default_company_id = conn.execute('SELECT id FROM companies LIMIT 1').fetchone()[0]
    for comp in conn.execute('SELECT id, industry_template FROM companies').fetchall():
        ensure_tenant_template(conn, comp['id'], dict(comp).get('industry_template') or 'Cleaning', force_reset=False)

    conn.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, username TEXT UNIQUE, password_hash TEXT,
        can_booking INTEGER DEFAULT 0, can_finance INTEGER DEFAULT 0, can_payroll INTEGER DEFAULT 0, is_superadmin INTEGER DEFAULT 0
    )''')
    
    try: conn.execute('ALTER TABLE users ADD COLUMN company_id INTEGER')
    except sqlite3.OperationalError: pass 
    try: conn.execute('ALTER TABLE users ADD COLUMN is_company_admin INTEGER DEFAULT 0')
    except sqlite3.OperationalError: pass 
    try: conn.execute('ALTER TABLE users ADD COLUMN can_invoicing INTEGER DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE users ADD COLUMN can_accounting INTEGER DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE users ADD COLUMN employee_id INTEGER')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE users ADD COLUMN is_staff INTEGER DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE users ADD COLUMN email TEXT')
    except sqlite3.OperationalError: pass

    conn.execute('''CREATE TABLE IF NOT EXISTS system_email_settings (
        setting_key TEXT PRIMARY KEY,
        setting_value TEXT
    )''')
    
    conn.execute('UPDATE users SET company_id = ? WHERE company_id IS NULL', (default_company_id,))

    if conn.execute('SELECT COUNT(*) FROM users WHERE is_superadmin=1').fetchone()[0] == 0:
        default_hash = generate_password_hash('Fawaaz!23')
        conn.execute('''INSERT INTO users (username, password_hash, company_id, can_booking, can_finance, can_payroll, can_invoicing, can_accounting, is_superadmin, is_company_admin) 
                        VALUES (?, ?, ?, 1, 1, 1, 1, 1, 1, 1)''', ('Marvellous', default_hash, default_company_id))

    conn.execute('''CREATE TABLE IF NOT EXISTS invoices (
        id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, client_name TEXT, 
        date TEXT, due_date TEXT, subtotal REAL, vat_amount REAL, total REAL, 
        status TEXT DEFAULT 'Unpaid'
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS invoice_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT, invoice_id INTEGER, 
        booking_id INTEGER, service_date TEXT, description TEXT, amount REAL
    )''')
    try: conn.execute('ALTER TABLE invoice_items ADD COLUMN quantity REAL DEFAULT 1')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE invoice_items ADD COLUMN unit_price REAL')
    except sqlite3.OperationalError: pass
    try: conn.execute('UPDATE invoice_items SET quantity=1 WHERE quantity IS NULL OR quantity<=0')
    except sqlite3.OperationalError: pass
    try: conn.execute('UPDATE invoice_items SET unit_price=amount WHERE unit_price IS NULL')
    except sqlite3.OperationalError: pass
    conn.execute('''CREATE TABLE IF NOT EXISTS invoice_payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        invoice_id INTEGER,
        payment_date TEXT,
        amount REAL DEFAULT 0,
        payment_method TEXT,
        reference TEXT,
        notes TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS invoice_credit_notes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        invoice_id INTEGER,
        credit_date TEXT,
        amount REAL DEFAULT 0,
        reason TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    # Accounting integration flags for invoices and credit notes.
    # These columns prevent duplicate postings and link billing documents to their GL journals.
    for sql in [
        "ALTER TABLE invoices ADD COLUMN accounting_status TEXT DEFAULT 'not_posted'",
        'ALTER TABLE invoices ADD COLUMN accounting_journal_id INTEGER',
        'ALTER TABLE invoices ADD COLUMN accounting_posted_at TEXT',
        'ALTER TABLE invoices ADD COLUMN accounting_posted_by TEXT',
        "ALTER TABLE invoice_credit_notes ADD COLUMN accounting_status TEXT DEFAULT 'not_posted'",
        'ALTER TABLE invoice_credit_notes ADD COLUMN accounting_journal_id INTEGER',
        'ALTER TABLE invoice_credit_notes ADD COLUMN accounting_posted_at TEXT',
        'ALTER TABLE invoice_credit_notes ADD COLUMN accounting_posted_by TEXT'
    ]:
        try: conn.execute(sql)
        except sqlite3.OperationalError: pass
    conn.execute('''CREATE TABLE IF NOT EXISTS projects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        client_id INTEGER,
        project_name TEXT NOT NULL,
        project_code TEXT,
        description TEXT,
        site_address TEXT,
        start_date TEXT,
        estimated_end_date TEXT,
        actual_end_date TEXT,
        fixed_price REAL DEFAULT 0,
        status TEXT DEFAULT 'Quoted',
        notes TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS project_employees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        project_id INTEGER NOT NULL,
        employee_id INTEGER NOT NULL,
        role TEXT,
        UNIQUE(company_id, project_id, employee_id)
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS project_costs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        project_id INTEGER NOT NULL,
        cost_date TEXT,
        cost_type TEXT DEFAULT 'Other',
        description TEXT,
        supplier TEXT,
        amount REAL DEFAULT 0,
        notes TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS attachments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        linked_type TEXT NOT NULL,
        linked_id INTEGER NOT NULL,
        original_filename TEXT NOT NULL,
        stored_filename TEXT NOT NULL,
        file_path TEXT NOT NULL,
        file_size INTEGER DEFAULT 0,
        mime_type TEXT,
        uploaded_by TEXT,
        uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS quotes (
        id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, client_name TEXT, 
        date TEXT, valid_until TEXT, subtotal REAL, vat_amount REAL, total REAL, 
        status TEXT DEFAULT 'Pending'
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS quote_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT, quote_id INTEGER, 
        service_date TEXT, description TEXT, amount REAL
    )''')
    try: conn.execute('ALTER TABLE quote_items ADD COLUMN quantity REAL DEFAULT 1')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE quote_items ADD COLUMN unit_price REAL')
    except sqlite3.OperationalError: pass
    try: conn.execute('UPDATE quote_items SET quantity=1 WHERE quantity IS NULL OR quantity<=0')
    except sqlite3.OperationalError: pass
    try: conn.execute('UPDATE quote_items SET unit_price=amount WHERE unit_price IS NULL')
    except sqlite3.OperationalError: pass

    conn.execute('''CREATE TABLE IF NOT EXISTS interviews (
        id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, name TEXT, email TEXT, phone TEXT, id_passport TEXT,
        address TEXT, interview_datetime TEXT, cv_file TEXT, id_file TEXT, scorecard_json TEXT, total_score INTEGER DEFAULT 0,
        final_decision TEXT DEFAULT 'Pending', interview_notes TEXT
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS audit_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        username TEXT,
        app_name TEXT,
        action TEXT,
        details TEXT,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS staff_leave_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        employee_id INTEGER NOT NULL,
        leave_type TEXT DEFAULT 'Annual Leave',
        start_date TEXT NOT NULL,
        end_date TEXT NOT NULL,
        days REAL DEFAULT 1,
        reason TEXT,
        status TEXT DEFAULT 'Pending',
        attachment_file TEXT,
        requested_at TEXT DEFAULT CURRENT_TIMESTAMP,
        reviewed_by TEXT,
        reviewed_at TEXT,
        admin_note TEXT,
        leave_record_id INTEGER
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS tax_brackets (
        id INTEGER PRIMARY KEY AUTOINCREMENT, tax_year INTEGER,
        min_income REAL, max_income REAL, base_tax REAL, rate REAL
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS tax_rebates (
        id INTEGER PRIMARY KEY AUTOINCREMENT, tax_year INTEGER, primary_rebate REAL,
        secondary_rebate REAL DEFAULT 0, tertiary_rebate REAL DEFAULT 0,
        threshold_under_65 REAL DEFAULT 0, threshold_65_to_74 REAL DEFAULT 0, threshold_75_plus REAL DEFAULT 0
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS public_holidays (
        id INTEGER PRIMARY KEY AUTOINCREMENT, year INTEGER, date_str TEXT, name TEXT
    )''')

    if conn.execute('SELECT COUNT(*) FROM tax_brackets').fetchone()[0] == 0:
        brackets = [
            (2026, 0, 245100, 0, 0.18),
            (2026, 245101, 383100, 44118, 0.26),
            (2026, 383101, 530200, 79998, 0.31),
            (2026, 530201, 695800, 125599, 0.36),
            (2026, 695801, 887000, 185215, 0.39),
            (2026, 887001, 1878600, 259783, 0.41),
            (2026, 1878601, 999999999, 666339, 0.45)
        ]
        for b in brackets:
            conn.execute('INSERT INTO tax_brackets (tax_year, min_income, max_income, base_tax, rate) VALUES (?, ?, ?, ?, ?)', b)
        conn.execute('''INSERT INTO tax_rebates (tax_year, primary_rebate, secondary_rebate, tertiary_rebate, threshold_under_65, threshold_65_to_74, threshold_75_plus)
                        VALUES (?, ?, ?, ?, ?, ?, ?)''', (2026, 17820, 9765, 3249, 99000, 153250, 171300))

    tables = {
        'bookings': 'CREATE TABLE IF NOT EXISTS bookings (id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT, start TEXT, employee TEXT)',
        'clients': 'CREATE TABLE IF NOT EXISTS clients (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT)',
        'employees': 'CREATE TABLE IF NOT EXISTS employees (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT)',
        'services': 'CREATE TABLE IF NOT EXISTS services (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, client_price REAL, company_cost REAL)',
        'expenses': 'CREATE TABLE IF NOT EXISTS expenses (id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT, category TEXT, supplier TEXT, description TEXT, amount REAL)',
        'payslips': 'CREATE TABLE IF NOT EXISTS payslips (id INTEGER PRIMARY KEY AUTOINCREMENT, employee_id INTEGER, date TEXT, gross_salary REAL, uif REAL, paye REAL, net_salary REAL)',
        'leave_records': 'CREATE TABLE IF NOT EXISTS leave_records (id INTEGER PRIMARY KEY AUTOINCREMENT, employee_id INTEGER, date_taken TEXT, days REAL)'
    }

    for table, create_sql in tables.items():
        conn.execute(create_sql)
        try:
            conn.execute(f'ALTER TABLE {table} ADD COLUMN company_id INTEGER')
            conn.execute(f'UPDATE {table} SET company_id = ? WHERE company_id IS NULL', (default_company_id,))
        except sqlite3.OperationalError: pass


    # --- Accounting Module (IFRS for SMEs) ---
    conn.execute("""CREATE TABLE IF NOT EXISTS accounting_accounts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        account_code TEXT NOT NULL,
        account_name TEXT NOT NULL,
        account_type TEXT NOT NULL,
        report_section TEXT NOT NULL,
        normal_balance TEXT NOT NULL DEFAULT 'debit',
        cash_flow_category TEXT DEFAULT 'operating',
        is_cash_equivalent INTEGER DEFAULT 0,
        active INTEGER DEFAULT 1,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(company_id, account_code)
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS accounting_journals (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        journal_date TEXT NOT NULL,
        reference TEXT,
        description TEXT,
        source_module TEXT DEFAULT 'manual',
        status TEXT DEFAULT 'draft',
        created_by TEXT,
        posted_by TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        posted_at TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS accounting_journal_lines (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        journal_id INTEGER NOT NULL,
        line_no INTEGER DEFAULT 0,
        account_id INTEGER NOT NULL,
        description TEXT,
        debit REAL DEFAULT 0,
        credit REAL DEFAULT 0,
        cash_flow_section TEXT DEFAULT 'operating'
    )""")
    for sql in [
        'ALTER TABLE accounting_journals ADD COLUMN source_record_type TEXT',
        'ALTER TABLE accounting_journals ADD COLUMN source_record_id INTEGER'
    ]:
        try: conn.execute(sql)
        except sqlite3.OperationalError: pass
    for sql in [
        'ALTER TABLE accounting_journal_lines ADD COLUMN vat_amount REAL DEFAULT 0',
        'ALTER TABLE accounting_journal_lines ADD COLUMN gross_amount REAL DEFAULT 0',
        'ALTER TABLE accounting_journal_lines ADD COLUMN net_amount REAL DEFAULT 0',
        'ALTER TABLE accounting_journal_lines ADD COLUMN vat_type TEXT'
    ]:
        try: conn.execute(sql)
        except Exception: pass
    conn.execute("""CREATE TABLE IF NOT EXISTS accounting_settings (
        company_id INTEGER PRIMARY KEY,
        reporting_framework TEXT DEFAULT 'IFRS for SMEs',
        financial_year_end_month INTEGER DEFAULT 2,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    for sql in [
        'ALTER TABLE accounting_settings ADD COLUMN receivables_account_id INTEGER',
        'ALTER TABLE accounting_settings ADD COLUMN sales_revenue_account_id INTEGER',
        'ALTER TABLE accounting_settings ADD COLUMN vat_output_account_id INTEGER',
        'ALTER TABLE accounting_settings ADD COLUMN vat_control_account_id INTEGER',
        'ALTER TABLE accounting_settings ADD COLUMN credit_notes_account_id INTEGER',
        'ALTER TABLE accounting_settings ADD COLUMN discount_account_id INTEGER',
        'ALTER TABLE accounting_settings ADD COLUMN rounding_account_id INTEGER'
    ]:
        try: conn.execute(sql)
        except Exception: pass
    conn.execute("""CREATE TABLE IF NOT EXISTS accounting_cashbook_batches (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        bank_account_id INTEGER NOT NULL,
        original_filename TEXT,
        status TEXT DEFAULT 'draft',
        imported_by TEXT,
        posted_by TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        posted_at TEXT,
        line_count INTEGER DEFAULT 0,
        debit_total REAL DEFAULT 0,
        credit_total REAL DEFAULT 0,
        bank_format TEXT DEFAULT 'auto',
        column_mapping_json TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS accounting_cashbook_lines (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        batch_id INTEGER NOT NULL,
        line_no INTEGER DEFAULT 0,
        transaction_date TEXT,
        description TEXT,
        debit REAL DEFAULT 0,
        credit REAL DEFAULT 0,
        balance REAL DEFAULT 0,
        allocated_account_id INTEGER,
        vat_amount REAL DEFAULT 0,
        notes TEXT,
        cash_flow_section TEXT DEFAULT 'operating',
        status TEXT DEFAULT 'draft',
        linked_journal_id INTEGER
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS accounting_transaction_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        linked_type TEXT NOT NULL,
        linked_id INTEGER NOT NULL,
        original_filename TEXT NOT NULL,
        stored_filename TEXT NOT NULL,
        file_path TEXT NOT NULL,
        file_size INTEGER DEFAULT 0,
        mime_type TEXT,
        uploaded_by TEXT,
        uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS accounting_bank_reconciliations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        bank_account_id INTEGER NOT NULL,
        start_date TEXT,
        recon_date TEXT NOT NULL,
        statement_balance REAL DEFAULT 0,
        gl_balance REAL DEFAULT 0,
        outstanding_deposits REAL DEFAULT 0,
        outstanding_payments REAL DEFAULT 0,
        adjusted_statement_balance REAL DEFAULT 0,
        difference_amount REAL DEFAULT 0,
        notes TEXT,
        created_by TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS accounting_bank_reconciliation_lines (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        reconciliation_id INTEGER NOT NULL,
        journal_line_id INTEGER NOT NULL,
        reconciled INTEGER DEFAULT 1,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(company_id, reconciliation_id, journal_line_id)
    )""")
    try: conn.execute("ALTER TABLE accounting_cashbook_batches ADD COLUMN bank_format TEXT DEFAULT 'auto'")
    except Exception: pass
    try: conn.execute("ALTER TABLE accounting_cashbook_batches ADD COLUMN column_mapping_json TEXT")
    except Exception: pass
    try: conn.execute("ALTER TABLE accounting_cashbook_lines ADD COLUMN vat_amount REAL DEFAULT 0")
    except Exception: pass

    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_accounting_accounts_company ON accounting_accounts(company_id, account_code)')
    except sqlite3.OperationalError: pass
    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_accounting_journals_company_date ON accounting_journals(company_id, journal_date, status)')
    except sqlite3.OperationalError: pass
    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_accounting_lines_company_journal ON accounting_journal_lines(company_id, journal_id, account_id)')
    except sqlite3.OperationalError: pass
    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_accounting_cashbook_batches_company ON accounting_cashbook_batches(company_id, status, created_at)')
    except sqlite3.OperationalError: pass
    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_accounting_cashbook_lines_batch ON accounting_cashbook_lines(company_id, batch_id, status)')
    except sqlite3.OperationalError: pass
    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_accounting_transaction_files_link ON accounting_transaction_files(company_id, linked_type, linked_id)')
    except sqlite3.OperationalError: pass
    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_accounting_bank_reconciliations_company ON accounting_bank_reconciliations(company_id, bank_account_id, recon_date)')
    except sqlite3.OperationalError: pass
    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_accounting_bank_recon_lines ON accounting_bank_reconciliation_lines(company_id, reconciliation_id, journal_line_id)')
    except sqlite3.OperationalError: pass

    if not table_exists(conn, 'settings'):
        conn.execute('CREATE TABLE IF NOT EXISTS settings (company_id INTEGER, key TEXT, value TEXT)')
    elif 'company_id' not in table_columns(conn, 'settings'):
        conn.execute('CREATE TABLE IF NOT EXISTS settings_new (company_id INTEGER, key TEXT, value TEXT)')
        try:
            old_data = conn.execute('SELECT * FROM settings').fetchall()
            for row in old_data:
                conn.execute('INSERT INTO settings_new (company_id, key, value) VALUES (?, ?, ?)', (default_company_id, row['key'], row['value']))
            conn.execute('DROP TABLE settings')
            conn.execute('ALTER TABLE settings_new RENAME TO settings')
        except Exception:
            pass

    loose_cols = {
        'bookings': [('client_id', 'INTEGER'), ('google_event_id', 'TEXT'), ('booking_type', 'TEXT DEFAULT "Standard Home Clean"'), ('transport', 'TEXT DEFAULT ""'), ('booking_notes', 'TEXT DEFAULT ""'), ('overtime_hours', 'REAL DEFAULT 0'), ('is_invoiced', 'INTEGER DEFAULT 0'), ('project_id', 'INTEGER')],
        'clients': [('surname', 'TEXT'), ('address', 'TEXT'), ('building_number', 'TEXT'), ('street_name', 'TEXT'), ('suburb', 'TEXT'), ('postal_code', 'TEXT'), ('phone', 'TEXT'), ('email', 'TEXT'), ('client_type', 'TEXT DEFAULT "Ad hoc"'), ('discount_percent', 'REAL DEFAULT 0'), ('company_name', 'TEXT'), ('registration_number', 'TEXT'), ('vat_number', 'TEXT')],
        'invoices': [('client_id', 'INTEGER')],
        'quotes': [('client_id', 'INTEGER')],
        'employees': [('start_date', 'TEXT'), ('inactive_date', 'TEXT'), ('gross_salary', 'REAL DEFAULT 0'), ('emp_number', 'TEXT'), ('id_passport', 'TEXT'), ('job_title', 'TEXT'), ('status', 'TEXT DEFAULT "Active"'), ('phone', 'TEXT'), ('email', 'TEXT'), ('address', 'TEXT'), ('emergency_contact', 'TEXT'), ('tax_number', 'TEXT'), ('paye_ref', 'TEXT'), ('bank_details', 'TEXT'), ('bank_name', 'TEXT'), ('account_holder', 'TEXT'), ('account_number', 'TEXT'), ('branch_code', 'TEXT'), ('account_type', 'TEXT'), ('payment_reference', 'TEXT'), ('google_event_id', 'TEXT'), ('emp_type', 'TEXT DEFAULT "Full-time (5 Days)"'), ('cv_file', 'TEXT'), ('id_file', 'TEXT'), ('contract_file', 'TEXT'), ('additional_leave', 'REAL DEFAULT 0'), ('workday_hours', 'REAL DEFAULT 7'), ('overtime_pay_treatment', 'TEXT DEFAULT "irregular"'), ('uif_contributor', 'TEXT DEFAULT "Yes"'), ('uif_non_contributor_reason', 'TEXT'), ('uif_termination_code', 'TEXT')],
        'payslips': [('transport', 'REAL DEFAULT 0'), ('overtime', 'REAL DEFAULT 0'), ('bonus', 'REAL DEFAULT 0'), ('reimbursable_expenses', 'REAL DEFAULT 0'), ('loan_repayment', 'REAL DEFAULT 0'), ('payslip_type', 'TEXT DEFAULT "regular"'), ('adjustment_of_payslip_id', 'INTEGER'), ('adjustment_reason', 'TEXT'), ('created_at', 'TEXT')],
        'leave_records': [('leave_type', 'TEXT DEFAULT "Annual Leave"'), ('document_file', 'TEXT')],
        'expenses': [('invoice_file', 'TEXT')]
    }
    for t_name, cols in loose_cols.items():
        for c_name, c_type in cols:
            try: conn.execute(f'ALTER TABLE {t_name} ADD COLUMN {c_name} {c_type}')
            except: pass

    for c_name, c_type in [
        ('company_id', 'INTEGER'), ('employee_id', 'INTEGER'), ('leave_type', 'TEXT DEFAULT "Annual Leave"'),
        ('start_date', 'TEXT'), ('end_date', 'TEXT'), ('days', 'REAL DEFAULT 1'), ('reason', 'TEXT'),
        ('status', 'TEXT DEFAULT "Pending"'), ('attachment_file', 'TEXT'), ('requested_at', 'TEXT DEFAULT CURRENT_TIMESTAMP'),
        ('reviewed_by', 'TEXT'), ('reviewed_at', 'TEXT'), ('admin_note', 'TEXT'), ('leave_record_id', 'INTEGER')
    ]:
        try: conn.execute(f'ALTER TABLE staff_leave_requests ADD COLUMN {c_name} {c_type}')
        except Exception: pass
    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_staff_leave_requests_company_status ON staff_leave_requests(company_id, status, start_date)')
    except sqlite3.OperationalError: pass
    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_staff_leave_requests_employee ON staff_leave_requests(company_id, employee_id, start_date)')
    except sqlite3.OperationalError: pass
    try: conn.execute('CREATE INDEX IF NOT EXISTS idx_users_staff_employee ON users(company_id, employee_id, is_staff)')
    except sqlite3.OperationalError: pass

    # Payroll compliance migrations: age-based PAYE rebates and employee date of birth.
    rebate_cols = [
        ('secondary_rebate', 'REAL DEFAULT 0'),
        ('tertiary_rebate', 'REAL DEFAULT 0'),
        ('threshold_under_65', 'REAL DEFAULT 0'),
        ('threshold_65_to_74', 'REAL DEFAULT 0'),
        ('threshold_75_plus', 'REAL DEFAULT 0')
    ]
    for c_name, c_type in rebate_cols:
        try: conn.execute(f'ALTER TABLE tax_rebates ADD COLUMN {c_name} {c_type}')
        except sqlite3.OperationalError: pass

    try: conn.execute('ALTER TABLE employees ADD COLUMN date_of_birth TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE employees ADD COLUMN inactive_date TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE employees ADD COLUMN uif_contributor TEXT DEFAULT "Yes"')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE employees ADD COLUMN uif_non_contributor_reason TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE employees ADD COLUMN uif_termination_code TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE employees ADD COLUMN overtime_pay_treatment TEXT DEFAULT "irregular"')
    except sqlite3.OperationalError: pass

    # Client/invoice discount support.
    try: conn.execute('ALTER TABLE clients ADD COLUMN discount_percent REAL DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE invoices ADD COLUMN discount_percent REAL DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE invoices ADD COLUMN discount_amount REAL DEFAULT 0')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE invoices ADD COLUMN amount_due_now REAL')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE invoices ADD COLUMN balance_remaining REAL')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE invoices ADD COLUMN invoice_type TEXT DEFAULT "standard"')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE invoices ADD COLUMN project_id INTEGER')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE bookings ADD COLUMN client_id INTEGER')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE invoices ADD COLUMN client_id INTEGER')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE quotes ADD COLUMN client_id INTEGER')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE quotes ADD COLUMN converted_invoice_id INTEGER')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE quotes ADD COLUMN converted_date TEXT')
    except sqlite3.OperationalError: pass

    # Backfill client_id where historical records can be matched safely to exactly one client.
    try:
        clients_for_backfill = [dict(c) for c in conn.execute('SELECT * FROM clients').fetchall()]
        def _bf_display(c):
            return ' '.join([str(c.get('name') or '').strip(), str(c.get('surname') or '').strip()]).strip() or (c.get('company_name') or c.get('name') or '')
        def _bf_match(company_id, stored_name):
            name = str(stored_name or '').strip()
            if not name:
                return None
            matches = []
            for c in clients_for_backfill:
                if c.get('company_id') != company_id:
                    continue
                if name in {str(c.get('name') or '').strip(), _bf_display(c), str(c.get('company_name') or '').strip()}:
                    matches.append(c.get('id'))
            matches = [m for m in matches if m]
            return matches[0] if len(set(matches)) == 1 else None
        for table_name, display_col in [('bookings', 'title'), ('invoices', 'client_name'), ('quotes', 'client_name')]:
            try:
                rows_to_backfill = conn.execute(f'SELECT id, company_id, {display_col} AS display_name FROM {table_name} WHERE client_id IS NULL OR client_id=0').fetchall()
                for row in rows_to_backfill:
                    cid_to_set = _bf_match(row['company_id'], row['display_name'])
                    if cid_to_set:
                        conn.execute(f'UPDATE {table_name} SET client_id=? WHERE id=? AND company_id=?', (cid_to_set, row['id'], row['company_id']))
            except Exception:
                pass
    except Exception:
        pass

    # Seed the SARS 2027 tax year if it has not yet been configured.
    if conn.execute('SELECT COUNT(*) FROM tax_brackets WHERE tax_year=?', (2027,)).fetchone()[0] == 0:
        brackets_2027 = [
            (2027, 0, 245100, 0, 0.18),
            (2027, 245101, 383100, 44118, 0.26),
            (2027, 383101, 530200, 79998, 0.31),
            (2027, 530201, 695800, 125599, 0.36),
            (2027, 695801, 887000, 185215, 0.39),
            (2027, 887001, 1878600, 259783, 0.41),
            (2027, 1878601, 999999999, 666339, 0.45)
        ]
        for b in brackets_2027:
            conn.execute('INSERT INTO tax_brackets (tax_year, min_income, max_income, base_tax, rate) VALUES (?, ?, ?, ?, ?)', b)

    rebate_2027_exists = conn.execute('SELECT 1 FROM tax_rebates WHERE tax_year=?', (2027,)).fetchone()
    if not rebate_2027_exists:
        conn.execute('''INSERT INTO tax_rebates (tax_year, primary_rebate, secondary_rebate, tertiary_rebate, threshold_under_65, threshold_65_to_74, threshold_75_plus)
                        VALUES (?, ?, ?, ?, ?, ?, ?)''', (2027, 17820, 9765, 3249, 99000, 153250, 171300))

    # Mobile PWA booking status fields. These keep mobile workflow updates
    # tenant-safe while preserving the existing desktop booking records.
    try: conn.execute('ALTER TABLE bookings ADD COLUMN mobile_status TEXT DEFAULT "Scheduled"')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE bookings ADD COLUMN mobile_status_updated_at TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE bookings ADD COLUMN mobile_status_updated_by TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE bookings ADD COLUMN mobile_started_at TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE bookings ADD COLUMN mobile_completed_at TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('UPDATE bookings SET mobile_status="Scheduled" WHERE mobile_status IS NULL OR mobile_status=""')
    except sqlite3.Error: pass

    # Payslip adjustment migrations.
    # Regular ledger payslips remain unique per employee/month, but adjustment payslips are allowed.
    try: conn.execute('ALTER TABLE payslips ADD COLUMN payslip_type TEXT DEFAULT "regular"')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE payslips ADD COLUMN adjustment_of_payslip_id INTEGER')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE payslips ADD COLUMN adjustment_reason TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute('ALTER TABLE payslips ADD COLUMN created_at TEXT')
    except sqlite3.OperationalError: pass
    try: conn.execute("UPDATE payslips SET payslip_type='regular' WHERE payslip_type IS NULL OR payslip_type='' ")
    except sqlite3.Error: pass
    try: conn.execute('DROP INDEX IF EXISTS idx_payslips_company_employee_month_unique')
    except sqlite3.Error: pass

    # Clean up any historical duplicate REGULAR payslip ledger rows before enforcing uniqueness.
    # Source of truth is one regular finalised payslip per company + employee + payroll month.
    # Adjustment payslips are separate rows and must not be deleted by this cleanup.
    try:
        conn.execute('''DELETE FROM payslips
                        WHERE COALESCE(payslip_type, 'regular')='regular'
                          AND id NOT IN (
                            SELECT MAX(id)
                            FROM payslips
                            WHERE COALESCE(payslip_type, 'regular')='regular'
                            GROUP BY company_id, employee_id, substr(date, 1, 7)
                        )''')
    except sqlite3.Error:
        pass

    index_statements = [
        'CREATE UNIQUE INDEX IF NOT EXISTS idx_companies_name_unique ON companies(name)',
        'CREATE UNIQUE INDEX IF NOT EXISTS idx_employees_company_emp_number_unique ON employees(company_id, emp_number)',
        'CREATE INDEX IF NOT EXISTS idx_employees_company_id ON employees(company_id, id)',
        'CREATE INDEX IF NOT EXISTS idx_bookings_company_start ON bookings(company_id, start)',
        'CREATE INDEX IF NOT EXISTS idx_bookings_company_employee ON bookings(company_id, employee)',
        'CREATE INDEX IF NOT EXISTS idx_bookings_company_project ON bookings(company_id, project_id)',
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_payslips_company_employee_month_unique ON payslips(company_id, employee_id, substr(date, 1, 7)) WHERE COALESCE(payslip_type, 'regular')='regular'",
        'CREATE INDEX IF NOT EXISTS idx_leave_company_employee_date ON leave_records(company_id, employee_id, date_taken)',
        'CREATE INDEX IF NOT EXISTS idx_expenses_company_date ON expenses(company_id, date)',
        'CREATE INDEX IF NOT EXISTS idx_invoices_company_date ON invoices(company_id, date)',
        'CREATE INDEX IF NOT EXISTS idx_invoices_company_client ON invoices(company_id, client_id)',
        'CREATE INDEX IF NOT EXISTS idx_quotes_company_date ON quotes(company_id, date)',
        'CREATE INDEX IF NOT EXISTS idx_quotes_company_client ON quotes(company_id, client_id)',
        'CREATE INDEX IF NOT EXISTS idx_invoice_payments_company_invoice ON invoice_payments(company_id, invoice_id)',
        'CREATE INDEX IF NOT EXISTS idx_invoice_credit_notes_company_invoice ON invoice_credit_notes(company_id, invoice_id)',
        'CREATE INDEX IF NOT EXISTS idx_invoices_company_accounting ON invoices(company_id, accounting_status, accounting_journal_id)',
        'CREATE INDEX IF NOT EXISTS idx_credit_notes_company_accounting ON invoice_credit_notes(company_id, accounting_status, accounting_journal_id)',
        'CREATE INDEX IF NOT EXISTS idx_projects_company_client ON projects(company_id, client_id)',
        'CREATE INDEX IF NOT EXISTS idx_projects_company_status ON projects(company_id, status)',
        'CREATE INDEX IF NOT EXISTS idx_project_employees_company_project ON project_employees(company_id, project_id)',
        'CREATE INDEX IF NOT EXISTS idx_project_costs_company_project ON project_costs(company_id, project_id)',
        'CREATE INDEX IF NOT EXISTS idx_attachments_company_link ON attachments(company_id, linked_type, linked_id)',
        'CREATE INDEX IF NOT EXISTS idx_clients_company_name ON clients(company_id, name)',
        'CREATE INDEX IF NOT EXISTS idx_services_company_name ON services(company_id, name)',
        'CREATE INDEX IF NOT EXISTS idx_interviews_company_datetime ON interviews(company_id, interview_datetime)'
    ]
    for sql in index_statements:
        try: conn.execute(sql)
        except sqlite3.Error: pass

    conn.commit()
    conn.close()

# --- Helpers ---
def log_action(app_name, action, details):
    if 'company_id' not in session or 'username' not in session: return
    conn = get_db_connection()
    conn.execute("DELETE FROM audit_logs WHERE timestamp < datetime('now', '-90 days')")
    conn.execute("INSERT INTO audit_logs (company_id, username, app_name, action, details, timestamp) VALUES (?, ?, ?, ?, ?, datetime('now', 'localtime'))",
                 (session['company_id'], session['username'], app_name, action, details))
    conn.commit()
    conn.close()

def get_setting(key, company_id=None):
    conn = get_db_connection()
    cid = company_id if company_id is not None else session.get('company_id', 0)
    res = conn.execute('SELECT value FROM settings WHERE key = ? AND company_id = ?', (key, cid)).fetchone()
    conn.close()
    return res['value'] if res else ""

SYSTEM_EMAIL_SETTING_KEYS = ['smtp_server', 'smtp_port', 'smtp_user', 'smtp_pass', 'sender_email']

def get_system_email_settings(conn=None):
    should_close = False
    if conn is None:
        conn = get_db_connection()
        should_close = True
    try:
        rows = conn.execute('SELECT setting_key, setting_value FROM system_email_settings').fetchall()
        return {row['setting_key']: row['setting_value'] for row in rows}
    except Exception:
        return {}
    finally:
        if should_close:
            conn.close()

def save_system_email_settings(values):
    conn = get_db_connection()
    try:
        for key in SYSTEM_EMAIL_SETTING_KEYS:
            val = (values.get(key) or '').strip()
            exists = conn.execute('SELECT 1 FROM system_email_settings WHERE setting_key=?', (key,)).fetchone()
            if exists:
                conn.execute('UPDATE system_email_settings SET setting_value=? WHERE setting_key=?', (val, key))
            else:
                conn.execute('INSERT INTO system_email_settings (setting_key, setting_value) VALUES (?, ?)', (key, val))
        conn.commit()
    finally:
        conn.close()

def _system_email_missing(settings):
    required = ['smtp_server', 'smtp_port', 'smtp_user', 'smtp_pass']
    return [k for k in required if not (settings.get(k) or '').strip()]

def _send_system_email(to_email, subject, body):
    settings = get_system_email_settings()
    missing = _system_email_missing(settings)
    if missing:
        raise RuntimeError('System Email Settings are incomplete.')
    server_host = settings.get('smtp_server', '').strip()
    port = int(settings.get('smtp_port') or 465)
    user = settings.get('smtp_user', '').strip()
    password = settings.get('smtp_pass', '')
    sender = (settings.get('sender_email') or user).strip()

    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = sender
    msg['To'] = to_email
    msg.set_content(body)

    if port == 587:
        with smtplib.SMTP(server_host, port, timeout=20) as smtp:
            smtp.starttls()
            smtp.login(user, password)
            smtp.send_message(msg)
    else:
        with smtplib.SMTP_SSL(server_host, port, timeout=20) as smtp:
            smtp.login(user, password)
            smtp.send_message(msg)

def generate_temporary_password(length=12):
    alphabet = string.ascii_letters + string.digits + '!@#$%&*?'
    while True:
        password = ''.join(secrets.choice(alphabet) for _ in range(length))
        if (any(c.islower() for c in password) and any(c.isupper() for c in password)
                and any(c.isdigit() for c in password) and any(c in '!@#$%&*?' for c in password)):
            return password

def _row_has_key(row, key):
    try:
        return key in row.keys()
    except Exception:
        return False

def _normalise_email(value):
    return (value or '').strip().lower()

def _account_email_for_password_reset(conn, user):
    user_dict = dict(user)
    account_email = user_dict.get('email') or ''
    if not account_email and user_dict.get('employee_id') and user_dict.get('company_id'):
        try:
            emp = conn.execute('SELECT email FROM employees WHERE id=? AND company_id=?', (user_dict.get('employee_id'), user_dict.get('company_id'))).fetchone()
            if emp:
                account_email = dict(emp).get('email') or ''
        except Exception:
            account_email = ''
    return account_email

def calculate_uif(monthly_gross): 
    return round(min(monthly_gross * 0.01, 177.12), 2)

def safe_money(value):
    try:
        amount = float(value or 0)
    except (TypeError, ValueError):
        amount = 0.0
    return max(0.0, amount)

def safe_adjustment_money(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0

def compose_client_address(building_number=None, street_name=None, suburb=None, postal_code=None, fallback_address=None):
    parts = []
    street_line = ' '.join([str(building_number or '').strip(), str(street_name or '').strip()]).strip()
    if street_line:
        parts.append(street_line)
    if suburb:
        parts.append(str(suburb).strip())
    if postal_code:
        parts.append(str(postal_code).strip())
    composed = ', '.join([p for p in parts if p])
    return composed if composed else (fallback_address or '')


def client_display_name(client_row):
    if not client_row:
        return ''
    c = dict(client_row)
    full = ' '.join([str(c.get('name') or '').strip(), str(c.get('surname') or '').strip()]).strip()
    return full or str(c.get('company_name') or c.get('name') or '').strip()


def client_option_label(client_row):
    if not client_row:
        return ''
    c = dict(client_row)
    label = client_display_name(c)
    extra = []
    if c.get('company_name') and c.get('company_name') not in label:
        extra.append(str(c.get('company_name')).strip())
    address = compose_client_address(c.get('building_number'), c.get('street_name'), c.get('suburb'), c.get('postal_code'), c.get('address'))
    if address:
        extra.append(address)
    if c.get('phone'):
        extra.append(str(c.get('phone')).strip())
    return label + ((' — ' + ' | '.join([x for x in extra if x])) if extra else '')


def prepare_client_options(clients):
    prepared = []
    for client in clients or []:
        d = dict(client)
        d['display_name'] = client_display_name(d)
        d['option_label'] = client_option_label(d)
        prepared.append(d)
    return prepared


def get_client_by_id(conn, company_id, client_id):
    try:
        cid_value = int(client_id or 0)
    except Exception:
        return None
    if not cid_value:
        return None
    return conn.execute('SELECT * FROM clients WHERE id=? AND company_id=?', (cid_value, company_id)).fetchone()


def find_client_by_display_name(conn, company_id, name_value):
    name_value = str(name_value or '').strip()
    if not name_value:
        return None
    rows = conn.execute('SELECT * FROM clients WHERE company_id=?', (company_id,)).fetchall()
    matches = []
    for row in rows:
        c = dict(row)
        if name_value in {str(c.get('name') or '').strip(), client_display_name(c), str(c.get('company_name') or '').strip()}:
            matches.append(row)
    return matches[0] if len(matches) == 1 else None


def resolve_client_from_payload(conn, company_id, data, id_key='client_id', name_key='client', required=True):
    data = data or {}
    client = get_client_by_id(conn, company_id, data.get(id_key))
    if not client and name_key:
        client = find_client_by_display_name(conn, company_id, data.get(name_key))
    if required and not client:
        raise ValueError('Client not found. Please select the client from the client list.')
    return client


def get_document_client(conn, company_id, doc_row):
    if not doc_row:
        return None
    d = dict(doc_row)
    return get_client_by_id(conn, company_id, d.get('client_id')) or find_client_by_display_name(conn, company_id, d.get('client_name'))

def sanitize_percent(value):
    try:
        pct = float(value or 0)
    except (TypeError, ValueError):
        return 0.0
    return max(0.0, min(100.0, round(pct, 2)))

def calculate_invoice_discount(subtotal, discount_percent):
    pct = sanitize_percent(discount_percent)
    base = float(subtotal or 0)
    discount_amount = round(base * (pct / 100.0), 2)
    discounted_subtotal = round(base - discount_amount, 2)
    return pct, discount_amount, discounted_subtotal

def sanitize_money(value, default=0.0):
    try:
        amount = float(value or default)
    except (TypeError, ValueError):
        amount = float(default or 0)
    return round(max(amount, 0.0), 2)

def sanitize_quantity(value, default=1.0):
    try:
        qty = float(value if value not in (None, '') else default)
    except (TypeError, ValueError):
        qty = float(default or 1)
    return round(max(qty, 0.0), 4)


def normalise_billing_item(item):
    item = item or {}
    quantity = sanitize_quantity(item.get('quantity', 1), 1)
    if quantity <= 0:
        quantity = 1.0

    # If a line total/amount is submitted, treat it as the final line amount.
    # This allows users to adjust the amount directly while the system derives
    # the unit price from amount ÷ quantity.
    submitted_amount = item.get('amount')
    if submitted_amount not in (None, '') and sanitize_money(submitted_amount, 0) > 0:
        amount = sanitize_money(submitted_amount, 0)
        unit_price = round(amount / quantity, 4) if quantity else amount
    else:
        unit_price = sanitize_money(item.get('unit_price'), 0)
        amount = round(quantity * unit_price, 2)
    return {
        'booking_id': item.get('booking_id'),
        'service_date': item.get('service_date'),
        'description': item.get('description') or '',
        'quantity': quantity,
        'unit_price': unit_price,
        'amount': amount
    }

def calculate_invoice_due_now(total, amount_due_now=None):
    total = sanitize_money(total)
    if amount_due_now is None or amount_due_now == '':
        due_now = total
    else:
        due_now = sanitize_money(amount_due_now)
    due_now = min(due_now, total)
    balance_remaining = round(total - due_now, 2)
    return due_now, balance_remaining

def get_invoice_financial_totals(conn, company_id, invoice_id):
    inv = conn.execute('SELECT total FROM invoices WHERE id=? AND company_id=?', (invoice_id, company_id)).fetchone()
    if not inv:
        return {'total': 0.0, 'paid': 0.0, 'credited': 0.0, 'outstanding': 0.0}
    total = float(inv['total'] or 0)
    paid = conn.execute('SELECT COALESCE(SUM(amount), 0) AS paid FROM invoice_payments WHERE invoice_id=? AND company_id=?', (invoice_id, company_id)).fetchone()['paid'] or 0
    credited = conn.execute('SELECT COALESCE(SUM(amount), 0) AS credited FROM invoice_credit_notes WHERE invoice_id=? AND company_id=?', (invoice_id, company_id)).fetchone()['credited'] or 0
    paid = float(paid or 0)
    credited = float(credited or 0)
    return {
        'total': round(total, 2),
        'paid': round(paid, 2),
        'credited': round(credited, 2),
        'outstanding': round(max(total - paid - credited, 0), 2)
    }


def update_invoice_payment_status(conn, company_id, invoice_id):
    totals = get_invoice_financial_totals(conn, company_id, invoice_id)
    total = totals['total']
    paid = totals['paid']
    credited = totals['credited']
    outstanding = totals['outstanding']
    if total <= 0:
        status = 'Unpaid'
    elif credited >= total:
        status = 'Credited'
    elif outstanding <= 0 and credited > 0:
        status = 'Settled with Credit'
    elif outstanding <= 0:
        status = 'Paid'
    elif credited > 0 and paid > 0:
        status = 'Partially Paid/Credited'
    elif credited > 0:
        status = 'Partially Credited'
    elif paid > 0:
        status = 'Partially Paid'
    else:
        status = 'Unpaid'
    conn.execute('UPDATE invoices SET status=? WHERE id=? AND company_id=?', (status, invoice_id, company_id))

class TaxTableNotSetError(Exception):
    pass

def get_sars_tax_year(check_date_str=None):
    tax_year = datetime.now().year
    if check_date_str:
        d = datetime.strptime(check_date_str[:10], '%Y-%m-%d')
        tax_year = d.year if d.month < 3 else d.year + 1
    return tax_year

def calculate_age_on_date(date_of_birth_str, on_date):
    if not date_of_birth_str:
        return None
    dob = datetime.strptime(date_of_birth_str[:10], '%Y-%m-%d')
    return on_date.year - dob.year - ((on_date.month, on_date.day) < (dob.month, dob.day))

def get_tax_rebate_and_threshold(rebate_row, age):
    primary = float(rebate_row['primary_rebate'] or 0)
    secondary = float(rebate_row['secondary_rebate'] or 0)
    tertiary = float(rebate_row['tertiary_rebate'] or 0)
    threshold_under_65 = float(rebate_row['threshold_under_65'] or 0)
    threshold_65_to_74 = float(rebate_row['threshold_65_to_74'] or 0)
    threshold_75_plus = float(rebate_row['threshold_75_plus'] or 0)

    if age is not None and age >= 75:
        return primary + secondary + tertiary, threshold_75_plus
    if age is not None and age >= 65:
        return primary + secondary, threshold_65_to_74
    return primary, threshold_under_65

def ensure_tax_tables_configured(tax_year):
    conn = get_db_connection()
    bracket_count = conn.execute('SELECT COUNT(*) FROM tax_brackets WHERE tax_year=?', (tax_year,)).fetchone()[0]
    rebate = conn.execute('''SELECT primary_rebate, secondary_rebate, tertiary_rebate, threshold_under_65, threshold_65_to_74, threshold_75_plus
                             FROM tax_rebates WHERE tax_year=?''', (tax_year,)).fetchone()
    conn.close()
    if bracket_count == 0 or not rebate:
        raise TaxTableNotSetError('Tax Tables not set')

def calculate_annual_tax(annual_income, check_date_str=None, date_of_birth=None):
    annual = float(annual_income or 0)
    if annual <= 0:
        return 0.0

    try:
        tax_year = get_sars_tax_year(check_date_str)
    except Exception:
        tax_year = get_sars_tax_year(None)

    conn = get_db_connection()
    bracket_count = conn.execute('SELECT COUNT(*) FROM tax_brackets WHERE tax_year=?', (tax_year,)).fetchone()[0]
    rebate_db = conn.execute("""SELECT primary_rebate, secondary_rebate, tertiary_rebate, threshold_under_65, threshold_65_to_74, threshold_75_plus
                                FROM tax_rebates WHERE tax_year=?""", (tax_year,)).fetchone()
    if bracket_count == 0 or not rebate_db:
        conn.close()
        raise TaxTableNotSetError('Tax Tables not set')

    tax_year_end = datetime(tax_year, 2, calendar.monthrange(tax_year, 2)[1])
    try:
        age = calculate_age_on_date(date_of_birth, tax_year_end)
    except Exception:
        age = None

    rebate, threshold = get_tax_rebate_and_threshold(rebate_db, age)
    if threshold and annual <= threshold:
        conn.close()
        return 0.0

    bracket = conn.execute('SELECT * FROM tax_brackets WHERE tax_year=? AND ? BETWEEN min_income AND max_income', (tax_year, annual)).fetchone()
    if not bracket:
        conn.close()
        raise TaxTableNotSetError('Tax Tables not set')

    base = float(bracket['base_tax'])
    rate = float(bracket['rate'])
    min_inc = float(bracket['min_income'])
    threshold_base = min_inc - 1 if min_inc > 0 else 0
    tax = base + ((annual - threshold_base) * rate)

    tax = max(0.0, tax - rebate)
    conn.close()
    return round(tax, 2)


def calculate_paye(monthly_gross, check_date_str=None, date_of_birth=None):
    annual_tax = calculate_annual_tax(float(monthly_gross or 0) * 12, check_date_str, date_of_birth)
    return round(annual_tax / 12, 2)


def calculate_paye_with_regular_irregular(regular_monthly_income, irregular_onceoff_income=0, check_date_str=None, date_of_birth=None):
    """Calculate monthly PAYE by annualising regular remuneration and adding once-off irregular remuneration once.

    This supports payroll treatment choices such as regular recurring overtime vs irregular/once-off overtime.
    """
    regular_monthly_income = float(regular_monthly_income or 0)
    irregular_onceoff_income = float(irregular_onceoff_income or 0)

    annual_regular = regular_monthly_income * 12
    regular_annual_tax = calculate_annual_tax(annual_regular, check_date_str, date_of_birth)
    normal_monthly_paye = regular_annual_tax / 12

    if irregular_onceoff_income <= 0:
        return round(normal_monthly_paye, 2)

    annual_tax_with_irregular = calculate_annual_tax(annual_regular + irregular_onceoff_income, check_date_str, date_of_birth)
    irregular_tax = max(0.0, annual_tax_with_irregular - regular_annual_tax)
    return round(normal_monthly_paye + irregular_tax, 2)

def add_months(sourcedate, months):
    month = sourcedate.month - 1 + months
    year = sourcedate.year + month // 12
    month = month % 12 + 1
    day = min(sourcedate.day, calendar.monthrange(year, month)[1])
    return datetime(year, month, day)

def get_employee_workday_hours(emp):
    try:
        hours = float(emp['workday_hours'] or 7)
    except Exception:
        hours = 7.0
    return max(0.01, hours)

def parse_date_safe(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value)[:10], '%Y-%m-%d')
    except Exception:
        return None

def get_month_start_end_from_date(date_str):
    base = parse_date_safe(date_str) or datetime.now()
    start = datetime(base.year, base.month, 1)
    end = datetime(base.year, base.month, calendar.monthrange(base.year, base.month)[1])
    return start, end

def get_employee_inactive_date(emp):
    try:
        return parse_date_safe(emp['inactive_date'])
    except Exception:
        return None

def get_employee_payroll_cutoff(emp, payroll_date_str):
    month_start, month_end = get_month_start_end_from_date(payroll_date_str)
    inactive_date = get_employee_inactive_date(emp)
    if inactive_date and inactive_date < month_start:
        return month_start, month_end, inactive_date, None
    cutoff = min(month_end, inactive_date) if inactive_date else month_end
    return month_start, month_end, inactive_date, cutoff

def count_ordinary_days_in_range(start_dt, end_dt, emp_type):
    if not start_dt or not end_dt or end_dt < start_dt:
        return 0
    ordinary_weekdays = get_contract_day_rules(emp_type).get('ordinary_weekdays', {0, 1, 2, 3, 4, 5})
    current = start_dt.date()
    end_date = end_dt.date()
    count = 0
    while current <= end_date:
        if current.weekday() in ordinary_weekdays:
            count += 1
        current += timedelta(days=1)
    return count

def prorate_monthly_salary_for_inactive_date(base_salary, emp_type, month_start, month_end, cutoff_dt):
    if not cutoff_dt or cutoff_dt >= month_end:
        return float(base_salary or 0.0), ''
    total_days = count_ordinary_days_in_range(month_start, month_end, emp_type)
    payable_days = count_ordinary_days_in_range(month_start, cutoff_dt, emp_type)
    if total_days <= 0:
        return 0.0, '(Inactive before payroll month)'
    prorated = float(base_salary or 0.0) * (payable_days / total_days)
    return prorated, f'(Prorated: {payable_days}/{total_days} ordinary days to inactive date {cutoff_dt.strftime("%Y-%m-%d")})'

def get_contract_day_rules(emp_type):
    emp_type = emp_type or ''
    if emp_type == 'Full-time (5 Days)':
        return {
            'ordinary_weekdays': {0, 1, 2, 3, 4},
            'saturday_multiplier': 1.5,
            'sunday_multiplier': 2.0,
            'public_holiday_multiplier': 2.0
        }
    if emp_type == 'Shift Worker':
        return {
            # True shift workers ordinarily work Monday to Sunday. Sunday is therefore
            # paid at the BCEA ordinary-Sunday premium of 1.5x unless it is also a public holiday.
            'ordinary_weekdays': {0, 1, 2, 3, 4, 5, 6},
            'saturday_multiplier': 1.0,
            'sunday_multiplier': 1.5,
            'public_holiday_multiplier': 2.0
        }
    if emp_type in ['Full-time (6 Days)', 'Contract >25 Hrs', 'Contract <25 Hrs']:
        return {
            'ordinary_weekdays': {0, 1, 2, 3, 4, 5},
            'saturday_multiplier': 1.0,
            'sunday_multiplier': 2.0,
            'public_holiday_multiplier': 2.0
        }
    return {
        'ordinary_weekdays': {0, 1, 2, 3, 4, 5},
        'saturday_multiplier': 1.0,
        'sunday_multiplier': 2.0,
        'public_holiday_multiplier': 2.0
    }

def analyse_booking_hours(bookings, emp_type, workday_hours, public_holiday_dates):
    rules = get_contract_day_rules(emp_type)
    public_holiday_dates = set(public_holiday_dates or [])
    ordinary_days = 0
    saturday_nonordinary_days = 0
    sunday_premium_days = 0
    public_holiday_days = 0
    explicit_overtime_hours = 0.0

    for b in bookings:
        if b['start']:
            b_date_str = b['start'][:10]
            b_date = datetime.strptime(b_date_str, "%Y-%m-%d")
            weekday = b_date.weekday()
            is_holiday = b_date_str in public_holiday_dates
            is_sunday = weekday == 6
            is_saturday = weekday == 5

            # Public holidays override ordinary/Sunday logic. A public holiday that
            # falls on a Sunday is paid at the public-holiday rate, not the Sunday rate.
            if is_holiday:
                public_holiday_days += 1
            elif is_sunday and rules['sunday_multiplier'] > 1.0:
                sunday_premium_days += 1
                if weekday in rules['ordinary_weekdays']:
                    ordinary_days += 1
            elif is_saturday and rules['saturday_multiplier'] > 1.0:
                saturday_nonordinary_days += 1
            elif weekday in rules['ordinary_weekdays']:
                ordinary_days += 1
            else:
                # Non-ordinary day not otherwise classified; treat as public-holiday style premium.
                public_holiday_days += 1
        try:
            explicit_overtime_hours += float(b['overtime_hours'] or 0.0)
        except Exception:
            pass

    hours_per_day = float(workday_hours or 0.0)
    saturday_nonordinary_hours = saturday_nonordinary_days * hours_per_day
    sunday_premium_hours = sunday_premium_days * hours_per_day
    public_holiday_hours = public_holiday_days * hours_per_day
    ordinary_hours = ordinary_days * hours_per_day

    # For weekly BCEA warnings, ordinary Sunday shift-worker hours are counted once as ordinary hours.
    # Non-ordinary/overtime warning hours include non-ordinary Saturdays, public holidays,
    # non-ordinary Sundays and explicit overtime captured on the booking.
    sunday_nonordinary_hours = 0.0 if emp_type == 'Shift Worker' else sunday_premium_hours
    nonordinary_hours = saturday_nonordinary_hours + public_holiday_hours + sunday_nonordinary_hours
    overtime_hours = nonordinary_hours + explicit_overtime_hours

    return {
        'ordinary_days': ordinary_days,
        'saturday_nonordinary_days': saturday_nonordinary_days,
        'sunday_premium_days': sunday_premium_days,
        'public_holiday_days': public_holiday_days,
        # Backwards-compatible aggregate for older display code.
        'sunday_holiday_days': sunday_premium_days + public_holiday_days,
        'ordinary_hours': ordinary_hours,
        'saturday_nonordinary_hours': saturday_nonordinary_hours,
        'sunday_premium_hours': sunday_premium_hours,
        'public_holiday_hours': public_holiday_hours,
        'sunday_holiday_hours': sunday_premium_hours + public_holiday_hours,
        'nonordinary_hours': nonordinary_hours,
        'explicit_overtime_hours': explicit_overtime_hours,
        'overtime_hours': overtime_hours,
        'total_hours': ordinary_hours + overtime_hours
    }

def get_bcea_hours_warning(ordinary_hours, overtime_hours, emp_type):
    ordinary_hours = float(ordinary_hours or 0)
    overtime_hours = float(overtime_hours or 0)
    total_hours = ordinary_hours + overtime_hours
    warnings = []
    status = 'green'

    if ordinary_hours > 45:
        status = 'red'
        warnings.append(f'BCEA maximum ordinary hours exceeded: {ordinary_hours:.2f}/45.00 hrs')
    elif ordinary_hours >= 40.5:
        status = 'yellow'
        warnings.append(f'BCEA ordinary hours near weekly maximum: {ordinary_hours:.2f}/45.00 hrs')

    if overtime_hours > 10:
        status = 'red'
        warnings.append(f'BCEA overtime / non-ordinary work limit exceeded: {overtime_hours:.2f}/10.00 hrs')
    elif overtime_hours >= 8:
        if status != 'red': status = 'yellow'
        warnings.append(f'BCEA overtime / non-ordinary work near weekly maximum: {overtime_hours:.2f}/10.00 hrs')

    if total_hours > 55:
        status = 'red'
        warnings.append(f'Combined ordinary and overtime hours exceed 55.00 hrs: {total_hours:.2f} hrs')

    if emp_type == 'Contract <25 Hrs':
        contract_hours = total_hours
        if contract_hours > 25:
            status = 'red'
            warnings.append(f'Contract <25 Hrs threshold exceeded: {contract_hours:.2f}/25.00 hrs')
        elif contract_hours >= 22:
            if status != 'red': status = 'yellow'
            warnings.append(f'Contract <25 Hrs is near 25-hour threshold: {contract_hours:.2f}/25.00 hrs')

    return {'status': status, 'ordinary_hours': round(ordinary_hours, 2), 'overtime_hours': round(overtime_hours, 2), 'total_hours': round(total_hours, 2), 'messages': warnings}


def employee_name_matches(booking_employee, employee_name):
    names = [n.strip() for n in (booking_employee or '').split(',') if n.strip()]
    return employee_name in names

def as_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value

def as_datetime_start(value):
    value = as_date(value)
    return datetime.combine(value, datetime.min.time())

def as_datetime_end(value):
    value = as_date(value)
    return datetime.combine(value, datetime.max.time().replace(microsecond=0))

def get_week_bounds(date_obj):
    # BCEA operational week is Monday to Sunday. Always calculate from the
    # selected booking date itself, not from browser/UTC conversions.
    date_only = as_date(date_obj)
    week_start_date = date_only - timedelta(days=date_only.weekday())
    week_end_date = week_start_date + timedelta(days=6)
    return as_datetime_start(week_start_date), as_datetime_end(week_end_date)

def get_month_bounds(date_obj):
    date_only = as_date(date_obj)
    month_start_date = date_only.replace(day=1)
    month_end_date = date_only.replace(day=calendar.monthrange(date_only.year, date_only.month)[1])
    return as_datetime_start(month_start_date), as_datetime_end(month_end_date)

def get_public_holiday_dates_between(conn, start_date, end_date):
    rows = conn.execute('''SELECT date_str FROM public_holidays
                           WHERE date_str BETWEEN ? AND ?''',
                        (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))).fetchall()
    return [r['date_str'] for r in rows]

def get_employee_leave_dates(conn, company_id, employee_id, start_date, end_date):
    rows = conn.execute('''SELECT date_taken, days FROM leave_records
                           WHERE company_id=? AND employee_id=? AND date_taken <= ?
                           ORDER BY date_taken ASC''',
                        (company_id, employee_id, end_date.strftime('%Y-%m-%d'))).fetchall()
    leave_dates = set()
    for r in rows:
        try:
            leave_start = datetime.strptime(r['date_taken'], '%Y-%m-%d').date()
            days = max(1, int(float(r['days'] or 1)))
            for offset in range(days):
                leave_day = leave_start + timedelta(days=offset)
                if start_date.date() <= leave_day <= end_date.date():
                    leave_dates.add(leave_day.strftime('%Y-%m-%d'))
        except Exception:
            continue
    return leave_dates

def fetch_employee_bookings_between(conn, company_id, employee_name, start_date, end_date, exclude_booking_id=None):
    rows = conn.execute('''SELECT * FROM bookings
                           WHERE company_id=? AND start >= ? AND start <= ?
                           ORDER BY start ASC''',
                        (company_id, f"{start_date.strftime('%Y-%m-%d')}T00:00", f"{end_date.strftime('%Y-%m-%d')}T23:59")).fetchall()
    bookings = []
    for row in rows:
        try:
            if exclude_booking_id and str(row['id']) == str(exclude_booking_id):
                continue
        except Exception:
            pass
        if employee_name_matches(row['employee'], employee_name):
            bookings.append(row)
    return bookings

def analyse_employee_hours_for_period(conn, company_id, employee, start_date, end_date, proposed_date=None, proposed_overtime=0, exclude_booking_id=None, include_proposed=False):
    emp_name = employee['name']
    workday_hours = get_employee_workday_hours(employee)
    bookings = fetch_employee_bookings_between(conn, company_id, emp_name, start_date, end_date, exclude_booking_id)

    proposed_date_only = as_date(proposed_date) if proposed_date else None
    if include_proposed and proposed_date_only and start_date.date() <= proposed_date_only <= end_date.date():
        bookings.append({
            'start': f"{proposed_date_only.strftime('%Y-%m-%d')}T00:00",
            'employee': emp_name,
            'overtime_hours': float(proposed_overtime or 0)
        })

    public_holidays = get_public_holiday_dates_between(conn, start_date, end_date)
    hours = analyse_booking_hours(bookings, employee['emp_type'], workday_hours, public_holidays)
    warning = get_bcea_hours_warning(hours['ordinary_hours'], hours['overtime_hours'], employee['emp_type'])
    return hours, warning

def get_booking_staff_hours_summary(conn, company_id, employee, target_date, proposed_overtime=0, exclude_booking_id=None, include_proposed=False):
    target_date_only = as_date(target_date)
    week_start, week_end = get_week_bounds(target_date_only)
    month_start, month_end = get_month_bounds(target_date_only)

    week_hours, week_warning = analyse_employee_hours_for_period(
        conn, company_id, employee, week_start, week_end, target_date_only, proposed_overtime, exclude_booking_id, include_proposed
    )

    month_hours, _month_warning = analyse_employee_hours_for_period(
        conn, company_id, employee, month_start, month_end, target_date_only, proposed_overtime, exclude_booking_id, include_proposed
    )

    # The month colour reflects the worst weekly BCEA status in the payroll month, because BCEA limits are weekly.
    worst_status = 'green'
    current_week_start = month_start - timedelta(days=month_start.weekday())
    while current_week_start <= month_end:
        current_week_end = current_week_start + timedelta(days=6)
        proposed_for_this_week = target_date_only if current_week_start.date() <= target_date_only <= current_week_end.date() else None
        _, week_check = analyse_employee_hours_for_period(
            conn, company_id, employee, current_week_start, current_week_end,
            proposed_for_this_week, proposed_overtime if proposed_for_this_week else 0, exclude_booking_id, include_proposed
        )
        if week_check['status'] == 'red':
            worst_status = 'red'
            break
        if week_check['status'] == 'yellow':
            worst_status = 'yellow'
        current_week_start += timedelta(days=7)

    leave_dates = get_employee_leave_dates(conn, company_id, employee['id'], target_date, target_date)

    return {
        'employee': employee['name'],
        'emp_type': employee['emp_type'],
        'workday_hours': get_employee_workday_hours(employee),
        'requested_date': target_date_only.strftime('%Y-%m-%d'),
        'is_on_leave': target_date_only.strftime('%Y-%m-%d') in leave_dates,
        'week_start': week_start.strftime('%Y-%m-%d'),
        'week_end': week_end.strftime('%Y-%m-%d'),
        'week_hours': round(week_hours['total_hours'], 2),
        'week_ordinary_hours': round(week_hours['ordinary_hours'], 2),
        'week_overtime_hours': round(week_hours['overtime_hours'], 2),
        'week_status': week_warning['status'],
        'week_messages': week_warning['messages'],
        'month_start': month_start.strftime('%Y-%m-%d'),
        'month_end': month_end.strftime('%Y-%m-%d'),
        'month_hours': round(month_hours['total_hours'], 2),
        'month_status': worst_status
    }

def validate_booking_employees_available(conn, company_id, assignments, booking_date, exclude_booking_id=None):
    employee_names = []
    for assignment in assignments:
        for name in (assignment.get('employee') or '').split(','):
            clean_name = name.strip()
            if clean_name and clean_name not in employee_names:
                employee_names.append(clean_name)

    unavailable = []
    for emp_name in employee_names:
        emp = conn.execute('SELECT * FROM employees WHERE company_id=? AND name=?', (company_id, emp_name)).fetchone()
        if not emp:
            continue
        leave_dates = get_employee_leave_dates(conn, company_id, emp['id'], booking_date, booking_date)
        if booking_date.strftime('%Y-%m-%d') in leave_dates:
            unavailable.append(emp_name)

    if unavailable:
        return False, f"Cannot book employee on leave: {', '.join(unavailable)}"
    return True, ''

def calculate_leave_balance(employee_id, start_date_str, emp_type, emp_name, ref_date_str=None):
    if not start_date_str: return 0.0
    start = datetime.strptime(start_date_str, '%Y-%m-%d')
    now = datetime.strptime(ref_date_str, '%Y-%m-%d') if ref_date_str else datetime.now()
    cid = session.get('company_id')
    conn = get_db_connection()
    emp_record = conn.execute('SELECT additional_leave, inactive_date FROM employees WHERE id=? AND company_id=?', (employee_id, cid)).fetchone()
    add_leave_per_year = float(emp_record['additional_leave']) if emp_record and 'additional_leave' in dict(emp_record) and emp_record['additional_leave'] else 0.0
    inactive_dt = parse_date_safe(emp_record['inactive_date']) if emp_record and 'inactive_date' in dict(emp_record) else None
    if inactive_dt and inactive_dt < now:
        now = inactive_dt

    safe_type = emp_type or 'Full-time (5 Days)'
    if 'Shift Worker' in safe_type: safe_type = 'Full-time (5 Days)'

    taken_rows = conn.execute('''SELECT date_taken, days FROM leave_records
                                 WHERE company_id=? AND employee_id=? AND leave_type="Annual Leave" AND date_taken <= ?
                                 ORDER BY date_taken ASC''', (cid, employee_id, now.strftime('%Y-%m-%d'))).fetchall()
    total_taken = sum(float(r['days'] or 0.0) for r in taken_rows)

    base_monthly_rate = 0.0
    if safe_type in ['Full-time', 'Full-time (5 Days)']: base_monthly_rate = 1.25
    elif safe_type == 'Full-time (6 Days)': base_monthly_rate = 1.5
    monthly_rate = base_monthly_rate + (add_leave_per_year / 12.0)

    valid_shift_dates = []
    if safe_type == 'Contract >25 Hrs':
        bookings = conn.execute('SELECT start FROM bookings WHERE company_id=? AND employee LIKE ?', (cid, f"%{emp_name}%")).fetchall()
        for b in bookings:
            try:
                b_date = datetime.strptime(b['start'][:10], '%Y-%m-%d')
                if start.date() <= b_date.date() <= now.date():
                    valid_shift_dates.append(b_date)
            except Exception:
                pass

    awards = []
    award_month = datetime(start.year, start.month, 1)
    end_month = datetime(now.year, now.month, 1)
    while award_month <= end_month:
        amount = 0.0
        if safe_type == 'Contract >25 Hrs':
            next_month = add_months(award_month, 1)
            shifts = sum(1 for d in valid_shift_dates if award_month.date() <= d.date() < next_month.date())
            amount = (shifts / 17.0) + (add_leave_per_year / 12.0)
        else:
            amount = monthly_rate

        # If employment ends during the award month, only award leave up to the last working day.
        if inactive_dt and award_month.year == inactive_dt.year and award_month.month == inactive_dt.month and amount > 0:
            month_start = datetime(award_month.year, award_month.month, 1)
            month_end = datetime(award_month.year, award_month.month, calendar.monthrange(award_month.year, award_month.month)[1])
            total_ordinary_days = count_ordinary_days_in_range(month_start, month_end, safe_type)
            payable_ordinary_days = count_ordinary_days_in_range(month_start, inactive_dt, safe_type)
            if total_ordinary_days > 0:
                amount = amount * (payable_ordinary_days / total_ordinary_days)

        expiry_date = add_months(award_month, 18)
        if now < expiry_date and amount > 0:
            awards.append({'award_month': award_month, 'expiry_date': expiry_date, 'remaining': amount})
        award_month = add_months(award_month, 1)

    for award in awards:
        if total_taken <= 0:
            break
        used = min(award['remaining'], total_taken)
        award['remaining'] -= used
        total_taken -= used

    valid_balance = sum(a['remaining'] for a in awards)
    conn.close()
    return round(valid_balance, 2)

def calculate_sick_leave_balance(employee_id, start_date_str, emp_type, emp_name, ref_date_str=None):
    if not start_date_str or emp_type in ['Supplier', 'Provider', 'Contract <25 Hrs']: return "N/A"
    start = datetime.strptime(start_date_str, '%Y-%m-%d')
    now = datetime.strptime(ref_date_str, '%Y-%m-%d') if ref_date_str else datetime.now()
    cid = session.get('company_id')
    conn = get_db_connection()
    emp_record = conn.execute('SELECT inactive_date FROM employees WHERE id=? AND company_id=?', (employee_id, cid)).fetchone()
    inactive_dt = parse_date_safe(emp_record['inactive_date']) if emp_record and 'inactive_date' in dict(emp_record) else None
    if inactive_dt and inactive_dt < now:
        now = inactive_dt
    months = (now.year - start.year) * 12 + (now.month - start.month)
    days = (now.date() - start.date()).days

    safe_type = emp_type or 'Full-time (5 Days)'
    if 'Shift Worker' in safe_type: safe_type = 'Full-time (5 Days)'

    cycle_start = add_months(start, (max(0, months) // 36) * 36)
    cycle_end = add_months(cycle_start, 36)

    earned = 0.0
    if months < 6:
        if safe_type in ['Full-time', 'Full-time (5 Days)']:
            earned = ((max(0, days) / 7.0) * 5.0) / 26.0
        elif safe_type == 'Full-time (6 Days)':
            earned = ((max(0, days) / 7.0) * 6.0) / 26.0
        elif safe_type == 'Contract >25 Hrs':
            valid = sum(1 for b in conn.execute('SELECT start FROM bookings WHERE company_id=? AND employee LIKE ?', (cid, f"%{emp_name}%")).fetchall() if start.date() <= datetime.strptime(b['start'][:10], '%Y-%m-%d').date() <= now.date())
            earned = valid / 26.0
    else:
        if safe_type in ['Full-time', 'Full-time (5 Days)']:
            earned = 30.0
        elif safe_type == 'Full-time (6 Days)':
            earned = 36.0
        elif safe_type == 'Contract >25 Hrs':
            cycle_bookings = conn.execute('SELECT start FROM bookings WHERE company_id=? AND employee LIKE ?', (cid, f"%{emp_name}%")).fetchall()
            valid = 0
            for b in cycle_bookings:
                try:
                    b_date = datetime.strptime(b['start'][:10], '%Y-%m-%d')
                    if cycle_start.date() <= b_date.date() < cycle_end.date() and b_date.date() <= now.date():
                        valid += 1
                except Exception:
                    pass
            weeks_in_cycle_to_date = max(1, (min(now, cycle_end).date() - cycle_start.date()).days / 7.0)
            avg_days_per_week = valid / weeks_in_cycle_to_date
            earned = avg_days_per_week * 6.0

    taken = conn.execute('''SELECT SUM(days) as total FROM leave_records
                            WHERE company_id=? AND employee_id=? AND leave_type="Sick Leave" AND date_taken >= ? AND date_taken <= ?''',
                         (cid, employee_id, cycle_start.strftime('%Y-%m-%d'), now.strftime('%Y-%m-%d'))).fetchone()
    conn.close()
    return round(max(0.0, earned - float(taken['total'] or 0.0)), 2)

def calculate_family_leave_balance(employee_id, start_date_str, emp_type, emp_name, ref_date_str=None):
    if not start_date_str or emp_type in ['Supplier', 'Provider', 'Contract <25 Hrs']: return "N/A"
    start, now = datetime.strptime(start_date_str, '%Y-%m-%d'), datetime.strptime(ref_date_str, '%Y-%m-%d') if ref_date_str else datetime.now()
    conn = get_db_connection()
    emp_record = conn.execute('SELECT inactive_date FROM employees WHERE id=? AND company_id=?', (employee_id, session.get('company_id'))).fetchone()
    inactive_dt = parse_date_safe(emp_record['inactive_date']) if emp_record and 'inactive_date' in dict(emp_record) else None
    if inactive_dt and inactive_dt < now:
        now = inactive_dt
    months = (now.year - start.year) * 12 + (now.month - start.month)
    if months < 4:
        conn.close()
        return 0.0
    if emp_type == 'Contract >25 Hrs':
        valid = sum(1 for b in conn.execute("SELECT start FROM bookings WHERE employee LIKE ? AND company_id=?", (f"%{emp_name}%", session['company_id'])).fetchall() if start.date() <= datetime.strptime(b['start'][:10], '%Y-%m-%d').date() <= now.date())
        if (valid / max(1, (now.date() - start.date()).days / 7.0)) < 4.0: return 0.0
    cyc = start.replace(year=start.year + (months // 12))
    taken = conn.execute('SELECT SUM(days) as total FROM leave_records WHERE employee_id=? AND leave_type="Family Responsibility" AND date_taken >= ? AND date_taken <= ? AND company_id=?', (employee_id, cyc.strftime('%Y-%m-%d'), now.strftime('%Y-%m-%d'), session['company_id'])).fetchone()
    conn.close()
    return round(3.0 - (taken['total'] or 0.0), 2)

def calculate_financials(start_date, end_date):
    conn = get_db_connection()
    cid = session['company_id']
    comp = conn.execute("SELECT transport_policy, transport_amount_per_lift FROM companies WHERE id=?", (cid,)).fetchone()
    t_policy = comp['transport_policy'] if comp else 'standard'
    transport_amount_per_lift = float(dict(comp).get('transport_amount_per_lift') or 25) if comp else 25.0
    
    bookings = conn.execute("SELECT booking_type, transport, employee FROM bookings WHERE company_id=? AND substr(start, 1, 10) BETWEEN ? AND ?", (cid, start_date, end_date)).fetchall()
    services_db = conn.execute("SELECT * FROM services WHERE company_id=?", (cid,)).fetchall()
    employees_db = conn.execute("SELECT name, emp_type FROM employees WHERE company_id=?", (cid,)).fetchall()
    s_dict = { (s['name'] or '').strip(): {'price': float(s['client_price'] or 0), 'cost': float(s['company_cost'] or 0)} for s in services_db }
    e_dict = { (e['name'] or '').strip(): (e['emp_type'] or 'Full-time (5 Days)') for e in employees_db }
    
    rev, e_cost, p_cost = 0.0, 0.0, 0.0
    for b in bookings:
        b_rev, b_cost = sum(s_dict[t]['price'] for t in (b['booking_type'] or '').split(', ') if t in s_dict), sum(s_dict[t]['cost'] for t in (b['booking_type'] or '').split(', ') if t in s_dict)
        rev += b_rev
        staff = [s.strip() for s in (b['employee'] or '').split(', ') if s.strip()]
        
        t_allw = 0
        if t_policy in ['standard', 'yes']:
            t_val = b['transport'] or ''
            company_lifts = 0
            if 'Pickup' in t_val: company_lifts += 1
            if 'Drop Off' in t_val: company_lifts += 1
            t_allw = max(0, (2 - company_lifts) * transport_amount_per_lift)

        if not staff: e_cost += b_cost + t_allw
        else:
            for s in staff:
                if e_dict.get(s) in ['Provider', 'Supplier']: p_cost += (b_cost / len(staff))
                else: e_cost += (b_cost / len(staff)) + t_allw

    exp_db = conn.execute("SELECT SUM(amount) as total FROM expenses WHERE company_id=? AND date BETWEEN ? AND ?", (cid, start_date, end_date)).fetchone()
    s_cost = float(exp_db['total'] or 0.0) if exp_db else 0.0
    conn.close()
    return { "revenue": round(rev, 2), "employee_cost": round(e_cost, 2), "provider_cost": round(p_cost, 2), "supplier_cost": round(s_cost, 2), "profit": round(rev - e_cost - p_cost - s_cost, 2), "jobs": len(bookings) }

# ==========================================================
# SESSION TIMEOUT HELPERS
# ==========================================================
def _request_expects_json():
    path = request.path or ''
    if path.startswith('/api/'):
        return True
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return True
    accept = request.headers.get('Accept', '') or ''
    return 'application/json' in accept and 'text/html' not in accept


def _session_timeout_response(message='Your session expired after 15 minutes of inactivity. Please log in again.'):
    session.clear()
    if _request_expects_json():
        return jsonify({'status': 'timeout', 'message': message, 'redirect': SESSION_TIMEOUT_LOGIN_URL}), 401
    return redirect(SESSION_TIMEOUT_LOGIN_URL)


def _touch_session_activity():
    session['last_activity_at'] = time.time()
    session.modified = True


def _is_mobile_no_timeout_path(path=None):
    path = path or (request.path or '')
    if path == '/mobile' or path.startswith('/mobile/') or path.startswith('/api/mobile'):
        return True
    if path == '/staff/mobile' or path.startswith('/staff/mobile/'):
        return True
    if session.get('easyadmin_client_context') in ('admin_mobile', 'staff_mobile'):
        if path.startswith('/api/mobile') or path.startswith('/api/session/'):
            return True
        if session.get('easyadmin_client_context') == 'staff_mobile' and path.startswith('/api/staff/'):
            return True
        if path.startswith('/staff/download_attachment/') or path.startswith('/staff/download_payslip/'):
            return True
    return False


def _update_client_context_from_request():
    path = request.path or ''
    if path == '/mobile' or path.startswith('/mobile/'):
        session['easyadmin_client_context'] = 'admin_mobile'
        session.modified = True
    elif path == '/staff/mobile' or path.startswith('/staff/mobile/'):
        session['easyadmin_client_context'] = 'staff_mobile'
        session.modified = True
    elif request.method == 'GET' and 'text/html' in (request.headers.get('Accept', '') or ''):
        if not (path.startswith('/api/') or path.startswith('/static/') or path.startswith('/uploads/') or path == '/manifest.webmanifest'):
            session['easyadmin_client_context'] = 'desktop'
            session.modified = True


@app.after_request
def inject_session_timeout_script(response):
    if 'logged_in' not in session:
        return response
    if _is_mobile_no_timeout_path():
        return response
    if response.status_code != 200 or response.is_streamed or response.mimetype != 'text/html':
        return response
    try:
        body = response.get_data(as_text=True)
    except Exception:
        return response
    if '</body>' not in body or 'session-timeout.js' in body:
        return response
    script = ('\n<script defer src="/static/session-timeout.js?v=20260625-timeout15-desktop" '
              'data-easyadmin-session-timeout data-timeout-seconds="%s"></script>\n') % DESKTOP_SESSION_IDLE_TIMEOUT_SECONDS
    body = body.replace('</body>', script + '</body>', 1)
    response.set_data(body)
    response.headers['Content-Length'] = str(len(response.get_data()))
    return response


@app.after_request
def prevent_dynamic_response_caching(response):
    """Keep live application data fresh across desktop, mobile PWA and staff portal.

    The PWA service worker and some browsers can otherwise reuse old JSON/HTML
    responses after a company, employee, booking, payslip, leave request or
    system setting is changed. Static assets are still allowed to cache via
    versioned filenames, but database-backed screens and API responses must
    always be re-read from the server.
    """
    try:
        path = request.path or ''
        mimetype = (response.mimetype or '').lower()
    except Exception:
        return response

    if path.startswith('/static/'):
        return response

    dynamic_prefixes = (
        '/api/', '/admin/', '/staff', '/mobile', '/hub', '/booking', '/bookings',
        '/payroll', '/invoicing', '/finance', '/accounting', '/clients',
        '/employees', '/projects', '/quotes', '/invoices', '/settings',
        '/download', '/export', '/uploads/'
    )
    is_dynamic = (
        request.method in ('GET', 'POST', 'PUT', 'PATCH', 'DELETE') and (
            mimetype in ('application/json', 'text/html', 'text/csv', 'application/pdf')
            or path.startswith(dynamic_prefixes)
        )
    )

    if is_dynamic:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        response.headers.setdefault('Vary', 'Cookie')

    return response

# ==========================================================
# 0. GLOBAL SECURITY
# ==========================================================
@app.before_request
def restrict_access():
    public_endpoints = ['landing', 'login', 'forgot_password', 'static', 'manifest_webmanifest', 'service_worker', 'mobile_offline', 'health_check']
    if request.endpoint in public_endpoints:
        return
    if 'logged_in' not in session:
        if _request_expects_json():
            return jsonify({'status': 'timeout', 'message': 'Please log in again.', 'redirect': url_for('login')}), 401
        return redirect(url_for('login'))

    _update_client_context_from_request()

    if _is_mobile_no_timeout_path():
        _touch_session_activity()
    else:
        now_ts = time.time()
        last_activity = float(session.get('last_activity_at') or now_ts)
        if now_ts - last_activity > DESKTOP_SESSION_IDLE_TIMEOUT_SECONDS:
            return _session_timeout_response()
        _touch_session_activity()
    
    if not session.get('company_id') and not session.get('is_superadmin'):
        return "Fatal Error: Account is not assigned to a company.", 403

    path = request.path

    if session.get('is_staff'):
        allowed_staff_paths = ('/staff', '/staff/mobile', '/staff/download_attachment/', '/staff/download_payslip/', '/api/staff/', '/api/session/', '/logout')
        if path == '/mobile':
            return redirect(url_for('staff_mobile'))
        if path == '/hub':
            return redirect(url_for('staff_portal'))
        if path.startswith('/uploads/leave/'):
            return
        if not any(path == p or path.startswith(p) for p in allowed_staff_paths):
            return "Access Denied: Staff portal access only.", 403
        return

    if session.get('is_superadmin'): return

    if path.startswith('/admin/'):
        if not session.get('is_company_admin'):
            return "Access Denied: Admin privileges required.", 403
        if path in ['/admin/companies', '/admin/companies/save', '/admin/switch_company', '/admin/tax_config', '/admin/holidays']:
            return "Access Denied: Superadmin privileges required.", 403
        return 

    if path == '/staff_admin' or path.startswith('/api/staff/admin'):
        if not (session.get('is_company_admin') or session.get('can_payroll')):
            return "Access Denied: Staff Portal administration requires Company Admin or HR & Payroll access.", 403
        return
        
    if path == '/update_client':
        if not (session.get('can_booking') or session.get('can_invoicing')):
            return "Access Denied: You do not have permissions to manage clients.", 403
    if path == '/booking' or path in ['/bookings', '/add', '/edit_booking', '/delete_booking', '/client_report', '/daily_route', '/export', '/generate_recurring', '/booking_staff_hours', '/export_bookings_range', '/api/booking_ops_report', '/export_booking_ops_report'] or path.startswith('/api/projects') or path.startswith('/api/attachments') or path.startswith('/download_attachment'):
        if not session.get('can_booking'): return "Access Denied: You do not have permissions to access Booking & Operations.", 403
    if path in ['/update_service', '/delete_service', '/api/services']:
        if not (session.get('can_finance') or session.get('can_invoicing')):
            return "Access Denied: You do not have permissions to manage service margins.", 403
    if path == '/finance' or path in ['/log_expense', '/finance_report', '/update_vendor']:
        if not session.get('can_finance'): return "Access Denied: You do not have permissions to access Finance.", 403
    if path == '/accounting' or path.startswith('/api/accounting'):
        if not session.get('can_accounting'): return "Access Denied: You do not have permissions to access Accounting.", 403
    if path == '/payroll' or path in ['/update_employee', '/generate_payslip', '/api/save_payslip', '/generate_irp5', '/generate_emp201', '/api/ui19_settings', '/api/ui19_data', '/export_ui19', '/email_payslip', '/record_leave', '/update_leave', '/delete_leave', '/generate_report', '/save_interview', '/delete_interview'] or path.startswith('/export_emp501') or path.startswith('/export_payroll_bank_file') or path.startswith('/uploads/'):
        if not session.get('can_payroll'): return "Access Denied: You do not have permissions to access Payroll & HR.", 403
    if path == '/invoicing' or path.startswith('/api/uninvoiced') or path.startswith('/api/save_invoice') or path.startswith('/api/invoice') or path.startswith('/api/save_quote') or path.startswith('/api/quote') or path.startswith('/download/invoice') or path.startswith('/download/quote') or path.startswith('/api/email_invoice_pdf') or path.startswith('/api/email_quote_pdf') or path == '/api/save_invoice_settings' or path == '/api/email_document' or path == '/api/client_statement' or path.startswith('/api/credit_note'):
        if not session.get('can_invoicing'): return "Access Denied: You do not have permissions to access Invoicing & Quotes.", 403
    
    if path == '/settings' or path == '/api/test_email_connection' or path == '/api/test_google_calendar_connection':
        if not session.get('is_company_admin'):
            return "Access Denied: Only Company Admins can access Email & Calendar Settings from the Hub.", 403

@app.route('/api/session/ping', methods=['POST'])
def api_session_ping():
    _touch_session_activity()
    return jsonify({'status': 'success', 'timeout_seconds': DESKTOP_SESSION_IDLE_TIMEOUT_SECONDS, 'mobile_timeout_disabled': _is_mobile_no_timeout_path()})


@app.route('/api/session/timeout', methods=['POST'])
def api_session_timeout():
    username = session.get('username')
    if username:
        try:
            log_action('System', 'Auto Logout', f'User {username} was logged out after inactivity.')
        except Exception:
            pass
    session.clear()
    return jsonify({'status': 'success', 'redirect': url_for('login', timeout=1)})


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = 'Your session expired after 15 minutes of inactivity. Please log in again.' if request.args.get('timeout') else None
    if request.method == 'POST':
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE username = ?', (request.form['username'],)).fetchone()
        
        if user and check_password_hash(user['password_hash'], request.form['password']):
            session.clear()
            session.permanent = True
            session['logged_in'] = True
            session['username'] = user['username']
            _touch_session_activity()
            session['is_superadmin'] = bool(user['is_superadmin'])
            session['is_company_admin'] = bool(dict(user).get('is_company_admin', 0))
            session['is_staff'] = bool(dict(user).get('is_staff', 0))
            session['employee_id'] = dict(user).get('employee_id')
            session['can_booking'] = bool(user['can_booking'])
            session['can_finance'] = bool(user['can_finance'])
            session['can_payroll'] = bool(user['can_payroll'])
            session['can_invoicing'] = bool(dict(user).get('can_invoicing', 0))
            session['can_accounting'] = bool(dict(user).get('can_accounting', 0))
            
            if user['is_superadmin']:
                comp = conn.execute('SELECT * FROM companies ORDER BY id ASC LIMIT 1').fetchone()
            else:
                comp = conn.execute('SELECT * FROM companies WHERE id = ?', (user['company_id'],)).fetchone()
                
            session['company_id'] = comp['id'] if comp else user['company_id']
            session['company_name'] = comp['name'] if comp else "Unknown Tenant"
            session['company_logo'] = comp['logo_file'] if comp else None
            
            session['comp_can_booking'] = bool(dict(comp).get('can_booking', 0)) if comp else False
            session['comp_can_finance'] = bool(dict(comp).get('can_finance', 0)) if comp else False
            session['comp_can_payroll'] = bool(dict(comp).get('can_payroll', 0)) if comp else False
            session['comp_can_invoicing'] = bool(dict(comp).get('can_invoicing', 0)) if comp else False
            session['comp_can_accounting'] = bool(dict(comp).get('can_accounting', 0)) if comp else False
            session['comp_google_calendar'] = bool(dict(comp).get('google_calendar_sync', 0)) if comp else False
                
            conn.close()
            log_action('System', 'Login', f"User {user['username']} logged in.")
            if session.get('is_staff'):
                return redirect(url_for('staff_portal'))
            return redirect(url_for('hub'))
        else: error = 'Invalid credentials.'
        conn.close()
    return render_template('login.html', error=error)

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    success_msg = None
    error_msg = None
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        email = _normalise_email(request.form.get('email'))
        generic_msg = 'If the username and email address match an account, a new temporary password will be emailed.'
        if not username or not email:
            error_msg = 'Please enter both your username and email address.'
        else:
            conn = get_db_connection()
            try:
                user = conn.execute('SELECT * FROM users WHERE LOWER(username)=LOWER(?)', (username,)).fetchone()
                if user:
                    account_email = _normalise_email(_account_email_for_password_reset(conn, user))
                    if account_email and account_email == email:
                        temp_password = generate_temporary_password()
                        body = (
                            'Good day,\n\n'
                            'A password reset was requested for your Easy Admin account.\n\n'
                            f'Username: {user["username"]}\n'
                            f'Temporary password: {temp_password}\n\n'
                            'Please log in and ask your administrator to change this password if required. '
                            'If you did not request this reset, please contact your administrator immediately.\n\n'
                            'Easy Admin System Email'
                        )
                        _send_system_email(account_email, 'Easy Admin password reset', body)
                        conn.execute('UPDATE users SET password_hash=? WHERE id=?', (generate_password_hash(temp_password), user['id']))
                        conn.commit()
                success_msg = generic_msg
            except Exception as exc:
                error_msg = f'Password reset could not be completed: {exc}'
            finally:
                conn.close()
    return render_template('forgot_password.html', success_msg=success_msg, error_msg=error_msg)

@app.route('/logout')
def logout(): 
    if 'username' in session:
        log_action('System', 'Logout', f"User {session['username']} logged out.")
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def landing():
    return render_template('landing.html')

@app.route('/hub')
def hub(): return render_template('hub.html', session=session)


# ==========================================================
# STAFF PORTAL ROUTES
# ==========================================================
STAFF_LEAVE_TYPES = ['Annual Leave', 'Sick Leave', 'Family Responsibility', 'Unpaid Leave', 'Other Leave']
STAFF_LEAVE_STATUSES = {'Pending', 'Approved', 'Declined'}


def _staff_is_admin():
    return bool(session.get('is_superadmin') or session.get('is_company_admin') or session.get('can_payroll'))


def _staff_is_user():
    return bool(session.get('is_staff') and session.get('employee_id'))


def _staff_json_error(message, status_code=400):
    return jsonify({'status': 'error', 'message': message}), status_code


def _staff_employee_row(conn):
    employee_id = session.get('employee_id')
    cid = session.get('company_id')
    if not employee_id or not cid:
        return None
    return conn.execute('SELECT * FROM employees WHERE id=? AND company_id=?', (employee_id, cid)).fetchone()


def _staff_date_range(default_days=30):
    today = datetime.now().strftime('%Y-%m-%d')
    start_date = (request.args.get('start_date') or today).strip()
    end_date = (request.args.get('end_date') or (datetime.now() + timedelta(days=default_days)).strftime('%Y-%m-%d')).strip()
    try:
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')
    except Exception:
        start_date = today
        end_date = (datetime.now() + timedelta(days=default_days)).strftime('%Y-%m-%d')
    return start_date, end_date


def _staff_leave_days(start_date, end_date):
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
        if end < start:
            return 0
        return float((end - start).days + 1)
    except Exception:
        return 0


def _staff_save_leave_attachment(file_obj):
    if not file_obj or not file_obj.filename:
        return None
    filename = secure_filename(file_obj.filename)
    if not filename:
        return None
    base, ext = os.path.splitext(filename)
    stored = f"staff_leave_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{secure_filename(base)[:40]}{ext.lower()}"
    os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'leave'), exist_ok=True)
    file_obj.save(os.path.join(app.config['UPLOAD_FOLDER'], 'leave', stored))
    return stored


def _staff_booking_rows(conn, cid, employee_name, start_date, end_date, limit=300):
    if not _mobile_table_exists(conn, 'bookings'):
        return []
    booking_cols, selects, joins = _mobile_booking_select_parts(conn, include_client_contact=True)
    if 'start' not in booking_cols or 'employee' not in booking_cols:
        return []
    params = []
    where_parts = [_mobile_company_where(booking_cols, 'b', cid, params)]
    where_parts.append("substr(COALESCE(CAST(b.start AS TEXT), ''), 1, 10) BETWEEN ? AND ?")
    params.extend([start_date, end_date])
    where_parts.append("LOWER(COALESCE(CAST(b.employee AS TEXT), '')) LIKE ?")
    params.append(f"%{str(employee_name or '').strip().lower()}%")
    sql = f'''SELECT {', '.join(selects)}
              FROM bookings b
              {' '.join(joins)}
              WHERE {' AND '.join(where_parts)}
              ORDER BY b.start ASC, b.id ASC
              LIMIT {int(limit)}'''
    return conn.execute(sql, params).fetchall()


def _staff_booking_is_assigned(employee_field, employee_name):
    target = str(employee_name or '').strip().lower()
    if not target:
        return False
    raw = str(employee_field or '').strip().lower()
    if not raw:
        return False
    parts = [p.strip() for chunk in raw.replace(';', ',').replace('|', ',').split(',') for p in chunk.split(' and ')]
    return target in parts or target in raw


def _staff_booking_detail_for_current_user(conn, cid, employee_name, booking_id):
    row = _mobile_booking_detail_row(conn, cid, booking_id)
    if not row:
        return None
    if not _staff_booking_is_assigned(dict(row).get('employee'), employee_name):
        return None
    return row


def _staff_update_booking_status(conn, cid, employee, booking_id, status):
    if status not in MOBILE_STATUS_VALUES:
        raise ValueError('Invalid job status.')
    row = _staff_booking_detail_for_current_user(conn, cid, employee['name'], booking_id)
    if not row:
        return None
    for sql in [
        'ALTER TABLE bookings ADD COLUMN mobile_status TEXT DEFAULT "Scheduled"',
        'ALTER TABLE bookings ADD COLUMN mobile_status_updated_at TEXT',
        'ALTER TABLE bookings ADD COLUMN mobile_status_updated_by TEXT',
        'ALTER TABLE bookings ADD COLUMN mobile_started_at TEXT',
        'ALTER TABLE bookings ADD COLUMN mobile_completed_at TEXT'
    ]:
        try:
            conn.execute(sql)
        except Exception:
            pass
    booking_cols = _mobile_columns(conn, 'bookings')
    params = []
    where_parts = [_mobile_company_where(booking_cols, 'bookings', cid, params), 'id=?']
    params.append(booking_id)
    update_cols = ['mobile_status=?']
    update_params = [status]
    updated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if 'mobile_status_updated_at' in booking_cols:
        update_cols.append('mobile_status_updated_at=?')
        update_params.append(updated_at)
    if 'mobile_status_updated_by' in booking_cols:
        update_cols.append('mobile_status_updated_by=?')
        update_params.append(session.get('username') or employee['name'] or 'Staff')
    if status == 'In Progress' and 'mobile_started_at' in booking_cols:
        update_cols.append('mobile_started_at=COALESCE(NULLIF(mobile_started_at, ?), ?)')
        update_params.extend(['', updated_at])
    if status == 'Completed' and 'mobile_completed_at' in booking_cols:
        update_cols.append('mobile_completed_at=?')
        update_params.append(updated_at)
    conn.execute(f"UPDATE bookings SET {', '.join(update_cols)} WHERE {' AND '.join(where_parts)}", update_params + params)
    return updated_at


def _staff_append_booking_note(conn, cid, employee, booking_id, note):
    row = _staff_booking_detail_for_current_user(conn, cid, employee['name'], booking_id)
    if not row:
        return None
    note = str(note or '').strip()
    if not note:
        raise ValueError('Please enter a booking note.')
    try:
        conn.execute('ALTER TABLE bookings ADD COLUMN booking_notes TEXT')
    except Exception:
        pass
    existing = dict(row).get('booking_notes') or ''
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    author = employee['name'] or session.get('username') or 'Staff'
    entry = f"[{stamp}] {author}: {note}"
    combined = (str(existing).rstrip() + '\n' + entry).strip() if existing else entry
    booking_cols = _mobile_columns(conn, 'bookings')
    params = []
    where_parts = [_mobile_company_where(booking_cols, 'bookings', cid, params), 'id=?']
    params.append(booking_id)
    conn.execute(f"UPDATE bookings SET booking_notes=? WHERE {' AND '.join(where_parts)}", [combined] + params)
    return combined


def _staff_save_booking_attachments(conn, cid, employee, booking_id, files):
    row = _staff_booking_detail_for_current_user(conn, cid, employee['name'], booking_id)
    if not row:
        return None
    files = files or []
    saved = []
    for f in files:
        if not f or not getattr(f, 'filename', ''):
            continue
        if not is_allowed_attachment(f.filename):
            raise ValueError(f'File type not allowed: {f.filename}')
        f.stream.seek(0, os.SEEK_END)
        size = f.stream.tell()
        f.stream.seek(0)
        if size > MAX_ATTACHMENT_SIZE:
            raise ValueError(f'File too large: {f.filename}. Maximum size is 20MB.')
        safe_original = secure_filename(f.filename) or 'attachment'
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
        stored_filename = f'{timestamp}_{safe_original}'
        folder = os.path.join(app.config['UPLOAD_FOLDER'], 'attachments', f'company_{cid}', 'bookings', f'booking_{booking_id}')
        os.makedirs(folder, exist_ok=True)
        target_path = os.path.join(folder, stored_filename)
        f.save(target_path)
        rel_path = os.path.relpath(target_path, app.config['UPLOAD_FOLDER'])
        cur = conn.execute('''INSERT INTO attachments
                              (company_id, linked_type, linked_id, original_filename, stored_filename, file_path, file_size, mime_type, uploaded_by)
                              VALUES (?, 'booking', ?, ?, ?, ?, ?, ?, ?)''',
                           (cid, booking_id, f.filename, stored_filename, rel_path, size, f.mimetype or '', employee['name'] or session.get('username', 'Staff')))
        saved.append(cur.lastrowid)
    return saved


def _staff_attachment_to_dict(row):
    data = attachment_to_dict(row)
    data['download_url'] = url_for('staff_download_attachment', attachment_id=row['id'])
    return data


def _staff_leave_request_dict(row, employee_name=''):
    r = dict(row)
    return {
        'id': r.get('id'),
        'employee_id': r.get('employee_id'),
        'employee_name': r.get('employee_name') or employee_name,
        'leave_type': r.get('leave_type') or 'Annual Leave',
        'start_date': format_display_date(r.get('start_date')),
        'end_date': format_display_date(r.get('end_date')),
        'days': float(r.get('days') or 0),
        'reason': r.get('reason') or '',
        'status': r.get('status') or 'Pending',
        'attachment_file': r.get('attachment_file') or '',
        'requested_at': format_display_date(r.get('requested_at')),
        'reviewed_by': r.get('reviewed_by') or '',
        'reviewed_at': format_display_date(r.get('reviewed_at')),
        'admin_note': r.get('admin_note') or '',
        'leave_record_id': r.get('leave_record_id') or ''
    }


def _staff_leave_balances(employee):
    balances = {'annual': '', 'sick': '', 'family': ''}
    try:
        balances['annual'] = round(float(calculate_leave_balance(employee['id'], employee['start_date'] or datetime.now().strftime('%Y-%m-%d'), employee['emp_type'] or 'Full-time (5 Days)', employee['name'] or '')), 2)
    except Exception:
        balances['annual'] = 'N/A'
    try:
        balances['sick'] = round(float(calculate_sick_leave_balance(employee['id'], employee['start_date'] or datetime.now().strftime('%Y-%m-%d'), employee['emp_type'] or 'Full-time (5 Days)', employee['name'] or '')), 2)
    except Exception:
        balances['sick'] = 'N/A'
    try:
        balances['family'] = round(float(calculate_family_leave_balance(employee['id'], employee['start_date'] or datetime.now().strftime('%Y-%m-%d'), employee['emp_type'] or 'Full-time (5 Days)', employee['name'] or '')), 2)
    except Exception:
        balances['family'] = 'N/A'
    return balances



def _staff_payslip_to_dict(row):
    r = dict(row)
    payslip_date = format_display_date(r.get('date') or '')
    payslip_type = (r.get('payslip_type') or 'regular').strip() or 'regular'
    is_adjustment = payslip_type.lower() == 'adjustment'
    title = 'Adjustment Payslip' if is_adjustment else 'Final Payslip'
    return {
        'id': r.get('id'),
        'employee_id': r.get('employee_id'),
        'date': payslip_date,
        'month': (payslip_date[:7] if payslip_date else ''),
        'payslip_type': payslip_type,
        'title': title,
        'adjustment_reason': r.get('adjustment_reason') or '',
        'gross_salary': float(r.get('gross_salary') or 0),
        'overtime': float(r.get('overtime') or 0),
        'bonus': float(r.get('bonus') or 0),
        'transport': float(r.get('transport') or 0),
        'reimbursable_expenses': float(r.get('reimbursable_expenses') or 0),
        'uif': float(r.get('uif') or 0),
        'paye': float(r.get('paye') or 0),
        'loan_repayment': float(r.get('loan_repayment') or 0),
        'net_salary': float(r.get('net_salary') or 0),
        'created_at': format_display_date(r.get('created_at') or ''),
        'download_url': url_for('staff_download_payslip', payslip_id=r.get('id'))
    }



def _staff_payslip_pdf_money(value):
    """Money format used on the HR & Payroll payslip PDF layout."""
    return _money(value)


def _build_staff_payslip_pdf(row, employee, company):
    """Create a staff downloadable payslip PDF matching the HR & Payroll payslip layout."""
    r = dict(row)
    emp = dict(employee)
    comp = dict(company) if company else {}

    page_w, page_h = 595.28, 841.89
    margin = 34.0
    content_x = 46.0
    dark = (0.12, 0.13, 0.16)
    muted = (0.42, 0.45, 0.48)
    line_col = (0.84, 0.85, 0.86)
    light = (0.965, 0.975, 0.985)
    green = (0.04, 0.52, 0.31)
    blue = (0.05, 0.43, 0.88)
    white = (1, 1, 1)

    cmds = []
    image_resources = {}

    def cmd(line):
        cmds.append(line)

    def color(c, op='rg'):
        return f"{c[0]:.3f} {c[1]:.3f} {c[2]:.3f} {op}"

    def text_width(value, size=9, bold=False):
        value = str(value or '')
        total = 0.0
        for ch in value:
            if ch.isdigit():
                total += 0.556
            elif ch in ',.':
                total += 0.278 if not bold else 0.333
            elif ch in ' -/:()':
                total += 0.278 if not bold else 0.333
            elif ch in 'ilI|!':
                total += 0.240 if not bold else 0.300
            elif ch in 'mwMW':
                total += 0.800 if not bold else 0.900
            else:
                total += 0.520 if not bold else 0.600
        return total * float(size)

    def text(x, y, value, size=9, bold=False, c=dark, align='left', italic=False):
        value = str(value or '').replace('\u00a0', ' ')
        tw = text_width(value, size, bold)
        if align == 'right':
            x -= tw
        elif align == 'center':
            x -= tw / 2
        # Base PDF fonts do not include Helvetica-Oblique in this lightweight writer.
        font = 'F2' if bold else 'F1'
        cmd(f"{color(c)} BT /{font} {float(size):.2f} Tf {float(x):.2f} {float(y):.2f} Td ({_pdf_text_escape(value)}) Tj ET")

    def rect(x, y, w, h, stroke=None, fill=None, width=0.55):
        if fill is not None:
            cmd(f"{color(fill)} {float(x):.2f} {float(y):.2f} {float(w):.2f} {float(h):.2f} re f")
        if stroke is not None:
            cmd(f"{float(width):.2f} w {color(stroke, 'RG')} {float(x):.2f} {float(y):.2f} {float(w):.2f} {float(h):.2f} re S")

    def line(x1, y1, x2, y2, stroke=line_col, width=0.55):
        cmd(f"{float(width):.2f} w {color(stroke, 'RG')} {float(x1):.2f} {float(y1):.2f} m {float(x2):.2f} {float(y2):.2f} l S")

    def draw_wrapped(value, x, y, max_chars=42, size=10, c=muted, leading=14, max_lines=4, bold=False):
        value = str(value or '').replace('\r', ' ').replace('\n', ' ')
        words = value.split()
        lines = []
        current = ''
        for word in words:
            trial = (current + ' ' + word).strip()
            if len(trial) <= max_chars:
                current = trial
            else:
                if current:
                    lines.append(current)
                current = word
                if len(lines) >= max_lines:
                    break
        if current and len(lines) < max_lines:
            lines.append(current)
        for i, ln in enumerate(lines[:max_lines]):
            text(x, y - i * leading, ln, size, bold, c)
        return y - len(lines[:max_lines]) * leading

    def draw_image(name, x, y, w, h):
        cmd(f"q {float(w):.2f} 0 0 {float(h):.2f} {float(x):.2f} {float(y):.2f} cm /{name} Do Q")

    # Outer document frame similar to the HR & Payroll payslip download.
    rect(12, 18, page_w - 24, page_h - 36, stroke=(0.20, 0.23, 0.27), fill=None, width=0.8)

    company_name = comp.get('name') or session.get('company_name') or 'Company'
    company_addr = comp.get('address') or ''
    company_reg = comp.get('registration_number') or ''
    period = (format_display_date(r.get('date') or '') or str(r.get('date') or ''))[:7]
    payslip_type = (r.get('payslip_type') or 'regular').strip().lower()
    doc_title = 'ADJUSTMENT PAYSLIP' if payslip_type == 'adjustment' else 'PAYSLIP'

    logo_resource_name = None
    try:
        logo_file = comp.get('logo_file') or ''
        if logo_file:
            logo_path = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), 'logos', os.path.basename(str(logo_file)))
            logo_info = _get_cached_pdf_image(logo_path) if os.path.exists(logo_path) else None
            if logo_info:
                logo_resource_name = 'ImLogo'
                image_resources[logo_resource_name] = logo_info
    except Exception:
        logo_resource_name = None

    top_y = page_h - 72
    if logo_resource_name:
        logo = image_resources[logo_resource_name]
        max_w, max_h = 120.0, 56.0
        ratio = min(max_w / max(1, float(logo.get('width', 1))), max_h / max(1, float(logo.get('height', 1))))
        logo_w = float(logo.get('width', 1)) * ratio
        logo_h = float(logo.get('height', 1)) * ratio
        draw_image(logo_resource_name, content_x, top_y - logo_h + 4, logo_w, logo_h)
        company_y = top_y - logo_h - 22
    else:
        company_y = top_y

    text(content_x, company_y, company_name, 22, True, dark)
    y = company_y - 22
    if company_reg:
        text(content_x, y, f"Reg No: {company_reg}", 10, False, muted)
        y -= 14
    if company_addr:
        y = draw_wrapped(company_addr, content_x, y, 38, 11, muted, 16, 4)

    text(page_w - content_x, page_h - 83, doc_title, 24, True, green, 'right')
    info_y = page_h - 142
    text(page_w - content_x - 85, info_y, 'Date:', 12, True, dark, 'right')
    text(page_w - content_x, info_y, period, 12, False, dark, 'right')
    info_y -= 18
    text(page_w - content_x - 113, info_y, 'Employee:', 12, True, dark, 'right')
    text(page_w - content_x, info_y, emp.get('name') or '', 12, False, dark, 'right')
    info_y -= 17
    if emp.get('emp_number'):
        text(page_w - content_x, info_y, f"({emp.get('emp_number')})", 12, False, dark, 'right')
        info_y -= 17
    text(page_w - content_x - 70, info_y, 'ID:', 12, True, dark, 'right')
    text(page_w - content_x, info_y, emp.get('id_passport') or '', 12, False, dark, 'right')

    line(content_x - 10, page_h - 238, page_w - content_x + 10, page_h - 238, line_col, 0.7)

    table_x = content_x
    table_y_top = page_h - 268
    row_h = 34.0
    table_w = page_w - (content_x * 2)
    desc_w = table_w * 0.66
    earn_w = table_w * 0.155
    ded_w = table_w - desc_w - earn_w
    x_desc = table_x
    x_earn = table_x + desc_w
    x_ded = x_earn + earn_w

    # Header and table frame.
    rect(table_x, table_y_top - row_h, table_w, row_h, stroke=None, fill=white)
    text(table_x + 7, table_y_top - 24, 'Description', 13, True, dark)
    text(x_earn + earn_w - 8, table_y_top - 24, 'Earnings', 13, True, dark, 'right')
    text(x_ded + ded_w - 8, table_y_top - 24, 'Deductions', 13, True, dark, 'right')
    line(x_earn, table_y_top, x_earn, table_y_top - row_h * 8, line_col, 0.9)
    line(x_ded, table_y_top, x_ded, table_y_top - row_h * 8, line_col, 0.9)
    line(table_x, table_y_top - row_h, table_x + table_w, table_y_top - row_h, (0.90,0.91,0.92), 0.4)

    def money_amount(value):
        return _staff_payslip_pdf_money(value)

    table_rows = [
        {'label': 'Calculated Gross', 'code': 'Code: 3601', 'earn': r.get('gross_salary') or 0, 'ded': None, 'show': True, 'muted': ''},
        {'label': 'Overtime / Double Rate', 'code': 'Code: 3605', 'earn': r.get('overtime') or 0, 'ded': None, 'show': abs(float(r.get('overtime') or 0)) > 0.004, 'muted': ''},
        {'label': 'Reimbursable Expenses (Non-taxable)', 'code': '', 'earn': r.get('reimbursable_expenses') or 0, 'ded': None, 'show': abs(float(r.get('reimbursable_expenses') or 0)) > 0.004, 'green': True},
        {'label': 'Bonus', 'code': 'Code: 3605/3601', 'earn': r.get('bonus') or 0, 'ded': None, 'show': abs(float(r.get('bonus') or 0)) > 0.004},
        {'label': 'Transport Reimbursement (Tax Free)', 'code': 'Code: 3702', 'earn': r.get('transport') or 0, 'ded': None, 'show': abs(float(r.get('transport') or 0)) > 0.004, 'green': True},
        {'label': 'PAYE Tax', 'code': 'Code: 4102', 'earn': None, 'ded': r.get('paye') or 0, 'show': True},
        {'label': 'UIF (Employee 1%)', 'code': 'Code: 4141', 'earn': None, 'ded': r.get('uif') or 0, 'show': True},
        {'label': 'Loan Repayment', 'code': '', 'earn': None, 'ded': r.get('loan_repayment') or 0, 'show': abs(float(r.get('loan_repayment') or 0)) > 0.004},
        {'label': 'UIF (Employer 1%)', 'code': '', 'earn': r.get('uif') or 0, 'ded': None, 'show': True, 'italic': True},
    ]
    visible_rows = [rr for rr in table_rows if rr.get('show')]
    y = table_y_top - row_h
    for rr in visible_rows:
        rect(table_x, y - row_h, table_w, row_h, stroke=(0.94,0.94,0.94), fill=white, width=0.35)
        line(x_earn, y, x_earn, y - row_h, line_col, 0.65)
        line(x_ded, y, x_ded, y - row_h, line_col, 0.65)
        label_col = muted if rr.get('italic') else dark
        text(table_x + 7, y - 22, rr['label'], 12, False, label_col)
        if rr.get('code'):
            text(x_earn - 22, y - 22, rr['code'], 9, False, muted, 'right')
        amount_col = green if rr.get('green') else dark
        if rr.get('earn') is not None:
            text(x_earn + earn_w - 8, y - 22, money_amount(rr.get('earn')), 12, False, amount_col, 'right')
        if rr.get('ded') is not None:
            text(x_ded + ded_w - 8, y - 22, money_amount(rr.get('ded')), 12, False, dark, 'right')
        y -= row_h

    # Net pay row.
    net_h = 39.0
    line(table_x, y, table_x + table_w, y, (0.10,0.10,0.10), 1.0)
    rect(table_x, y - net_h, table_w, net_h, stroke=(0.10,0.10,0.10), fill=white, width=0.8)
    line(x_earn, y, x_earn, y - net_h, (0.10,0.10,0.10), 1.0)
    text(table_x + 7, y - 25, 'NET PAY', 13, True, dark)
    text(x_ded + ded_w - 8, y - 25, money_amount(r.get('net_salary') or 0), 16, True, green, 'right')
    y -= net_h + 38

    if r.get('adjustment_reason'):
        rect(content_x, y - 46, table_w, 46, stroke=line_col, fill=light, width=0.5)
        text(content_x + 8, y - 16, 'Adjustment reason', 10, True, dark)
        draw_wrapped(r.get('adjustment_reason'), content_x + 8, y - 31, 84, 9, muted, 11, 2)
        y -= 62

    # Leave balances panel.
    panel_h = 112
    rect(content_x, y - panel_h, table_w, panel_h, stroke=(0.86,0.88,0.91), fill=light, width=0.55)
    text(content_x + 13, y - 23, 'Statutory Leave Balances', 13, False, blue)
    line(content_x + 13, y - 34, content_x + table_w - 13, y - 34, (0.82,0.84,0.87), 0.5)
    try:
        leave_ref = r.get('date') or datetime.now().strftime('%Y-%m-%d')
        annual = calculate_leave_balance(emp.get('id'), emp.get('start_date') or datetime.now().strftime('%Y-%m-%d'), emp.get('emp_type') or 'Full-time (5 Days)', emp.get('name') or '', leave_ref)
        sick = calculate_sick_leave_balance(emp.get('id'), emp.get('start_date') or datetime.now().strftime('%Y-%m-%d'), emp.get('emp_type') or 'Full-time (5 Days)', emp.get('name') or '', leave_ref)
        family = calculate_family_leave_balance(emp.get('id'), emp.get('start_date') or datetime.now().strftime('%Y-%m-%d'), emp.get('emp_type') or 'Full-time (5 Days)', emp.get('name') or '', leave_ref)
    except Exception:
        annual, sick, family = 'N/A', 'N/A', 'N/A'

    def leave_val(v):
        try:
            f = float(v)
            return f"{f:g}"
        except Exception:
            return str(v)

    line_y = y - 55
    text(content_x + 13, line_y, 'Annual Leave Balance:', 10, False, dark)
    text(content_x + 133, line_y, f"{leave_val(annual)} Days", 10, True, dark)
    line_y -= 18
    text(content_x + 13, line_y, 'Sick Leave Balance:', 10, False, dark)
    text(content_x + 128, line_y, f"{leave_val(sick)} Days left in 36-month cycle", 10, True, dark)
    line_y -= 18
    text(content_x + 13, line_y, 'Family Responsibility Leave:', 10, False, dark)
    text(content_x + 166, line_y, f"{leave_val(family)} Days left in annual cycle", 10, True, dark)

    if image_resources:
        pdf_bytes = _build_raw_pdf(['\n'.join(cmds)], page_w, page_h, image_resources)
    else:
        pdf_bytes = _build_raw_pdf(['\n'.join(cmds)], page_w, page_h, {})
    return io.BytesIO(pdf_bytes)


@app.route('/staff')
def staff_portal():
    if not _staff_is_user():
        return redirect(url_for('hub'))
    return render_template('staff_portal.html', session=session)


@app.route('/staff/mobile')
def staff_mobile():
    if not _staff_is_user():
        return redirect(url_for('hub'))
    return render_template('staff_portal.html', session=session, mobile_only=True)


@app.route('/staff_admin')
def staff_admin_page():
    if not _staff_is_admin():
        return "Forbidden", 403
    return render_template('staff_admin.html', session=session)


@app.route('/api/staff/dashboard')
def api_staff_dashboard():
    if not _staff_is_user():
        return _staff_json_error('Staff portal access is required.', 403)
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return _staff_json_error('Your staff account is not linked to an active employee record.', 404)
        today = datetime.now().strftime('%Y-%m-%d')
        future = (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d')
        today_rows = _staff_booking_rows(conn, cid, employee['name'], today, today)
        upcoming_rows = _staff_booking_rows(conn, cid, employee['name'], today, future)
        leave_rows = conn.execute('''SELECT * FROM staff_leave_requests
                                     WHERE company_id=? AND employee_id=?
                                     ORDER BY requested_at DESC, id DESC LIMIT 20''',
                                  (cid, employee['id'])).fetchall()
        return jsonify({
            'status': 'success',
            'employee': {
                'id': employee['id'], 'name': employee['name'] or '', 'job_title': employee['job_title'] or '',
                'emp_number': employee['emp_number'] or '', 'status': employee['status'] or '',
                'start_date': format_display_date(employee['start_date'] or '')
            },
            'balances': _staff_leave_balances(employee),
            'today_bookings': [_mobile_booking_to_dict(r) for r in today_rows],
            'upcoming_bookings': [_mobile_booking_to_dict(r, include_notes=False) for r in upcoming_rows[:20]],
            'leave_requests': [_staff_leave_request_dict(r, employee['name']) for r in leave_rows],
            'leave_types': STAFF_LEAVE_TYPES
        })
    except Exception as exc:
        return _staff_json_error(f'Staff dashboard could not be loaded: {exc}', 500)
    finally:
        conn.close()


@app.route('/api/staff/bookings')
def api_staff_bookings():
    if not _staff_is_user():
        return _staff_json_error('Staff portal access is required.', 403)
    start_date, end_date = _staff_date_range(default_days=30)
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return _staff_json_error('Your staff account is not linked to an active employee record.', 404)
        rows = _staff_booking_rows(conn, cid, employee['name'], start_date, end_date)
        return jsonify({'status': 'success', 'bookings': [_mobile_booking_to_dict(r) for r in rows], 'start_date': start_date, 'end_date': end_date})
    except Exception as exc:
        return _staff_json_error(f'Staff bookings could not be loaded: {exc}', 500)
    finally:
        conn.close()


@app.route('/api/staff/bookings/<int:booking_id>')
def api_staff_booking_detail(booking_id):
    if not _staff_is_user():
        return _staff_json_error('Staff portal access is required.', 403)
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return _staff_json_error('Your staff account is not linked to an active employee record.', 404)
        row = _staff_booking_detail_for_current_user(conn, cid, employee['name'], booking_id)
        if not row:
            return _staff_json_error('Booking not found for your staff profile.', 404)
        item = _mobile_booking_to_dict(row)
        item['client_phone'] = dict(row).get('client_phone') or ''
        item['client_email'] = dict(row).get('client_email') or ''
        item['client_address'] = dict(row).get('client_address') or ''
        attachments = _mobile_attachment_rows(conn, cid, 'booking', booking_id)
        item['attachments'] = [_staff_attachment_to_dict(a) for a in attachments]
        return jsonify({'status': 'success', 'booking': item})
    except Exception as exc:
        return _staff_json_error(f'Booking details could not be loaded: {exc}', 500)
    finally:
        conn.close()


@app.route('/api/staff/bookings/<int:booking_id>/start', methods=['POST'])
def api_staff_start_booking(booking_id):
    if not _staff_is_user():
        return _staff_json_error('Staff portal access is required.', 403)
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return _staff_json_error('Your staff account is not linked to an active employee record.', 404)
        updated_at = _staff_update_booking_status(conn, cid, employee, booking_id, 'In Progress')
        if not updated_at:
            return _staff_json_error('Booking not found for your staff profile.', 404)
        conn.commit()
        log_action('Staff Portal', 'Started Job', f"{employee['name']} started booking ID {booking_id}.")
        return jsonify({'status': 'success', 'mobile_status': 'In Progress', 'updated_at': updated_at, 'message': 'Job started.'})
    except Exception as exc:
        return _staff_json_error(f'Job could not be started: {exc}', 500)
    finally:
        conn.close()


@app.route('/api/staff/bookings/<int:booking_id>/complete', methods=['POST'])
def api_staff_complete_booking(booking_id):
    if not _staff_is_user():
        return _staff_json_error('Staff portal access is required.', 403)
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return _staff_json_error('Your staff account is not linked to an active employee record.', 404)
        updated_at = _staff_update_booking_status(conn, cid, employee, booking_id, 'Completed')
        if not updated_at:
            return _staff_json_error('Booking not found for your staff profile.', 404)
        conn.commit()
        log_action('Staff Portal', 'Completed Job', f"{employee['name']} completed booking ID {booking_id}.")
        return jsonify({'status': 'success', 'mobile_status': 'Completed', 'updated_at': updated_at, 'message': 'Job completed.'})
    except Exception as exc:
        return _staff_json_error(f'Job could not be completed: {exc}', 500)
    finally:
        conn.close()


@app.route('/api/staff/bookings/<int:booking_id>/notes', methods=['POST'])
def api_staff_booking_notes(booking_id):
    if not _staff_is_user():
        return _staff_json_error('Staff portal access is required.', 403)
    data = request.get_json(silent=True) or {}
    note = (data.get('note') or data.get('notes') or '').strip()
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return _staff_json_error('Your staff account is not linked to an active employee record.', 404)
        notes = _staff_append_booking_note(conn, cid, employee, booking_id, note)
        if notes is None:
            return _staff_json_error('Booking not found for your staff profile.', 404)
        conn.commit()
        log_action('Staff Portal', 'Added Booking Note', f"{employee['name']} added a note to booking ID {booking_id}.")
        return jsonify({'status': 'success', 'notes': notes, 'message': 'Booking note added.'})
    except ValueError as exc:
        return _staff_json_error(str(exc), 400)
    except Exception as exc:
        return _staff_json_error(f'Booking note could not be saved: {exc}', 500)
    finally:
        conn.close()


@app.route('/api/staff/bookings/<int:booking_id>/attachments', methods=['GET', 'POST'])
def api_staff_booking_attachments(booking_id):
    if not _staff_is_user():
        return _staff_json_error('Staff portal access is required.', 403)
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return _staff_json_error('Your staff account is not linked to an active employee record.', 404)
        row = _staff_booking_detail_for_current_user(conn, cid, employee['name'], booking_id)
        if not row:
            return _staff_json_error('Booking not found for your staff profile.', 404)
        if request.method == 'POST':
            files = request.files.getlist('files')
            if not files:
                return _staff_json_error('No files selected.', 400)
            saved = _staff_save_booking_attachments(conn, cid, employee, booking_id, files)
            if saved is None:
                return _staff_json_error('Booking not found for your staff profile.', 404)
            conn.commit()
            log_action('Staff Portal', 'Uploaded Booking Attachment', f"{employee['name']} uploaded {len(saved)} file(s) to booking ID {booking_id}.")
        attachments = _mobile_attachment_rows(conn, cid, 'booking', booking_id)
        return jsonify({'status': 'success', 'attachments': [_staff_attachment_to_dict(a) for a in attachments], 'message': 'File upload complete.' if request.method == 'POST' else ''})
    except ValueError as exc:
        return _staff_json_error(str(exc), 400)
    except Exception as exc:
        return _staff_json_error(f'Booking files could not be processed: {exc}', 500)
    finally:
        conn.close()


@app.route('/staff/download_attachment/<int:attachment_id>')
def staff_download_attachment(attachment_id):
    if not _staff_is_user():
        return redirect(url_for('hub'))
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return 'Staff profile not found.', 404
        row = conn.execute("SELECT * FROM attachments WHERE id=? AND company_id=? AND linked_type='booking'", (attachment_id, cid)).fetchone()
        if not row:
            return 'Attachment not found.', 404
        booking = _staff_booking_detail_for_current_user(conn, cid, employee['name'], row['linked_id'])
        if not booking:
            return 'Forbidden', 403
        abs_path = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], row['file_path']))
        uploads_root = os.path.abspath(app.config['UPLOAD_FOLDER'])
        if not abs_path.startswith(uploads_root) or not os.path.exists(abs_path):
            return 'Attachment file missing.', 404
        return send_file(abs_path, as_attachment=True, download_name=row['original_filename'])
    finally:
        conn.close()



@app.route('/api/staff/payslips')
def api_staff_payslips():
    if not _staff_is_user():
        return _staff_json_error('Staff portal access is required.', 403)
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return _staff_json_error('Your staff account is not linked to an active employee record.', 404)
        rows = conn.execute('''SELECT * FROM payslips
                               WHERE company_id=? AND employee_id=?
                               ORDER BY date DESC, id DESC
                               LIMIT 60''', (cid, employee['id'])).fetchall()
        return jsonify({'status': 'success', 'payslips': [_staff_payslip_to_dict(r) for r in rows]})
    except Exception as exc:
        return _staff_json_error(f'Payslips could not be loaded: {exc}', 500)
    finally:
        conn.close()


@app.route('/api/staff/payslips/<int:payslip_id>')
def api_staff_payslip_detail(payslip_id):
    if not _staff_is_user():
        return _staff_json_error('Staff portal access is required.', 403)
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return _staff_json_error('Your staff account is not linked to an active employee record.', 404)
        row = conn.execute('''SELECT * FROM payslips
                              WHERE id=? AND company_id=? AND employee_id=?''',
                           (payslip_id, cid, employee['id'])).fetchone()
        if not row:
            return _staff_json_error('Payslip not found for your staff profile.', 404)
        return jsonify({'status': 'success', 'payslip': _staff_payslip_to_dict(row)})
    except Exception as exc:
        return _staff_json_error(f'Payslip could not be loaded: {exc}', 500)
    finally:
        conn.close()


@app.route('/staff/download_payslip/<int:payslip_id>')
def staff_download_payslip(payslip_id):
    if not _staff_is_user():
        return redirect(url_for('login'))
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return 'Staff profile not found.', 404
        row = conn.execute('''SELECT * FROM payslips
                              WHERE id=? AND company_id=? AND employee_id=?''',
                           (payslip_id, cid, employee['id'])).fetchone()
        if not row:
            return 'Payslip not found.', 404
        company = conn.execute('SELECT * FROM companies WHERE id=?', (cid,)).fetchone()
        pdf_bytes = _build_staff_payslip_pdf(row, employee, company)
        safe_name = secure_filename(employee['name'] or 'employee') or 'employee'
        period = (format_display_date(row['date']) or str(row['date'] or 'payslip'))[:10]
        filename = f"Payslip_{safe_name}_{period}.pdf"
        log_action('Staff Portal', 'Downloaded Payslip', f"{employee['name']} downloaded payslip ID {payslip_id}.")
        return send_file(pdf_bytes, mimetype='application/pdf', as_attachment=True, download_name=filename)
    finally:
        conn.close()


@app.route('/api/staff/leave_requests', methods=['GET', 'POST'])
def api_staff_leave_requests():
    if not _staff_is_user():
        return _staff_json_error('Staff portal access is required.', 403)
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        employee = _staff_employee_row(conn)
        if not employee:
            return _staff_json_error('Your staff account is not linked to an active employee record.', 404)
        if request.method == 'POST':
            data = request.form if request.form else (request.get_json(silent=True) or {})
            leave_type = (data.get('leave_type') or 'Annual Leave').strip()
            if leave_type not in STAFF_LEAVE_TYPES:
                leave_type = 'Other Leave'
            start_date = (data.get('start_date') or '').strip()
            end_date = (data.get('end_date') or start_date).strip()
            days = _staff_leave_days(start_date, end_date)
            if days <= 0:
                return _staff_json_error('Please select a valid leave start and end date.', 400)
            reason = (data.get('reason') or '').strip()
            attachment_file = _staff_save_leave_attachment(request.files.get('attachment')) if request.files else None
            conn.execute('''INSERT INTO staff_leave_requests
                            (company_id, employee_id, leave_type, start_date, end_date, days, reason, status, attachment_file, requested_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, 'Pending', ?, ?)''',
                         (cid, employee['id'], leave_type, start_date, end_date, days, reason, attachment_file, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            conn.commit()
            log_action('Staff Portal', 'Submitted Leave Request', f"{employee['name']} requested {days:g} day(s) of {leave_type}.")
            return jsonify({'status': 'success', 'message': 'Leave request submitted for approval.'})
        rows = conn.execute('''SELECT * FROM staff_leave_requests
                               WHERE company_id=? AND employee_id=?
                               ORDER BY requested_at DESC, id DESC LIMIT 100''',
                            (cid, employee['id'])).fetchall()
        return jsonify({'status': 'success', 'leave_requests': [_staff_leave_request_dict(r, employee['name']) for r in rows]})
    except Exception as exc:
        return _staff_json_error(f'Leave requests could not be processed: {exc}', 500)
    finally:
        conn.close()


@app.route('/api/staff/admin/employees')
def api_staff_admin_employees():
    if not _staff_is_admin():
        return _staff_json_error('Staff Portal administration access is required.', 403)
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        rows = conn.execute('''SELECT e.id, e.name, e.emp_number, e.job_title, e.status, e.email,
                                      u.id AS staff_user_id, u.username AS staff_username, u.is_staff AS staff_enabled
                               FROM employees e
                               LEFT JOIN users u ON u.employee_id=e.id AND u.company_id=e.company_id AND COALESCE(u.is_staff,0)=1
                               WHERE e.company_id=?
                               ORDER BY e.name ASC''', (cid,)).fetchall()
        return jsonify({'status': 'success', 'employees': [dict(r) for r in rows]})
    except Exception as exc:
        return _staff_json_error(f'Staff employee list could not be loaded: {exc}', 500)
    finally:
        conn.close()


@app.route('/api/staff/admin/accounts/save', methods=['POST'])
def api_staff_admin_account_save():
    if not _staff_is_admin():
        return _staff_json_error('Staff Portal administration access is required.', 403)
    data = request.get_json(silent=True) or {}
    cid = session.get('company_id')
    try:
        employee_id = int(data.get('employee_id') or 0)
    except Exception:
        return _staff_json_error('Invalid employee selected.', 400)
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    enabled = bool(data.get('enabled', True))
    if enabled and not username:
        return _staff_json_error('Please enter a username for this staff account.', 400)
    conn = get_db_connection()
    try:
        employee = conn.execute('SELECT * FROM employees WHERE id=? AND company_id=?', (employee_id, cid)).fetchone()
        if not employee:
            return _staff_json_error('Employee not found for this company.', 404)
        existing = conn.execute('SELECT * FROM users WHERE company_id=? AND employee_id=? AND COALESCE(is_staff,0)=1', (cid, employee_id)).fetchone()
        try:
            if existing:
                employee_email = (dict(employee).get('email') or '').strip()
                if password:
                    conn.execute('''UPDATE users SET username=?, email=?, password_hash=?, employee_id=?, is_staff=?, can_booking=0, can_finance=0, can_payroll=0, can_invoicing=0, can_accounting=0, is_company_admin=0
                                    WHERE id=? AND company_id=?''',
                                 (username or existing['username'], employee_email, generate_password_hash(password), employee_id, 1 if enabled else 0, existing['id'], cid))
                else:
                    conn.execute('''UPDATE users SET username=?, email=?, employee_id=?, is_staff=?, can_booking=0, can_finance=0, can_payroll=0, can_invoicing=0, can_accounting=0, is_company_admin=0
                                    WHERE id=? AND company_id=?''',
                                 (username or existing['username'], employee_email, employee_id, 1 if enabled else 0, existing['id'], cid))
            elif enabled:
                employee_email = (dict(employee).get('email') or '').strip()
                conn.execute('''INSERT INTO users (username, email, company_id, password_hash, employee_id, is_staff, can_booking, can_finance, can_payroll, can_invoicing, can_accounting, is_superadmin, is_company_admin)
                                VALUES (?, ?, ?, ?, ?, 1, 0, 0, 0, 0, 0, 0, 0)''',
                             (username, employee_email, cid, generate_password_hash(password or 'Password123'), employee_id))
            conn.commit()
        except sqlite3.IntegrityError:
            return _staff_json_error('That username already exists. Please choose another username.', 400)
        log_action('Staff Portal', 'Updated Staff Account', f"Updated portal access for {employee['name']}.")
        return jsonify({'status': 'success', 'message': 'Staff portal account saved.'})
    except Exception as exc:
        return _staff_json_error(f'Staff account could not be saved: {exc}', 500)
    finally:
        conn.close()


@app.route('/api/staff/admin/leave_requests')
def api_staff_admin_leave_requests():
    if not _staff_is_admin():
        return _staff_json_error('Staff Portal administration access is required.', 403)
    cid = session.get('company_id')
    status = (request.args.get('status') or 'Pending').strip()
    if status not in STAFF_LEAVE_STATUSES and status != 'All':
        status = 'Pending'
    conn = get_db_connection()
    try:
        params = [cid]
        where = 'r.company_id=?'
        if status != 'All':
            where += " AND COALESCE(r.status, 'Pending')=?"
            params.append(status)
        rows = conn.execute(f'''SELECT r.*, e.name AS employee_name, e.emp_number AS emp_number, e.job_title AS job_title
                                FROM staff_leave_requests r
                                JOIN employees e ON e.id=r.employee_id AND e.company_id=r.company_id
                                WHERE {where}
                                ORDER BY r.requested_at DESC, r.id DESC
                                LIMIT 300''', params).fetchall()
        return jsonify({'status': 'success', 'leave_requests': [_staff_leave_request_dict(r) for r in rows]})
    except Exception as exc:
        return _staff_json_error(f'Leave requests could not be loaded: {exc}', 500)
    finally:
        conn.close()


@app.route('/api/staff/admin/leave_requests/<int:request_id>/review', methods=['POST'])
def api_staff_admin_review_leave(request_id):
    if not _staff_is_admin():
        return _staff_json_error('Staff Portal administration access is required.', 403)
    data = request.get_json(silent=True) or {}
    decision = (data.get('decision') or '').strip().title()
    if decision not in ('Approve', 'Decline'):
        return _staff_json_error('Please choose Approve or Decline.', 400)
    admin_note = (data.get('admin_note') or '').strip()
    cid = session.get('company_id')
    conn = get_db_connection()
    try:
        req = conn.execute('''SELECT r.*, e.name AS employee_name
                              FROM staff_leave_requests r
                              JOIN employees e ON e.id=r.employee_id AND e.company_id=r.company_id
                              WHERE r.id=? AND r.company_id=?''', (request_id, cid)).fetchone()
        if not req:
            return _staff_json_error('Leave request not found.', 404)
        if (req['status'] or 'Pending') != 'Pending':
            return _staff_json_error('Only pending leave requests can be reviewed.', 400)
        reviewed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        new_status = 'Approved' if decision == 'Approve' else 'Declined'
        leave_record_id = req['leave_record_id'] or None
        if new_status == 'Approved' and not leave_record_id:
            cur = conn.execute('''INSERT INTO leave_records (company_id, employee_id, date_taken, days, leave_type, document_file)
                                  VALUES (?, ?, ?, ?, ?, ?)''',
                               (cid, req['employee_id'], req['start_date'], req['days'], req['leave_type'], req['attachment_file']))
            leave_record_id = getattr(cur, 'lastrowid', None)
        conn.execute('''UPDATE staff_leave_requests
                        SET status=?, reviewed_by=?, reviewed_at=?, admin_note=?, leave_record_id=?
                        WHERE id=? AND company_id=?''',
                     (new_status, session.get('username', ''), reviewed_at, admin_note, leave_record_id, request_id, cid))
        conn.commit()
        log_action('Staff Portal', f'{new_status} Leave Request', f"{new_status} leave request for {req['employee_name']}.")
        return jsonify({'status': 'success', 'message': f'Leave request {new_status.lower()}.'})
    except Exception as exc:
        return _staff_json_error(f'Leave request could not be reviewed: {exc}', 500)
    finally:
        conn.close()

# ==========================================================
# EASY ADMIN MOBILE PWA ROUTES
# ==========================================================
MOBILE_STATUS_VALUES = {'Scheduled', 'In Progress', 'Completed', 'Cancelled'}


def _mobile_has_booking_access():
    return bool(session.get('can_booking') or session.get('is_superadmin'))


def _mobile_has_invoicing_access():
    return bool(session.get('can_invoicing') or session.get('is_superadmin'))


def _mobile_forbidden(message='You do not have permission to access this mobile function.'):
    return jsonify({'status': 'error', 'message': message}), 403


def _mobile_error(message, status_code=500):
    return jsonify({'status': 'error', 'message': message}), status_code


def _mobile_date_range(default_days=14):
    start_date = (request.args.get('start_date') or datetime.now().strftime('%Y-%m-%d')).strip()
    end_date = (request.args.get('end_date') or (datetime.now() + timedelta(days=default_days)).strftime('%Y-%m-%d')).strip()
    try:
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')
    except Exception:
        start_date = datetime.now().strftime('%Y-%m-%d')
        end_date = (datetime.now() + timedelta(days=default_days)).strftime('%Y-%m-%d')
    return start_date, end_date


def _mobile_table_exists(conn, table_name):
    try:
        return compat_table_exists(conn, table_name)
    except Exception:
        return False


def _mobile_columns(conn, table_name):
    try:
        return set(compat_table_columns(conn, table_name) or [])
    except Exception:
        return set()


def _mobile_sql_literal(value):
    return "'" + str(value or '').replace("'", "''") + "'"


def _mobile_select_column(columns, table_alias, column_name, alias, default=''):
    if column_name in columns:
        return f'{table_alias}.{column_name} AS {alias}'
    return f'{_mobile_sql_literal(default)} AS {alias}'


def _mobile_company_where(columns, table_alias, cid, params):
    if 'company_id' in columns:
        params.append(cid)
        return f'{table_alias}.company_id=?'
    return '1=1'


def _mobile_client_display(row):
    r = dict(row)
    name = ' '.join([str(x).strip() for x in [r.get('client_first_name'), r.get('client_surname')] if x]).strip()
    return name or r.get('client_company_name') or r.get('title') or 'Unknown Client'


def _mobile_booking_to_dict(row, include_notes=True):
    r = dict(row)
    start_value = str(r.get('start') or '')
    date_part = start_value[:10]
    time_part = ''
    if 'T' in start_value:
        time_part = start_value.split('T', 1)[1][:5]
    elif len(start_value) >= 16:
        time_part = start_value[11:16]
    try:
        day_name = datetime.strptime(date_part, '%Y-%m-%d').strftime('%A') if date_part else ''
    except Exception:
        day_name = ''
    try:
        overtime_hours = float(r.get('overtime_hours') or 0)
    except Exception:
        overtime_hours = 0.0
    item = {
        'id': r.get('id'),
        'client': _mobile_client_display(r),
        'client_id': r.get('client_id') or '',
        'date': date_part,
        'day': day_name,
        'time': time_part,
        'employee': r.get('employee') or 'Unassigned',
        'service': r.get('booking_type') or '',
        'transport': r.get('transport') or '',
        'project_id': r.get('project_id') or '',
        'project_name': r.get('project_name') or '',
        'project_code': r.get('project_code') or '',
        'mobile_status': r.get('mobile_status') or 'Scheduled',
        'mobile_status_updated_at': r.get('mobile_status_updated_at') or '',
        'mobile_status_updated_by': r.get('mobile_status_updated_by') or '',
        'mobile_started_at': r.get('mobile_started_at') or '',
        'mobile_completed_at': r.get('mobile_completed_at') or '',
        'invoiced': bool(r.get('is_invoiced')),
        'attachment_count': r.get('attachment_count') or 0,
        'overtime_hours': overtime_hours,
    }
    if include_notes:
        item['notes'] = r.get('booking_notes') or ''
    return item


def _mobile_project_to_dict(row):
    r = dict(row)
    client = ' '.join([str(x).strip() for x in [r.get('client_first_name'), r.get('client_surname')] if x]).strip() or r.get('client_company_name') or ''
    try:
        fixed_price = float(r.get('fixed_price') or 0)
    except Exception:
        fixed_price = 0.0
    return {
        'id': r.get('id'),
        'project_name': r.get('project_name') or '',
        'project_code': r.get('project_code') or '',
        'client': client,
        'description': r.get('description') or '',
        'site_address': r.get('site_address') or '',
        'start_date': r.get('start_date') or '',
        'estimated_end_date': r.get('estimated_end_date') or '',
        'actual_end_date': r.get('actual_end_date') or '',
        'fixed_price': fixed_price,
        'status': r.get('status') or '',
        'notes': r.get('notes') or '',
        'attachment_count': r.get('attachment_count') or 0,
        'booking_count': r.get('booking_count') or 0,
    }


def _mobile_attachment_count_expr(conn, linked_type, owner_alias, owner_cols):
    attachment_cols = _mobile_columns(conn, 'attachments') if _mobile_table_exists(conn, 'attachments') else set()
    required = {'linked_type', 'linked_id'}
    if not required.issubset(attachment_cols):
        return '0 AS attachment_count'
    company_filter = ''
    if 'company_id' in attachment_cols and 'company_id' in owner_cols:
        company_filter = f' AND a.company_id={owner_alias}.company_id'
    return f"(SELECT COUNT(*) FROM attachments a WHERE a.linked_type='{linked_type}' AND a.linked_id={owner_alias}.id{company_filter}) AS attachment_count"


def _mobile_booking_select_parts(conn, include_client_contact=False):
    booking_cols = _mobile_columns(conn, 'bookings')
    project_cols = _mobile_columns(conn, 'projects') if _mobile_table_exists(conn, 'projects') else set()
    client_cols = _mobile_columns(conn, 'clients') if _mobile_table_exists(conn, 'clients') else set()

    selects = ['b.*']
    joins = []

    if project_cols and 'project_id' in booking_cols and 'id' in project_cols:
        project_join = 'LEFT JOIN projects p ON p.id=b.project_id'
        if 'company_id' in project_cols and 'company_id' in booking_cols:
            project_join += ' AND p.company_id=b.company_id'
        joins.append(project_join)
        selects.append(_mobile_select_column(project_cols, 'p', 'project_name', 'project_name'))
        selects.append(_mobile_select_column(project_cols, 'p', 'project_code', 'project_code'))
    else:
        selects.append("'' AS project_name")
        selects.append("'' AS project_code")

    if client_cols and 'client_id' in booking_cols and 'id' in client_cols:
        client_join = 'LEFT JOIN clients c ON c.id=b.client_id'
        if 'company_id' in client_cols and 'company_id' in booking_cols:
            client_join += ' AND c.company_id=b.company_id'
        joins.append(client_join)
        selects.append(_mobile_select_column(client_cols, 'c', 'name', 'client_first_name'))
        selects.append(_mobile_select_column(client_cols, 'c', 'surname', 'client_surname'))
        selects.append(_mobile_select_column(client_cols, 'c', 'company_name', 'client_company_name'))
        if include_client_contact:
            selects.append(_mobile_select_column(client_cols, 'c', 'phone', 'client_phone'))
            selects.append(_mobile_select_column(client_cols, 'c', 'email', 'client_email'))
            selects.append(_mobile_select_column(client_cols, 'c', 'address', 'client_address'))
    else:
        selects.append("'' AS client_first_name")
        selects.append("'' AS client_surname")
        selects.append("'' AS client_company_name")
        if include_client_contact:
            selects.append("'' AS client_phone")
            selects.append("'' AS client_email")
            selects.append("'' AS client_address")

    selects.append(_mobile_attachment_count_expr(conn, 'booking', 'b', booking_cols))
    return booking_cols, selects, joins


def _mobile_booking_rows(conn, cid, start_date, end_date, status_filter=''):
    if not _mobile_table_exists(conn, 'bookings'):
        return []
    booking_cols, selects, joins = _mobile_booking_select_parts(conn)
    if 'start' not in booking_cols:
        return []
    params = []
    where_parts = [_mobile_company_where(booking_cols, 'b', cid, params)]
    where_parts.append("substr(COALESCE(CAST(b.start AS TEXT), ''), 1, 10) BETWEEN ? AND ?")
    params.extend([start_date, end_date])
    if status_filter and 'mobile_status' in booking_cols:
        where_parts.append("COALESCE(b.mobile_status, 'Scheduled')=?")
        params.append(status_filter)
    where_sql = ' AND '.join(where_parts)
    join_sql = ' '.join(joins)
    order_sql = 'b.start ASC, b.id ASC' if 'id' in booking_cols else 'b.start ASC'
    sql = f'''SELECT {', '.join(selects)}
              FROM bookings b
              {join_sql}
              WHERE {where_sql}
              ORDER BY {order_sql}
              LIMIT 300'''
    return conn.execute(sql, params).fetchall()


def _mobile_booking_detail_row(conn, cid, booking_id):
    if not _mobile_table_exists(conn, 'bookings'):
        return None
    booking_cols, selects, joins = _mobile_booking_select_parts(conn, include_client_contact=True)
    if 'id' not in booking_cols:
        return None
    params = []
    where_parts = [_mobile_company_where(booking_cols, 'b', cid, params), 'b.id=?']
    params.append(booking_id)
    sql = f'''SELECT {', '.join(selects)}
              FROM bookings b
              {' '.join(joins)}
              WHERE {' AND '.join(where_parts)}'''
    return conn.execute(sql, params).fetchone()


def _mobile_project_select_parts(conn):
    project_cols = _mobile_columns(conn, 'projects')
    client_cols = _mobile_columns(conn, 'clients') if _mobile_table_exists(conn, 'clients') else set()
    booking_cols = _mobile_columns(conn, 'bookings') if _mobile_table_exists(conn, 'bookings') else set()
    selects = ['p.*']
    joins = []
    if client_cols and 'client_id' in project_cols and 'id' in client_cols:
        client_join = 'LEFT JOIN clients c ON c.id=p.client_id'
        if 'company_id' in client_cols and 'company_id' in project_cols:
            client_join += ' AND c.company_id=p.company_id'
        joins.append(client_join)
        selects.append(_mobile_select_column(client_cols, 'c', 'name', 'client_first_name'))
        selects.append(_mobile_select_column(client_cols, 'c', 'surname', 'client_surname'))
        selects.append(_mobile_select_column(client_cols, 'c', 'company_name', 'client_company_name'))
    else:
        selects.append("'' AS client_first_name")
        selects.append("'' AS client_surname")
        selects.append("'' AS client_company_name")
    selects.append(_mobile_attachment_count_expr(conn, 'project', 'p', project_cols))
    if booking_cols and 'project_id' in booking_cols:
        company_filter = ' AND b.company_id=p.company_id' if 'company_id' in booking_cols and 'company_id' in project_cols else ''
        selects.append(f'(SELECT COUNT(*) FROM bookings b WHERE b.project_id=p.id{company_filter}) AS booking_count')
    else:
        selects.append('0 AS booking_count')
    return project_cols, booking_cols, selects, joins


def _mobile_active_project_rows(conn, cid, limit=8):
    if not _mobile_table_exists(conn, 'projects'):
        return []
    project_cols, booking_cols, selects, joins = _mobile_project_select_parts(conn)
    if 'id' not in project_cols:
        return []
    params = []
    where_parts = [_mobile_company_where(project_cols, 'p', cid, params)]
    if 'status' in project_cols:
        where_parts.append("COALESCE(p.status, '') NOT IN ('Completed', 'Cancelled')")
    # PostgreSQL does not allow COALESCE across mixed types such as
    # TEXT start_date and TIMESTAMP created_at. Order by each available
    # field separately instead, which works on both SQLite and PostgreSQL.
    order_terms = []
    if 'start_date' in project_cols:
        order_terms.append('p.start_date DESC')
    if 'created_at' in project_cols:
        order_terms.append('p.created_at DESC')
    order_terms.append('p.id DESC')
    order_clause = ', '.join(order_terms)
    sql = f'''SELECT {', '.join(selects)}
              FROM projects p
              {' '.join(joins)}
              WHERE {' AND '.join(where_parts)}
              ORDER BY {order_clause}
              LIMIT {int(limit)}'''
    return conn.execute(sql, params).fetchall()


def _mobile_project_rows(conn, cid, status_filter='', q='', limit=200):
    if not _mobile_table_exists(conn, 'projects'):
        return []
    project_cols, booking_cols, selects, joins = _mobile_project_select_parts(conn)
    if 'id' not in project_cols:
        return []
    params = []
    where_parts = [_mobile_company_where(project_cols, 'p', cid, params)]
    if status_filter and 'status' in project_cols:
        where_parts.append('COALESCE(p.status, \'\')=?')
        params.append(status_filter)
    if q:
        like = f'%{q}%'
        search_parts = []
        for col in ['project_name', 'project_code']:
            if col in project_cols:
                search_parts.append(f"COALESCE(p.{col}, '') LIKE ?")
                params.append(like)
        for col in ['name', 'surname', 'company_name']:
            if col in _mobile_columns(conn, 'clients'):
                search_parts.append(f"COALESCE(c.{col}, '') LIKE ?")
                params.append(like)
        if search_parts:
            where_parts.append('(' + ' OR '.join(search_parts) + ')')
    # PostgreSQL does not allow COALESCE across mixed types such as
    # TEXT start_date and TIMESTAMP created_at. Order by each available
    # field separately instead, which works on both SQLite and PostgreSQL.
    order_terms = []
    if 'start_date' in project_cols:
        order_terms.append('p.start_date DESC')
    if 'created_at' in project_cols:
        order_terms.append('p.created_at DESC')
    order_terms.append('p.id DESC')
    order_clause = ', '.join(order_terms)
    sql = f'''SELECT {', '.join(selects)}
              FROM projects p
              {' '.join(joins)}
              WHERE {' AND '.join(where_parts)}
              ORDER BY {order_clause}
              LIMIT {int(limit)}'''
    return conn.execute(sql, params).fetchall()


def _mobile_project_detail_row(conn, cid, project_id):
    if not _mobile_table_exists(conn, 'projects'):
        return None
    project_cols, booking_cols, selects, joins = _mobile_project_select_parts(conn)
    if 'id' not in project_cols:
        return None
    params = []
    where_parts = [_mobile_company_where(project_cols, 'p', cid, params), 'p.id=?']
    params.append(project_id)
    sql = f'''SELECT {', '.join(selects)}
              FROM projects p
              {' '.join(joins)}
              WHERE {' AND '.join(where_parts)}'''
    return conn.execute(sql, params).fetchone()


def _mobile_linked_project_booking_rows(conn, cid, project_id):
    if not _mobile_table_exists(conn, 'bookings'):
        return []
    booking_cols = _mobile_columns(conn, 'bookings')
    if 'project_id' not in booking_cols:
        return []
    _, selects, joins = _mobile_booking_select_parts(conn)
    params = []
    where_parts = [_mobile_company_where(booking_cols, 'b', cid, params), 'b.project_id=?']
    params.append(project_id)
    order_sql = 'b.start ASC, b.id ASC' if 'start' in booking_cols else 'b.id ASC'
    sql = f'''SELECT {', '.join(selects)}
              FROM bookings b
              {' '.join(joins)}
              WHERE {' AND '.join(where_parts)}
              ORDER BY {order_sql}
              LIMIT 100'''
    return conn.execute(sql, params).fetchall()


def _mobile_attachment_rows(conn, cid, linked_type, linked_id):
    if not _mobile_table_exists(conn, 'attachments'):
        return []
    attachment_cols = _mobile_columns(conn, 'attachments')
    if not {'linked_type', 'linked_id'}.issubset(attachment_cols):
        return []
    params = [linked_type, linked_id]
    where = 'linked_type=? AND linked_id=?'
    if 'company_id' in attachment_cols:
        where += ' AND company_id=?'
        params.append(cid)
    order_cols = []
    if 'uploaded_at' in attachment_cols:
        order_cols.append('uploaded_at DESC')
    if 'id' in attachment_cols:
        order_cols.append('id DESC')
    order_sql = ', '.join(order_cols) if order_cols else 'linked_id DESC'
    return conn.execute(f'SELECT * FROM attachments WHERE {where} ORDER BY {order_sql}', params).fetchall()


@app.route('/mobile')
def mobile_index():
    return render_template('mobile.html', session=session)


@app.route('/mobile/offline')
def mobile_offline():
    return render_template('mobile_offline.html')


@app.route('/manifest.webmanifest')
def manifest_webmanifest():
    return send_from_directory('static', 'manifest.webmanifest', mimetype='application/manifest+json')


@app.route('/service-worker.js')
def service_worker():
    response = send_from_directory('static', 'service-worker.js', mimetype='application/javascript')
    response.headers['Cache-Control'] = 'no-cache'
    return response


@app.route('/api/mobile/dashboard')
def api_mobile_dashboard():
    cid = session.get('company_id')
    if not cid:
        return _mobile_error('Your session does not have an active company. Please log out and log back in.', 401)
    today = datetime.now().strftime('%Y-%m-%d')
    tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    two_weeks = (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d')
    conn = get_db_connection()
    try:
        payload = {
            'status': 'success',
            'company': session.get('company_name') or 'Easy Admin',
            'username': session.get('username') or '',
            'modules': {
                'booking': _mobile_has_booking_access(),
                'projects': _mobile_has_booking_access() and _mobile_table_exists(conn, 'projects'),
                'invoicing': _mobile_has_invoicing_access(),
                'finance': bool(session.get('can_finance') or session.get('is_superadmin')),
                'payroll': bool(session.get('can_payroll') or session.get('is_superadmin')),
                'accounting': bool(session.get('can_accounting') or session.get('is_superadmin')),
            },
            'today': today,
            'stats': {'today_bookings': 0, 'upcoming_bookings': 0, 'active_projects': 0, 'unpaid_invoices': 0},
            'today_bookings': [],
            'upcoming_bookings': [],
            'active_projects': [],
            'warnings': []
        }
        if _mobile_has_booking_access():
            try:
                today_rows = _mobile_booking_rows(conn, cid, today, today)
                upcoming_rows = _mobile_booking_rows(conn, cid, tomorrow, two_weeks)
                payload['today_bookings'] = [_mobile_booking_to_dict(r) for r in today_rows]
                payload['upcoming_bookings'] = [_mobile_booking_to_dict(r, include_notes=False) for r in upcoming_rows[:10]]
                payload['stats']['today_bookings'] = len(payload['today_bookings'])
                payload['stats']['upcoming_bookings'] = len(upcoming_rows)
            except Exception as exc:
                payload['warnings'].append(f'Bookings could not be loaded: {exc}')
            try:
                project_rows = _mobile_active_project_rows(conn, cid, limit=8)
                payload['active_projects'] = [_mobile_project_to_dict(r) for r in project_rows]
                payload['stats']['active_projects'] = len(payload['active_projects'])
            except Exception as exc:
                payload['warnings'].append(f'Projects could not be loaded: {exc}')
        if _mobile_has_invoicing_access():
            try:
                if _mobile_table_exists(conn, 'invoices') and 'company_id' in _mobile_columns(conn, 'invoices'):
                    unpaid = conn.execute("SELECT COUNT(*) AS total FROM invoices WHERE company_id=? AND COALESCE(status, '') NOT IN ('Paid', 'Cancelled')", (cid,)).fetchone()
                    payload['stats']['unpaid_invoices'] = unpaid['total'] if unpaid else 0
            except Exception as exc:
                payload['warnings'].append(f'Invoices could not be counted: {exc}')
        return jsonify(payload)
    except Exception as exc:
        return _mobile_error(f'Mobile dashboard could not be loaded: {exc}')
    finally:
        conn.close()


@app.route('/api/mobile/bookings')
def api_mobile_bookings():
    if not _mobile_has_booking_access():
        return _mobile_forbidden('Booking permission is required.')
    cid = session.get('company_id')
    if not cid:
        return _mobile_error('Your session does not have an active company. Please log out and log back in.', 401)
    start_date, end_date = _mobile_date_range(default_days=14)
    status_filter = (request.args.get('status') or '').strip()
    if status_filter and status_filter not in MOBILE_STATUS_VALUES:
        status_filter = ''
    conn = get_db_connection()
    try:
        rows = _mobile_booking_rows(conn, cid, start_date, end_date, status_filter)
        return jsonify({'status': 'success', 'bookings': [_mobile_booking_to_dict(r) for r in rows], 'start_date': start_date, 'end_date': end_date})
    except Exception as exc:
        return _mobile_error(f'Bookings could not be loaded: {exc}')
    finally:
        conn.close()


@app.route('/api/mobile/bookings/<int:booking_id>')
def api_mobile_booking_detail(booking_id):
    if not _mobile_has_booking_access():
        return _mobile_forbidden('Booking permission is required.')
    cid = session.get('company_id')
    if not cid:
        return _mobile_error('Your session does not have an active company. Please log out and log back in.', 401)
    conn = get_db_connection()
    try:
        rows = _mobile_booking_detail_row(conn, cid, booking_id)
        if not rows:
            return jsonify({'status': 'error', 'message': 'Booking not found.'}), 404
        item = _mobile_booking_to_dict(rows)
        item['client_phone'] = rows.get('client_phone') or ''
        item['client_email'] = rows.get('client_email') or ''
        item['client_address'] = rows.get('client_address') or ''
        attachments = _mobile_attachment_rows(conn, cid, 'booking', booking_id)
        item['attachments'] = [attachment_to_dict(a) for a in attachments]
        return jsonify({'status': 'success', 'booking': item})
    except Exception as exc:
        return _mobile_error(f'Booking details could not be loaded: {exc}')
    finally:
        conn.close()


@app.route('/api/mobile/bookings/<int:booking_id>/status', methods=['POST'])
def api_mobile_booking_status(booking_id):
    if not _mobile_has_booking_access():
        return _mobile_forbidden('Booking permission is required.')
    data = request.get_json(silent=True) or {}
    new_status = (data.get('status') or '').strip()
    if new_status not in MOBILE_STATUS_VALUES:
        return jsonify({'status': 'error', 'message': 'Invalid mobile status.'}), 400
    cid = session.get('company_id')
    if not cid:
        return _mobile_error('Your session does not have an active company. Please log out and log back in.', 401)
    updated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db_connection()
    try:
        booking_cols = _mobile_columns(conn, 'bookings')
        if 'id' not in booking_cols:
            return jsonify({'status': 'error', 'message': 'Booking table is not available.'}), 404
        params = []
        where_parts = [_mobile_company_where(booking_cols, 'bookings', cid, params), 'id=?']
        params.append(booking_id)
        row = conn.execute(f"SELECT id FROM bookings WHERE {' AND '.join(where_parts)}", params).fetchone()
        if not row:
            return jsonify({'status': 'error', 'message': 'Booking not found.'}), 404
        for sql in [
            'ALTER TABLE bookings ADD COLUMN mobile_status TEXT DEFAULT "Scheduled"',
            'ALTER TABLE bookings ADD COLUMN mobile_status_updated_at TEXT',
            'ALTER TABLE bookings ADD COLUMN mobile_status_updated_by TEXT',
            'ALTER TABLE bookings ADD COLUMN mobile_started_at TEXT',
            'ALTER TABLE bookings ADD COLUMN mobile_completed_at TEXT'
        ]:
            try:
                conn.execute(sql)
            except Exception:
                pass
        booking_cols = _mobile_columns(conn, 'bookings')
        update_cols = ['mobile_status=?']
        update_params = [new_status]
        if 'mobile_status_updated_at' in booking_cols:
            update_cols.append('mobile_status_updated_at=?')
            update_params.append(updated_at)
        if 'mobile_status_updated_by' in booking_cols:
            update_cols.append('mobile_status_updated_by=?')
            update_params.append(session.get('username', ''))
        if new_status == 'In Progress' and 'mobile_started_at' in booking_cols:
            update_cols.append('mobile_started_at=COALESCE(NULLIF(mobile_started_at, ?), ?)')
            update_params.extend(['', updated_at])
        if new_status == 'Completed' and 'mobile_completed_at' in booking_cols:
            update_cols.append('mobile_completed_at=?')
            update_params.append(updated_at)
        update_params.extend(params)
        conn.execute(f"UPDATE bookings SET {', '.join(update_cols)} WHERE {' AND '.join(where_parts)}", update_params)
        conn.commit()
        log_action('Mobile PWA', 'Updated Booking Status', f"Booking ID {booking_id} set to {new_status}.")
        return jsonify({'status': 'success', 'mobile_status': new_status, 'updated_at': updated_at})
    except Exception as exc:
        return _mobile_error(f'Booking status could not be updated: {exc}')
    finally:
        conn.close()


@app.route('/api/mobile/bookings/<int:booking_id>/notes', methods=['POST'])
def api_mobile_booking_notes(booking_id):
    if not _mobile_has_booking_access():
        return _mobile_forbidden('Booking permission is required.')
    data = request.get_json(silent=True) or {}
    notes = (data.get('notes') or '').strip()
    cid = session.get('company_id')
    if not cid:
        return _mobile_error('Your session does not have an active company. Please log out and log back in.', 401)
    conn = get_db_connection()
    try:
        booking_cols = _mobile_columns(conn, 'bookings')
        if 'booking_notes' not in booking_cols:
            return jsonify({'status': 'error', 'message': 'Booking notes are not enabled in this database yet.'}), 400
        params = []
        where_parts = [_mobile_company_where(booking_cols, 'bookings', cid, params), 'id=?']
        params.append(booking_id)
        row = conn.execute(f"SELECT id FROM bookings WHERE {' AND '.join(where_parts)}", params).fetchone()
        if not row:
            return jsonify({'status': 'error', 'message': 'Booking not found.'}), 404
        update_params = [notes] + params
        conn.execute(f"UPDATE bookings SET booking_notes=? WHERE {' AND '.join(where_parts)}", update_params)
        conn.commit()
        log_action('Mobile PWA', 'Updated Booking Notes', f"Booking ID {booking_id} notes updated.")
        return jsonify({'status': 'success', 'notes': notes})
    except Exception as exc:
        return _mobile_error(f'Booking notes could not be updated: {exc}')
    finally:
        conn.close()


@app.route('/api/mobile/projects')
def api_mobile_projects():
    if not _mobile_has_booking_access():
        return _mobile_forbidden('Booking permission is required.')
    cid = session.get('company_id')
    if not cid:
        return _mobile_error('Your session does not have an active company. Please log out and log back in.', 401)
    status_filter = (request.args.get('status') or '').strip()
    q = (request.args.get('q') or '').strip()
    conn = get_db_connection()
    try:
        rows = _mobile_project_rows(conn, cid, status_filter=status_filter, q=q)
        return jsonify({'status': 'success', 'projects': [_mobile_project_to_dict(r) for r in rows]})
    except Exception as exc:
        return _mobile_error(f'Projects could not be loaded: {exc}')
    finally:
        conn.close()


@app.route('/api/mobile/projects/<int:project_id>')
def api_mobile_project_detail(project_id):
    if not _mobile_has_booking_access():
        return _mobile_forbidden('Booking permission is required.')
    cid = session.get('company_id')
    if not cid:
        return _mobile_error('Your session does not have an active company. Please log out and log back in.', 401)
    conn = get_db_connection()
    try:
        row = _mobile_project_detail_row(conn, cid, project_id)
        if not row:
            return jsonify({'status': 'error', 'message': 'Project not found.'}), 404
        project = _mobile_project_to_dict(row)
        bookings = _mobile_linked_project_booking_rows(conn, cid, project_id)
        attachments = _mobile_attachment_rows(conn, cid, 'project', project_id)
        project['bookings'] = [_mobile_booking_to_dict(b, include_notes=False) for b in bookings]
        project['attachments'] = [attachment_to_dict(a) for a in attachments]
        return jsonify({'status': 'success', 'project': project})
    except Exception as exc:
        return _mobile_error(f'Project details could not be loaded: {exc}')
    finally:
        conn.close()


@app.route('/api/mobile/invoices')
def api_mobile_invoices():
    if not _mobile_has_invoicing_access():
        return _mobile_forbidden('Invoicing permission is required.')
    cid = session.get('company_id')
    if not cid:
        return _mobile_error('Your session does not have an active company. Please log out and log back in.', 401)
    start_date, end_date = _mobile_date_range(default_days=30)
    conn = get_db_connection()
    try:
        if not _mobile_table_exists(conn, 'invoices'):
            return jsonify({'status': 'success', 'invoices': [], 'start_date': start_date, 'end_date': end_date})
        invoice_cols = _mobile_columns(conn, 'invoices')
        selects = []
        for col, default in [('id', 0), ('client_name', ''), ('date', ''), ('due_date', ''), ('total', 0), ('status', '')]:
            selects.append(_mobile_select_column(invoice_cols, 'i', col, col, default))
        params = []
        where_parts = [_mobile_company_where(invoice_cols, 'i', cid, params)]
        if 'date' in invoice_cols:
            where_parts.append("COALESCE(i.date, '') BETWEEN ? AND ?")
            params.extend([start_date, end_date])
        order_sql = 'i.date DESC, i.id DESC' if {'date', 'id'}.issubset(invoice_cols) else ('i.id DESC' if 'id' in invoice_cols else '1')
        rows = conn.execute(f'''SELECT {', '.join(selects)}
                               FROM invoices i
                               WHERE {' AND '.join(where_parts)}
                               ORDER BY {order_sql}
                               LIMIT 100''', params).fetchall()
        invoices = [dict(r) for r in rows]
        return jsonify({'status': 'success', 'invoices': invoices, 'start_date': start_date, 'end_date': end_date})
    except Exception as exc:
        return _mobile_error(f'Invoices could not be loaded: {exc}')
    finally:
        conn.close()


# ==========================================================
# SAAS OPTIMISATION SUPPORT ROUTES
# ==========================================================
@app.route('/api/background_jobs', methods=['GET'])
def api_background_jobs():
    return jsonify({"jobs": job_manager.list()})

@app.route('/api/background_jobs/<job_id>', methods=['GET'])
def api_background_job(job_id):
    job = job_manager.get(job_id)
    if not job:
        return jsonify({"status": "error", "message": "Job not found"}), 404
    return jsonify(job)

@app.route('/api/paged/payroll_employees', methods=['GET'])
def api_paged_payroll_employees():
    if not session.get('can_payroll') and not session.get('is_superadmin'):
        return jsonify({"status": "error", "message": "Forbidden"}), 403
    conn = get_db_connection()
    cid = session['company_id']
    page, per_page, offset = get_page_args(request.args, default_per_page=50, max_per_page=100)
    q = (request.args.get('q') or '').strip()
    where = "company_id=? AND (emp_type != 'Supplier' OR emp_type IS NULL)"
    params = [cid]
    if q:
        like = f"%{q}%"
        where += " AND (COALESCE(name,'') LIKE ? OR COALESCE(emp_number,'') LIKE ? OR COALESCE(job_title,'') LIKE ? OR COALESCE(status,'') LIKE ? OR COALESCE(emp_type,'') LIKE ?)"
        params.extend([like, like, like, like, like])
    total = conn.execute(f"SELECT COUNT(*) FROM employees WHERE {where}", params).fetchone()[0]
    rows = conn.execute(f"SELECT id, emp_number, name, job_title, emp_type, status, start_date, inactive_date FROM employees WHERE {where} ORDER BY name ASC LIMIT ? OFFSET ?", params + [per_page, offset]).fetchall()
    conn.close()
    return jsonify({"items": [dict(r) for r in rows], "pagination": pagination_meta(total, page, per_page), "query": q})

@app.route('/api/paged/bookings', methods=['GET'])
def api_paged_bookings():
    if not session.get('can_booking') and not session.get('is_superadmin'):
        return jsonify({"status": "error", "message": "Forbidden"}), 403
    conn = get_db_connection()
    cid = session['company_id']
    page, per_page, offset = get_page_args(request.args, default_per_page=50, max_per_page=100)
    q = (request.args.get('q') or '').strip()
    start_date = (request.args.get('start_date') or '').strip()
    end_date = (request.args.get('end_date') or '').strip()
    where = 'b.company_id=?'
    params = [cid]
    if start_date:
        where += ' AND COALESCE(b.start, "") >= ?'
        params.append(start_date)
    if end_date:
        where += ' AND COALESCE(b.start, "") <= ?'
        params.append(end_date)
    if q:
        like = f'%{q}%'
        where += ' AND (COALESCE(b.title, "") LIKE ? OR COALESCE(b.employee, "") LIKE ? OR COALESCE(b.booking_type, "") LIKE ? OR COALESCE(p.project_name, "") LIKE ?)'
        params.extend([like, like, like, like])
    total = conn.execute(f'''SELECT COUNT(*) FROM bookings b LEFT JOIN projects p ON p.id=b.project_id AND p.company_id=b.company_id WHERE {where}''', params).fetchone()[0]
    rows = conn.execute(f'''SELECT b.id, b.title, b.employee, b.booking_type, b.start, b.transport, b.project_id, p.project_name, p.project_code
                            FROM bookings b
                            LEFT JOIN projects p ON p.id=b.project_id AND p.company_id=b.company_id
                            WHERE {where}
                            ORDER BY b.start DESC
                            LIMIT ? OFFSET ?''', params + [per_page, offset]).fetchall()
    conn.close()
    return jsonify({"items": [dict(r) for r in rows], "pagination": pagination_meta(total, page, per_page), "query": q})

@app.route('/api/paged/invoices', methods=['GET'])
def api_paged_invoices():
    if not session.get('can_invoicing') and not session.get('is_superadmin'):
        return jsonify({"status": "error", "message": "Forbidden"}), 403
    conn = get_db_connection()
    cid = session['company_id']
    page, per_page, offset = get_page_args(request.args, default_per_page=50, max_per_page=100)
    q = (request.args.get('q') or '').strip()
    where = 'company_id=?'
    params = [cid]
    if q:
        like = f'%{q}%'
        where += " AND (COALESCE(client_name,'') LIKE ? OR COALESCE(status,'') LIKE ? OR COALESCE(date,'') LIKE ? OR COALESCE(due_date,'') LIKE ?)"
        params.extend([like, like, like, like])
    total = conn.execute(f'SELECT COUNT(*) FROM invoices WHERE {where}', params).fetchone()[0]
    rows = conn.execute(f'SELECT id, client_name, date, due_date, total, status, invoice_type, project_id FROM invoices WHERE {where} ORDER BY id DESC LIMIT ? OFFSET ?', params + [per_page, offset]).fetchall()
    conn.close()
    return jsonify({"items": [dict(r) for r in rows], "pagination": pagination_meta(total, page, per_page), "query": q})

@app.route('/api/paged/finance_expenses', methods=['GET'])
def api_paged_finance_expenses():
    if not session.get('can_finance') and not session.get('is_superadmin'):
        return jsonify({"status": "error", "message": "Forbidden"}), 403
    conn = get_db_connection()
    cid = session['company_id']
    page, per_page, offset = get_page_args(request.args, default_per_page=25, max_per_page=100)
    q = (request.args.get('q') or '').strip()
    where = 'company_id=?'
    params = [cid]
    if q:
        like = f'%{q}%'
        where += " AND (COALESCE(category,'') LIKE ? OR COALESCE(supplier,'') LIKE ? OR COALESCE(description,'') LIKE ? OR COALESCE(date,'') LIKE ?)"
        params.extend([like, like, like, like])
    total = conn.execute(f'SELECT COUNT(*) FROM expenses WHERE {where}', params).fetchone()[0]
    rows = conn.execute(f'SELECT id, date, category, supplier, description, amount, invoice_file FROM expenses WHERE {where} ORDER BY date DESC, id DESC LIMIT ? OFFSET ?', params + [per_page, offset]).fetchall()
    conn.close()
    return jsonify({"items": [dict(r) for r in rows], "pagination": pagination_meta(total, page, per_page), "query": q})

# ==========================================================
# SUPER ADMIN / COMPANY ADMIN ROUTES
# ==========================================================

@app.route('/admin/tax_config/<year>', methods=['GET'])
def get_tax_config(year):
    if not session.get('is_superadmin'): return "Forbidden", 403
    conn = get_db_connection()
    brackets = conn.execute("SELECT * FROM tax_brackets WHERE tax_year=? ORDER BY min_income ASC", (year,)).fetchall()
    rebates = conn.execute('''SELECT primary_rebate, secondary_rebate, tertiary_rebate, threshold_under_65, threshold_65_to_74, threshold_75_plus
                              FROM tax_rebates WHERE tax_year=?''', (year,)).fetchone()
    conn.close()
    rebate_defaults = {
        "primary_rebate": 17820.0,
        "secondary_rebate": 9765.0,
        "tertiary_rebate": 3249.0,
        "threshold_under_65": 99000.0,
        "threshold_65_to_74": 153250.0,
        "threshold_75_plus": 171300.0
    }
    if rebates:
        rebate_defaults.update(dict(rebates))
    return jsonify({
        "brackets": [dict(b) for b in brackets],
        **rebate_defaults
    })

@app.route('/admin/tax_config/bracket/save', methods=['POST'])
def save_tax_bracket():
    if not session.get('is_superadmin'): return "Forbidden", 403
    data = request.json
    conn = get_db_connection()
    if data.get('id'):
        conn.execute("UPDATE tax_brackets SET min_income=?, max_income=?, base_tax=?, rate=? WHERE id=?", 
                     (data['min_income'], data['max_income'], data['base_tax'], data['rate'], data['id']))
    else:
        conn.execute("INSERT INTO tax_brackets (tax_year, min_income, max_income, base_tax, rate) VALUES (?, ?, ?, ?, ?)", 
                     (data['tax_year'], data['min_income'], data['max_income'], data['base_tax'], data['rate']))
    conn.commit(); conn.close()
    return jsonify({"status": "success"})

@app.route('/admin/tax_config/bracket/delete', methods=['POST'])
def delete_tax_bracket():
    if not session.get('is_superadmin'): return "Forbidden", 403
    conn = get_db_connection()
    conn.execute("DELETE FROM tax_brackets WHERE id=?", (request.json['id'],))
    conn.commit(); conn.close()
    return jsonify({"status": "success"})

@app.route('/admin/tax_config/rebate/save', methods=['POST'])
def save_tax_rebate():
    if not session.get('is_superadmin'): return "Forbidden", 403
    data = request.json
    conn = get_db_connection()
    values = (
        data.get('primary_rebate', 0), data.get('secondary_rebate', 0), data.get('tertiary_rebate', 0),
        data.get('threshold_under_65', 0), data.get('threshold_65_to_74', 0), data.get('threshold_75_plus', 0),
        data['tax_year']
    )
    exists = conn.execute("SELECT 1 FROM tax_rebates WHERE tax_year=?", (data['tax_year'],)).fetchone()
    if exists:
        conn.execute('''UPDATE tax_rebates SET primary_rebate=?, secondary_rebate=?, tertiary_rebate=?,
                       threshold_under_65=?, threshold_65_to_74=?, threshold_75_plus=? WHERE tax_year=?''', values)
    else:
        conn.execute('''INSERT INTO tax_rebates (primary_rebate, secondary_rebate, tertiary_rebate, threshold_under_65, threshold_65_to_74, threshold_75_plus, tax_year)
                        VALUES (?, ?, ?, ?, ?, ?, ?)''', values)
    conn.commit(); conn.close()
    return jsonify({"status": "success"})

@app.route('/admin/tax_config/copy', methods=['POST'])
def copy_tax_config():
    if not session.get('is_superadmin'): return "Forbidden", 403
    data = request.json
    from_year = data.get('from_year')
    to_year = data.get('to_year')

    conn = get_db_connection()
    conn.execute("DELETE FROM tax_brackets WHERE tax_year=?", (to_year,))
    conn.execute("DELETE FROM tax_rebates WHERE tax_year=?", (to_year,))

    brackets = conn.execute("SELECT min_income, max_income, base_tax, rate FROM tax_brackets WHERE tax_year=?", (from_year,)).fetchall()
    for b in brackets:
        conn.execute("INSERT INTO tax_brackets (tax_year, min_income, max_income, base_tax, rate) VALUES (?, ?, ?, ?, ?)",
                     (to_year, b['min_income'], b['max_income'], b['base_tax'], b['rate']))

    rebate = conn.execute('''SELECT primary_rebate, secondary_rebate, tertiary_rebate, threshold_under_65, threshold_65_to_74, threshold_75_plus
                             FROM tax_rebates WHERE tax_year=?''', (from_year,)).fetchone()
    if rebate:
        conn.execute('''INSERT INTO tax_rebates (tax_year, primary_rebate, secondary_rebate, tertiary_rebate, threshold_under_65, threshold_65_to_74, threshold_75_plus)
                        VALUES (?, ?, ?, ?, ?, ?, ?)''', (to_year, rebate['primary_rebate'], rebate['secondary_rebate'], rebate['tertiary_rebate'], rebate['threshold_under_65'], rebate['threshold_65_to_74'], rebate['threshold_75_plus']))

    conn.commit()
    conn.close()
    log_action('System Admin', 'Copied Tax Config', f"Copied tax configuration from {from_year} to {to_year}")
    return jsonify({"status": "success"})

@app.route('/admin/holidays/<year>', methods=['GET'])
def get_holidays(year):
    if not session.get('is_superadmin'): return "Forbidden", 403
    conn = get_db_connection()
    h = conn.execute("SELECT * FROM public_holidays WHERE year=? ORDER BY date_str ASC", (year,)).fetchall()
    conn.close()
    return jsonify([dict(x) for x in h])

@app.route('/admin/holidays/save', methods=['POST'])
def save_holiday():
    if not session.get('is_superadmin'): return "Forbidden", 403
    data = request.json
    conn = get_db_connection()
    if data.get('id'):
        conn.execute("UPDATE public_holidays SET date_str=?, name=? WHERE id=?", (data['date_str'], data['name'], data['id']))
    else:
        conn.execute("INSERT INTO public_holidays (year, date_str, name) VALUES (?, ?, ?)", (data['year'], data['date_str'], data['name']))
    conn.commit(); conn.close()
    return jsonify({"status": "success"})

@app.route('/admin/holidays/delete', methods=['POST'])
def delete_holiday():
    if not session.get('is_superadmin'): return "Forbidden", 403
    conn = get_db_connection()
    conn.execute("DELETE FROM public_holidays WHERE id=?", (request.json['id'],))
    conn.commit(); conn.close()
    return jsonify({"status": "success"})

@app.route('/admin/holidays/generate', methods=['POST'])
def generate_holidays():
    if not session.get('is_superadmin'): return "Forbidden", 403
    year = int(request.json['year'])
    conn = get_db_connection()
    existing = conn.execute("SELECT COUNT(*) FROM public_holidays WHERE year=?", (year,)).fetchone()[0]
    if existing == 0:
        standard = [
            (f"{year}-01-01", "New Year's Day"),
            (f"{year}-03-21", "Human Rights Day"),
            (f"{year}-04-27", "Freedom Day"),
            (f"{year}-05-01", "Workers' Day"),
            (f"{year}-06-16", "Youth Day"),
            (f"{year}-08-09", "National Women's Day"),
            (f"{year}-09-24", "Heritage Day"),
            (f"{year}-12-16", "Day of Reconciliation"),
            (f"{year}-12-25", "Christmas Day"),
            (f"{year}-12-26", "Day of Goodwill")
        ]
        for d, n in standard:
            conn.execute("INSERT INTO public_holidays (year, date_str, name) VALUES (?, ?, ?)", (year, d, n))
        conn.commit()
    conn.close()
    return jsonify({"status": "success"})

@app.route('/admin/audit_logs', methods=['GET'])
def get_audit_logs():
    if not session.get('is_superadmin') and not session.get('is_company_admin'): return "Forbidden", 403
    conn = get_db_connection()
    logs = conn.execute('''SELECT * FROM audit_logs 
                           WHERE company_id=? AND timestamp >= datetime('now', '-90 days') 
                           ORDER BY timestamp DESC''', (session['company_id'],)).fetchall()
    conn.close()
    return jsonify([dict(l) for l in logs])

@app.route('/admin/switch_company', methods=['POST'])
def switch_company():
    if not session.get('is_superadmin'): return "Forbidden", 403
    conn = get_db_connection()
    comp = conn.execute('SELECT * FROM companies WHERE id = ?', (request.get_json()['company_id'],)).fetchone()
    conn.close()
    if comp:
        session['company_id'] = comp['id']
        session['company_name'] = comp['name']
        session['company_logo'] = comp['logo_file']
        session['comp_can_booking'] = bool(comp['can_booking'])
        session['comp_can_finance'] = bool(comp['can_finance'])
        session['comp_can_payroll'] = bool(comp['can_payroll'])
        session['comp_can_invoicing'] = bool(dict(comp).get('can_invoicing', 0))
        session['comp_can_accounting'] = bool(dict(comp).get('can_accounting', 0))
        session['comp_google_calendar'] = bool(dict(comp).get('google_calendar_sync', 0))
    return jsonify({"status": "success"})

@app.route('/admin/companies', methods=['GET'])
def get_companies():
    if not session.get('is_superadmin'): return "Forbidden", 403
    conn = get_db_connection()
    comps = conn.execute('SELECT * FROM companies ORDER BY name ASC').fetchall()
    conn.close()
    return jsonify([dict(c) for c in comps])

@app.route('/admin/companies/save', methods=['POST'])
def save_company():
    if not session.get('is_superadmin'): return "Forbidden", 403
    c_id = (request.form.get('id') or '').strip()
    c_name = (request.form.get('name') or '').strip()
    if not c_name:
        return jsonify({"status": "error", "message": "Company name is required."}), 400

    c_trans = request.form.get('transport_policy', 'yes')
    c_cb = 1 if request.form.get('can_booking') == 'true' else 0
    c_cf = 1 if request.form.get('can_finance') == 'true' else 0
    c_cp = 1 if request.form.get('can_payroll') == 'true' else 0
    c_ci = 1 if request.form.get('can_invoicing') == 'true' else 0
    c_ca = 1 if request.form.get('can_accounting') == 'true' else 0
    c_gcal = 1 if request.form.get('google_calendar_sync') == 'true' else 0
    try:
        c_transport_per_lift = max(0.0, float(request.form.get('transport_amount_per_lift') or 25))
    except (TypeError, ValueError):
        c_transport_per_lift = 25.0
    c_industry = get_valid_industry_template_name(request.form.get('industry_template', 'Cleaning'))
    c_address = request.form.get('address', '')
    c_reg_no = request.form.get('registration_number', '')
    c_vat_no = request.form.get('vat_number', '')

    logo = request.files.get('logo')
    filename = None
    saved_logo_path = None
    if logo and logo.filename:
        original_name = secure_filename(logo.filename)
        if original_name:
            logo_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'logos')
            os.makedirs(logo_dir, exist_ok=True)
            filename = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{original_name}"
            saved_logo_path = os.path.join(logo_dir, filename)
            logo.save(saved_logo_path)
            if not os.path.exists(saved_logo_path) or os.path.getsize(saved_logo_path) <= 0:
                return jsonify({"status": "error", "message": "Logo upload failed. Please try again."}), 500

    conn = get_db_connection()
    target_company_id = c_id or None
    try:
        if c_id:
            if filename:
                conn.execute('UPDATE companies SET name=?, logo_file=?, transport_policy=?, transport_amount_per_lift=?, can_booking=?, can_finance=?, can_payroll=?, can_invoicing=?, can_accounting=?, google_calendar_sync=?, address=?, registration_number=?, vat_number=?, industry_template=? WHERE id=?', (c_name, filename, c_trans, c_transport_per_lift, c_cb, c_cf, c_cp, c_ci, c_ca, c_gcal, c_address, c_reg_no, c_vat_no, c_industry, c_id))
            else:
                conn.execute('UPDATE companies SET name=?, transport_policy=?, transport_amount_per_lift=?, can_booking=?, can_finance=?, can_payroll=?, can_invoicing=?, can_accounting=?, google_calendar_sync=?, address=?, registration_number=?, vat_number=?, industry_template=? WHERE id=?', (c_name, c_trans, c_transport_per_lift, c_cb, c_cf, c_cp, c_ci, c_ca, c_gcal, c_address, c_reg_no, c_vat_no, c_industry, c_id))
        else:
            cur = conn.execute('INSERT INTO companies (name, logo_file, transport_policy, transport_amount_per_lift, can_booking, can_finance, can_payroll, can_invoicing, can_accounting, google_calendar_sync, address, registration_number, vat_number, industry_template) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', (c_name, filename, c_trans, c_transport_per_lift, c_cb, c_cf, c_cp, c_ci, c_ca, c_gcal, c_address, c_reg_no, c_vat_no, c_industry))
            target_company_id = getattr(cur, 'lastrowid', None)
            if not target_company_id:
                row = conn.execute('SELECT id FROM companies WHERE name=? ORDER BY id DESC LIMIT 1', (c_name,)).fetchone()
                target_company_id = row['id'] if row else None

        ensure_tenant_template(conn, target_company_id, c_industry, force_reset=False)
        conn.commit()

        saved_company = conn.execute('SELECT * FROM companies WHERE id=?', (target_company_id,)).fetchone() if target_company_id else None
    except Exception as exc:
        try: conn.rollback()
        except Exception: pass
        # Remove orphaned file if the database update failed.
        if saved_logo_path and os.path.exists(saved_logo_path):
            try: os.remove(saved_logo_path)
            except Exception: pass
        msg = str(exc)
        if 'unique' in msg.lower() or 'duplicate' in msg.lower():
            return jsonify({"status": "error", "message": "Company name already exists."})
        return jsonify({"status": "error", "message": f"Company could not be saved: {msg}"}), 500
    finally:
        conn.close()

    if filename:
        _clear_pdf_logo_cache()

    # Refresh the session whenever the edited/new company is the active company.
    if str(session.get('company_id')) == str(target_company_id):
        session['company_logo'] = filename or (dict(saved_company).get('logo_file') if saved_company else None)
        session['comp_can_booking'] = bool(c_cb)
        session['comp_can_finance'] = bool(c_cf)
        session['comp_can_payroll'] = bool(c_cp)
        session['comp_can_invoicing'] = bool(c_ci)
        session['comp_can_accounting'] = bool(c_ca)
        session['comp_google_calendar'] = bool(c_gcal)

    logo_file = dict(saved_company).get('logo_file') if saved_company else filename
    return jsonify({
        "status": "success",
        "company_id": target_company_id,
        "logo_file": logo_file or '',
        "logo_url": _safe_logo_url(logo_file) if logo_file else ''
    })

@app.route('/admin/industry_templates', methods=['GET'])
def admin_industry_templates():
    if not session.get('is_superadmin') and not session.get('is_company_admin'):
        return "Forbidden", 403
    conn = get_db_connection()
    templates = get_safe_industry_templates(conn)
    conn.close()
    return jsonify(templates)


@app.route('/admin/industry_templates/upload', methods=['POST'])
def admin_upload_industry_template():
    if not session.get('is_superadmin'):
        return "Forbidden", 403
    upload = request.files.get('template_file')
    raw_text = ''
    if upload and upload.filename:
        if not upload.filename.lower().endswith('.json'):
            return jsonify({"status": "error", "message": "Please upload a JSON template file."}), 400
        raw_text = upload.read().decode('utf-8')
    else:
        raw_text = request.form.get('template_json') or ''
    if not raw_text.strip():
        return jsonify({"status": "error", "message": "Template JSON is required."}), 400
    try:
        payload = json.loads(raw_text)
    except Exception as exc:
        return jsonify({"status": "error", "message": f"Invalid JSON: {exc}"}), 400

    template_name = (request.form.get('template_name') or payload.get('template_name') or payload.get('name') or '').strip()
    if not template_name:
        return jsonify({"status": "error", "message": "Template name is required in the file or form."}), 400
    if template_name in ['Cleaning', 'Transportation', 'Construction', 'Finance Auditing', 'Custom']:
        return jsonify({"status": "error", "message": "Built-in templates cannot be overwritten. Use a different template name."}), 400

    try:
        normalised = _normalise_template_config(payload)
    except Exception as exc:
        return jsonify({"status": "error", "message": f"Could not read template: {exc}"}), 400

    safe_payload = {
        "labels": normalised.get("labels", {}),
        "booking_fields": [
            {"field_key": f[0], "field_label": f[1], "field_type": f[2], "required": f[3], "visible": f[4]}
            for f in normalised.get("booking_fields", [])
        ],
        "finance_categories": normalised.get("finance_categories", []),
        "evaluation_scorecard": normalised.get("evaluation_scorecard", [])
    }
    conn = get_db_connection()
    conn.execute("""INSERT INTO industry_template_uploads (template_name, template_json, active, created_by, created_at)
                    VALUES (?, ?, 1, ?, ?)
                    ON CONFLICT(template_name) DO UPDATE SET
                        template_json=excluded.template_json, active=1, created_by=excluded.created_by, created_at=excluded.created_at""",
                 (template_name, json.dumps(safe_payload), session.get('username'), datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "Template uploaded successfully.", "template_name": template_name})

@app.route('/admin/tenant_setup/<int:company_id>', methods=['GET'])
def admin_get_tenant_setup(company_id):
    if not session.get('is_superadmin') and not session.get('is_company_admin'):
        return "Forbidden", 403
    if not session.get('is_superadmin') and int(session.get('company_id')) != int(company_id):
        return "Forbidden", 403
    conn = get_db_connection()
    comp = conn.execute('SELECT * FROM companies WHERE id=?', (company_id,)).fetchone()
    if not comp:
        conn.close()
        return jsonify({"status": "error", "message": "Company not found"}), 404
    ensure_tenant_template(conn, company_id, dict(comp).get('industry_template') or 'Cleaning', force_reset=False)
    conn.commit()
    labels = conn.execute('SELECT label_key, label_value FROM tenant_labels WHERE company_id=? ORDER BY label_key', (company_id,)).fetchall()
    fields = conn.execute("""SELECT id, module_name, field_key, field_label, field_type, required, visible, sort_order, options_json
                             FROM tenant_custom_fields WHERE company_id=? AND module_name='booking' ORDER BY sort_order, id""", (company_id,)).fetchall()
    cats = conn.execute("""SELECT id, category_name, category_type, active, sort_order
                           FROM finance_categories WHERE company_id=? ORDER BY sort_order, category_name""", (company_id,)).fetchall()
    scorecard = get_tenant_scorecard_template(conn, company_id)
    conn.close()
    return jsonify({
        "status": "success",
        "company": dict(comp),
        "labels": {r['label_key']: r['label_value'] for r in labels},
        "booking_fields": [dict(r) for r in fields],
        "finance_categories": [dict(r) for r in cats],
        "evaluation_scorecard": scorecard
    })

@app.route('/admin/tenant_setup/save', methods=['POST'])
def admin_save_tenant_setup():
    if not session.get('is_superadmin') and not session.get('is_company_admin'):
        return "Forbidden", 403
    data = request.get_json() or {}
    company_id = int(data.get('company_id') or 0)
    if not company_id:
        return jsonify({"status": "error", "message": "Company is required"}), 400
    if not session.get('is_superadmin') and int(session.get('company_id')) != company_id:
        return "Forbidden", 403

    industry = data.get('industry_template') or 'Cleaning'

    conn = get_db_connection()
    industry = get_valid_industry_template_name(industry, conn)
    if not conn.execute('SELECT id FROM companies WHERE id=?', (company_id,)).fetchone():
        conn.close()
        return jsonify({"status": "error", "message": "Company not found"}), 404

    apply_defaults = bool(data.get('apply_defaults'))
    ensure_tenant_template(conn, company_id, industry, force_reset=apply_defaults)
    conn.execute('UPDATE companies SET industry_template=?, setup_complete=1 WHERE id=?', (industry, company_id))

    labels = data.get('labels') or {}
    for key, value in labels.items():
        key = (key or '').strip()
        value = (value or '').strip()
        if not key or not value:
            continue
        conn.execute("""INSERT INTO tenant_labels (company_id, module_name, label_key, label_value)
                        VALUES (?, 'global', ?, ?)
                        ON CONFLICT(company_id, label_key) DO UPDATE SET label_value=excluded.label_value""",
                     (company_id, key, value))

    existing_dropdown_options = {}
    existing_field_rows = conn.execute("""SELECT id, field_key, options_json
                                        FROM tenant_custom_fields
                                        WHERE company_id=? AND module_name='booking'""", (company_id,)).fetchall()
    for existing_field in existing_field_rows:
        option_rows = get_custom_field_options(conn, company_id, existing_field['id'], include_inactive=True)
        if option_rows:
            existing_dropdown_options[existing_field['field_key']] = option_rows
        else:
            fallback_opts = parse_options_json(existing_field['options_json'])
            if fallback_opts:
                existing_dropdown_options[existing_field['field_key']] = [
                    {"option_label": opt, "option_value": opt, "sort_order": idx, "active": 1}
                    for idx, opt in enumerate(fallback_opts, start=1)
                ]

    conn.execute("DELETE FROM tenant_custom_field_options WHERE company_id=? AND field_id IN (SELECT id FROM tenant_custom_fields WHERE company_id=? AND module_name='booking')", (company_id, company_id))
    conn.execute("DELETE FROM tenant_custom_fields WHERE company_id=? AND module_name='booking'", (company_id,))
    for idx, field in enumerate(data.get('booking_fields') or [], start=1):
        label = (field.get('field_label') or '').strip()
        if not label:
            continue
        key = (field.get('field_key') or label.lower().replace(' ', '_').replace('/', '_').replace('-', '_'))[:80]
        ftype = field.get('field_type') or 'text'
        if ftype not in ['text', 'number', 'date', 'time', 'dropdown', 'checkbox', 'textarea']:
            ftype = 'text'
        options = field.get('options_json') or '[]'
        if isinstance(options, list):
            options = json.dumps(options)
        cur_field = conn.execute("""INSERT INTO tenant_custom_fields
                        (company_id, module_name, field_key, field_label, field_type, required, visible, sort_order, options_json, industry_template)
                        VALUES (?, 'booking', ?, ?, ?, ?, ?, ?, ?, ?)""",
                     (company_id, key, label, ftype, 1 if field.get('required') else 0, 1 if field.get('visible') else 0, idx, options, industry))
        if ftype == 'dropdown' and key in existing_dropdown_options:
            for opt_idx, opt in enumerate(existing_dropdown_options.get(key) or [], start=1):
                opt_label = (opt.get('option_label') or opt.get('option_value') or '').strip()
                if not opt_label:
                    continue
                opt_value = normalise_dropdown_option_value(opt.get('option_value') or opt_label)
                conn.execute("""INSERT OR IGNORE INTO tenant_custom_field_options
                                (company_id, field_id, option_label, option_value, sort_order, active)
                                VALUES (?, ?, ?, ?, ?, ?)""",
                             (company_id, cur_field.lastrowid, opt_label, opt_value, int(opt.get('sort_order') or opt_idx), 1 if opt.get('active', 1) else 0))

    conn.execute('DELETE FROM finance_categories WHERE company_id=?', (company_id,))
    for idx, cat in enumerate(data.get('finance_categories') or [], start=1):
        name = (cat.get('category_name') if isinstance(cat, dict) else str(cat)).strip()
        if not name:
            continue
        ctype = cat.get('category_type', 'expense') if isinstance(cat, dict) else 'expense'
        active = 1 if (not isinstance(cat, dict) or cat.get('active', True)) else 0
        conn.execute("""INSERT OR IGNORE INTO finance_categories (company_id, category_name, category_type, active, industry_template, sort_order)
                        VALUES (?, ?, ?, ?, ?, ?)""", (company_id, name, ctype, active, industry, idx))

    old_blocks = conn.execute('SELECT id FROM tenant_scorecard_blocks WHERE company_id=?', (company_id,)).fetchall()
    for block in old_blocks:
        conn.execute('DELETE FROM tenant_scorecard_questions WHERE block_id=?', (block['id'],))
    conn.execute('DELETE FROM tenant_scorecard_blocks WHERE company_id=?', (company_id,))
    for block_idx, block in enumerate(data.get('evaluation_scorecard') or [], start=1):
        block_name = (block.get('block_name') or '').strip()
        if not block_name:
            continue
        cur = conn.execute("""INSERT INTO tenant_scorecard_blocks (company_id, block_name, sort_order, active, industry_template)
                              VALUES (?, ?, ?, ?, ?)""", (company_id, block_name, block_idx, 1 if block.get('active', True) else 0, industry))
        block_id = cur.lastrowid
        for q_idx, question in enumerate(block.get('questions') or [], start=1):
            q_text = (question.get('question_text') or '').strip()
            if not q_text:
                continue
            try:
                max_score = int(question.get('max_score') or 10)
            except Exception:
                max_score = 10
            max_score = max(1, min(max_score, 100))
            conn.execute("""INSERT INTO tenant_scorecard_questions (company_id, block_id, question_key, question_text, max_score, sort_order, active)
                            VALUES (?, ?, ?, ?, ?, ?, ?)""", (company_id, block_id, f"q{block_idx}_{q_idx}", q_text, max_score, q_idx, 1 if question.get('active', True) else 0))

    switch_company_after_setup = False
    if session.get('is_superadmin'):
        comp = conn.execute('SELECT * FROM companies WHERE id=?', (company_id,)).fetchone()
        if comp:
            session['company_id'] = comp['id']
            session['company_name'] = comp['name']
            session['company_logo'] = comp['logo_file']
            session['comp_can_booking'] = bool(dict(comp).get('can_booking', 0))
            session['comp_can_finance'] = bool(dict(comp).get('can_finance', 0))
            session['comp_can_payroll'] = bool(dict(comp).get('can_payroll', 0))
            session['comp_can_invoicing'] = bool(dict(comp).get('can_invoicing', 0))
            session['comp_can_accounting'] = bool(dict(comp).get('can_accounting', 0))
            session['comp_google_calendar'] = bool(dict(comp).get('google_calendar_sync', 0))
            switch_company_after_setup = True

    conn.commit()
    conn.close()
    log_action('System Admin', 'Updated Tenant Setup', f"Updated industry setup for company ID {company_id}")
    return jsonify({"status": "success", "switched_company": switch_company_after_setup})

@app.route('/admin/custom_field_settings/data', methods=['GET'])
def admin_custom_field_settings_data():
    if not session.get('is_superadmin') and not session.get('is_company_admin'):
        return "Forbidden", 403
    conn = get_db_connection()
    if session.get('is_superadmin'):
        companies = [dict(c) for c in conn.execute('SELECT id, name, industry_template FROM companies ORDER BY name').fetchall()]
        requested_company_id = request.args.get('company_id')
        current_company_id = session.get('company_id')
        valid_company_ids = {int(c['id']) for c in companies}
        if requested_company_id:
            company_id = int(requested_company_id)
        elif current_company_id and int(current_company_id) in valid_company_ids:
            company_id = int(current_company_id)
        else:
            company_id = companies[0]['id'] if companies else 0
    else:
        company_id = int(session.get('company_id'))
        comp = conn.execute('SELECT id, name, industry_template FROM companies WHERE id=?', (company_id,)).fetchone()
        companies = [dict(comp)] if comp else []
    fields = []
    if company_id:
        ensure_tenant_template(conn, company_id, None, force_reset=False)
        conn.commit()
        fields = [dict(f) for f in conn.execute("""SELECT id, field_key, field_label, field_type, required, visible, sort_order
                                                 FROM tenant_custom_fields
                                                 WHERE company_id=? AND module_name='booking'
                                                 ORDER BY sort_order, field_label""", (company_id,)).fetchall()]
    conn.close()
    return jsonify({"status": "success", "companies": companies, "company_id": company_id, "fields": fields})


@app.route('/admin/custom_field_options/<int:field_id>', methods=['GET'])
def admin_get_custom_field_options(field_id):
    if not session.get('is_superadmin') and not session.get('is_company_admin'):
        return "Forbidden", 403
    conn = get_db_connection()
    field = conn.execute("""SELECT id, company_id, field_label, field_type
                            FROM tenant_custom_fields
                            WHERE id=? AND module_name='booking'""", (field_id,)).fetchone()
    if not field:
        conn.close()
        return jsonify({"status": "error", "message": "Custom field not found"}), 404
    if not session.get('is_superadmin') and int(field['company_id']) != int(session.get('company_id')):
        conn.close()
        return "Forbidden", 403
    if field['field_type'] != 'dropdown':
        conn.close()
        return jsonify({"status": "error", "message": "Only dropdown fields have options"}), 400
    options = get_custom_field_options(conn, field['company_id'], field_id, include_inactive=True)
    conn.close()
    return jsonify({"status": "success", "field": dict(field), "options": options})


@app.route('/admin/custom_field_options/save', methods=['POST'])
def admin_save_custom_field_options():
    if not session.get('is_superadmin') and not session.get('is_company_admin'):
        return "Forbidden", 403
    data = request.get_json() or {}
    field_id = int(data.get('field_id') or 0)
    options = data.get('options') or []
    conn = get_db_connection()
    field = conn.execute("""SELECT id, company_id, field_label, field_type
                            FROM tenant_custom_fields
                            WHERE id=? AND module_name='booking'""", (field_id,)).fetchone()
    if not field:
        conn.close()
        return jsonify({"status": "error", "message": "Custom field not found"}), 404
    if not session.get('is_superadmin') and int(field['company_id']) != int(session.get('company_id')):
        conn.close()
        return "Forbidden", 403
    if field['field_type'] != 'dropdown':
        conn.close()
        return jsonify({"status": "error", "message": "Only dropdown fields have options"}), 400

    cleaned = []
    seen = set()
    for idx, opt in enumerate(options, start=1):
        label = (opt.get('option_label') if isinstance(opt, dict) else str(opt)).strip()
        if not label:
            continue
        value = normalise_dropdown_option_value(label)
        key = value.lower()
        if key in seen:
            continue
        seen.add(key)
        active = 1 if (not isinstance(opt, dict) or opt.get('active', True)) else 0
        cleaned.append((field['company_id'], field_id, label, value, idx, active))

    conn.execute('DELETE FROM tenant_custom_field_options WHERE company_id=? AND field_id=?', (field['company_id'], field_id))
    for row in cleaned:
        conn.execute("""INSERT INTO tenant_custom_field_options
                        (company_id, field_id, option_label, option_value, sort_order, active)
                        VALUES (?, ?, ?, ?, ?, ?)""", row)
    # Keep options_json as a fallback for older code paths and older patch installs.
    conn.execute('UPDATE tenant_custom_fields SET options_json=? WHERE id=? AND company_id=?',
                 (json.dumps([r[2] for r in cleaned if r[5]]), field_id, field['company_id']))
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "Dropdown options saved."})


@app.route('/admin/export_company_data', methods=['GET'])
def admin_export_company_data():
    if not session.get('is_company_admin') and not session.get('is_superadmin'):
        return "Forbidden: only Company Admins or Super Admins can export company data.", 403

    try:
        cid = get_active_company_id_for_admin()
    except Exception:
        return "Invalid company selected.", 400
    if not cid:
        return "No active company selected.", 400

    conn = get_db_connection()
    company = conn.execute('SELECT * FROM companies WHERE id=?', (cid,)).fetchone()
    if not company:
        conn.close()
        return "Company not found.", 404

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        direct_tables = [
            'companies', 'clients', 'employees', 'services', 'bookings', 'expenses', 'payslips',
            'leave_records', 'interviews', 'invoices', 'quotes', 'settings', 'users', 'audit_logs',
            'tenant_labels', 'tenant_custom_fields', 'custom_field_values', 'tenant_custom_field_options',
            'finance_categories', 'tenant_scorecard_blocks'
        ]
        for table in direct_tables:
            if not table_exists(conn, table):
                continue
            if table == 'companies':
                rows = [dict(r) for r in conn.execute('SELECT * FROM companies WHERE id=?', (cid,)).fetchall()]
            elif table == 'users':
                rows = [dict(r) for r in conn.execute('SELECT id, username, company_id, can_booking, can_finance, can_payroll, can_invoicing, can_accounting, is_superadmin, is_company_admin FROM users WHERE company_id=?', (cid,)).fetchall()]
            else:
                rows = query_company_table(conn, table, cid)
            zf.writestr(f'{table}.csv', rows_to_csv_bytes(rows))

        if table_exists(conn, 'invoice_items') and table_exists(conn, 'invoices'):
            rows = [dict(r) for r in conn.execute("""SELECT ii.* FROM invoice_items ii
                                                    JOIN invoices i ON i.id = ii.invoice_id
                                                    WHERE i.company_id=?""", (cid,)).fetchall()]
            zf.writestr('invoice_items.csv', rows_to_csv_bytes(rows))
        if table_exists(conn, 'quote_items') and table_exists(conn, 'quotes'):
            rows = [dict(r) for r in conn.execute("""SELECT qi.* FROM quote_items qi
                                                    JOIN quotes q ON q.id = qi.quote_id
                                                    WHERE q.company_id=?""", (cid,)).fetchall()]
            zf.writestr('quote_items.csv', rows_to_csv_bytes(rows))
        if table_exists(conn, 'tenant_scorecard_questions') and table_exists(conn, 'tenant_scorecard_blocks'):
            rows = [dict(r) for r in conn.execute("""SELECT q.* FROM tenant_scorecard_questions q
                                                    JOIN tenant_scorecard_blocks b ON b.id = q.block_id
                                                    WHERE b.company_id=?""", (cid,)).fetchall()]
            zf.writestr('tenant_scorecard_questions.csv', rows_to_csv_bytes(rows))

        readme = f"Company-scoped export for {company['name']} (company_id={cid}) generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}. This export intentionally excludes other tenants' data and does not include the raw database file.\n"
        zf.writestr('README.txt', readme.encode('utf-8'))

    conn.close()
    log_action('Hub', 'Exported Company Data', f"Exported company-scoped database data for company_id={cid}")
    zip_buffer.seek(0)
    safe_name = ''.join(c if c.isalnum() or c in ('-', '_') else '_' for c in (company['name'] or 'company'))
    response = Response(zip_buffer.getvalue(), mimetype='application/zip')
    response.headers['Content-Disposition'] = f'attachment; filename={safe_name}_company_data_export.zip'
    return response


@app.route('/admin/import_template/<import_type>', methods=['GET'])
def admin_import_template(import_type):
    if not session.get('is_company_admin') and not session.get('is_superadmin'):
        return "Forbidden", 403
    rows = get_import_template_rows(import_type)
    if rows is None:
        return "Invalid import type.", 400
    response = Response(rows_to_csv_bytes(rows), mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename={import_type}_import_template.csv'
    return response


@app.route('/admin/import_company_data', methods=['POST'])
def admin_import_company_data():
    if not session.get('is_company_admin') and not session.get('is_superadmin'):
        return jsonify({"status": "error", "message": "Forbidden: only Company Admins or Super Admins can import company data."}), 403

    import_type = request.form.get('import_type', '').strip()
    validate_only = request.form.get('validate_only') in ('1', 'true', 'on', 'yes')
    if import_type not in COMPANY_IMPORT_TABLES:
        return jsonify({"status": "error", "message": "Invalid import type."}), 400
    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({"status": "error", "message": "Please upload a CSV file."}), 400

    try:
        cid = get_active_company_id_for_admin()
    except Exception:
        return jsonify({"status": "error", "message": "Invalid target company selected."}), 400
    if not cid:
        return jsonify({"status": "error", "message": "No target company selected."}), 400

    uploaded = request.files['file']
    try:
        content = uploaded.read().decode('utf-8-sig')
    except UnicodeDecodeError:
        return jsonify({"status": "error", "message": "Could not read CSV. Please save it as UTF-8 CSV and try again."}), 400

    reader = csv.DictReader(io.StringIO(content))
    if not reader.fieldnames:
        return jsonify({"status": "error", "message": "CSV has no header row."}), 400

    allowed_fields = COMPANY_IMPORT_TABLES[import_type]
    present_fields = [f for f in allowed_fields if f in reader.fieldnames]
    if not present_fields:
        return jsonify({"status": "error", "message": f"CSV does not contain any valid fields for {import_type}. Download the template and try again."}), 400

    conn = get_db_connection()
    if not conn.execute('SELECT id FROM companies WHERE id=?', (cid,)).fetchone():
        conn.close()
        return jsonify({"status": "error", "message": "Target company not found."}), 404

    errors = []
    rows_to_insert = []
    for line_no, row in enumerate(reader, start=2):
        cleaned = {field: normalise_import_value(row.get(field)) for field in present_fields}
        if not any(cleaned.values()):
            continue
        if import_type in ('clients', 'employees', 'services') and not cleaned.get('name'):
            errors.append(f"Line {line_no}: name is required.")
            continue
        if import_type == 'bookings' and (not cleaned.get('title') or not cleaned.get('start')):
            errors.append(f"Line {line_no}: title and start are required.")
            continue
        if import_type == 'expenses' and (not cleaned.get('date') or not cleaned.get('amount')):
            errors.append(f"Line {line_no}: date and amount are required.")
            continue
        if import_type == 'leave_records' and (not cleaned.get('employee_id') or not cleaned.get('date_taken')):
            errors.append(f"Line {line_no}: employee_id and date_taken are required.")
            continue
        rows_to_insert.append(cleaned)

    if validate_only:
        conn.close()
        return jsonify({"status": "success", "message": "Validation complete.", "valid_rows": len(rows_to_insert), "errors": errors})
    if errors:
        conn.close()
        return jsonify({"status": "error", "message": "Import has validation errors. Fix the CSV and try again.", "errors": errors[:50]}), 400

    try:
        table = safe_table_name(import_type)
        table_cols = set(table_columns(conn, table))
        inserted = 0
        for row in rows_to_insert:
            cols = [c for c in row.keys() if c in table_cols]
            if 'company_id' in table_cols:
                cols.append('company_id')
            values = [row[c] for c in row.keys() if c in table_cols]
            if 'company_id' in table_cols:
                values.append(cid)
            placeholders = ','.join(['?'] * len(cols))
            conn.execute(f"INSERT INTO {table} ({','.join(cols)}) VALUES ({placeholders})", values)
            inserted += 1
        conn.commit()
    except Exception as exc:
        conn.rollback()
        conn.close()
        return jsonify({"status": "error", "message": f"Import failed: {exc}"}), 500

    conn.close()
    log_action('Hub', 'Imported Company Data', f"Imported {inserted} {import_type} rows for company_id={cid}")
    return jsonify({"status": "success", "message": f"Imported {inserted} row(s).", "inserted": inserted})


@app.route('/admin/export_full_database', methods=['GET'])
def admin_export_full_database():
    if not session.get('is_superadmin'):
        return "Forbidden: only Super Admins can export the full database.", 403
    if is_postgres_enabled():
        return "Full raw database export is disabled when Easy Admin is connected to Supabase PostgreSQL. Use Supabase backups/exports instead.", 400
    db_path = get_database_path()
    if not os.path.exists(db_path):
        return "Database file not found.", 404

    zip_buffer = io.BytesIO()
    conn = sqlite3.connect(db_path)
    schema_rows = conn.execute("SELECT sql FROM sqlite_master WHERE sql IS NOT NULL ORDER BY type, name").fetchall()
    schema_sql = '\n\n'.join(row[0] + ';' for row in schema_rows)
    dump_sql = '\n'.join(conn.iterdump())
    conn.close()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(db_path, 'database.db')
        zf.writestr('schema.sql', schema_sql.encode('utf-8'))
        zf.writestr('full_dump.sql', dump_sql.encode('utf-8'))
        zf.writestr('README.txt', f"Full raw database backup generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}. Contains all tenants and system data. Super Admin only.\n".encode('utf-8'))

    log_action('Hub', 'Exported Full Database Backup', 'Super Admin exported full raw database backup.')
    zip_buffer.seek(0)
    response = Response(zip_buffer.getvalue(), mimetype='application/zip')
    response.headers['Content-Disposition'] = f"attachment; filename=full_database_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return response


@app.route('/admin/import_full_database', methods=['POST'])
def admin_import_full_database():
    if not session.get('is_superadmin'):
        return jsonify({"status": "error", "message": "Forbidden: only Super Admins can restore the full database."}), 403
    if is_postgres_enabled():
        return jsonify({"status": "error", "message": "Full raw database restore is disabled when Easy Admin is connected to Supabase PostgreSQL. Use Supabase backup/restore tools instead."}), 400
    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({"status": "error", "message": "Please upload a .db file or a backup ZIP containing database.db."}), 400

    uploaded = request.files['file']
    backup_dir = 'backups'
    os.makedirs(backup_dir, exist_ok=True)
    safety_backup = os.path.join(backup_dir, f"database_before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
    if os.path.exists(get_database_path()):
        shutil.copy2(get_database_path(), safety_backup)

    temp_dir = os.path.join(backup_dir, f"restore_upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    os.makedirs(temp_dir, exist_ok=True)
    temp_upload = os.path.join(temp_dir, secure_filename(uploaded.filename))
    uploaded.save(temp_upload)

    try:
        restore_db = temp_upload
        if zipfile.is_zipfile(temp_upload):
            with zipfile.ZipFile(temp_upload, 'r') as zf:
                db_members = [m for m in zf.namelist() if os.path.basename(m) == 'database.db' or m.lower().endswith('.db')]
                if not db_members:
                    return jsonify({"status": "error", "message": "Backup ZIP does not contain a .db file."}), 400
                zf.extract(db_members[0], temp_dir)
                restore_db = os.path.join(temp_dir, db_members[0])
        test_conn = sqlite3.connect(restore_db)
        test_conn.execute('PRAGMA integrity_check').fetchone()
        test_conn.close()
        shutil.copy2(restore_db, get_database_path())
    except Exception as exc:
        return jsonify({"status": "error", "message": f"Restore failed: {exc}. Safety backup: {safety_backup if os.path.exists(safety_backup) else 'none'}"}), 500

    log_action('Hub', 'Restored Full Database Backup', f"Super Admin restored full database. Safety backup saved to {safety_backup}.")
    return jsonify({"status": "success", "message": "Full database restored. Log out and back in, or restart the app, to ensure all sessions use the restored data.", "safety_backup": safety_backup})


@app.route('/admin/users/save', methods=['POST'])
def save_user():
    if not session.get('is_superadmin') and not session.get('is_company_admin'): return "Forbidden", 403
    data = request.get_json()
    user_email = (data.get('email') or '').strip()
    conn = get_db_connection()
    if not session.get('is_superadmin'):
        cid = session['company_id']
        is_comp_admin = 0 
        comp = conn.execute('SELECT can_booking, can_finance, can_payroll, can_invoicing, can_accounting FROM companies WHERE id=?', (cid,)).fetchone()
        comp_dict = dict(comp)
        can_b = 1 if data.get('can_booking') and comp_dict.get('can_booking') else 0
        can_f = 1 if data.get('can_finance') and comp_dict.get('can_finance') else 0
        can_p = 1 if data.get('can_payroll') and comp_dict.get('can_payroll') else 0
        can_i = 1 if data.get('can_invoicing') and comp_dict.get('can_invoicing') else 0
        can_a = 1 if data.get('can_accounting') and comp_dict.get('can_accounting') else 0
    else:
        cid = data.get('company_id')
        is_comp_admin = 1 if data.get('is_company_admin') else 0
        can_b = 1 if data.get('can_booking') else 0
        can_f = 1 if data.get('can_finance') else 0
        can_p = 1 if data.get('can_payroll') else 0
        can_i = 1 if data.get('can_invoicing') else 0
        can_a = 1 if data.get('can_accounting') else 0

    action_msg = None
    try:
        if data.get('id'):
            if data.get('password'): conn.execute('UPDATE users SET username=?, email=?, company_id=?, password_hash=?, can_booking=?, can_finance=?, can_payroll=?, can_invoicing=?, can_accounting=?, is_company_admin=? WHERE id=? AND is_superadmin=0', (data['username'], user_email, cid, generate_password_hash(data['password']), can_b, can_f, can_p, can_i, can_a, is_comp_admin, data['id']))
            else: conn.execute('UPDATE users SET username=?, email=?, company_id=?, can_booking=?, can_finance=?, can_payroll=?, can_invoicing=?, can_accounting=?, is_company_admin=? WHERE id=? AND is_superadmin=0', (data['username'], user_email, cid, can_b, can_f, can_p, can_i, can_a, is_comp_admin, data['id']))
            action_msg = ('System Admin', 'Updated User', f"Updated settings for user: {data['username']}")
        else:
            conn.execute('INSERT INTO users (username, email, company_id, password_hash, can_booking, can_finance, can_payroll, can_invoicing, can_accounting, is_superadmin, is_company_admin) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?)', (data['username'], user_email, cid, generate_password_hash(data['password'] or 'Password123'), can_b, can_f, can_p, can_i, can_a, is_comp_admin))
            action_msg = ('System Admin', 'Created User', f"Created new user account: {data['username']}")
        conn.commit()
    except sqlite3.IntegrityError: return jsonify({"status": "error", "message": "Username already exists."})
    finally: conn.close()
    
    if action_msg: log_action(action_msg[0], action_msg[1], action_msg[2])
    return jsonify({"status": "success"})

@app.route('/admin/users', methods=['GET'])
def get_users():
    if not session.get('is_superadmin') and not session.get('is_company_admin'): return "Forbidden", 403
    conn = get_db_connection()
    if session.get('is_superadmin'):
        users = conn.execute('''SELECT u.id, u.username, u.email, u.can_booking, u.can_finance, u.can_payroll, u.can_invoicing, u.can_accounting, u.is_superadmin, u.is_company_admin, u.company_id, c.name as company_name 
                                FROM users u LEFT JOIN companies c ON u.company_id = c.id 
                                ORDER BY u.is_superadmin DESC, u.username ASC''').fetchall()
    else:
        users = conn.execute('''SELECT u.id, u.username, u.email, u.can_booking, u.can_finance, u.can_payroll, u.can_invoicing, u.can_accounting, u.is_superadmin, u.is_company_admin, u.company_id, c.name as company_name 
                                FROM users u LEFT JOIN companies c ON u.company_id = c.id 
                                WHERE u.company_id = ? ORDER BY u.is_company_admin DESC, u.username ASC''', (session['company_id'],)).fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])

@app.route('/admin/users/delete', methods=['POST'])
def delete_user():
    if not session.get('is_superadmin') and not session.get('is_company_admin'): return "Forbidden", 403
    uid = request.get_json()['id']
    conn = get_db_connection()
    deleted_user = conn.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
    if session.get('is_superadmin'):
        conn.execute('DELETE FROM users WHERE id=? AND is_superadmin=0', (uid,))
    else:
        conn.execute('DELETE FROM users WHERE id=? AND is_superadmin=0 AND company_id=?', (uid, session['company_id']))
    conn.commit(); conn.close()
    if deleted_user: log_action('System Admin', 'Deleted User', f"Deleted account: {deleted_user['username']}")
    return jsonify({"status": "success"})


# ==========================================================
# 1. BOOKING & OPERATIONS ROUTES
# ==========================================================
@app.route('/booking')
def booking_index():
    conn = get_db_connection()
    cid = session['company_id']
    ensure_tenant_template(conn, cid)
    clients = prepare_client_options(conn.execute('SELECT * FROM clients WHERE company_id=? ORDER BY name ASC, surname ASC, id ASC', (cid,)).fetchall())
    employees = conn.execute("SELECT * FROM employees WHERE company_id=? AND (emp_type != 'Supplier' OR emp_type IS NULL)", (cid,)).fetchall()
    services = [dict(row) for row in conn.execute('SELECT * FROM services WHERE company_id=? ORDER BY name ASC', (cid,)).fetchall()]
    projects = [dict(row) for row in conn.execute('''SELECT p.*, c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name
                                                     FROM projects p
                                                     LEFT JOIN clients c ON c.id=p.client_id AND c.company_id=p.company_id
                                                     WHERE p.company_id=?
                                                     ORDER BY p.created_at DESC, p.project_name ASC''', (cid,)).fetchall()]
    tenant_labels = get_tenant_labels(conn, cid)
    booking_custom_fields = get_tenant_custom_fields(conn, cid, 'booking', visible_only=True)
    conn.commit()
    conn.close()
    return render_template('booking_index.html', clients=clients, employees=employees, services=services, projects=projects, session=session, tenant_labels=tenant_labels, booking_custom_fields=booking_custom_fields)

@app.route('/bookings')
def bookings():
    conn = get_db_connection()
    cid = session['company_id']
    start_range = (request.args.get('start') or '').split('T')[0]
    end_range = (request.args.get('end') or '').split('T')[0]
    search_q = (request.args.get('q') or '').strip()
    where = 'b.company_id=?'
    params = [cid]
    if start_range:
        where += ' AND COALESCE(b.start, "") >= ?'
        params.append(start_range)
    else:
        default_start = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        where += ' AND COALESCE(b.start, "") >= ?'
        params.append(default_start)
    if end_range:
        where += ' AND COALESCE(b.start, "") < ?'
        params.append(end_range)
    else:
        default_end = (datetime.now() + timedelta(days=90)).strftime('%Y-%m-%d')
        where += ' AND COALESCE(b.start, "") <= ?'
        params.append(default_end)
    if search_q:
        like = f'%{search_q}%'
        where += ' AND (COALESCE(b.title, "") LIKE ? OR COALESCE(c.name, "") LIKE ? OR COALESCE(c.surname, "") LIKE ? OR COALESCE(c.company_name, "") LIKE ? OR COALESCE(b.employee, "") LIKE ? OR COALESCE(b.booking_type, "") LIKE ? OR COALESCE(p.project_name, "") LIKE ?)'
        params.extend([like, like, like, like, like, like, like])
    bookings = conn.execute(f'''SELECT b.*, p.project_name, p.project_code,
                                      c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name,
                                      c.address AS client_address, c.building_number AS client_building_number, c.street_name AS client_street_name,
                                      c.suburb AS client_suburb, c.postal_code AS client_postal_code, c.phone AS client_phone, c.email AS client_email,
                                      (SELECT COUNT(*) FROM attachments a WHERE a.company_id=b.company_id AND a.linked_type='booking' AND a.linked_id=b.id) AS attachment_count
                              FROM bookings b
                              LEFT JOIN projects p ON p.id=b.project_id AND p.company_id=b.company_id
                              LEFT JOIN clients c ON c.id=b.client_id AND c.company_id=b.company_id
                              WHERE {where}
                              ORDER BY b.start ASC
                              LIMIT 1000''', params).fetchall()
    leave_where = 'lr.company_id=?'
    leave_params = [cid]
    if start_range:
        leave_where += ' AND COALESCE(lr.date_taken, "") >= ?'
        leave_params.append(start_range)
    if end_range:
        leave_where += ' AND COALESCE(lr.date_taken, "") < ?'
        leave_params.append(end_range)
    leave_rows = conn.execute(f'''SELECT lr.*, e.name AS employee_name
                                 FROM leave_records lr
                                 JOIN employees e ON e.id = lr.employee_id AND e.company_id = lr.company_id
                                 WHERE {leave_where}
                                 ORDER BY lr.date_taken ASC
                                 LIMIT 1000''', leave_params).fetchall()
    
    events = []
    for r in bookings:
        emp_display = r['employee'] or 'Unassigned'
        client_name = ' '.join([x for x in [r['client_first_name'], r['client_surname']] if x]).strip() or r['client_company_name'] or r['title'] or 'Unknown'
        project_suffix = f" | {r['project_name']}" if r['project_name'] else ''
        title = f"{client_name} - {emp_display} ({r['booking_type'] or 'Clean'}){project_suffix}"
        custom_values = get_custom_field_values_for_record(conn, cid, 'booking', r['id'])
        r_dict = dict(r)
        client_address = compose_client_address(
            r_dict.get('client_building_number'),
            r_dict.get('client_street_name'),
            r_dict.get('client_suburb'),
            r_dict.get('client_postal_code'),
            r_dict.get('client_address')
        )
        is_invoiced_raw = r_dict.get('is_invoiced')
        is_invoiced = str(is_invoiced_raw).strip().lower() in ('1', 'true', 'yes', 'y')
        
        events.append({
            "id": r["id"], 
            "title": title, 
            "start": r['start'] or '', 
            "color": "#0d6efd",
            "extendedProps": { 
                "event_type": "booking",
                "client": client_name, 
                "client_id": r['client_id'] or '',
                "client_address": client_address or '',
                "client_phone": r_dict.get('client_phone') or '',
                "client_email": r_dict.get('client_email') or '',
                "booking_title": r_dict.get('title') or client_name,
                "is_invoiced": is_invoiced,
                "employee": r['employee'] or 'Unassigned', 
                "booking_type": r['booking_type'], 
                "transport": r['transport'] or '', 
                "booking_notes": r['booking_notes'] or '', 
                "overtime_hours": float(r['overtime_hours'] or 0),
                "project_id": r['project_id'] or '',
                "project_name": r['project_name'] or '',
                "project_code": r['project_code'] or '',
                "mobile_status": r_dict.get('mobile_status') or 'Scheduled',
                "mobile_status_updated_at": r_dict.get('mobile_status_updated_at') or '',
                "mobile_status_updated_by": r_dict.get('mobile_status_updated_by') or '',
                "mobile_started_at": r_dict.get('mobile_started_at') or '',
                "mobile_completed_at": r_dict.get('mobile_completed_at') or '',
                "attachment_count": r_dict.get('attachment_count') or 0,
                "custom_fields": custom_values
            }
        })

    for lr in leave_rows:
        try:
            leave_start = datetime.strptime(lr['date_taken'], '%Y-%m-%d')
            days = max(1, int(float(lr['days'] or 1)))
        except Exception:
            continue
        for offset in range(days):
            leave_day = leave_start + timedelta(days=offset)
            events.append({
                "id": f"leave-{lr['id']}-{offset}",
                "title": f"{lr['employee_name']} - LEAVE",
                "start": leave_day.strftime('%Y-%m-%d'),
                "allDay": True,
                "color": "#fd7e14",
                "extendedProps": {
                    "event_type": "leave",
                    "employee": lr['employee_name'],
                    "leave_type": lr['leave_type'] or 'Leave'
                }
            })
    conn.close()
    response = jsonify(events)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    return response

@app.route('/booking_staff_hours')
def booking_staff_hours():
    date_str = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
    employee_name = request.args.get('employee') or ''
    exclude_booking_id = request.args.get('exclude_booking_id') or None
    try:
        overtime_hours = float(request.args.get('overtime_hours') or 0)
    except Exception:
        overtime_hours = 0.0

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d')
    except Exception:
        return jsonify({"status": "error", "message": "Invalid date"}), 400

    conn = get_db_connection()
    cid = session['company_id']

    if employee_name:
        emp = conn.execute('''SELECT * FROM employees
                              WHERE company_id=? AND name=?
                              AND (emp_type != 'Supplier' OR emp_type IS NULL)''',
                           (cid, employee_name)).fetchone()
        if not emp:
            conn.close()
            return jsonify({"status": "error", "message": "Employee not found"}), 404
        summary = get_booking_staff_hours_summary(conn, cid, emp, target_date, overtime_hours, exclude_booking_id, True)
        conn.close()
        return jsonify({"status": "success", "employee": summary})

    employees = conn.execute('''SELECT * FROM employees
                                WHERE company_id=?
                                AND (emp_type != 'Supplier' OR emp_type IS NULL)
                                ORDER BY name ASC''', (cid,)).fetchall()
    summaries = [get_booking_staff_hours_summary(conn, cid, emp, target_date, 0, exclude_booking_id, False) for emp in employees]
    conn.close()
    return jsonify({"status": "success", "employees": summaries})

def normalise_booking_project_id(conn, company_id, project_id):
    """Return a tenant-safe project id or None for unlinked bookings."""
    if project_id in (None, '', 0, '0'):
        return None
    try:
        project_id = int(project_id)
    except Exception:
        raise ValueError('Invalid project selected.')
    project = conn.execute('SELECT id, status FROM projects WHERE id=? AND company_id=?', (project_id, company_id)).fetchone()
    if not project:
        raise ValueError('Selected project was not found for this company.')
    if (project['status'] or '').lower() == 'cancelled':
        raise ValueError('Cancelled projects cannot be linked to new bookings.')
    return project_id

@app.route('/add', methods=['POST'])
def add_booking():
    data = request.get_json()
    conn = get_db_connection()
    try:
        project_id = normalise_booking_project_id(conn, session['company_id'], data.get('project_id'))
    except ValueError as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)}), 400
    try:
        client = resolve_client_from_payload(conn, session['company_id'], data, id_key='client_id', name_key='client')
        client_id = client['id']
        client_name = client_display_name(client)
    except ValueError as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)}), 400
    dt_str = f"{data['date']}T{data['time']}"
    try:
        booking_date = datetime.strptime(data['date'], '%Y-%m-%d')
    except Exception:
        conn.close()
        return jsonify({"status": "error", "message": "Invalid booking date"}), 400
    custom_fields = data.get('custom_fields') or {}
    booking_custom_fields = get_tenant_custom_fields(conn, session['company_id'], 'booking', visible_only=True)
    valid_custom, custom_message = validate_custom_field_payload(booking_custom_fields, custom_fields)
    if not valid_custom:
        conn.close()
        return jsonify({"status": "error", "message": custom_message}), 400

    assignments_to_check = data.get('assignments') if 'assignments' in data else [{"employee": data.get('employee', '')}]
    is_available, availability_message = validate_booking_employees_available(conn, session['company_id'], assignments_to_check, booking_date)
    if not is_available:
        conn.close()
        return jsonify({"status": "error", "message": availability_message}), 400
    
    booking_ids = []
    if 'assignments' in data:
        for req in data['assignments']:
            google_id = None
            if session.get('comp_google_calendar'):
                try: 
                    google_id = create_google_event(client_name, data['date'], data['time'], req['employee'], data.get('booking_type'), req.get('transport'), session['company_name'], session.get('company_id'))
                except Exception as e: 
                    print(f"Google Sync Error: {_format_google_calendar_error(e)}")

            cur = conn.execute('INSERT INTO bookings (company_id, client_id, title, start, employee, google_event_id, booking_type, transport, booking_notes, overtime_hours, is_invoiced, project_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                         (session['company_id'], client_id, client_name, dt_str, req['employee'], google_id, data.get('booking_type'), req.get('transport'), data.get('booking_notes', ''), float(req.get('overtime_hours', 0)), 0, project_id))
            booking_ids.append(cur.lastrowid)
            save_custom_field_values(conn, session['company_id'], 'booking', cur.lastrowid, custom_fields)
    else:
        google_id = None
        if session.get('comp_google_calendar'):
            try: google_id = create_google_event(client_name, data['date'], data['time'], data['employee'], data.get('booking_type'), data.get('transport'), session['company_name'], session.get('company_id'))
            except Exception as e: print(f"Google Sync Error: {_format_google_calendar_error(e)}")
                
        cur = conn.execute('INSERT INTO bookings (company_id, client_id, title, start, employee, google_event_id, booking_type, transport, booking_notes, overtime_hours, is_invoiced, project_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', 
                     (session['company_id'], client_id, client_name, dt_str, data['employee'], google_id, data.get('booking_type'), data.get('transport'), data.get('booking_notes', ''), float(data.get('overtime_hours', 0)), 0, project_id))
        booking_ids.append(cur.lastrowid)
        save_custom_field_values(conn, session['company_id'], 'booking', cur.lastrowid, custom_fields)
                     
    conn.commit()
    conn.close()
    
    log_action('Booking & Ops', 'Created Booking', f"Client: {client_name}, Date: {data['date']}")
    return jsonify({"status": "success", "booking_ids": booking_ids})

@app.route('/edit_booking', methods=['POST'])
def edit_booking():
    data = request.get_json(silent=True) or {}
    required_payload = ['id', 'client', 'date', 'time', 'booking_type', 'employee']
    missing_payload = [k for k in required_payload if str(data.get(k) or '').strip() == '']
    if missing_payload:
        return jsonify({"status": "error", "message": "Missing booking data: " + ", ".join(missing_payload)}), 400

    conn = get_db_connection()
    b = conn.execute("SELECT google_event_id FROM bookings WHERE id=? AND company_id=?", (data['id'], session['company_id'])).fetchone()
    if not b:
        conn.close()
        return jsonify({"status": "error", "message": "Booking not found for this company."}), 404

    try:
        client = resolve_client_from_payload(conn, session['company_id'], data, id_key='client_id', name_key='client')
        client_id = client['id']
        client_name = client_display_name(client)
        project_id = normalise_booking_project_id(conn, session['company_id'], data.get('project_id'))
        booking_date = datetime.strptime(data['date'], '%Y-%m-%d')
        dt_str = f"{data['date']}T{data['time']}"
        overtime_hours = float(data.get('overtime_hours', 0) or 0)
    except ValueError as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)}), 400
    except Exception:
        conn.close()
        return jsonify({"status": "error", "message": "Invalid booking date, time, or overtime value."}), 400

    custom_fields_submitted = 'custom_fields' in data
    custom_fields = data.get('custom_fields') or {}
    booking_custom_fields = get_tenant_custom_fields(conn, session['company_id'], 'booking', visible_only=True)
    # Legacy bookings created before templates/custom fields may be edited through older paths
    # without custom_fields in the payload. Do not block those edits on newly-required fields.
    valid_custom, custom_message = validate_custom_field_payload(booking_custom_fields, custom_fields, require_missing=custom_fields_submitted)
    if not valid_custom:
        conn.close()
        return jsonify({"status": "error", "message": custom_message}), 400

    is_available, availability_message = validate_booking_employees_available(conn, session['company_id'], [{"employee": data.get('employee', '')}], booking_date, data.get('id'))
    if not is_available:
        conn.close()
        return jsonify({"status": "error", "message": availability_message}), 400
    
    new_google_event_id = b['google_event_id'] if b else None
    if session.get('comp_google_calendar'):
        try:
            if b and b['google_event_id']:
                updated_google_id = update_google_event(b['google_event_id'], client_name, data['date'], data['time'], data['employee'], data['booking_type'], data.get('transport'), session['company_name'], session.get('company_id'))
                if updated_google_id:
                    new_google_event_id = updated_google_id
            else:
                new_google_event_id = create_google_event(client_name, data['date'], data['time'], data['employee'], data['booking_type'], data.get('transport'), session['company_name'], session.get('company_id'))
        except Exception as e:
            print(f"Google Sync Error: {_format_google_calendar_error(e)}")
            
    conn.execute('UPDATE bookings SET client_id=?, title=?, start=?, employee=?, google_event_id=?, booking_type=?, transport=?, booking_notes=?, overtime_hours=?, project_id=? WHERE id=? AND company_id=?', (client_id, client_name, dt_str, data['employee'], new_google_event_id, data['booking_type'], data.get('transport'), data.get('booking_notes', ''), overtime_hours, project_id, data['id'], session['company_id']))
    if custom_fields_submitted:
        save_custom_field_values(conn, session['company_id'], 'booking', data['id'], custom_fields)
    conn.commit(); conn.close()
    
    log_action('Booking & Ops', 'Updated Booking', f"Updated assignment for {client_name} on {data['date']}.")
    return jsonify({"status": "success", "booking_id": int(data['id'])})

@app.route('/delete_booking', methods=['POST'])
def delete_booking():
    id = request.get_json()['id']
    conn = get_db_connection()
    b = conn.execute("SELECT google_event_id, title, start FROM bookings WHERE id=? AND company_id=?", (id, session['company_id'])).fetchone()
    
    if session.get('comp_google_calendar'):
        try:
            if b and b['google_event_id']: delete_google_event(b['google_event_id'], session.get('company_id'))
        except Exception as e: print(f"Google Sync Error: {_format_google_calendar_error(e)}")
            
    conn.execute("DELETE FROM custom_field_values WHERE company_id=? AND module_name='booking' AND record_id=?", (session['company_id'], id))
    conn.execute('DELETE FROM bookings WHERE id=? AND company_id=?', (id, session['company_id']))
    conn.commit(); conn.close()
    
    if b: log_action('Booking & Ops', 'Deleted Booking', f"Removed assignment for {b['title']} on {b['start'][:10]}")
    return jsonify({"status": "success"})

def async_google_sync(bookings_list, company_name, company_id=None):
    conn = get_db_connection()

    for b in bookings_list:
        try:
            cid = company_id or b.get('company_id')
            g_id = create_google_event(b['client'], b['date'], b['time'], b['employee'], b['booking_type'], b['transport'], company_name, cid)
            if g_id:
                conn.execute('UPDATE bookings SET google_event_id=? WHERE id=? AND company_id=?', (g_id, b['db_id'], cid))
                conn.commit()
        except Exception as e:
            print(f"Google Sync Error on recurring booking: {_format_google_calendar_error(e)}")

    conn.close()

@app.route('/generate_recurring', methods=['POST'])
def generate_recurring():
    if not session.get('can_booking') and not session.get('is_superadmin'): 
        return "Forbidden", 403
        
    data = request.get_json()
    conn = get_db_connection()
    try:
        client = resolve_client_from_payload(conn, session['company_id'], data, id_key='client_id', name_key='client')
        client_id = client['id']
        client_name = client_display_name(client)
        project_id = normalise_booking_project_id(conn, session['company_id'], data.get('project_id'))
    except ValueError as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)}), 400
    dates_to_schedule = []
    
    start_str = data.get('start_date')
    end_str = data.get('end_date')
    days_list = data.get('days') 
    
    if start_str and end_str:
        try:
            start_dt = datetime.strptime(start_str[:10], '%Y-%m-%d')
            end_dt = datetime.strptime(end_str[:10], '%Y-%m-%d')
            allowed_days = [int(d) for d in days_list] if days_list else list(range(7))
            curr = start_dt
            while curr <= end_dt:
                if curr.weekday() in allowed_days:
                    dates_to_schedule.append(curr.strftime('%Y-%m-%d'))
                curr += timedelta(days=1)
        except Exception as e:
            print(f"Error parsing start/end dates: {e}")

    dates_to_schedule = sorted(list(set(dates_to_schedule)))
    
    if not dates_to_schedule:
        conn.close()
        return jsonify({"status": "error", "message": "No valid dates found in that range matching those days."})

    sync_payload = []
    cursor = conn.cursor()
    
    assignments = data.get('assignments', [])
    if not assignments:
        assignments = [{
            'employee': data.get('employee', ''),
            'transport': data.get('transport', ''),
            'overtime_hours': data.get('overtime_hours', 0)
        }]

    custom_fields = data.get('custom_fields') or {}
    booking_custom_fields = get_tenant_custom_fields(conn, session['company_id'], 'booking', visible_only=True)
    valid_custom, custom_message = validate_custom_field_payload(booking_custom_fields, custom_fields)
    if not valid_custom:
        conn.close()
        return jsonify({"status": "error", "message": custom_message}), 400
    
    for d in dates_to_schedule:
        booking_date = datetime.strptime(d, '%Y-%m-%d')
        is_available, availability_message = validate_booking_employees_available(conn, session['company_id'], assignments, booking_date)
        if not is_available:
            conn.close()
            return jsonify({"status": "error", "message": f"{availability_message} on {d}"}), 400

    for d in dates_to_schedule:
        dt_str = f"{d}T{data.get('time', '08:00')}"
        
        for req in assignments:
            cursor.execute('''
                INSERT INTO bookings (company_id, client_id, title, start, employee, booking_type, transport, booking_notes, overtime_hours, is_invoiced, project_id) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?)
            ''', (
                session['company_id'], client_id, client_name, dt_str, req.get('employee'), 
                data.get('booking_type'), req.get('transport'), data.get('booking_notes', ''), 
                float(req.get('overtime_hours', 0)), project_id
            ))
            db_id = cursor.lastrowid
            save_custom_field_values(conn, session['company_id'], 'booking', db_id, custom_fields)
            
            sync_payload.append({
                'db_id': db_id, 'client': client_name, 'date': d, 'time': data.get('time', '08:00'),
                'employee': req.get('employee'), 'booking_type': data.get('booking_type'),
                'transport': req.get('transport')
            })
        
    conn.commit()
    conn.close()
    
    if session.get('comp_google_calendar'):
        threading.Thread(target=async_google_sync, args=(sync_payload, session.get('company_name'), session.get('company_id'))).start()
    
    log_action('Booking & Ops', 'Created Contract', f"Auto-scheduled {len(dates_to_schedule)} days for {client_name}")
    
    return jsonify({
        "status": "success", 
        "message": f"Successfully generated {len(dates_to_schedule)} matching days!"
    })



# --- ATTACHMENT HELPERS ---
ALLOWED_ATTACHMENT_EXTENSIONS = {
    'pdf', 'doc', 'docx', 'xls', 'xlsx', 'csv', 'txt',
    'jpg', 'jpeg', 'png', 'gif', 'webp', 'heic'
}
BLOCKED_ATTACHMENT_EXTENSIONS = {'exe', 'bat', 'cmd', 'com', 'js', 'php', 'msi', 'sh', 'ps1'}
MAX_ATTACHMENT_SIZE = 20 * 1024 * 1024


def attachment_extension(filename):
    return filename.rsplit('.', 1)[1].lower() if filename and '.' in filename else ''


def is_allowed_attachment(filename):
    ext = attachment_extension(filename)
    return ext and ext in ALLOWED_ATTACHMENT_EXTENSIONS and ext not in BLOCKED_ATTACHMENT_EXTENSIONS


def validate_attachment_link(conn, company_id, linked_type, linked_id):
    if linked_type not in ('booking', 'project'):
        return False
    try:
        linked_id = int(linked_id)
    except Exception:
        return False
    if linked_type == 'booking':
        row = conn.execute('SELECT id FROM bookings WHERE id=? AND company_id=?', (linked_id, company_id)).fetchone()
    else:
        row = conn.execute('SELECT id FROM projects WHERE id=? AND company_id=?', (linked_id, company_id)).fetchone()
    return bool(row)


def attachment_to_dict(row):
    return {
        'id': row['id'],
        'linked_type': row['linked_type'],
        'linked_id': row['linked_id'],
        'original_filename': row['original_filename'],
        'file_size': row['file_size'] or 0,
        'mime_type': row['mime_type'] or '',
        'uploaded_by': row['uploaded_by'] or '',
        'uploaded_at': row['uploaded_at'] or ''
    }

@app.route('/api/attachments/<linked_type>/<int:linked_id>', methods=['GET'])
def api_list_attachments(linked_type, linked_id):
    cid = session['company_id']
    conn = get_db_connection()
    if not validate_attachment_link(conn, cid, linked_type, linked_id):
        conn.close()
        return jsonify({'status': 'error', 'message': 'Linked record not found for this company.'}), 404
    rows = conn.execute('''SELECT * FROM attachments
                           WHERE company_id=? AND linked_type=? AND linked_id=?
                           ORDER BY uploaded_at DESC, id DESC''', (cid, linked_type, linked_id)).fetchall()
    conn.close()
    return jsonify({'status': 'success', 'attachments': [attachment_to_dict(r) for r in rows]})

@app.route('/api/attachments/upload', methods=['POST'])
def api_upload_attachments():
    cid = session['company_id']
    linked_type = (request.form.get('linked_type') or '').strip().lower()
    linked_ids_raw = request.form.getlist('linked_id') or []
    if not linked_ids_raw and request.form.get('linked_id'):
        linked_ids_raw = [request.form.get('linked_id')]
    try:
        linked_ids = [int(x) for x in linked_ids_raw if str(x).strip()]
    except Exception:
        return jsonify({'status': 'error', 'message': 'Invalid linked record.'}), 400
    if not linked_ids:
        return jsonify({'status': 'error', 'message': 'No linked record selected for attachments.'}), 400
    files = request.files.getlist('files')
    if not files:
        return jsonify({'status': 'error', 'message': 'No files selected.'}), 400
    conn = get_db_connection()
    for linked_id in linked_ids:
        if not validate_attachment_link(conn, cid, linked_type, linked_id):
            conn.close()
            return jsonify({'status': 'error', 'message': f'{linked_type.title()} {linked_id} was not found for this company.'}), 404
    saved = []
    for f in files:
        if not f or not f.filename:
            continue
        if not is_allowed_attachment(f.filename):
            conn.close()
            return jsonify({'status': 'error', 'message': f'File type not allowed: {f.filename}'}), 400
        f.stream.seek(0, os.SEEK_END)
        size = f.stream.tell()
        f.stream.seek(0)
        if size > MAX_ATTACHMENT_SIZE:
            conn.close()
            return jsonify({'status': 'error', 'message': f'File too large: {f.filename}. Maximum size is 20MB.'}), 400
        safe_original = secure_filename(f.filename) or 'attachment'
        ext = attachment_extension(safe_original)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
        base_stored = f'{timestamp}_{safe_original}'
        for linked_id in linked_ids:
            folder = os.path.join(app.config['UPLOAD_FOLDER'], 'attachments', f'company_{cid}', f'{linked_type}s', f'{linked_type}_{linked_id}')
            os.makedirs(folder, exist_ok=True)
            stored_filename = base_stored
            target_path = os.path.join(folder, stored_filename)
            # Rewind and save a copy for each linked record.
            f.stream.seek(0)
            f.save(target_path)
            rel_path = os.path.relpath(target_path, app.config['UPLOAD_FOLDER'])
            cur = conn.execute('''INSERT INTO attachments
                                  (company_id, linked_type, linked_id, original_filename, stored_filename, file_path, file_size, mime_type, uploaded_by)
                                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                               (cid, linked_type, linked_id, f.filename, stored_filename, rel_path, size, f.mimetype or '', session.get('username', '')))
            saved.append(cur.lastrowid)
            log_action('Booking & Ops', 'Uploaded Attachment', f'{linked_type.title()} ID {linked_id}: {f.filename}')
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'attachment_ids': saved})

@app.route('/download_attachment/<int:attachment_id>')
def download_attachment(attachment_id):
    cid = session['company_id']
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM attachments WHERE id=? AND company_id=?', (attachment_id, cid)).fetchone()
    conn.close()
    if not row:
        return 'Attachment not found.', 404
    abs_path = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], row['file_path']))
    uploads_root = os.path.abspath(app.config['UPLOAD_FOLDER'])
    if not abs_path.startswith(uploads_root) or not os.path.exists(abs_path):
        return 'Attachment file missing.', 404
    return send_file(abs_path, as_attachment=True, download_name=row['original_filename'])

@app.route('/api/attachments/delete', methods=['POST'])
def api_delete_attachment():
    cid = session['company_id']
    data = request.get_json(silent=True) or {}
    attachment_id = data.get('id')
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM attachments WHERE id=? AND company_id=?', (attachment_id, cid)).fetchone()
    if not row:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Attachment not found for this company.'}), 404
    abs_path = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], row['file_path']))
    uploads_root = os.path.abspath(app.config['UPLOAD_FOLDER'])
    conn.execute('DELETE FROM attachments WHERE id=? AND company_id=?', (attachment_id, cid))
    conn.commit()
    conn.close()
    if abs_path.startswith(uploads_root) and os.path.exists(abs_path):
        try:
            os.remove(abs_path)
        except Exception as exc:
            print(f'Attachment file delete warning: {exc}')
    log_action('Booking & Ops', 'Deleted Attachment', f"{row['linked_type'].title()} ID {row['linked_id']}: {row['original_filename']}")
    return jsonify({'status': 'success'})


# --- ACCOUNTING TRANSACTION FILE HELPERS ---
def accounting_file_to_dict(row):
    if not row:
        return None
    return {
        'id': row['id'],
        'linked_type': row['linked_type'],
        'linked_id': row['linked_id'],
        'original_filename': row['original_filename'],
        'file_size': row['file_size'] or 0,
        'mime_type': row['mime_type'] or '',
        'uploaded_by': row['uploaded_by'] or '',
        'uploaded_at': row['uploaded_at'] or '',
        'download_url': url_for('download_accounting_transaction_file', file_id=row['id']),
        'view_url': url_for('view_accounting_transaction_file', file_id=row['id'])
    }


def _request_payload_with_optional_files():
    if request.content_type and 'multipart/form-data' in request.content_type:
        raw = request.form.get('payload') or '{}'
        try:
            return json.loads(raw)
        except Exception:
            return {}
    return request.get_json(silent=True) or {}


def _delete_accounting_transaction_file(conn, company_id, linked_type, linked_id):
    row = conn.execute('''SELECT * FROM accounting_transaction_files
                          WHERE company_id=? AND linked_type=? AND linked_id=?
                          ORDER BY id DESC LIMIT 1''', (company_id, linked_type, linked_id)).fetchone()
    if not row:
        return
    abs_path = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], row['file_path']))
    uploads_root = os.path.abspath(app.config['UPLOAD_FOLDER'])
    conn.execute('DELETE FROM accounting_transaction_files WHERE id=? AND company_id=?', (row['id'], company_id))
    if abs_path.startswith(uploads_root) and os.path.exists(abs_path):
        try:
            os.remove(abs_path)
        except Exception as exc:
            print(f'Accounting file delete warning: {exc}')


def _save_accounting_transaction_file(conn, company_id, linked_type, linked_id, file_storage):
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return None
    if linked_type not in ('journal_line', 'cashbook_line'):
        raise ValueError('Invalid accounting file link type.')
    if not is_allowed_attachment(file_storage.filename):
        raise ValueError(f'File type not allowed: {file_storage.filename}')
    file_storage.stream.seek(0, os.SEEK_END)
    size = file_storage.stream.tell()
    file_storage.stream.seek(0)
    if size > MAX_ATTACHMENT_SIZE:
        raise ValueError(f'File too large: {file_storage.filename}. Maximum size is 20MB.')
    _delete_accounting_transaction_file(conn, company_id, linked_type, linked_id)
    safe_original = secure_filename(file_storage.filename) or 'accounting_file'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
    stored_filename = f'{timestamp}_{safe_original}'
    folder = os.path.join(app.config['UPLOAD_FOLDER'], 'accounting', f'company_{company_id}', linked_type, f'{linked_type}_{linked_id}')
    os.makedirs(folder, exist_ok=True)
    target_path = os.path.join(folder, stored_filename)
    file_storage.save(target_path)
    rel_path = os.path.relpath(target_path, app.config['UPLOAD_FOLDER'])
    cur = conn.execute('''INSERT INTO accounting_transaction_files
                          (company_id, linked_type, linked_id, original_filename, stored_filename, file_path, file_size, mime_type, uploaded_by)
                          VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                       (company_id, linked_type, linked_id, file_storage.filename, stored_filename, rel_path, size, file_storage.mimetype or '', session.get('username', '')))
    return cur.lastrowid


def _attach_accounting_transaction_files(conn, company_id, linked_type, rows):
    out = []
    for r in rows or []:
        d = dict(r) if not isinstance(r, dict) else dict(r)
        file_row = conn.execute('''SELECT * FROM accounting_transaction_files
                                   WHERE company_id=? AND linked_type=? AND linked_id=?
                                   ORDER BY id DESC LIMIT 1''', (company_id, linked_type, d.get('id'))).fetchone()
        d['transaction_file'] = accounting_file_to_dict(file_row) if file_row else None
        out.append(d)
    return out


def _accounting_transaction_file_response(file_id, as_attachment=False):
    cid = session.get('company_id')
    if not cid:
        return redirect(url_for('login'))
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM accounting_transaction_files WHERE id=? AND company_id=?', (file_id, cid)).fetchone()
    conn.close()
    if not row:
        return 'Accounting file not found.', 404
    abs_path = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], row['file_path']))
    uploads_root = os.path.abspath(app.config['UPLOAD_FOLDER'])
    if not abs_path.startswith(uploads_root) or not os.path.exists(abs_path):
        return 'Accounting file missing.', 404
    return send_file(
        abs_path,
        as_attachment=as_attachment,
        download_name=row['original_filename'],
        mimetype=row['mime_type'] or None
    )


@app.route('/view_accounting_file/<int:file_id>')
def view_accounting_transaction_file(file_id):
    return _accounting_transaction_file_response(file_id, as_attachment=False)


@app.route('/download_accounting_file/<int:file_id>')
def download_accounting_transaction_file(file_id):
    return _accounting_transaction_file_response(file_id, as_attachment=True)


def project_to_dict(row, employees=None, costs_total=0.0):
    d = dict(row)
    client_bits = [d.get('client_first_name') or '', d.get('client_surname') or '']
    client_name = ' '.join([b for b in client_bits if b]).strip()
    if d.get('client_company_name'):
        client_name = f"{client_name} ({d.get('client_company_name')})" if client_name else d.get('client_company_name')
    d['client_display'] = client_name or 'No client selected'
    d['employees'] = employees or []
    d['costs_total'] = round(float(costs_total or 0), 2)
    d['profit_estimate'] = round(float(d.get('fixed_price') or 0) - float(costs_total or 0), 2)
    return d

@app.route('/api/projects', methods=['GET'])
def api_projects_list():
    conn = get_db_connection()
    cid = session['company_id']
    rows = conn.execute('''SELECT p.*, c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name,
                                  COALESCE((SELECT SUM(amount) FROM project_costs pc WHERE pc.company_id=p.company_id AND pc.project_id=p.id), 0) AS costs_total
                           FROM projects p
                           LEFT JOIN clients c ON c.id=p.client_id AND c.company_id=p.company_id
                           WHERE p.company_id=?
                           ORDER BY CASE p.status WHEN 'In Progress' THEN 1 WHEN 'Approved' THEN 2 WHEN 'Quoted' THEN 3 WHEN 'On Hold' THEN 4 WHEN 'Completed' THEN 5 WHEN 'Cancelled' THEN 6 ELSE 7 END,
                                    p.start_date IS NULL, p.start_date DESC, p.project_name ASC''', (cid,)).fetchall()
    projects = []
    for row in rows:
        emps = conn.execute('''SELECT e.id, e.name, pe.role
                               FROM project_employees pe
                               JOIN employees e ON e.id=pe.employee_id AND e.company_id=pe.company_id
                               WHERE pe.company_id=? AND pe.project_id=?
                               ORDER BY e.name ASC''', (cid, row['id'])).fetchall()
        projects.append(project_to_dict(row, [dict(e) for e in emps], row['costs_total']))
    conn.close()
    return jsonify({"status": "success", "projects": projects})

@app.route('/api/projects/<int:project_id>', methods=['GET'])
def api_project_get(project_id):
    conn = get_db_connection()
    cid = session['company_id']
    row = conn.execute('''SELECT p.*, c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name,
                                  COALESCE((SELECT SUM(amount) FROM project_costs pc WHERE pc.company_id=p.company_id AND pc.project_id=p.id), 0) AS costs_total
                           FROM projects p
                           LEFT JOIN clients c ON c.id=p.client_id AND c.company_id=p.company_id
                           WHERE p.company_id=? AND p.id=?''', (cid, project_id)).fetchone()
    if not row:
        conn.close()
        return jsonify({"status": "error", "message": "Project not found."}), 404
    emps = conn.execute('''SELECT e.id, e.name, pe.role
                           FROM project_employees pe
                           JOIN employees e ON e.id=pe.employee_id AND e.company_id=pe.company_id
                           WHERE pe.company_id=? AND pe.project_id=?
                           ORDER BY e.name ASC''', (cid, project_id)).fetchall()
    conn.close()
    return jsonify({"status": "success", "project": project_to_dict(row, [dict(e) for e in emps], row['costs_total'])})

@app.route('/api/projects/save', methods=['POST'])
def api_project_save():
    data = request.get_json(silent=True) or {}
    cid = session['company_id']
    project_name = (data.get('project_name') or '').strip()
    if not project_name:
        return jsonify({"status": "error", "message": "Project name is required."}), 400
    try:
        client_id = int(data.get('client_id') or 0) or None
    except Exception:
        client_id = None
    fixed_price = sanitize_money(data.get('fixed_price'))
    employee_ids = data.get('employee_ids') or []
    try:
        employee_ids = [int(eid) for eid in employee_ids if str(eid).strip()]
    except Exception:
        return jsonify({"status": "error", "message": "Invalid employee assignment."}), 400

    conn = get_db_connection()
    if client_id:
        client = conn.execute('SELECT id FROM clients WHERE id=? AND company_id=?', (client_id, cid)).fetchone()
        if not client:
            conn.close()
            return jsonify({"status": "error", "message": "Selected client does not belong to this company."}), 400
    valid_employee_ids = []
    if employee_ids:
        placeholders = ','.join(['?'] * len(employee_ids))
        rows = conn.execute(f'''SELECT id FROM employees
                                WHERE company_id=? AND id IN ({placeholders})
                                AND (emp_type != 'Supplier' OR emp_type IS NULL)''', [cid] + employee_ids).fetchall()
        valid_employee_ids = [int(r['id']) for r in rows]
        if len(valid_employee_ids) != len(set(employee_ids)):
            conn.close()
            return jsonify({"status": "error", "message": "One or more assigned employees/providers are invalid."}), 400

    project_id = data.get('id')
    fields = (
        client_id,
        project_name,
        (data.get('project_code') or '').strip(),
        (data.get('description') or '').strip(),
        (data.get('site_address') or '').strip(),
        data.get('start_date') or None,
        data.get('estimated_end_date') or None,
        data.get('actual_end_date') or None,
        fixed_price,
        data.get('status') or 'Quoted',
        (data.get('notes') or '').strip()
    )
    if project_id:
        existing = conn.execute('SELECT id FROM projects WHERE id=? AND company_id=?', (project_id, cid)).fetchone()
        if not existing:
            conn.close()
            return jsonify({"status": "error", "message": "Project not found for this company."}), 404
        conn.execute('''UPDATE projects
                        SET client_id=?, project_name=?, project_code=?, description=?, site_address=?,
                            start_date=?, estimated_end_date=?, actual_end_date=?, fixed_price=?, status=?, notes=?, updated_at=CURRENT_TIMESTAMP
                        WHERE id=? AND company_id=?''', fields + (project_id, cid))
        action = 'Updated Project'
    else:
        cur = conn.execute('''INSERT INTO projects
                              (company_id, client_id, project_name, project_code, description, site_address,
                               start_date, estimated_end_date, actual_end_date, fixed_price, status, notes)
                              VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (cid,) + fields)
        project_id = cur.lastrowid
        action = 'Created Project'

    conn.execute('DELETE FROM project_employees WHERE company_id=? AND project_id=?', (cid, project_id))
    for employee_id in valid_employee_ids:
        conn.execute('''INSERT OR IGNORE INTO project_employees (company_id, project_id, employee_id, role)
                        VALUES (?, ?, ?, ?)''', (cid, project_id, employee_id, 'Assigned'))
    conn.commit()
    conn.close()
    log_action('Booking & Ops', action, f"{project_name} (ID {project_id})")
    return jsonify({"status": "success", "project_id": project_id})

@app.route('/api/projects/delete', methods=['POST'])
def api_project_delete():
    data = request.get_json(silent=True) or {}
    project_id = data.get('id')
    conn = get_db_connection()
    cid = session['company_id']
    project = conn.execute('SELECT project_name FROM projects WHERE id=? AND company_id=?', (project_id, cid)).fetchone()
    if not project:
        conn.close()
        return jsonify({"status": "error", "message": "Project not found."}), 404
    conn.execute("UPDATE projects SET status='Cancelled', updated_at=CURRENT_TIMESTAMP WHERE id=? AND company_id=?", (project_id, cid))
    conn.commit()
    conn.close()
    log_action('Booking & Ops', 'Cancelled Project', project['project_name'])
    return jsonify({"status": "success"})

@app.route('/update_client', methods=['POST'])
def update_client():
    data = request.get_json()
    source_app = data.get('_source_app') or 'Booking & Ops'
    if source_app not in ['Booking & Ops', 'Invoicing & Quotes']:
        source_app = 'Booking & Ops'
    conn = get_db_connection()
    action_msg = None
    address = compose_client_address(
        data.get('building_number'),
        data.get('street_name'),
        data.get('suburb'),
        data.get('postal_code'),
        data.get('address')
    )
    if data.get('id'):
        conn.execute('''UPDATE clients
                        SET name=?, surname=?, client_type=?, phone=?, email=?, address=?,
                            building_number=?, street_name=?, suburb=?, postal_code=?, discount_percent=?,
                            company_name=?, registration_number=?, vat_number=?
                        WHERE id=? AND company_id=?''',
                     (data['name'], data.get('surname'), data.get('client_type'), data.get('phone'), data.get('email'), address,
                      data.get('building_number'), data.get('street_name'), data.get('suburb'), data.get('postal_code'), sanitize_percent(data.get('discount_percent')),
                      data.get('company_name'), data.get('registration_number'), data.get('vat_number'), data['id'], session['company_id']))
        action_msg = (source_app, 'Updated Client', f"Updated client profile: {data['name']}")
    else:
        conn.execute('''INSERT INTO clients
                        (company_id, name, surname, client_type, phone, email, address,
                         building_number, street_name, suburb, postal_code, discount_percent,
                         company_name, registration_number, vat_number)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                     (session['company_id'], data['name'], data.get('surname'), data.get('client_type'), data.get('phone'), data.get('email'), address,
                      data.get('building_number'), data.get('street_name'), data.get('suburb'), data.get('postal_code'), sanitize_percent(data.get('discount_percent')),
                      data.get('company_name'), data.get('registration_number'), data.get('vat_number')))
        action_msg = (source_app, 'Created Client', f"Added new client: {data['name']}")
    conn.commit()
    conn.close()

    if action_msg: log_action(action_msg[0], action_msg[1], action_msg[2])
    return jsonify({"status": "success"})

@app.route('/client_report', methods=['POST'])
def client_report():
    data = request.get_json()
    conn = get_db_connection()
    try:
        client_id = int(data.get('client_id') or 0)
    except Exception:
        client_id = 0
    if client_id:
        b = conn.execute("SELECT start, booking_type, employee, booking_notes FROM bookings WHERE company_id=? AND client_id=? AND substr(start, 1, 10) BETWEEN ? AND ? ORDER BY start ASC", (session['company_id'], client_id, data['start_date'], data['end_date'])).fetchall()
    else:
        b = conn.execute("SELECT start, booking_type, employee, booking_notes FROM bookings WHERE company_id=? AND title=? AND substr(start, 1, 10) BETWEEN ? AND ? ORDER BY start ASC", (session['company_id'], data.get('client_name'), data['start_date'], data['end_date'])).fetchall()
    conn.close()
    return jsonify({"bookings": [{"date": r['start'][:10], "type": r['booking_type'], "staff": r['employee'], "notes": r['booking_notes'] or ""} for r in b]})



@app.route('/api/booking_ops_report', methods=['POST'])
def api_booking_ops_report():
    data = request.get_json(silent=True) or {}
    report_type = (data.get('report_type') or 'bookings').strip().lower()
    start_date = (data.get('start_date') or '').strip()
    end_date = (data.get('end_date') or '').strip()
    if report_type not in ('bookings', 'customers', 'projects'):
        return jsonify({'status': 'error', 'message': 'Invalid report type.'}), 400
    if not start_date or not end_date:
        return jsonify({'status': 'error', 'message': 'Start date and end date are required.'}), 400
    try:
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')
    except Exception:
        return jsonify({'status': 'error', 'message': 'Invalid date range.'}), 400

    cid = session['company_id']
    conn = get_db_connection()

    if report_type == 'bookings':
        rows = conn.execute("""SELECT b.id, b.title, b.client_id, b.start, b.employee, b.booking_type, b.transport,
                                      b.overtime_hours, b.booking_notes, b.is_invoiced, b.project_id,
                                      p.project_name, c.name AS client_first_name, c.surname AS client_surname,
                                      c.company_name AS client_company_name,
                                      (SELECT COUNT(*) FROM attachments a WHERE a.company_id=b.company_id AND a.linked_type='booking' AND a.linked_id=b.id) AS attachment_count
                               FROM bookings b
                               LEFT JOIN projects p ON p.id=b.project_id AND p.company_id=b.company_id
                               LEFT JOIN clients c ON c.id=b.client_id AND c.company_id=b.company_id
                               WHERE b.company_id=? AND substr(b.start, 1, 10) BETWEEN ? AND ?
                               ORDER BY b.start ASC, b.id ASC""", (cid, start_date, end_date)).fetchall()
        bookings = []
        for r in rows:
            date_part = (r['start'] or '')[:10]
            time_part = (r['start'] or '').split('T')[1] if 'T' in (r['start'] or '') else ''
            try:
                day_name = datetime.strptime(date_part, '%Y-%m-%d').strftime('%A')
            except Exception:
                day_name = ''
            client = ' '.join([x for x in [r['client_first_name'], r['client_surname']] if x]).strip() or r['client_company_name'] or r['title'] or ''
            bookings.append({
                'id': r['id'], 'date': date_part, 'day': day_name, 'time': time_part,
                'client_id': r['client_id'] or '', 'client': client, 'project': r['project_name'] or '',
                'service': r['booking_type'] or '', 'staff': r['employee'] or '',
                'transport': r['transport'] or '', 'overtime_hours': r['overtime_hours'] or 0,
                'notes': r['booking_notes'] or '', 'invoiced': bool(r['is_invoiced']),
                'attachment_count': r['attachment_count'] or 0
            })
        conn.close()
        return jsonify({'status': 'success', 'report_type': report_type, 'bookings': bookings})

    if report_type == 'customers':
        rows = conn.execute("""SELECT b.id, b.title, b.client_id, b.start, b.employee, b.booking_type, b.booking_notes,
                                      b.is_invoiced, p.project_name, c.name AS client_first_name,
                                      c.surname AS client_surname, c.company_name AS client_company_name,
                                      (SELECT COUNT(*) FROM attachments a WHERE a.company_id=b.company_id AND a.linked_type='booking' AND a.linked_id=b.id) AS attachment_count
                               FROM bookings b
                               LEFT JOIN projects p ON p.id=b.project_id AND p.company_id=b.company_id
                               LEFT JOIN clients c ON c.id=b.client_id AND c.company_id=b.company_id
                               WHERE b.company_id=? AND substr(b.start, 1, 10) BETWEEN ? AND ?
                               ORDER BY c.name ASC, c.surname ASC, b.title ASC, b.start ASC, b.id ASC""", (cid, start_date, end_date)).fetchall()
        grouped = {}
        for r in rows:
            client = ' '.join([x for x in [r['client_first_name'], r['client_surname']] if x]).strip() or r['client_company_name'] or r['title'] or 'Unknown Client'
            group_key = str(r['client_id'] or ('legacy:' + client))
            date_part = (r['start'] or '')[:10]
            try:
                day_name = datetime.strptime(date_part, '%Y-%m-%d').strftime('%A')
            except Exception:
                day_name = ''
            grouped.setdefault(group_key, {'client_id': r['client_id'] or '', 'client': client, 'booking_count': 0, 'last_booking_date': '', 'staff_used': set(), 'projects': set(), 'bookings': []})
            item = grouped[group_key]
            item['booking_count'] += 1
            item['last_booking_date'] = max(item['last_booking_date'], date_part)
            for staff in [x.strip() for x in (r['employee'] or '').split(',') if x.strip()]:
                item['staff_used'].add(staff)
            if r['project_name']:
                item['projects'].add(r['project_name'])
            item['bookings'].append({
                'date': date_part, 'day': day_name, 'service': r['booking_type'] or '', 'staff': r['employee'] or '',
                'project': r['project_name'] or '', 'notes': r['booking_notes'] or '',
                'invoiced': bool(r['is_invoiced']), 'attachment_count': r['attachment_count'] or 0
            })
        customers = []
        for item in grouped.values():
            item['staff_used'] = sorted(item['staff_used'])
            item['projects'] = sorted(item['projects'])
            customers.append(item)
        customers.sort(key=lambda x: x['client'].lower())
        conn.close()
        return jsonify({'status': 'success', 'report_type': report_type, 'customers': customers})

    project_rows = conn.execute("""SELECT p.*, c.name AS client_name, c.surname AS client_surname,
                                          (SELECT COUNT(*) FROM attachments a WHERE a.company_id=p.company_id AND a.linked_type='project' AND a.linked_id=p.id) AS attachment_count
                                   FROM projects p
                                   LEFT JOIN clients c ON c.id=p.client_id AND c.company_id=p.company_id
                                   WHERE p.company_id=?
                                     AND (
                                        EXISTS (SELECT 1 FROM bookings b WHERE b.company_id=p.company_id AND b.project_id=p.id AND substr(b.start, 1, 10) BETWEEN ? AND ?)
                                        OR (COALESCE(p.start_date, '') <= ? AND COALESCE(NULLIF(p.actual_end_date, ''), NULLIF(p.estimated_end_date, ''), '9999-12-31') >= ?)
                                     )
                                   ORDER BY p.start_date ASC, p.project_name ASC""", (cid, start_date, end_date, end_date, start_date)).fetchall()
    projects = []
    for pr in project_rows:
        b_rows = conn.execute("""SELECT b.id, b.start, b.employee, b.booking_type, b.overtime_hours, b.booking_notes,
                                        (SELECT COUNT(*) FROM attachments a WHERE a.company_id=b.company_id AND a.linked_type='booking' AND a.linked_id=b.id) AS attachment_count
                                 FROM bookings b
                                 WHERE b.company_id=? AND b.project_id=? AND substr(b.start, 1, 10) BETWEEN ? AND ?
                                 ORDER BY b.start ASC, b.id ASC""", (cid, pr['id'], start_date, end_date)).fetchall()
        bookings = []
        staff_set = set()
        total_overtime = 0.0
        for b in b_rows:
            date_part = (b['start'] or '')[:10]
            try:
                day_name = datetime.strptime(date_part, '%Y-%m-%d').strftime('%A')
            except Exception:
                day_name = ''
            for staff in [x.strip() for x in (b['employee'] or '').split(',') if x.strip()]:
                staff_set.add(staff)
            try:
                total_overtime += float(b['overtime_hours'] or 0)
            except Exception:
                pass
            bookings.append({
                'date': date_part, 'day': day_name, 'staff': b['employee'] or '',
                'service': b['booking_type'] or '', 'overtime_hours': b['overtime_hours'] or 0,
                'notes': b['booking_notes'] or '', 'attachment_count': b['attachment_count'] or 0
            })
        client_name = ' '.join([x for x in [pr['client_name'], pr['client_surname']] if x])
        projects.append({
            'id': pr['id'], 'project_name': pr['project_name'] or '', 'project_code': pr['project_code'] or '',
            'client': client_name, 'status': pr['status'] or '', 'fixed_price': pr['fixed_price'] or 0,
            'start_date': pr['start_date'] or '', 'estimated_end_date': pr['estimated_end_date'] or '',
            'actual_end_date': pr['actual_end_date'] or '', 'site_address': pr['site_address'] or '',
            'notes': pr['notes'] or '', 'attachment_count': pr['attachment_count'] or 0,
            'booking_count': len(bookings), 'total_days_worked': len(set([b['date'] for b in bookings if b['date']])),
            'total_overtime_hours': total_overtime, 'staff_used': sorted(staff_set), 'bookings': bookings
        })
    conn.close()
    return jsonify({'status': 'success', 'report_type': report_type, 'projects': projects})


@app.route('/export_booking_ops_report', methods=['POST'])
def export_booking_ops_report():
    data = request.get_json(silent=True) or {}
    report_type = (data.get('report_type') or 'bookings').strip().lower()
    start_date = (data.get('start_date') or '').strip()
    end_date = (data.get('end_date') or '').strip()
    if report_type not in ('bookings', 'customers', 'projects'):
        return jsonify({'status': 'error', 'message': 'Invalid report type.'}), 400
    if not start_date or not end_date:
        return jsonify({'status': 'error', 'message': 'Start and end dates are required.'}), 400
    cid = session['company_id']
    conn = get_db_connection()
    output = io.StringIO()
    writer = csv.writer(output)
    if report_type == 'bookings':
        writer.writerow(['Date', 'Day of Week', 'Time', 'Client', 'Project', 'Service', 'Staff', 'Transport', 'Overtime Hours', 'Booking Notes', 'Invoiced', 'Attachments'])
        rows = conn.execute("""SELECT b.*, p.project_name, c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name,
                                      (SELECT COUNT(*) FROM attachments a WHERE a.company_id=b.company_id AND a.linked_type='booking' AND a.linked_id=b.id) AS attachment_count
                               FROM bookings b
                               LEFT JOIN projects p ON p.id=b.project_id AND p.company_id=b.company_id
                               LEFT JOIN clients c ON c.id=b.client_id AND c.company_id=b.company_id
                               WHERE b.company_id=? AND substr(b.start, 1, 10) BETWEEN ? AND ?
                               ORDER BY b.start ASC, b.id ASC""", (cid, start_date, end_date)).fetchall()
        for r in rows:
            date_part = (r['start'] or '')[:10]
            time_part = (r['start'] or '').split('T')[1] if 'T' in (r['start'] or '') else ''
            day_name = datetime.strptime(date_part, '%Y-%m-%d').strftime('%A') if date_part else ''
            client = ' '.join([x for x in [r['client_first_name'], r['client_surname']] if x]).strip() or r['client_company_name'] or r['title'] or ''
            writer.writerow([date_part, day_name, time_part, client, r['project_name'] or '', r['booking_type'] or '', r['employee'] or '', r['transport'] or '', r['overtime_hours'] or 0, r['booking_notes'] or '', 'Yes' if r['is_invoiced'] else 'No', r['attachment_count'] or 0])
    elif report_type == 'customers':
        writer.writerow(['Client', 'Date', 'Day of Week', 'Service', 'Staff', 'Project', 'Booking Notes', 'Invoiced', 'Attachments'])
        rows = conn.execute("""SELECT b.*, p.project_name, c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name,
                                      (SELECT COUNT(*) FROM attachments a WHERE a.company_id=b.company_id AND a.linked_type='booking' AND a.linked_id=b.id) AS attachment_count
                               FROM bookings b
                               LEFT JOIN projects p ON p.id=b.project_id AND p.company_id=b.company_id
                               LEFT JOIN clients c ON c.id=b.client_id AND c.company_id=b.company_id
                               WHERE b.company_id=? AND substr(b.start, 1, 10) BETWEEN ? AND ?
                               ORDER BY c.name ASC, c.surname ASC, b.title ASC, b.start ASC, b.id ASC""", (cid, start_date, end_date)).fetchall()
        for r in rows:
            date_part = (r['start'] or '')[:10]
            day_name = datetime.strptime(date_part, '%Y-%m-%d').strftime('%A') if date_part else ''
            client = ' '.join([x for x in [r['client_first_name'], r['client_surname']] if x]).strip() or r['client_company_name'] or r['title'] or ''
            writer.writerow([client, date_part, day_name, r['booking_type'] or '', r['employee'] or '', r['project_name'] or '', r['booking_notes'] or '', 'Yes' if r['is_invoiced'] else 'No', r['attachment_count'] or 0])
    else:
        writer.writerow(['Project', 'Project Code', 'Client', 'Status', 'Fixed Price', 'Start Date', 'Estimated End Date', 'Actual End Date', 'Site Address', 'Booking Date', 'Day of Week', 'Staff', 'Service', 'Overtime Hours', 'Booking Notes', 'Project Attachments', 'Booking Attachments'])
        rows = conn.execute("""SELECT p.*, c.name AS client_name, c.surname AS client_surname,
                                      b.start AS booking_start, b.employee AS booking_employee, b.booking_type, b.overtime_hours, b.booking_notes,
                                      (SELECT COUNT(*) FROM attachments a WHERE a.company_id=p.company_id AND a.linked_type='project' AND a.linked_id=p.id) AS project_attachment_count,
                                      (SELECT COUNT(*) FROM attachments a WHERE a.company_id=b.company_id AND a.linked_type='booking' AND a.linked_id=b.id) AS booking_attachment_count
                               FROM projects p
                               LEFT JOIN clients c ON c.id=p.client_id AND c.company_id=p.company_id
                               LEFT JOIN bookings b ON b.company_id=p.company_id AND b.project_id=p.id AND substr(b.start, 1, 10) BETWEEN ? AND ?
                               WHERE p.company_id=?
                                 AND (b.id IS NOT NULL OR (COALESCE(p.start_date, '') <= ? AND COALESCE(NULLIF(p.actual_end_date, ''), NULLIF(p.estimated_end_date, ''), '9999-12-31') >= ?))
                               ORDER BY p.project_name ASC, b.start ASC""", (start_date, end_date, cid, end_date, start_date)).fetchall()
        for r in rows:
            date_part = (r['booking_start'] or '')[:10]
            day_name = datetime.strptime(date_part, '%Y-%m-%d').strftime('%A') if date_part else ''
            client_name = ' '.join([x for x in [r['client_name'], r['client_surname']] if x])
            writer.writerow([r['project_name'] or '', r['project_code'] or '', client_name, r['status'] or '', r['fixed_price'] or 0, r['start_date'] or '', r['estimated_end_date'] or '', r['actual_end_date'] or '', r['site_address'] or '', date_part, day_name, r['booking_employee'] or '', r['booking_type'] or '', r['overtime_hours'] or 0, r['booking_notes'] or '', r['project_attachment_count'] or 0, r['booking_attachment_count'] or 0])
    conn.close()
    csv_data = output.getvalue()
    filename = f"Booking_Ops_{report_type.title()}_{start_date}_to_{end_date}.csv"
    log_action('Booking & Ops', 'Exported Report', f"Exported {report_type} report from {start_date} to {end_date}")
    response = Response(csv_data, mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

@app.route('/export_bookings_range', methods=['POST'])
def export_bookings_range():
    data = request.get_json(silent=True) or {}
    start_date = (data.get('start_date') or '').strip()
    end_date = (data.get('end_date') or '').strip()
    selected_fields = data.get('fields') or []
    if not start_date or not end_date:
        return jsonify({"status": "error", "message": "Start and end dates are required."}), 400
    if not selected_fields:
        return jsonify({"status": "error", "message": "Select at least one field to export."}), 400
    try:
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')
    except Exception:
        return jsonify({"status": "error", "message": "Invalid date range."}), 400

    cid = session['company_id']
    conn = get_db_connection()
    custom_fields = get_tenant_custom_fields(conn, cid, 'booking', visible_only=True)
    custom_map = {f['field_key']: f for f in custom_fields}

    core_field_defs = {
        'id': ('Booking ID', lambda r, cv: r['id']),
        'client': ('Client', lambda r, cv: (' '.join([x for x in [r['client_first_name'], r['client_surname']] if x]).strip() or r['client_company_name'] or r['title'] or '')),
        'date': ('Date', lambda r, cv: (r['start'] or '')[:10]),
        'day': ('Day of Week', lambda r, cv: datetime.strptime((r['start'] or '')[:10], '%Y-%m-%d').strftime('%A') if (r['start'] or '')[:10] else ''),
        'time': ('Time', lambda r, cv: (r['start'] or '').split('T')[1] if 'T' in (r['start'] or '') else ''),
        'employee': ('Employee(s)', lambda r, cv: r['employee'] or ''),
        'service': ('Service Type', lambda r, cv: r['booking_type'] or ''),
        'transport': ('Transport', lambda r, cv: r['transport'] or ''),
        'overtime_hours': ('Overtime Hours', lambda r, cv: r['overtime_hours'] or 0),
        'booking_notes': ('Booking Notes', lambda r, cv: r['booking_notes'] or ''),
        'project': ('Project', lambda r, cv: r['project_name'] or ''),
        'is_invoiced': ('Invoiced', lambda r, cv: 'Yes' if r['is_invoiced'] else 'No'),
    }

    columns = []
    for field in selected_fields:
        if field in core_field_defs:
            columns.append((core_field_defs[field][0], field, core_field_defs[field][1], False))
        elif field.startswith('custom:'):
            key = field.split(':', 1)[1]
            if key in custom_map:
                columns.append((custom_map[key]['field_label'], key, None, True))

    if not columns:
        conn.close()
        return jsonify({"status": "error", "message": "No valid fields selected."}), 400

    rows = conn.execute("""SELECT b.*, p.project_name, c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name
                         FROM bookings b
                         LEFT JOIN projects p ON p.id=b.project_id AND p.company_id=b.company_id
                         LEFT JOIN clients c ON c.id=b.client_id AND c.company_id=b.company_id
                         WHERE b.company_id=? AND substr(b.start, 1, 10) BETWEEN ? AND ?
                         ORDER BY b.start ASC, b.id ASC""", (cid, start_date, end_date)).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([c[0] for c in columns])
    for r in rows:
        custom_values = get_custom_field_values_for_record(conn, cid, 'booking', r['id'])
        out = []
        for label, key, getter, is_custom in columns:
            if is_custom:
                out.append(custom_values.get(key, ''))
            else:
                try:
                    out.append(getter(r, custom_values))
                except Exception:
                    out.append('')
        writer.writerow(out)
    conn.close()

    csv_data = output.getvalue()
    log_action('Booking & Ops', 'Exported Booking Date Range', f"Exported bookings from {start_date} to {end_date}")
    response = Response(csv_data, mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename=Bookings_{start_date}_to_{end_date}.csv'
    return response

@app.route('/daily_route', methods=['POST'])
def daily_route():
    conn = get_db_connection()
    date = request.get_json()['date']
    b = conn.execute('''SELECT b.title, b.start, b.client_id, c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name,
                               c.address, c.building_number, c.street_name, c.suburb, c.postal_code
                        FROM bookings b
                        LEFT JOIN clients c ON c.id=b.client_id AND c.company_id=b.company_id
                        WHERE b.company_id=? AND b.start LIKE ? ORDER BY b.start ASC''', (session['company_id'], f"{date}%")).fetchall()
    route = []
    for row in b:
        c = row
        address = ""
        if c:
            address = compose_client_address(c['building_number'], c['street_name'], c['suburb'], c['postal_code'], c['address'])
        route.append({"time": row['start'].split('T')[1], "client": (' '.join([x for x in [row['client_first_name'], row['client_surname']] if x]).strip() or row['client_company_name'] or row['title']), "address": address if address else "No address on file"})
    conn.close()
    return jsonify({"route": route})

# ==========================================================
# 2. FINANCE ROUTES
# ==========================================================
@app.route('/finance')
def finance_index():
    today = datetime.now()
    cid = session['company_id']
    t_str, w_str = today.strftime('%Y-%m-%d'), (today + timedelta(days=7)).strftime('%Y-%m-%d')
    m_s, m_e = today.replace(day=1).strftime('%Y-%m-%d'), ((today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)).strftime('%Y-%m-%d')

    conn = get_db_connection()
    services = [dict(row) for row in conn.execute('SELECT * FROM services WHERE company_id=? ORDER BY name ASC', (cid,)).fetchall()]
    expense_page, expense_per_page, expense_offset = get_page_args(request.args, prefix='expense_', default_per_page=25, max_per_page=100)
    expense_q = (request.args.get('expense_q') or '').strip()
    exp_where = 'company_id=?'
    exp_params = [cid]
    if expense_q:
        like = f'%{expense_q}%'
        exp_where += " AND (COALESCE(category,'') LIKE ? OR COALESCE(supplier,'') LIKE ? OR COALESCE(description,'') LIKE ? OR COALESCE(date,'') LIKE ?)"
        exp_params.extend([like, like, like, like])
    expense_total = conn.execute(f'SELECT COUNT(*) FROM expenses WHERE {exp_where}', exp_params).fetchone()[0]
    expenses = [dict(row) for row in conn.execute(f'SELECT * FROM expenses WHERE {exp_where} ORDER BY date DESC, id DESC LIMIT ? OFFSET ?', exp_params + [expense_per_page, expense_offset]).fetchall()]
    expense_pagination = pagination_meta(expense_total, expense_page, expense_per_page)
    vendors = [dict(row) for row in conn.execute("SELECT * FROM employees WHERE company_id=? AND emp_type IN ('Provider', 'Supplier') ORDER BY name ASC", (cid,)).fetchall()]
    conn.close()
    return render_template('finance_index.html', today=calculate_financials(t_str, t_str), week=calculate_financials(t_str, w_str), month=calculate_financials(m_s, m_e), services=services, expenses=expenses, vendors=vendors, expense_pagination=expense_pagination, expense_q=expense_q, session=session)

@app.route('/update_vendor', methods=['POST'])
def update_vendor():
    data = request.get_json()
    conn = get_db_connection()
    action_msg = None
    if data.get('id'): 
        conn.execute('UPDATE employees SET name=?, emp_type=?, status=?, phone=?, email=?, bank_details=?, address=? WHERE id=? AND company_id=?', (data['name'], data['emp_type'], data['status'], data.get('phone'), data.get('email'), data.get('bank_details'), data.get('address'), data['id'], session['company_id']))
        action_msg = ('Finance', 'Updated Vendor', f"Updated supplier/vendor: {data['name']}")
    else: 
        conn.execute('INSERT INTO employees (company_id, name, emp_type, status, phone, email, bank_details, address) VALUES (?, ?, ?, ?, ?, ?, ?, ?)', (session['company_id'], data['name'], data['emp_type'], data['status'], data.get('phone'), data.get('email'), data.get('bank_details'), data.get('address')))
        action_msg = ('Finance', 'Created Vendor', f"Created new supplier/vendor: {data['name']}")
    conn.commit()
    conn.close()
    
    if action_msg: log_action(action_msg[0], action_msg[1], action_msg[2])
    return jsonify({"status": "success"})

@app.route('/update_service', methods=['POST'])
def update_service():
    data = request.get_json()
    source_app = 'Invoicing' if request.headers.get('X-App-Source') == 'Invoicing' else 'Finance'
    conn = get_db_connection()
    action_msg = None
    if data.get('id'):
        conn.execute('UPDATE services SET name=?, client_price=?, company_cost=? WHERE id=? AND company_id=?', (data['name'], data['price'], data['cost'], data['id'], session['company_id']))
        action_msg = (source_app, 'Updated Service Margin', f"Updated service pricing: {data['name']}")
    else:
        conn.execute('INSERT INTO services (company_id, name, client_price, company_cost) VALUES (?, ?, ?, ?)', (session['company_id'], data['name'], data['price'], data['cost']))
        action_msg = (source_app, 'Created Service Margin', f"Added new service to pricing map: {data['name']}")
    conn.commit()
    conn.close()

    if action_msg: log_action(action_msg[0], action_msg[1], action_msg[2])
    return jsonify({"status": "success"})


@app.route('/api/services', methods=['GET'])
def api_services():
    conn = get_db_connection()
    services = [dict(s) for s in conn.execute("SELECT id, name, client_price, company_cost FROM services WHERE company_id=? ORDER BY name ASC", (session['company_id'],)).fetchall()]
    conn.close()
    return jsonify({"status": "success", "services": services})


@app.route('/delete_service', methods=['POST'])
def delete_service():
    data = request.get_json() or {}
    service_id = data.get('id')
    if not service_id:
        return jsonify({"status": "error", "message": "Service ID is required."}), 400

    source_app = 'Invoicing' if request.headers.get('X-App-Source') == 'Invoicing' else 'Finance'
    conn = get_db_connection()
    service = conn.execute('SELECT id, name FROM services WHERE id=? AND company_id=?', (service_id, session['company_id'])).fetchone()
    if not service:
        conn.close()
        return jsonify({"status": "error", "message": "Service not found or already deleted."}), 404

    conn.execute('DELETE FROM services WHERE id=? AND company_id=?', (service_id, session['company_id']))
    conn.commit()
    deleted_name = service['name']
    conn.close()

    log_action(source_app, 'Deleted Service Margin', f"Deleted service pricing: {deleted_name}")
    return jsonify({"status": "success", "message": "Service deleted successfully."})

@app.route('/log_expense', methods=['POST'])
def log_expense():
    conn = get_db_connection()
    
    if request.content_type and request.content_type.startswith('multipart/form-data'):
        data = request.form
        invoice_file = request.files.get('invoice_file')
        filename = None
        if invoice_file and invoice_file.filename != '':
            filename = secure_filename(invoice_file.filename)
            filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
            invoice_file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'expenses', filename))
    else:
        data = request.get_json()
        filename = None
        
    conn.execute('INSERT INTO expenses (company_id, date, category, supplier, description, amount, invoice_file) VALUES (?, ?, ?, ?, ?, ?, ?)', 
                 (session['company_id'], data['date'], data['category'], data['supplier'], data['description'], data['amount'], filename))
    conn.commit()
    conn.close()
    
    log_action('Finance', 'Logged Expense', f"Recorded an expense of R{data['amount']} for {data['category']}")
    return jsonify({"status": "success"})

@app.route('/finance_report', methods=['POST'])
def finance_report():
    data = request.get_json()
    stats = calculate_financials(data['start_date'], data['end_date'])
    return jsonify({"stats": stats})

# ==========================================================
# 3. PAYROLL & HR ROUTES
# ==========================================================
@app.route('/payroll')
def payroll_index():
    conn = get_db_connection()
    cid = session['company_id']

    # Ensure older/existing tenants have their setup defaults available before rendering HR screens.
    # Patch 5.3 introduced tenant scorecards, but the payroll page must explicitly load the
    # configured scorecard for the active tenant; otherwise Jinja renders the empty warning.
    comp = conn.execute('SELECT industry_template FROM companies WHERE id=?', (cid,)).fetchone()
    ensure_tenant_template(conn, cid, (dict(comp).get('industry_template') if comp else 'Cleaning') or 'Cleaning', force_reset=False)
    conn.commit()

    payroll_page, payroll_per_page, payroll_offset = get_page_args(request.args, prefix='payroll_', default_per_page=50, max_per_page=100)
    payroll_q = (request.args.get('payroll_q') or '').strip()
    emp_where = "company_id=? AND (emp_type != 'Supplier' OR emp_type IS NULL)"
    emp_params = [cid]
    if payroll_q:
        like = f"%{payroll_q}%"
        emp_where += " AND (COALESCE(name,'') LIKE ? OR COALESCE(emp_number,'') LIKE ? OR COALESCE(job_title,'') LIKE ? OR COALESCE(status,'') LIKE ? OR COALESCE(emp_type,'') LIKE ?)"
        emp_params.extend([like, like, like, like, like])
    employee_total = conn.execute(f"SELECT COUNT(*) FROM employees WHERE {emp_where}", emp_params).fetchone()[0]
    employees = conn.execute(f"SELECT * FROM employees WHERE {emp_where} ORDER BY name ASC LIMIT ? OFFSET ?", emp_params + [payroll_per_page, payroll_offset]).fetchall()
    employee_pagination = pagination_meta(employee_total, payroll_page, payroll_per_page)
    
    emp_data = []
    for emp in employees:
        d = dict(emp)
        d['leave_balance'] = calculate_leave_balance(emp['id'], emp['start_date'], emp['emp_type'], emp['name'])
        d['sick_leave_balance'] = calculate_sick_leave_balance(emp['id'], emp['start_date'], emp['emp_type'], emp['name'])
        d['hours_worked'] = round(conn.execute("SELECT COUNT(*) as c FROM bookings WHERE company_id=? AND start LIKE ? AND employee LIKE ?", (cid, f"{datetime.now().strftime('%Y-%m')}%", f"%{emp['name']}%")).fetchone()['c'] * get_employee_workday_hours(emp), 2)
        emp_data.append(d)

    interviews = conn.execute("SELECT * FROM interviews WHERE company_id=? ORDER BY interview_datetime DESC", (cid,)).fetchall()
    interview_scorecard = get_tenant_scorecard_template(conn, cid)
    interview_scorecard_max_score = get_scorecard_max_score(interview_scorecard)
    
    conn.close()
    return render_template(
        'payroll_index.html',
        employees=emp_data,
        interviews=[dict(i) for i in interviews],
        interview_scorecard=interview_scorecard,
        interview_scorecard_max_score=interview_scorecard_max_score,
        employee_pagination=employee_pagination,
        payroll_q=payroll_q,
        session=session
    )

@app.route('/system_email_settings', methods=['GET', 'POST'])
def system_email_settings():
    if not session.get('is_superadmin'):
        return 'Access Denied: Only Super Admins can manage System Email Settings.', 403
    success_msg = session.pop('settings_success_msg', None)
    if request.method == 'POST':
        save_system_email_settings(request.form)
        success_msg = 'System Email Settings saved successfully!'
        log_action('System', 'Updated System Email Settings', 'Updated global SMTP settings for system generated emails.')
    settings_data = get_system_email_settings()
    return render_template('system_email_settings.html', settings=settings_data, session=session, success_msg=success_msg)

@app.route('/api/test_system_email_connection', methods=['POST'])
def test_system_email_connection():
    if not session.get('is_superadmin'):
        return jsonify({'status': 'error', 'message': 'Forbidden: Only Super Admins can test System Email Settings.'}), 403
    data = request.get_json(silent=True) or {}
    server_host = (data.get('smtp_server') or '').strip()
    port_raw = (data.get('smtp_port') or '').strip()
    user = (data.get('smtp_user') or '').strip()
    password = data.get('smtp_pass') or ''
    if not server_host or not port_raw or not user or not password:
        return jsonify({'status': 'error', 'message': 'Failed: Please fill in Server, Port, Username, and Password.'})
    try:
        port = int(port_raw)
        if port == 587:
            with smtplib.SMTP(server_host, port, timeout=10) as smtp:
                smtp.starttls()
                smtp.login(user, password)
        else:
            with smtplib.SMTP_SSL(server_host, port, timeout=10) as smtp:
                smtp.login(user, password)
        return jsonify({'status': 'success', 'message': 'Successful: system email SMTP connection established and authenticated.'})
    except Exception as exc:
        return jsonify({'status': 'error', 'message': f'Failed: {exc}'})

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if not session.get('is_company_admin') and not session.get('is_superadmin'):
        return "Access Denied: Only Company Admins or Super Admins can access Email & Calendar Settings from the Hub.", 403
    conn = get_db_connection()
    cid = session['company_id']
    success_msg = session.pop('settings_success_msg', None)
    if request.method == 'POST':
        for key in ['smtp_server', 'smtp_port', 'smtp_user', 'smtp_pass', 'sender_email', 'gcal_calendar_id']:
            val = request.form.get(key, '')
            _set_company_setting(conn, cid, key, val)

        # Service account JSON is not used in this workflow. Each company stores
        # its own OAuth Client JSON and its own OAuth token file. The JSON is
        # uploaded as a file, saved server-side, and never echoed back to screen.
        conn.execute("DELETE FROM settings WHERE company_id=? AND key='gcal_service_account_json'", (cid,))
        uploaded_json = request.files.get('gcal_credentials_file')
        legacy_pasted_json = (request.form.get('gcal_credentials_json') or '').strip()
        try:
            calendar_json = _read_uploaded_google_credentials_json(uploaded_json) if uploaded_json and uploaded_json.filename else legacy_pasted_json
            if calendar_json:
                _save_company_google_credentials(cid, calendar_json)
                _set_company_setting(conn, cid, 'gcal_credentials_saved', '1')
                _set_company_setting(conn, cid, 'gcal_credentials_saved_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                if uploaded_json and uploaded_json.filename:
                    _set_company_setting(conn, cid, 'gcal_credentials_original_filename', os.path.basename(uploaded_json.filename))
                _clear_company_google_token(cid, conn)
                success_msg = "Settings saved successfully. Google Calendar JSON was uploaded for this company. Please reconnect Google Calendar to create a new token."
            else:
                success_msg = "Settings saved successfully!"
            conn.commit()
        except Exception as exc:
            conn.rollback()
            rows = conn.execute('SELECT key, value FROM settings WHERE company_id=?', (cid,)).fetchall()
            s_dict = {row['key']: row['value'] for row in rows}
            try:
                s_dict['gcal_credentials_file_exists'] = '1' if os.path.exists(_company_google_credentials_path(cid)) else ''
                s_dict['gcal_token_file_exists'] = '1' if os.path.exists(_company_google_token_path(cid)) else ''
            except Exception:
                pass
            conn.close()
            success_msg = "Settings were not saved: " + _format_google_calendar_error(exc)
            return render_template('settings.html', settings=s_dict, session=session, success_msg=success_msg)

    rows = conn.execute('SELECT key, value FROM settings WHERE company_id=?', (cid,)).fetchall()
    s_dict = {row['key']: row['value'] for row in rows}
    try:
        s_dict['gcal_credentials_file_exists'] = '1' if os.path.exists(_company_google_credentials_path(cid)) else ''
        s_dict['gcal_token_file_exists'] = '1' if os.path.exists(_company_google_token_path(cid)) else ''
    except Exception:
        pass
    conn.close()
    
    if request.method == 'POST':
        log_action('System', 'Updated Settings', "Modified company integrations & email configurations.")
        
    return render_template('settings.html', settings=s_dict, session=session, success_msg=success_msg)


@app.route('/settings/google_calendar/connect')
def google_calendar_oauth_connect():
    if not session.get('is_company_admin') and not session.get('is_superadmin'):
        return "Access Denied: Only Company Admins or Super Admins can connect Google Calendar.", 403
    try:
        redirect_uri = _google_redirect_uri()
        flow = _build_google_oauth_flow(session.get('company_id'), redirect_uri)
        settings = _company_calendar_settings(session.get('company_id'))
        auth_kwargs = {
            'access_type': 'offline',
            'include_granted_scopes': 'true',
            'prompt': 'consent'
        }
        smtp_user = str(settings.get('smtp_user') or '').strip()
        if smtp_user:
            auth_kwargs['login_hint'] = smtp_user
        authorization_url, state = flow.authorization_url(**auth_kwargs)
        session['google_oauth_state'] = state
        session['google_oauth_company_id'] = session.get('company_id')
        return redirect(authorization_url)
    except Exception as exc:
        session['settings_success_msg'] = 'Google Calendar connection could not start: ' + _format_google_calendar_error(exc)
        return redirect(url_for('settings'))


@app.route('/settings/google_calendar/callback')
def google_calendar_oauth_callback():
    if not session.get('is_company_admin') and not session.get('is_superadmin'):
        return "Access Denied: Only Company Admins or Super Admins can connect Google Calendar.", 403
    if request.args.get('error'):
        session['settings_success_msg'] = 'Google Calendar connection cancelled or failed: ' + request.args.get('error')
        return redirect(url_for('settings'))
    try:
        state = session.get('google_oauth_state')
        oauth_company_id = session.get('google_oauth_company_id') or session.get('company_id')
        redirect_uri = _google_redirect_uri()
        flow = _build_google_oauth_flow(oauth_company_id, redirect_uri)
        if state:
            flow.oauth2session.state = state
        authorization_response = request.url
        if os.environ.get('RENDER') and authorization_response.startswith('http://'):
            authorization_response = 'https://' + authorization_response[len('http://'):]
        flow.fetch_token(authorization_response=authorization_response)
        creds = flow.credentials
        token_path = _preferred_google_token_path(oauth_company_id)
        token_dir = os.path.dirname(token_path)
        if token_dir:
            os.makedirs(token_dir, exist_ok=True)
        with open(token_path, 'w', encoding='utf-8') as token:
            token.write(creds.to_json())
        token_id = _create_google_token_id()
        conn = get_db_connection()
        try:
            _set_company_setting(conn, oauth_company_id, 'gcal_token_id', token_id)
            _set_company_setting(conn, oauth_company_id, 'gcal_token_saved_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            conn.commit()
        finally:
            conn.close()
        session.pop('google_oauth_state', None)
        session.pop('google_oauth_company_id', None)
        session['settings_success_msg'] = 'Google Calendar connected successfully for this company. Token ID: ' + token_id + '. Bookings will use the Target Google Calendar ID, or the connected account primary calendar if the field is blank.'
        return redirect(url_for('settings'))
    except Exception as exc:
        session['settings_success_msg'] = 'Google Calendar connection failed: ' + _format_google_calendar_error(exc)
        return redirect(url_for('settings'))


@app.route('/api/test_email_connection', methods=['POST'])
def test_email_connection():
    if not session.get('is_company_admin') and not session.get('is_superadmin'):
        return jsonify({"status": "error", "message": "Forbidden: Only Company Admins or Super Admins can test Email & Calendar Settings."}), 403

    data = request.json
    server = data.get('smtp_server')
    port = data.get('smtp_port')
    user = data.get('smtp_user')
    password = data.get('smtp_pass')

    if not server or not port or not user or not password:
        return jsonify({"status": "error", "message": "Failed: Please fill in the Server, Port, Username, and Password fields."})

    try:
        port = int(port)
        if port == 587:
            with smtplib.SMTP(server, port, timeout=10) as smtp:
                smtp.starttls()
                smtp.login(user, password)
        else:
            with smtplib.SMTP_SSL(server, port, timeout=10) as smtp:
                smtp.login(user, password)
        
        return jsonify({"status": "success", "message": "Successful: Connection established and authenticated!"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Failed: {str(e)}"})


@app.route('/api/test_google_calendar_connection', methods=['POST'])
def test_google_calendar_connection():
    if not session.get('is_company_admin') and not session.get('is_superadmin'):
        return jsonify({"status": "error", "message": "Forbidden: Only Company Admins or Super Admins can test Google Calendar Settings."}), 403
    try:
        result = test_google_calendar_sync(session.get('company_id'))
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": "Failed: " + _format_google_calendar_error(e)})


@app.route('/api/sync_existing_bookings_google_calendar', methods=['POST'])
def sync_existing_bookings_google_calendar():
    if not session.get('is_company_admin') and not session.get('is_superadmin'):
        return jsonify({"status": "error", "message": "Forbidden: Only Company Admins or Super Admins can sync existing bookings."}), 403
    data = request.get_json(silent=True) or {}
    start_date = (data.get('start_date') or '').strip()
    end_date = (data.get('end_date') or '').strip()
    if not start_date or not end_date:
        return jsonify({"status": "error", "message": "Failed: Please select both From Date and To Date."}), 400
    try:
        result = sync_existing_bookings_to_google_calendar(session.get('company_id'), start_date, end_date)
        msg = f"Sync complete. Created {result['created']} calendar event(s), updated {result['updated']} event(s)."
        if result.get('skipped'):
            msg += f" Skipped {result['skipped']} booking(s)."
        if result.get('failed'):
            msg += f" Failed {result['failed']} booking(s)."
            if result.get('errors'):
                msg += " First errors: " + " | ".join(result['errors'])
        log_action('System', 'Synced Existing Bookings to Google Calendar', f"Synced bookings from {start_date} to {end_date}: created {result['created']}, updated {result['updated']}, failed {result['failed']}.")
        return jsonify({"status": "success" if not result.get('failed') else "warning", "message": msg, "result": result})
    except Exception as exc:
        return jsonify({"status": "error", "message": "Failed: " + _format_google_calendar_error(exc)}), 400


@app.route('/email_payslip', methods=['POST'])
def email_payslip():
    if not session.get('can_payroll') and not session.get('is_superadmin'): return jsonify({"message": "Forbidden"}), 403
    emp_id, date_str, pdf_file = request.form.get('employee_id'), request.form.get('date'), request.files.get('pdf')
    if not pdf_file: return jsonify({"message": "No PDF provided."}), 400
    conn = get_db_connection()
    emp = conn.execute('SELECT name, email FROM employees WHERE id=? AND company_id=?', (emp_id, session['company_id'])).fetchone()
    settings_rows = conn.execute('SELECT key, value FROM settings WHERE company_id=?', (session['company_id'],)).fetchall()
    conn.close()
    
    s_dict = {r['key']: r['value'] for r in settings_rows}
    if not emp or not emp['email']: return jsonify({"message": "Employee has no email address saved in their profile."}), 400
    if not s_dict.get('smtp_server') or not s_dict.get('smtp_user') or not s_dict.get('smtp_pass'): return jsonify({"message": "SMTP settings are incomplete. Please configure them in Settings."}), 400
        
    try:
        msg = EmailMessage()
        msg['Subject'] = f"Payslip - {date_str}"
        msg['From'] = s_dict.get('sender_email', s_dict.get('smtp_user'))
        msg['To'] = emp['email']
        msg.set_content(f"Dear {emp['name']},\n\nPlease find your attached payslip for {date_str}.\n\nKind regards,\n{session.get('company_name')}")
        msg.add_attachment(pdf_file.read(), maintype='application', subtype='pdf', filename=f"Payslip_{emp['name']}_{date_str}.pdf")
        
        port = int(s_dict.get('smtp_port', 465))
        if port == 587:
            with smtplib.SMTP(s_dict['smtp_server'], port, timeout=10) as server:
                server.starttls()
                server.login(s_dict['smtp_user'], s_dict['smtp_pass'])
                server.send_message(msg)
        else:
            with smtplib.SMTP_SSL(s_dict['smtp_server'], port, timeout=10) as server:
                server.login(s_dict['smtp_user'], s_dict['smtp_pass'])
                server.send_message(msg)
            
        log_action('HR & Payroll', 'Emailed Payslip', f"Successfully sent digital payslip to {emp['name']} for {date_str}")
        return jsonify({"message": "Email sent successfully!"})
    except Exception as e: 
        return jsonify({"message": f"Error sending email: {str(e)}"}), 500


def compose_bank_details(data):
    parts = [
        data.get('bank_name') or '',
        data.get('account_holder') or '',
        data.get('account_number') or '',
        data.get('branch_code') or '',
        data.get('account_type') or '',
        data.get('payment_reference') or ''
    ]
    return ' | '.join([str(p).strip() for p in parts if str(p or '').strip()])

def normalise_account_type(value):
    value = (value or '').strip().lower()
    if value in ['cheque', 'current', 'transmission']:
        return 'Current'
    if value in ['savings', 'save']:
        return 'Savings'
    if value in ['credit']:
        return 'Credit'
    return (value.title() if value else '')

def get_bank_export_templates():
    return {
        'generic_csv': 'Generic CSV',
        'fnb_enterprise_csv': 'FNB Enterprise CSV',
        'absa_bio_csv': 'ABSA BIO CSV',
        'standard_bank_bol_csv': 'Standard Bank Business Online CSV',
        'nedbank_acb_csv': 'Nedbank / ACB-style CSV'
    }

def payroll_payment_rows(conn, company_id, month_str):
    # Bank exports use finalised ledger rows only, including adjustment payslips.
    # Multiple ledger rows for an employee/month are combined into one net payment amount.
    return conn.execute('''
        SELECT MAX(p.id) AS payslip_id, MAX(p.date) AS date, SUM(p.net_salary) AS net_salary,
               e.emp_number, e.name, e.id_passport, e.bank_name, e.account_holder,
               e.account_number, e.branch_code, e.account_type, e.payment_reference
        FROM payslips p
        JOIN employees e ON e.id = p.employee_id AND e.company_id = p.company_id
        WHERE p.company_id=? AND p.date LIKE ?
        GROUP BY e.id
        ORDER BY e.name ASC
    ''', (company_id, f"{month_str}%")).fetchall()

def validate_payroll_bank_rows(rows):
    missing = []
    for r in rows:
        required = ['bank_name', 'account_holder', 'account_number', 'branch_code', 'account_type']
        empty = [f for f in required if not str(r[f] or '').strip()]
        if empty:
            missing.append(f"{r['name']}: " + ', '.join(empty))
    return missing

def build_payroll_bank_csv(rows, template_key, month_str, company_name):
    si = io.StringIO()
    cw = csv.writer(si)
    label = get_bank_export_templates().get(template_key, 'Generic CSV')

    if template_key == 'fnb_enterprise_csv':
        cw.writerow(['Recipient Name', 'Bank Name', 'Account Number', 'Branch Code', 'Account Type', 'Amount', 'Recipient Reference', 'Own Reference'])
        for r in rows:
            cw.writerow([r['account_holder'] or r['name'], r['bank_name'], r['account_number'], r['branch_code'], normalise_account_type(r['account_type']), f"{float(r['net_salary'] or 0):.2f}", r['payment_reference'] or f"Salary {month_str}", f"{company_name} Payroll"])
    elif template_key == 'absa_bio_csv':
        cw.writerow(['Account Holder', 'Bank', 'Branch Code', 'Account Number', 'Account Type', 'Amount', 'Statement Reference', 'Employee Number'])
        for r in rows:
            cw.writerow([r['account_holder'] or r['name'], r['bank_name'], r['branch_code'], r['account_number'], normalise_account_type(r['account_type']), f"{float(r['net_salary'] or 0):.2f}", r['payment_reference'] or f"Salary {month_str}", r['emp_number']])
    elif template_key == 'standard_bank_bol_csv':
        cw.writerow(['Beneficiary Name', 'Beneficiary Bank', 'Branch Code', 'Account Number', 'Account Type', 'Payment Amount', 'Reference'])
        for r in rows:
            cw.writerow([r['account_holder'] or r['name'], r['bank_name'], r['branch_code'], r['account_number'], normalise_account_type(r['account_type']), f"{float(r['net_salary'] or 0):.2f}", r['payment_reference'] or f"Salary {month_str}"])
    elif template_key == 'nedbank_acb_csv':
        cw.writerow(['Record Type', 'Account Name', 'Bank', 'Branch Code', 'Account Number', 'Account Type', 'Amount', 'Reference'])
        for r in rows:
            cw.writerow(['PAYMENT', r['account_holder'] or r['name'], r['bank_name'], r['branch_code'], r['account_number'], normalise_account_type(r['account_type']), f"{float(r['net_salary'] or 0):.2f}", r['payment_reference'] or f"Salary {month_str}"])
    else:
        cw.writerow(['Employee Number', 'Employee Name', 'Bank Name', 'Account Holder', 'Account Number', 'Branch Code', 'Account Type', 'Amount', 'Payment Reference'])
        for r in rows:
            cw.writerow([r['emp_number'], r['name'], r['bank_name'], r['account_holder'] or r['name'], r['account_number'], r['branch_code'], normalise_account_type(r['account_type']), f"{float(r['net_salary'] or 0):.2f}", r['payment_reference'] or f"Salary {month_str}"])

    return si.getvalue(), label

@app.route('/update_employee', methods=['POST'])
def update_employee():
    if not session.get('can_payroll') and not session.get('is_superadmin'): return "Forbidden", 403
    data = request.form
    emp_id, cid = data.get('id'), session['company_id']
    cv_file, id_file, contract_file = request.files.get('cv_file'), request.files.get('id_file'), request.files.get('contract_file')
    cv_filename, id_filename, contract_filename = None, None, None
    
    if cv_file and cv_file.filename != '':
        cv_filename = secure_filename(cv_file.filename)
        cv_file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'cvs', cv_filename))
    if id_file and id_file.filename != '':
        id_filename = secure_filename(id_file.filename)
        id_file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'ids', id_filename))
    if contract_file and contract_file.filename != '':
        contract_filename = secure_filename(contract_file.filename)
        contract_file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'contracts', contract_filename))

    required_fields = {
        'name': 'Full Name',
        'job_title': 'Job Title',
        'status': 'Status',
        'emp_type': 'Contract Type',
        'gross_salary': 'Monthly Base Salary / Rate (R)',
        'id_passport': 'ID / Passport Number',
        'date_of_birth': 'Date of Birth',
        'start_date': 'Start Date',
        'phone': 'Phone Number',
        'emergency_contact': 'Emergency Contact',
        'address': 'Physical Address',
        'overtime_pay_treatment': 'Overtime PAYE Treatment'
    }
    missing_fields = [label for key, label in required_fields.items() if not str(data.get(key) or '').strip()]
    if missing_fields:
        return jsonify({"status": "error", "message": "Missing required fields: " + ", ".join(missing_fields)}), 400
    try:
        emp_workday_hours = float(data.get('workday_hours') or 7)
        if emp_workday_hours <= 0:
            raise ValueError()
    except Exception:
        return jsonify({"status": "error", "message": "Working Hours per Day must be greater than 0."}), 400
    overtime_pay_treatment = (data.get('overtime_pay_treatment') or '').strip().lower()
    if overtime_pay_treatment not in ['regular', 'irregular']:
        return jsonify({"status": "error", "message": "Overtime PAYE Treatment must be Regular recurring or Irregular/once-off."}), 400
    inactive_date = (data.get('inactive_date') or '').strip()
    if inactive_date:
        try:
            inactive_dt = datetime.strptime(inactive_date, '%Y-%m-%d')
            start_dt = datetime.strptime(data.get('start_date'), '%Y-%m-%d')
            if inactive_dt < start_dt:
                return jsonify({"status": "error", "message": "Inactive Date / Last Working Day cannot be before Start Date."}), 400
        except Exception:
            return jsonify({"status": "error", "message": "Inactive Date / Last Working Day must be a valid date."}), 400
    emp_add_leave = data.get('additional_leave', 0)

    conn = get_db_connection()
    action_msg = None
    if emp_id:
        existing = conn.execute("SELECT cv_file, id_file, contract_file FROM employees WHERE id=? AND company_id=?", (emp_id, session['company_id'])).fetchone()
        if not existing:
            conn.close()
            return jsonify({"status": "error", "message": "Employee not found."}), 404
        final_cv = cv_filename if cv_filename else existing['cv_file']
        final_id = id_filename if id_filename else existing['id_file']
        final_contract = contract_filename if contract_filename else existing['contract_file']
        if not final_id or not final_contract:
            conn.close()
            return jsonify({"status": "error", "message": "Upload Signed Contract and Upload ID/Passport Copy are required."}), 400
        
        conn.execute('''UPDATE employees SET name=?, job_title=?, emp_type=?, status=?, start_date=?, inactive_date=?, date_of_birth=?, gross_salary=?, id_passport=?, phone=?, email=?, emergency_contact=?, tax_number=?, paye_ref=?, bank_details=?, bank_name=?, account_holder=?, account_number=?, branch_code=?, account_type=?, payment_reference=?, address=?, cv_file=?, id_file=?, contract_file=?, additional_leave=?, workday_hours=?, overtime_pay_treatment=?, uif_contributor=?, uif_non_contributor_reason=?, uif_termination_code=? WHERE id=? AND company_id=?''', (data.get('name'), data.get('job_title'), data.get('emp_type'), data.get('status'), data.get('start_date'), inactive_date, data.get('date_of_birth'), data.get('gross_salary'), data.get('id_passport'), data.get('phone'), data.get('email'), data.get('emergency_contact'), data.get('tax_number'), data.get('paye_ref'), compose_bank_details(data), data.get('bank_name'), data.get('account_holder'), data.get('account_number'), data.get('branch_code'), data.get('account_type'), data.get('payment_reference'), data.get('address'), final_cv, final_id, final_contract, emp_add_leave, emp_workday_hours, overtime_pay_treatment, data.get('uif_contributor') or 'Yes', data.get('uif_non_contributor_reason') or '', data.get('uif_termination_code') or '', emp_id, cid))
        action_msg = ('HR & Payroll', 'Updated Employee', f"Updated profile information for {data.get('name')}")
    else:
        count = conn.execute("SELECT COUNT(*) FROM employees WHERE company_id=?", (cid,)).fetchone()[0]
        emp_num = f"EMP{count+1:03d}"
        if not id_filename or not contract_filename:
            conn.close()
            return jsonify({"status": "error", "message": "Upload Signed Contract and Upload ID/Passport Copy are required."}), 400
        conn.execute('''INSERT INTO employees (company_id, name, emp_number, job_title, emp_type, status, start_date, inactive_date, date_of_birth, gross_salary, id_passport, phone, email, emergency_contact, tax_number, paye_ref, bank_details, bank_name, account_holder, account_number, branch_code, account_type, payment_reference, address, cv_file, id_file, contract_file, additional_leave, workday_hours, overtime_pay_treatment, uif_contributor, uif_non_contributor_reason, uif_termination_code) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (cid, data.get('name'), emp_num, data.get('job_title'), data.get('emp_type'), data.get('status'), data.get('start_date'), inactive_date, data.get('date_of_birth'), data.get('gross_salary'), data.get('id_passport'), data.get('phone'), data.get('email'), data.get('emergency_contact'), data.get('tax_number'), data.get('paye_ref'), compose_bank_details(data), data.get('bank_name'), data.get('account_holder'), data.get('account_number'), data.get('branch_code'), data.get('account_type'), data.get('payment_reference'), data.get('address'), cv_filename, id_filename, contract_filename, emp_add_leave, emp_workday_hours, overtime_pay_treatment, data.get('uif_contributor') or 'Yes', data.get('uif_non_contributor_reason') or '', data.get('uif_termination_code') or ''))
        action_msg = ('HR & Payroll', 'Created Employee', f"Onboarded new employee: {data.get('name')}")
    conn.commit()
    conn.close()
    
    if action_msg: log_action(action_msg[0], action_msg[1], action_msg[2])
    return jsonify({"status": "success"})

@app.route('/save_interview', methods=['POST'])
def save_interview():
    if not session.get('can_payroll') and not session.get('is_superadmin'): return "Forbidden", 403
    data = request.form
    int_id = data.get('id')
    cid = session['company_id']
    
    cv_file, id_file = request.files.get('cv_file'), request.files.get('id_file')
    cv_filename, id_filename = None, None
    if cv_file and cv_file.filename != '':
        cv_filename = secure_filename(cv_file.filename)
        cv_file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'cvs', cv_filename))
    if id_file and id_file.filename != '':
        id_filename = secure_filename(id_file.filename)
        id_file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'ids', id_filename))

    conn = get_db_connection()
    action_msg = None
    if int_id:
        existing = conn.execute("SELECT cv_file, id_file FROM interviews WHERE id=? AND company_id=?", (int_id, session['company_id'])).fetchone()
        final_cv = cv_filename if cv_filename else existing['cv_file']
        final_id = id_filename if id_filename else existing['id_file']
        conn.execute('''UPDATE interviews SET name=?, email=?, phone=?, id_passport=?, address=?, interview_datetime=?, cv_file=?, id_file=?, scorecard_json=?, total_score=?, final_decision=?, interview_notes=? WHERE id=? AND company_id=?''',
                     (data.get('name'), data.get('email'), data.get('phone'), data.get('id_passport'), data.get('address'), data.get('interview_datetime'), final_cv, final_id, data.get('scorecard_json'), data.get('total_score', 0), data.get('final_decision'), data.get('interview_notes'), int_id, cid))
        action_msg = ('HR & Payroll', 'Updated Interview', f"Modified candidate interview for {data.get('name')}")
    else:
        conn.execute('''INSERT INTO interviews (company_id, name, email, phone, id_passport, address, interview_datetime, cv_file, id_file, scorecard_json, total_score, final_decision, interview_notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                     (cid, data.get('name'), data.get('email'), data.get('phone'), data.get('id_passport'), data.get('address'), data.get('interview_datetime'), cv_filename, id_filename, data.get('scorecard_json'), data.get('total_score', 0), data.get('final_decision', 'Pending'), data.get('interview_notes')))
        action_msg = ('HR & Payroll', 'Scheduled Interview', f"Logged new candidate: {data.get('name')}")
    conn.commit()
    conn.close()
    
    if action_msg: log_action(action_msg[0], action_msg[1], action_msg[2])
    return jsonify({"status": "success"})

@app.route('/delete_interview', methods=['POST'])
def delete_interview():
    if not session.get('can_payroll') and not session.get('is_superadmin'): return "Forbidden", 403
    conn = get_db_connection()
    conn.execute('DELETE FROM interviews WHERE id=? AND company_id=?', (request.get_json().get('id'), session['company_id']))
    conn.commit()
    conn.close()
    
    log_action('HR & Payroll', 'Deleted Interview', "Removed a candidate interview record.")
    return jsonify({"status": "success"})

@app.route('/record_leave', methods=['POST'])
def record_leave():
    data, doc_file = request.form, request.files.get('leave_doc')
    doc_filename = secure_filename(doc_file.filename) if doc_file and doc_file.filename else None
    if doc_filename: doc_file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'leave', doc_filename))
    conn = get_db_connection()
    conn.execute('INSERT INTO leave_records (company_id, employee_id, date_taken, days, leave_type, document_file) VALUES (?, ?, ?, ?, ?, ?)', (session['company_id'], data.get('employee_id'), data.get('date'), data.get('days'), data.get('leave_type'), doc_filename))
    conn.commit()
    conn.close()
    
    log_action('HR & Payroll', 'Logged Leave', f"Recorded {data.get('days')} days of {data.get('leave_type')}")
    return jsonify({"status": "success"})

@app.route('/update_leave', methods=['POST'])
def update_leave():
    data, doc_file = request.form, request.files.get('leave_doc')
    conn = get_db_connection()
    if doc_file and doc_file.filename:
        doc_filename = secure_filename(doc_file.filename)
        doc_file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'leave', doc_filename))
        conn.execute('UPDATE leave_records SET date_taken=?, days=?, leave_type=?, document_file=? WHERE id=? AND company_id=?', (data.get('date'), data.get('days'), data.get('leave_type'), doc_filename, data.get('leave_id'), session['company_id']))
    else: conn.execute('UPDATE leave_records SET date_taken=?, days=?, leave_type=? WHERE id=? AND company_id=?', (data.get('date'), data.get('days'), data.get('leave_type'), data.get('leave_id'), session['company_id']))
    conn.commit()
    conn.close()
    
    log_action('HR & Payroll', 'Updated Leave', f"Modified a leave record ({data.get('days')} days).")
    return jsonify({"status": "success"})

@app.route('/delete_leave', methods=['POST'])
def delete_leave():
    conn = get_db_connection()
    conn.execute('DELETE FROM leave_records WHERE id=? AND company_id=?', (request.get_json().get('leave_id'), session['company_id']))
    conn.commit()
    conn.close()
    
    log_action('HR & Payroll', 'Deleted Leave', "Deleted a leave record entry.")
    return jsonify({"status": "success"})

@app.route('/api/save_payslip', methods=['POST'])
def save_payslip():
    if not session.get('can_payroll') and not session.get('is_superadmin'): return "Forbidden", 403
    data = request.json
    target_month = data['date'][:7]
    try:
        ensure_tax_tables_configured(get_sars_tax_year(data['date']))
    except TaxTableNotSetError as e:
        return jsonify({"message": str(e)}), 400

    conn = get_db_connection()
    employee = conn.execute("SELECT * FROM employees WHERE id=? AND company_id=?", (data['employee_id'], session['company_id'])).fetchone()
    if not employee:
        conn.close()
        return jsonify({"message": "Employee not found."}), 404
    _month_start, _month_end, _inactive_date, payroll_cutoff = get_employee_payroll_cutoff(employee, data['date'])
    if payroll_cutoff is None:
        conn.close()
        return jsonify({"message": "Employee inactive before this payroll month. Payroll cannot be saved beyond the inactive date."}), 400

    existing_rows = conn.execute("SELECT id FROM payslips WHERE company_id=? AND employee_id=? AND date LIKE ? AND COALESCE(payslip_type, 'regular')='regular' ORDER BY id DESC", (session['company_id'], data['employee_id'], f"{target_month}%")).fetchall()
    exists = existing_rows[0] if existing_rows else None
    
    if exists:
        conn.execute('''UPDATE payslips SET date=?, gross_salary=?, overtime=?, transport=?, bonus=?, reimbursable_expenses=?, loan_repayment=?, uif=?, paye=?, net_salary=? WHERE id=? AND company_id=?''',
                     (data['date'], data['gross'], data['overtime'], data.get('transport', 0), data.get('bonus', 0), data.get('reimbursable_expenses', 0), data.get('loan_repayment', 0), data['uif'], data['paye'], data['net'], exists['id'], session['company_id']))
        # Defensive duplicate cleanup: keep the newest ledger row for this employee/month only.
        if len(existing_rows) > 1:
            duplicate_ids = [row['id'] for row in existing_rows[1:]]
            conn.executemany('DELETE FROM payslips WHERE id=? AND company_id=?', [(dup_id, session['company_id']) for dup_id in duplicate_ids])
    else:
        conn.execute('''INSERT INTO payslips (company_id, employee_id, date, gross_salary, overtime, transport, bonus, reimbursable_expenses, loan_repayment, uif, paye, net_salary, payslip_type, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'regular', ?)''',
                     (session['company_id'], data['employee_id'], data['date'], data['gross'], data['overtime'], data.get('transport', 0), data.get('bonus', 0), data.get('reimbursable_expenses', 0), data.get('loan_repayment', 0), data['uif'], data['paye'], data['net'], datetime.now().isoformat(timespec='seconds')))
                     
    conn.commit()
    conn.close()
    log_action('HR & Payroll', 'Saved Payslip', f"Saved payslip ledger for Employee ID {data['employee_id']} for {target_month}")
    return jsonify({"status": "success"})

@app.route('/generate_payslip', methods=['POST'])
def generate_payslip():
    data = request.get_json()
    emp_id, date_str = data.get('employee_id'), data.get('date')
    bonus_amount = safe_money(data.get('bonus'))
    reimbursable_amount = safe_money(data.get('reimbursable_expenses'))
    loan_repayment_amount = safe_money(data.get('loan_repayment'))
    target_month = date_str[:7] 

    conn = get_db_connection()
    emp = conn.execute('SELECT * FROM employees WHERE id=? AND company_id=?', (emp_id, session['company_id'])).fetchone()
    if not emp:
        conn.close()
        return jsonify({"message": "Employee not found."}), 404

    company = conn.execute('SELECT * FROM companies WHERE id=?', (session['company_id'],)).fetchone()
    existing_ledger = conn.execute('''SELECT * FROM payslips
                                      WHERE company_id=? AND employee_id=? AND date LIKE ?
                                        AND COALESCE(payslip_type, 'regular')='regular'
                                      ORDER BY id DESC LIMIT 1''',
                                   (session['company_id'], emp_id, f"{target_month}%")).fetchone()
    if existing_ledger:
        payload = build_saved_payslip_payload(conn, emp, company, existing_ledger)
        adj_rows = conn.execute('''SELECT id, date, gross_salary, overtime, transport, bonus, reimbursable_expenses, loan_repayment, uif, paye, net_salary, adjustment_reason
                                   FROM payslips
                                   WHERE company_id=? AND employee_id=? AND date LIKE ?
                                     AND COALESCE(payslip_type, 'regular')='adjustment'
                                   ORDER BY id ASC''',
                                (session['company_id'], emp_id, f"{target_month}%")).fetchall()
        payload['payslip']['adjustment_count'] = len(adj_rows)
        payload['payslip']['adjustment_net_total'] = f"R {sum(float(r['net_salary'] or 0) for r in adj_rows):.2f}"
        payload['payslip']['regular_payslip_id'] = existing_ledger['id']
        conn.close()
        return jsonify(payload)

    tax_year = get_sars_tax_year(date_str)
    try:
        ensure_tax_tables_configured(tax_year)
    except TaxTableNotSetError as e:
        conn.close()
        return jsonify({"message": str(e)}), 400

    month_start, month_end, inactive_date, payroll_cutoff = get_employee_payroll_cutoff(emp, date_str)
    if payroll_cutoff is None:
        conn.close()
        return jsonify({"message": "Employee inactive before this payroll month. Payroll cannot be generated beyond the inactive date."}), 400

    bookings = conn.execute("""SELECT * FROM bookings
                            WHERE company_id=? AND start LIKE ? AND employee LIKE ?
                            AND substr(start, 1, 10) <= ?""",
                            (session['company_id'], f"{target_month}%", f"%{emp['name']}%", payroll_cutoff.strftime('%Y-%m-%d'))).fetchall()
    
    days_worked = len(bookings)
    base_salary = float(emp['gross_salary'] or 0)
    
    emp_type = emp['emp_type'] or ''
    is_daily_rate = emp_type in ['Contract <25 Hrs', 'Contract >25 Hrs'] and base_salary < 2000
    
    payable_base_salary, inactive_salary_note = prorate_monthly_salary_for_inactive_date(base_salary, emp_type, month_start, month_end, payroll_cutoff)
    divisor = 26.0 if '(6 Days)' in emp_type else 22.0
    daily_rate = base_salary if is_daily_rate else (base_salary / divisor)
    workday_hours = get_employee_workday_hours(emp)
    
    overtime_amount = 0.0
    sundays_display = ""
    
    target_year = date_str[:4]
    hol_db = conn.execute("SELECT date_str FROM public_holidays WHERE year=?", (target_year,)).fetchall()
    sa_holidays = [h['date_str'] for h in hol_db]
    booking_hours = analyse_booking_hours(bookings, emp_type, workday_hours, sa_holidays)
    saturday_count = booking_hours['saturday_nonordinary_days']
    sunday_premium_count = booking_hours['sunday_premium_days']
    public_holiday_count = booking_hours['public_holiday_days']
    total_overtime_hours = booking_hours['explicit_overtime_hours']

    hourly_rate = daily_rate / workday_hours
    rules = get_contract_day_rules(emp_type)

    if is_daily_rate:
        # Daily-rate workers already receive 1.0x in gross for each booking. Add only the extra
        # premium portion required to reach the statutory/applicable multiplier.
        overtime_amount += daily_rate * max(0.0, rules['saturday_multiplier'] - 1.0) * saturday_count
        overtime_amount += daily_rate * max(0.0, rules['sunday_multiplier'] - 1.0) * sunday_premium_count
        overtime_amount += daily_rate * max(0.0, rules['public_holiday_multiplier'] - 1.0) * public_holiday_count
    else:
        # Monthly-rate employees receive the applicable premium amount as an additional earning.
        overtime_amount += (hourly_rate * workday_hours * rules['saturday_multiplier']) * saturday_count
        overtime_amount += (hourly_rate * workday_hours * rules['sunday_multiplier']) * sunday_premium_count
        overtime_amount += (hourly_rate * workday_hours * rules['public_holiday_multiplier']) * public_holiday_count

    if total_overtime_hours > 0:
        overtime_amount += (hourly_rate * 1.5) * total_overtime_hours
        
    display_parts = []
    if public_holiday_count > 0: display_parts.append(f"{public_holiday_count} Public Holiday @2.0x")
    if sunday_premium_count > 0: display_parts.append(f"{sunday_premium_count} Sun @{rules['sunday_multiplier']:.1f}x")
    if saturday_count > 0: display_parts.append(f"{saturday_count} Sat @{rules['saturday_multiplier']:.1f}x")
    if total_overtime_hours > 0: display_parts.append(f"{total_overtime_hours} hrs extra OT @1.5x")
    sundays_display = f"({', '.join(display_parts)})" if display_parts else ""

    days_worked_display = f"({days_worked} shifts @ R{base_salary:.2f})" if is_daily_rate else inactive_salary_note
    gross = (base_salary * days_worked) if is_daily_rate else payable_base_salary

    transport_amount = 0.0
    if company and dict(company).get('transport_policy') in ['standard', 'yes']:
        transport_amount_per_lift = float(dict(company).get('transport_amount_per_lift') or 25)
        for b in bookings:
            t_val = b['transport'] or ''
            company_lifts = 0
            if 'Pickup' in t_val: company_lifts += 1
            if 'Drop Off' in t_val: company_lifts += 1
            transport_amount += max(0, (2 - company_lifts) * transport_amount_per_lift)

    overtime_pay_treatment = (emp['overtime_pay_treatment'] if 'overtime_pay_treatment' in emp.keys() else 'irregular') or 'irregular'
    overtime_pay_treatment = overtime_pay_treatment.lower()
    if overtime_pay_treatment == 'regular':
        regular_taxable = gross + overtime_amount
        irregular_taxable = bonus_amount
    else:
        regular_taxable = gross
        irregular_taxable = overtime_amount + bonus_amount

    total_taxable = gross + overtime_amount + bonus_amount
    uif = calculate_uif(total_taxable)
    paye = calculate_paye_with_regular_irregular(regular_taxable, irregular_taxable, date_str, emp['date_of_birth'])
    net = total_taxable - uif - paye + transport_amount + reimbursable_amount - loan_repayment_amount
    
    benefit_ref_date = payroll_cutoff.strftime('%Y-%m-%d')
    leave = calculate_leave_balance(emp['id'], emp['start_date'], emp['emp_type'], emp['name'], benefit_ref_date)
    sick = calculate_sick_leave_balance(emp['id'], emp['start_date'], emp['emp_type'], emp['name'], benefit_ref_date)
    family = calculate_family_leave_balance(emp['id'], emp['start_date'], emp['emp_type'], emp['name'], benefit_ref_date)
    bcea_warning = get_bcea_hours_warning(booking_hours['ordinary_hours'], booking_hours['overtime_hours'], emp_type)
    
    conn.close()

    return jsonify({
        "payslip": { 
            "date": target_month, 
            "name": emp['name'], 
            "emp_num": emp['emp_number'] or 'N/A', 
            "id_num": emp['id_passport'] or 'N/A', 
            "gross": f"R {gross:.2f}", 
            "days_worked_display": days_worked_display, 
            "overtime": f"R {overtime_amount:.2f}", 
            "sundays_display": sundays_display, 
            "bonus": f"R {bonus_amount:.2f}",
            "reimbursable_expenses": f"R {reimbursable_amount:.2f}",
            "loan_repayment": f"R {loan_repayment_amount:.2f}",
            "transport": f"R {transport_amount:.2f}", 
            "uif_emp": f"R {uif:.2f}", 
            "uif_er": f"R {uif:.2f}", 
            "paye": f"R {paye:.2f}", 
            "net": f"R {net:.2f}", 
            "leave": leave, 
            "sick_leave": sick, 
            "family_leave": family,
            "company_name": company['name'] if company else '',
            "company_logo": dict(company).get('logo_file', '') if company else '',
            "company_address": dict(company).get('address', '') if company else '',
            "company_reg": dict(company).get('registration_number', '') if company else '',
            "workday_hours": workday_hours,
            "inactive_date": emp['inactive_date'] or '',
            "payroll_cutoff": payroll_cutoff.strftime('%Y-%m-%d'),
            "bcea_warning": bcea_warning,
            "overtime_pay_treatment": "Regular recurring" if overtime_pay_treatment == "regular" else "Irregular / once-off",
            "payslip_status": "Draft Payslip",
            "is_finalized": False,
            "source": "Draft calculation - not saved to ledger"
        },
        "raw": {
            "employee_id": emp_id,
            "date": date_str,
            "gross": gross,
            "overtime": overtime_amount,
            "transport": transport_amount,
            "bonus": bonus_amount,
            "reimbursable_expenses": reimbursable_amount,
            "loan_repayment": loan_repayment_amount,
            "uif": uif,
            "paye": paye,
            "net": net
        }
    })

def build_saved_payslip_payload(conn, emp, company, ledger_row):
    """Return a payslip response payload using only the saved payslip ledger row."""
    date_str = ledger_row['date'] or ''
    target_month = date_str[:7] if date_str else ''
    gross = float(ledger_row['gross_salary'] or 0)
    overtime_amount = float(ledger_row['overtime'] or 0)
    transport_amount = float(ledger_row['transport'] or 0)
    bonus_amount = float(ledger_row['bonus'] or 0)
    reimbursable_amount = float(ledger_row['reimbursable_expenses'] or 0)
    loan_repayment_amount = float(ledger_row['loan_repayment'] or 0)
    uif = float(ledger_row['uif'] or 0)
    paye = float(ledger_row['paye'] or 0)
    net = float(ledger_row['net_salary'] or 0)
    benefit_ref_date = date_str or datetime.now().strftime('%Y-%m-%d')
    try:
        leave = calculate_leave_balance(emp['id'], emp['start_date'], emp['emp_type'], emp['name'], benefit_ref_date)
        sick = calculate_sick_leave_balance(emp['id'], emp['start_date'], emp['emp_type'], emp['name'], benefit_ref_date)
        family = calculate_family_leave_balance(emp['id'], emp['start_date'], emp['emp_type'], emp['name'], benefit_ref_date)
    except Exception:
        leave, sick, family = 'N/A', 'N/A', 'N/A'
    return {
        "payslip": {
            "date": target_month,
            "name": emp['name'],
            "emp_num": emp['emp_number'] or 'N/A',
            "id_num": emp['id_passport'] or 'N/A',
            "gross": f"R {gross:.2f}",
            "days_worked_display": "(Saved to Ledger)",
            "overtime": f"R {overtime_amount:.2f}",
            "sundays_display": "",
            "bonus": f"R {bonus_amount:.2f}",
            "reimbursable_expenses": f"R {reimbursable_amount:.2f}",
            "loan_repayment": f"R {loan_repayment_amount:.2f}",
            "transport": f"R {transport_amount:.2f}",
            "uif_emp": f"R {uif:.2f}",
            "uif_er": f"R {uif:.2f}",
            "paye": f"R {paye:.2f}",
            "net": f"R {net:.2f}",
            "leave": leave,
            "sick_leave": sick,
            "family_leave": family,
            "company_name": company['name'] if company else '',
            "company_logo": dict(company).get('logo_file', '') if company else '',
            "company_address": dict(company).get('address', '') if company else '',
            "company_reg": dict(company).get('registration_number', '') if company else '',
            "payslip_status": "Final Payslip",
            "is_finalized": True,
            "source": "Saved Payslip Ledger",
            "regular_payslip_id": ledger_row['id']
        },
        "raw": {
            "employee_id": emp['id'],
            "date": date_str,
            "gross": gross,
            "overtime": overtime_amount,
            "transport": transport_amount,
            "bonus": bonus_amount,
            "reimbursable_expenses": reimbursable_amount,
            "loan_repayment": loan_repayment_amount,
            "uif": uif,
            "paye": paye,
            "net": net
        }
    }

@app.route('/api/payslip_for_distribution', methods=['POST'])
def payslip_for_distribution():
    if not session.get('can_payroll') and not session.get('is_superadmin'):
        return jsonify({"message": "Forbidden"}), 403
    data = request.get_json() or {}
    emp_id, date_str = data.get('employee_id'), data.get('date')
    if not emp_id or not date_str:
        return jsonify({"message": "Employee and payroll date are required."}), 400
    target_month = date_str[:7]
    conn = get_db_connection()
    emp = conn.execute('SELECT * FROM employees WHERE id=? AND company_id=?', (emp_id, session['company_id'])).fetchone()
    if not emp:
        conn.close()
        return jsonify({"message": "Employee not found."}), 404
    company = conn.execute('SELECT * FROM companies WHERE id=?', (session['company_id'],)).fetchone()
    ledger = conn.execute('''SELECT * FROM payslips
                             WHERE company_id=? AND employee_id=? AND date LIKE ?
                               AND COALESCE(payslip_type, 'regular')='regular'
                             ORDER BY id DESC LIMIT 1''',
                          (session['company_id'], emp_id, f"{target_month}%")).fetchone()
    if not ledger:
        conn.close()
        return jsonify({"status": "draft_required", "message": "No finalised ledger payslip found for this employee/month.", "employee_name": emp['name']})
    payload = build_saved_payslip_payload(conn, emp, company, ledger)
    conn.close()
    payload['status'] = 'finalized'
    return jsonify(payload)


@app.route('/api/save_adjustment_payslip', methods=['POST'])
def save_adjustment_payslip():
    if not session.get('can_payroll') and not session.get('is_superadmin'):
        return jsonify({"message": "Forbidden"}), 403
    data = request.get_json() or {}
    emp_id = data.get('employee_id')
    date_str = data.get('date')
    reason = (data.get('adjustment_reason') or '').strip()
    if not emp_id or not date_str:
        return jsonify({"message": "Employee and payroll date are required."}), 400
    if not reason:
        return jsonify({"message": "Adjustment reason is required."}), 400
    target_month = date_str[:7]

    gross = safe_adjustment_money(data.get('gross'))
    overtime = safe_adjustment_money(data.get('overtime'))
    transport = safe_adjustment_money(data.get('transport'))
    bonus = safe_adjustment_money(data.get('bonus'))
    reimbursable = safe_adjustment_money(data.get('reimbursable_expenses'))
    loan_repayment = safe_adjustment_money(data.get('loan_repayment'))
    uif = safe_adjustment_money(data.get('uif'))
    paye = safe_adjustment_money(data.get('paye'))
    net = gross + overtime + bonus + transport + reimbursable - uif - paye - loan_repayment

    if all(abs(v) < 0.005 for v in [gross, overtime, transport, bonus, reimbursable, loan_repayment, uif, paye]):
        return jsonify({"message": "Enter at least one adjustment amount."}), 400

    conn = get_db_connection()
    emp = conn.execute('SELECT id, name FROM employees WHERE id=? AND company_id=?', (emp_id, session['company_id'])).fetchone()
    if not emp:
        conn.close()
        return jsonify({"message": "Employee not found."}), 404
    regular = conn.execute('''SELECT id FROM payslips
                              WHERE company_id=? AND employee_id=? AND date LIKE ?
                                AND COALESCE(payslip_type, 'regular')='regular'
                              ORDER BY id DESC LIMIT 1''',
                           (session['company_id'], emp_id, f"{target_month}%")).fetchone()
    if not regular:
        conn.close()
        return jsonify({"message": "A finalised regular payslip must exist before an adjustment payslip can be created."}), 400

    conn.execute('''INSERT INTO payslips
                    (company_id, employee_id, date, gross_salary, overtime, transport, bonus, reimbursable_expenses, loan_repayment, uif, paye, net_salary, payslip_type, adjustment_of_payslip_id, adjustment_reason, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'adjustment', ?, ?, ?)''',
                 (session['company_id'], emp_id, date_str, gross, overtime, transport, bonus, reimbursable, loan_repayment, uif, paye, net, regular['id'], reason, datetime.now().isoformat(timespec='seconds')))
    conn.commit()
    conn.close()
    log_action('HR & Payroll', 'Created Adjustment Payslip', f"Created adjustment payslip for Employee ID {emp_id} for {target_month}: {reason}")
    return jsonify({"status": "success", "net": round(net, 2)})


UI19_SETTING_KEYS = [
    'ui19_uif_ref', 'ui19_branch_no', 'ui19_paye_ref', 'ui19_trading_name',
    'ui19_physical_address', 'ui19_work_address', 'ui19_postal_address',
    'ui19_email', 'ui19_fax', 'ui19_phone', 'ui19_authorised_person', 'ui19_authorised_person_id'
]


def get_company_settings_dict(conn, company_id):
    rows = conn.execute('SELECT key, value FROM settings WHERE company_id=?', (company_id,)).fetchall()
    return {r['key']: r['value'] for r in rows}


def upsert_company_setting(conn, company_id, key, value):
    exists = conn.execute('SELECT 1 FROM settings WHERE company_id=? AND key=?', (company_id, key)).fetchone()
    if exists:
        conn.execute('UPDATE settings SET value=? WHERE company_id=? AND key=?', (value, company_id, key))
    else:
        conn.execute('INSERT INTO settings (company_id, key, value) VALUES (?, ?, ?)', (company_id, key, value))


def month_bounds_from_yyyy_mm(month_str):
    try:
        start = datetime.strptime(month_str, '%Y-%m')
    except Exception:
        raise ValueError('Month must be in YYYY-MM format.')
    last_day = calendar.monthrange(start.year, start.month)[1]
    end = start.replace(day=last_day)
    return start, end


def ui19_date_digits(date_str):
    if not date_str:
        return ''
    try:
        dt = datetime.strptime(date_str[:10], '%Y-%m-%d')
        return dt.strftime('%d%m%y')
    except Exception:
        return ''


def split_employee_name_for_ui19(full_name):
    parts = [p for p in (full_name or '').replace(',', ' ').split() if p]
    if not parts:
        return '', ''
    surname = parts[-1]
    initials = ''.join(p[0].upper() for p in parts[:-1]) or surname[:1].upper()
    return surname, initials


def default_termination_code(emp):
    explicit = (dict(emp).get('uif_termination_code') or '').strip()
    if explicit:
        return explicit
    status = (dict(emp).get('status') or '').strip().lower()
    if status == 'resigned':
        return '6'
    if status == 'terminated':
        return '4'
    return ''


def parse_ui19_employee_ids(raw_ids):
    if raw_ids is None:
        return None
    if isinstance(raw_ids, str):
        raw_parts = raw_ids.replace(';', ',').split(',')
    elif isinstance(raw_ids, (list, tuple, set)):
        raw_parts = list(raw_ids)
    else:
        raw_parts = [raw_ids]
    parsed = []
    for item in raw_parts:
        try:
            val = int(str(item).strip())
        except Exception:
            continue
        if val > 0 and val not in parsed:
            parsed.append(val)
    return parsed


def build_ui19_data(conn, company_id, month_str, employee_ids=None):
    selected_employee_ids = parse_ui19_employee_ids(employee_ids)
    month_start, month_end = month_bounds_from_yyyy_mm(month_str)
    company = conn.execute('SELECT * FROM companies WHERE id=?', (company_id,)).fetchone()
    settings = get_company_settings_dict(conn, company_id)
    company_dict = dict(company) if company else {}
    month_label = month_start.strftime('%B %Y')

    employer = {
        'month': month_label,
        'uif_ref': settings.get('ui19_uif_ref', ''),
        'branch_no': settings.get('ui19_branch_no', ''),
        'paye_ref': settings.get('ui19_paye_ref', ''),
        'trading_name': settings.get('ui19_trading_name') or company_dict.get('name', ''),
        'physical_address': settings.get('ui19_physical_address') or company_dict.get('address', ''),
        'work_address': settings.get('ui19_work_address') or settings.get('ui19_physical_address') or company_dict.get('address', ''),
        'postal_address': settings.get('ui19_postal_address') or settings.get('ui19_physical_address') or company_dict.get('address', ''),
        'company_reg': company_dict.get('registration_number', ''),
        'email': settings.get('ui19_email') or settings.get('sender_email') or settings.get('smtp_user', ''),
        'fax': settings.get('ui19_fax', ''),
        'phone': settings.get('ui19_phone', ''),
        'authorised_person': settings.get('ui19_authorised_person') or session.get('username', ''),
        'authorised_person_id': settings.get('ui19_authorised_person_id', '')
    }

    if selected_employee_ids is not None:
        if not selected_employee_ids:
            employees = []
        else:
            placeholders = ','.join(['?'] * len(selected_employee_ids))
            employees = conn.execute(f"""SELECT * FROM employees
                                      WHERE company_id=?
                                        AND id IN ({placeholders})
                                        AND COALESCE(emp_type, '') NOT IN ('Supplier', 'Provider')
                                      ORDER BY name ASC""", [company_id] + selected_employee_ids).fetchall()
    else:
        employees = conn.execute("""SELECT * FROM employees
                                    WHERE company_id=?
                                      AND COALESCE(emp_type, '') NOT IN ('Supplier', 'Provider')
                                    ORDER BY name ASC""", (company_id,)).fetchall()
    rows = []
    for emp in employees:
        emp_dict = dict(emp)
        start_dt = parse_date_safe(emp_dict.get('start_date'))
        inactive_dt = parse_date_safe(emp_dict.get('inactive_date'))
        ledger = conn.execute("""SELECT SUM(COALESCE(gross_salary,0) + COALESCE(overtime,0) + COALESCE(bonus,0)) AS gross,
                                        COUNT(*) AS payslip_count
                                 FROM payslips
                                 WHERE company_id=? AND employee_id=? AND date LIKE ?""",
                              (company_id, emp['id'], f'{month_str}%')).fetchone()
        payslip_count = int(ledger['payslip_count'] or 0)
        contributor = (emp_dict.get('uif_contributor') or 'Yes').strip()
        if contributor.lower() not in ['yes', 'no']:
            contributor = 'Yes'
        non_contributor_reason = (emp_dict.get('uif_non_contributor_reason') or '').strip()

        # UI-19 inclusion rule:
        # Include only employees with UIF activity for the selected month:
        # 1) finalised/saved payslip in the selected month,
        # 2) commencement date in the selected month,
        # 3) inactive/last working day in the selected month,
        # 4) explicit non-contributor status and reason.
        started_in_month = bool(start_dt and month_start.date() <= start_dt.date() <= month_end.date())
        terminated_in_month = bool(inactive_dt and month_start.date() <= inactive_dt.date() <= month_end.date())
        has_finalised_payroll = payslip_count > 0
        has_explicit_non_contribution = contributor.lower() == 'no' and bool(non_contributor_reason)
        if not (has_finalised_payroll or started_in_month or terminated_in_month or has_explicit_non_contribution):
            continue

        try:
            hours_info, _warning = analyse_employee_hours_for_period(conn, company_id, emp, month_start, month_end)
            total_hours = float(hours_info.get('total_hours') or 0)
        except Exception:
            total_hours = 0.0
        gross = float(ledger['gross'] or 0)
        termination_date = ''
        termination_code = ''
        if inactive_dt and month_start.date() <= inactive_dt.date() <= month_end.date():
            termination_date = ui19_date_digits(emp_dict.get('inactive_date'))
            termination_code = default_termination_code(emp)

        surname, initials = split_employee_name_for_ui19(emp_dict.get('name'))
        rows.append({
            'surname': surname,
            'initials': initials,
            'id_number': emp_dict.get('id_passport') or '',
            'gross': round(gross, 2),
            'hours': round(total_hours, 2),
            'commencement': ui19_date_digits(emp_dict.get('start_date')),
            'termination': termination_date,
            'termination_code': termination_code,
            'contributor': 'No' if contributor.lower() == 'no' else 'Yes',
            'non_contributor_reason': non_contributor_reason if contributor.lower() == 'no' else ''
        })
    return {'month': month_str, 'month_label': month_label, 'employer': employer, 'employees': rows}


@app.route('/api/ui19_settings', methods=['GET', 'POST'])
def ui19_settings():
    if not session.get('can_payroll') and not session.get('is_superadmin'):
        return jsonify({'message': 'Forbidden'}), 403
    conn = get_db_connection()
    cid = session['company_id']
    company = conn.execute('SELECT * FROM companies WHERE id=?', (cid,)).fetchone()
    settings = get_company_settings_dict(conn, cid)
    if request.method == 'POST':
        data = request.get_json() or {}
        for key in UI19_SETTING_KEYS:
            upsert_company_setting(conn, cid, key, (data.get(key) or '').strip())
        conn.commit()
        settings = get_company_settings_dict(conn, cid)
        log_action('HR & Payroll', 'Updated UI-19 Settings', 'Updated UI-19 employer declaration settings')
    comp = dict(company) if company else {}
    payload = {
        'ui19_uif_ref': settings.get('ui19_uif_ref', ''),
        'ui19_branch_no': settings.get('ui19_branch_no', ''),
        'ui19_paye_ref': settings.get('ui19_paye_ref', ''),
        'ui19_trading_name': settings.get('ui19_trading_name') or comp.get('name', ''),
        'ui19_physical_address': settings.get('ui19_physical_address') or comp.get('address', ''),
        'ui19_work_address': settings.get('ui19_work_address') or settings.get('ui19_physical_address') or comp.get('address', ''),
        'ui19_postal_address': settings.get('ui19_postal_address') or settings.get('ui19_physical_address') or comp.get('address', ''),
        'ui19_email': settings.get('ui19_email') or settings.get('sender_email') or settings.get('smtp_user', ''),
        'ui19_fax': settings.get('ui19_fax', ''),
        'ui19_phone': settings.get('ui19_phone', ''),
        'ui19_authorised_person': settings.get('ui19_authorised_person') or session.get('username', ''),
        'ui19_authorised_person_id': settings.get('ui19_authorised_person_id', ''),
        'company_registration_number': comp.get('registration_number', '')
    }
    conn.close()
    return jsonify(payload)


@app.route('/api/ui19_data', methods=['POST'])
def ui19_data():
    if not session.get('can_payroll') and not session.get('is_superadmin'):
        return jsonify({'message': 'Forbidden'}), 403
    data = request.get_json() or {}
    month_str = (data.get('month') or '').strip()
    employee_ids = parse_ui19_employee_ids(data.get('employee_ids'))
    if not month_str:
        return jsonify({'message': 'Month is required.'}), 400
    if employee_ids is not None and not employee_ids:
        return jsonify({'message': 'Select at least one employee for the UI-19 declaration.'}), 400
    conn = get_db_connection()
    try:
        payload = build_ui19_data(conn, session['company_id'], month_str, employee_ids)
    except ValueError as exc:
        conn.close()
        return jsonify({'message': str(exc)}), 400
    conn.close()
    return jsonify(payload)



def _pdf_escape(value):
    text = str(value or '')
    text = text.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')
    return text.encode('latin-1', 'replace').decode('latin-1')


class _SimplePdf:
    """Tiny PDF writer used for UI-19 output without third-party dependencies."""
    def __init__(self, width=841.89, height=595.28):
        self.width = float(width)
        self.height = float(height)
        self.pages = []
        self.current = []
        self.font_alias = {'Helvetica': 'F1', 'Helvetica-Bold': 'F2'}
        self.current_font = ('Helvetica', 8)

    def new_page(self):
        if self.current:
            self.pages.append('\n'.join(self.current))
        self.current = []
        self.set_font('Helvetica', 8)
        self.set_line_width(0.6)

    def set_font(self, font='Helvetica', size=8):
        if font not in self.font_alias:
            font = 'Helvetica'
        self.current_font = (font, float(size))
        self.current.append(f"/{self.font_alias[font]} {float(size):.2f} Tf")

    def set_line_width(self, width):
        self.current.append(f"{float(width):.2f} w")

    def rect(self, x, y, w, h):
        self.current.append(f"{float(x):.2f} {float(y):.2f} {float(w):.2f} {float(h):.2f} re S")

    def line(self, x1, y1, x2, y2):
        self.current.append(f"{float(x1):.2f} {float(y1):.2f} m {float(x2):.2f} {float(y2):.2f} l S")

    def text_width(self, text, font='Helvetica', size=8):
        text = str(text or '')
        # Conservative Helvetica approximation. Good enough for wrapping in form cells.
        total = 0.0
        for ch in text:
            if ch in 'il.,:;|! ':
                total += 0.25
            elif ch in 'MW@#%&':
                total += 0.75
            elif ch.isupper() or ch.isdigit():
                total += 0.58
            else:
                total += 0.50
        return total * float(size)

    def draw_string(self, x, y, text, font=None, size=None):
        if font or size:
            self.set_font(font or self.current_font[0], size or self.current_font[1])
        self.current.append(f"BT 1 0 0 1 {float(x):.2f} {float(y):.2f} Tm ({_pdf_escape(text)}) Tj ET")

    def draw_centered(self, x, y, text, font=None, size=None):
        font = font or self.current_font[0]
        size = float(size or self.current_font[1])
        self.draw_string(float(x) - self.text_width(text, font, size) / 2, y, text, font, size)

    def draw_right(self, x, y, text, font=None, size=None):
        font = font or self.current_font[0]
        size = float(size or self.current_font[1])
        self.draw_string(float(x) - self.text_width(text, font, size), y, text, font, size)

    def wrap_text(self, text, max_width, font='Helvetica', size=7, max_lines=2):
        text = str(text or '').replace('\r', ' ').replace('\n', ' ')
        words = text.split()
        if not words:
            return []
        lines = []
        current = ''
        for word in words:
            trial = (current + ' ' + word).strip()
            if not current or self.text_width(trial, font, size) <= max_width:
                current = trial
            else:
                lines.append(current)
                current = word
                if len(lines) >= max_lines:
                    break
        if current and len(lines) < max_lines:
            lines.append(current)
        return lines[:max_lines]

    def draw_wrapped(self, text, x, y, max_width, font='Helvetica', size=7, leading=8, max_lines=2):
        for idx, line in enumerate(self.wrap_text(text, max_width, font, size, max_lines)):
            self.draw_string(x, y - idx * leading, line, font, size)

    def draw_box_text(self, label, value, x, y, w, h, label_w=90, font_size=6.5):
        self.rect(x, y, w, h)
        self.draw_string(x + 3, y + h - font_size - 3, label, 'Helvetica-Bold', font_size)
        self.draw_wrapped(value, x + label_w, y + h - font_size - 3, max(10, w - label_w - 6), 'Helvetica', font_size, font_size + 1, 2)

    def finish(self):
        if self.current:
            self.pages.append('\n'.join(self.current))
            self.current = []
        objects = []
        catalog_id = 1
        pages_id = 2
        font1_id = 3
        font2_id = 4
        next_id = 5
        page_ids = []
        content_ids = []
        for content in self.pages:
            page_ids.append(next_id); next_id += 1
            content_ids.append(next_id); next_id += 1
        objects.append((catalog_id, f"<< /Type /Catalog /Pages {pages_id} 0 R >>".encode('latin-1')))
        kids = ' '.join(f'{pid} 0 R' for pid in page_ids)
        objects.append((pages_id, f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>".encode('latin-1')))
        objects.append((font1_id, b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"))
        objects.append((font2_id, b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>"))
        for pid, cid, content in zip(page_ids, content_ids, self.pages):
            page_obj = (f"<< /Type /Page /Parent {pages_id} 0 R /MediaBox [0 0 {self.width:.2f} {self.height:.2f}] "
                        f"/Resources << /Font << /F1 {font1_id} 0 R /F2 {font2_id} 0 R >> >> /Contents {cid} 0 R >>")
            stream = content.encode('latin-1', 'replace')
            content_obj = b"<< /Length " + str(len(stream)).encode('ascii') + b" >>\nstream\n" + stream + b"\nendstream"
            objects.append((pid, page_obj.encode('latin-1')))
            objects.append((cid, content_obj))
        out = io.BytesIO()
        out.write(b'%PDF-1.4\n%\xe2\xe3\xcf\xd3\n')
        offsets = [0]
        for obj_id, body in objects:
            offsets.append(out.tell())
            out.write(f"{obj_id} 0 obj\n".encode('ascii'))
            out.write(body)
            out.write(b"\nendobj\n")
        xref_pos = out.tell()
        max_id = max(obj_id for obj_id, _ in objects)
        obj_by_id = {obj_id: offsets[i + 1] for i, (obj_id, _) in enumerate(objects)}
        out.write(f"xref\n0 {max_id + 1}\n".encode('ascii'))
        out.write(b"0000000000 65535 f \n")
        for obj_id in range(1, max_id + 1):
            out.write(f"{obj_by_id.get(obj_id, 0):010d} 00000 n \n".encode('ascii'))
        out.write(f"trailer\n<< /Size {max_id + 1} /Root {catalog_id} 0 R >>\nstartxref\n{xref_pos}\n%%EOF".encode('ascii'))
        out.seek(0)
        return out


def _fmt_amount(value):
    try:
        return f"{float(value or 0):.2f}"
    except Exception:
        return "0.00"


def export_ui19_pdf(payload):
    pdf = _SimplePdf()
    page_w, page_h = pdf.width, pdf.height
    margin = 18
    emp_rows = payload.get('employees') or []
    rows_per_page = 8
    chunks = [emp_rows[i:i + rows_per_page] for i in range(0, len(emp_rows), rows_per_page)] or [[]]

    def draw_footer(page_idx, total_pages):
        pdf.draw_right(page_w - margin, 9, f"Generated by Easy Admin | Page {page_idx + 1} of {total_pages}", 'Helvetica', 6)

    for page_idx, chunk in enumerate(chunks):
        pdf.new_page()
        y = page_h - 20
        pdf.draw_centered(page_w / 2, y, 'UNEMPLOYMENT INSURANCE ACT 63 OF 2001 AS AMENDED', 'Helvetica-Bold', 11)
        pdf.draw_right(page_w - margin, y, 'UI-19', 'Helvetica-Bold', 8)
        y -= 14
        pdf.draw_centered(page_w / 2, y, 'Employers Declaration of Employees for the month of', 'Helvetica-Bold', 9)
        pdf.rect(page_w / 2 + 140, y - 5, 126, 17)
        pdf.draw_string(page_w / 2 + 148, y, payload.get('month_label', ''), 'Helvetica', 10)
        y -= 12
        pdf.draw_centered(page_w / 2, y, 'Information to be supplied in terms of Section 56(1&3) read with Regulation 13(1&2)', 'Helvetica-Bold', 6)
        y -= 8
        pdf.draw_centered(page_w / 2, y, 'An employer must declare employee remuneration details, new appointments and termination of service for the previous month.', 'Helvetica', 5.5)
        y -= 13
        pdf.draw_string(margin, y, '1. EMPLOYER DETAILS', 'Helvetica-Bold', 8)
        y -= 7

        employer = payload.get('employer') or {}
        left_x = margin
        gap = 6
        box_w = (page_w - 2 * margin - gap) / 2
        right_x = left_x + box_w + gap
        box_h = 15
        pdf.draw_box_text('1.1 UIF Employer Ref / Branch No', f"{employer.get('uif_ref','')} / {employer.get('branch_no','')}", left_x, y - box_h, box_w, box_h, 113, 6.3)
        pdf.draw_box_text('1.2 PAYE Reference No', employer.get('paye_ref',''), right_x, y - box_h, box_w, box_h, 90, 6.3)
        y -= box_h
        pdf.draw_box_text('1.3 Trading name of business', employer.get('trading_name',''), left_x, y - box_h, box_w, box_h, 100, 6.3)
        pdf.draw_box_text('1.4 Physical Address', employer.get('physical_address',''), right_x, y - box_h, box_w, box_h, 85, 6.3)
        y -= box_h
        pdf.draw_box_text('1.5 Work Address', employer.get('work_address',''), left_x, y - box_h, box_w, box_h, 80, 6.3)
        pdf.draw_box_text('1.6 Postal Address', employer.get('postal_address',''), right_x, y - box_h, box_w, box_h, 82, 6.3)
        y -= box_h
        pdf.draw_box_text('1.7 Co. Reg.No', employer.get('company_reg',''), left_x, y - box_h, box_w, box_h, 78, 6.3)
        pdf.draw_box_text('1.8 Email / 1.9 Fax / 1.10 Phone', f"{employer.get('email','')} | {employer.get('fax','')} | {employer.get('phone','')}", right_x, y - box_h, box_w, box_h, 128, 5.8)
        y -= box_h
        pdf.draw_box_text('1.11 Authorised person', employer.get('authorised_person',''), left_x, y - box_h, box_w, box_h, 95, 6.3)
        pdf.draw_box_text('Authorised person ID', employer.get('authorised_person_id',''), right_x, y - box_h, box_w, box_h, 90, 6.3)
        y -= box_h + 9

        pdf.draw_string(margin, y, '2. EMPLOYEE DETAILS', 'Helvetica-Bold', 8)
        y -= 5
        table_x = margin
        table_top = y
        col_widths = [120, 40, 185, 75, 60, 78, 78, 55, 58, 57]
        headers = [
            ['A', 'Surname'], ['B', 'Initials'], ['C', 'Identity Document Number'],
            ['D*', 'Total Gross', 'Remuneration paid', 'per Month', 'R c'],
            ['E*', 'Total Hours', 'Worked'], ['F', 'Commencement', 'Date', 'DDMMYY'],
            ['G', 'Termination Date', 'DDMMYY'], ['H', 'Reason Code'],
            ['I', 'Contributor', 'YES/NO'], ['J***', 'Reason']
        ]
        header_h = 45
        row_h = 17
        x = table_x
        for w, lines in zip(col_widths, headers):
            pdf.rect(x, table_top - header_h, w, header_h)
            for li, line in enumerate(lines):
                pdf.draw_centered(x + w / 2, table_top - 9 - li * 7, line, 'Helvetica-Bold', 5.4)
            x += w
        y = table_top - header_h
        for r_i in range(rows_per_page):
            row = chunk[r_i] if r_i < len(chunk) else {}
            values = [
                row.get('surname',''), row.get('initials',''), row.get('id_number',''),
                _fmt_amount(row.get('gross')) if row else '', _fmt_amount(row.get('hours')) if row else '',
                row.get('commencement',''), row.get('termination',''), row.get('termination_code',''),
                row.get('contributor',''), row.get('non_contributor_reason','')
            ]
            x = table_x
            for i, (w, val) in enumerate(zip(col_widths, values)):
                pdf.rect(x, y - row_h, w, row_h)
                if i in [3, 4]:
                    pdf.draw_right(x + w - 4, y - 11, val, 'Helvetica', 7)
                elif i == 2:
                    pdf.draw_string(x + 3, y - 11, str(val)[:28], 'Helvetica', 6.6)
                else:
                    pdf.draw_string(x + 3, y - 11, str(val)[:24], 'Helvetica', 7)
                x += w
            y -= row_h

        y -= 8
        if page_idx == len(chunks) - 1:
            employer_name = employer.get('authorised_person','') or employer.get('trading_name','')
            employer_id = employer.get('authorised_person_id','')
            pdf.draw_string(margin, y, 'I,', 'Helvetica', 7)
            pdf.line(margin + 12, y - 1, margin + 176, y - 1)
            pdf.draw_string(margin + 18, y + 1, employer_name, 'Helvetica', 7)
            pdf.draw_string(margin + 182, y, '(Name of Employer), ID No', 'Helvetica', 6.3)
            pdf.line(margin + 266, y - 1, margin + 410, y - 1)
            pdf.draw_string(margin + 272, y + 1, employer_id, 'Helvetica', 7)
            pdf.draw_string(margin + 416, y, 'declare that the above information is true and correct. I understand that it is an offence to make a false statement.', 'Helvetica', 5.8)
            y -= 21
            pdf.draw_string(margin, y, 'EMPLOYER SIGNATURE', 'Helvetica-Bold', 8)
            pdf.line(margin + 100, y - 1, margin + 330, y - 1)
            pdf.draw_string(page_w / 2 + 100, y, 'DATE', 'Helvetica-Bold', 8)
            pdf.line(page_w / 2 + 132, y - 1, page_w - margin, y - 1)
            y -= 17
            left_code_w = 300
            mid_code_w = 300
            stamp_w = page_w - 2 * margin - left_code_w - mid_code_w
            pdf.rect(margin, y - 52, left_code_w, 52)
            pdf.rect(margin + left_code_w, y - 52, mid_code_w, 52)
            pdf.rect(margin + left_code_w + mid_code_w, y - 52, stamp_w, 52)
            pdf.draw_centered(margin + left_code_w / 2, y - 8, 'DESCRIPTIONS', 'Helvetica-Bold', 6)
            pdf.draw_wrapped('D* Remuneration means actual gross salary paid. If paid weekly, convert weekly wages X 52/12. E* Total Hours Worked during the month.', margin + 4, y - 18, left_code_w - 8, 'Helvetica', 5.4, 7, 4)
            pdf.draw_string(margin + left_code_w + 4, y - 8, 'Reason for Non-Contribution', 'Helvetica-Bold', 6)
            pdf.draw_string(margin + left_code_w + 4, y - 18, '1 Temporary employees less than 24 hours per month', 'Helvetica', 5.4)
            pdf.draw_string(margin + left_code_w + 4, y - 27, '2 Commission-only employees', 'Helvetica', 5.4)
            pdf.draw_string(margin + left_code_w + 4, y - 36, '3 No income paid for the payroll period', 'Helvetica', 5.4)
            pdf.draw_string(margin + left_code_w + mid_code_w + 5, y - 10, 'Employer Stamp', 'Helvetica-Bold', 6)
            pdf.draw_string(margin + left_code_w + mid_code_w + 5, y - 19, '(if available)', 'Helvetica', 5.4)
            y -= 62
            pdf.draw_centered(page_w / 2, y, 'REASON FOR TERMINATION CODES', 'Helvetica-Bold', 6)
            y -= 9
            codes = '2 Deceased   3 Retired   4 Dismissed   5 Contract Expired   6 Resigned   7 Constructive Dismissal   8 Insolvency/Liquidation   9 Maternity/Adoption   10 Illness/Medically boarded   11 Retrenched/Staff Reduction   12 Transfer to another Branch   13 Absconded   14 Business Closed   15 Death of Domestic Employer   16 Voluntary Severance Package   17 Reduced Work Time   18 Commissioning Parental   19 Parental Leave'
            pdf.draw_wrapped(codes, margin, y, page_w - 2 * margin, 'Helvetica', 5.2, 7, 3)
        else:
            pdf.draw_string(margin, y, f'Continued on next page... ({page_idx + 1}/{len(chunks)})', 'Helvetica-Bold', 8)
        draw_footer(page_idx, len(chunks))

    return pdf.finish()


@app.route('/export_ui19')
def export_ui19():
    if not session.get('can_payroll') and not session.get('is_superadmin'):
        return 'Forbidden', 403
    month_str = (request.args.get('month') or '').strip()
    employee_ids = parse_ui19_employee_ids(request.args.get('employee_ids'))
    if not month_str:
        return jsonify({'message': 'Month is required.'}), 400
    if not employee_ids:
        return jsonify({'message': 'Select at least one employee for the UI-19 declaration.'}), 400
    conn = get_db_connection()
    try:
        payload = build_ui19_data(conn, session['company_id'], month_str, employee_ids)
        pdf_bytes = export_ui19_pdf(payload)
    except ValueError as exc:
        conn.close()
        return jsonify({'message': str(exc)}), 400
    conn.close()
    log_action('HR & Payroll', 'Generated UI-19 Declaration', f'Generated UI-19 declaration for {month_str} using {len(employee_ids)} selected employee(s)')
    filename = f"UI-19_{session.get('company_name', 'Company').replace(' ', '_')}_{month_str}.pdf"
    return send_file(pdf_bytes, mimetype='application/pdf', as_attachment=True, download_name=filename)


@app.route('/generate_emp201', methods=['POST'])
def generate_emp201():
    if not session.get('can_payroll') and not session.get('is_superadmin'):
        return jsonify({"message": "Forbidden"}), 403

    data = request.get_json()
    month_str = data.get('month')
    if not month_str: return jsonify({"message": "Month is required"}), 400

    conn = get_db_connection()
    cid = session['company_id']
    # EMP201 is based strictly on payslips saved to the payroll ledger.
    # Unsaved payslip previews/calculations are not included.
    ledger = conn.execute('''SELECT SUM(gross_salary + overtime + COALESCE(bonus,0)) as total_taxable,
                                    SUM(paye) as total_paye,
                                    SUM(uif) as total_uif_emp,
                                    COUNT(*) as payslip_count
                             FROM payslips
                             WHERE company_id=? AND date LIKE ?''', (cid, f"{month_str}%")).fetchone()

    total_taxable = float(ledger['total_taxable'] or 0)
    total_paye = float(ledger['total_paye'] or 0)
    uif_emp = float(ledger['total_uif_emp'] or 0)
    total_uif = uif_emp * 2.0
    payslip_count = int(ledger['payslip_count'] or 0)

    conn.close()
    log_action('HR & Payroll', 'Generated Document', f"Generated Internal Use EMP201 report for {month_str}")

    return jsonify({
        "internal_use": "Internal Use",
        "month": month_str,
        "gross": round(total_taxable, 2),
        "paye": round(total_paye, 2),
        "uif": round(total_uif, 2),
        "total": round(total_paye + total_uif, 2),
        "source": "Saved Payslip Ledger Only",
        "payslip_count": payslip_count
    })

@app.route('/generate_irp5', methods=['POST'])
def generate_irp5():
    data = request.get_json()
    emp_id = data.get('employee_id')
    tax_year = int(data.get('tax_year'))
    
    start_date = f"{tax_year-1}-03-01"
    end_date = f"{tax_year}-02-28" 
    
    conn = get_db_connection()
    emp = conn.execute('SELECT * FROM employees WHERE id=? AND company_id=?', (emp_id, session['company_id'])).fetchone()
    if not emp:
        conn.close()
        return jsonify({"message": "Employee not found."}), 404
    
    # IRP5 / IT3(a) internal-use totals are based strictly on saved payslip ledger records.
    # Unsaved payslip previews/calculations are not included.
    ledger = conn.execute('''SELECT SUM(gross_salary) as gross, SUM(overtime) as ot, SUM(transport) as travel, SUM(COALESCE(bonus,0)) as bonus, SUM(COALESCE(reimbursable_expenses,0)) as reimbursable_expenses, 
                                    SUM(paye) as paye, SUM(uif) as uif, COUNT(*) as payslip_count 
                             FROM payslips 
                             WHERE company_id=? AND employee_id=? AND date >= ? AND date <= ?''', 
                          (session['company_id'], emp_id, start_date, end_date)).fetchone()
    conn.close()
    
    gross = float(ledger['gross'] or 0)
    ot = float(ledger['ot'] or 0)
    travel = float(ledger['travel'] or 0)
    bonus = float(ledger['bonus'] or 0)
    reimbursable_expenses = float(ledger['reimbursable_expenses'] or 0)
    paye = float(ledger['paye'] or 0)
    uif = float(ledger['uif'] or 0)
    taxable_regular_income = gross + bonus
    total_gross = taxable_regular_income + ot + travel
    payslip_count = int(ledger['payslip_count'] or 0)
    
    log_action('HR & Payroll', 'Generated Document', f"Generated Ledger-backed IRP5 for {emp['name']} (Year: {tax_year})")
    
    return jsonify({"irp5": { "internal_use": "Internal Use", "tax_year": tax_year, "period": f"01 Mar {tax_year-1} - 28 Feb {tax_year}", "name": emp['name'], "emp_num": emp['emp_number'] or 'N/A', "id_num": emp['id_passport'] or 'N/A', "tax_number": emp['tax_number'] or 'N/A', "code_3601": f"{taxable_regular_income:.2f}", "code_3605": f"{ot:.2f}", "code_3702": f"{travel:.2f}", "code_3699": f"{total_gross:.2f}", "code_4102": f"{paye:.2f}", "code_4141": f"{uif:.2f}", "source": "Saved Payslip Ledger Only", "payslip_count": payslip_count }})

@app.route('/export_emp501/<year>')
def export_emp501(year):
    tax_year = int(year)
    start_date = f"{tax_year-1}-03-01"
    end_date = f"{tax_year}-02-28"
    
    conn = get_db_connection()
    employees = conn.execute("SELECT * FROM employees WHERE company_id=?", (session['company_id'],)).fetchall()
    
    si = io.StringIO()
    cw = csv.writer(si)
    cw.writerow(['Internal Use', 'Emp Number', 'Name', 'ID/Passport', 'Tax Number', 'PAYE Ref', 'Code 3601 (Income)', 'Code 3605 (Overtime)', 'Code 3702 (Travel)', 'Code 3699 (Gross)', 'Code 4102 (PAYE)', 'Code 4141 (UIF)'])
    
    for emp in employees:
        # EMP501 is based strictly on saved payslip ledger records.
        # Unsaved payslip previews/calculations are not included.
        ledger = conn.execute('''SELECT SUM(gross_salary) as gross, SUM(overtime) as ot, SUM(transport) as travel, SUM(COALESCE(bonus,0)) as bonus, SUM(COALESCE(reimbursable_expenses,0)) as reimbursable_expenses, 
                                        SUM(paye) as paye, SUM(uif) as uif, COUNT(*) as payslip_count 
                                 FROM payslips 
                                 WHERE company_id=? AND employee_id=? AND date >= ? AND date <= ?''', 
                              (session['company_id'], emp['id'], start_date, end_date)).fetchone()
                              
        gross = float(ledger['gross'] or 0)
        ot = float(ledger['ot'] or 0)
        travel = float(ledger['travel'] or 0)
        bonus = float(ledger['bonus'] or 0)
        paye = float(ledger['paye'] or 0)
        uif = float(ledger['uif'] or 0)
        payslip_count = int(ledger['payslip_count'] or 0)
        taxable_regular_income = gross + bonus
        total_gross = taxable_regular_income + ot + travel
        
        if payslip_count > 0 and total_gross > 0:
            cw.writerow(['Internal Use', emp['emp_number'], emp['name'], emp['id_passport'], emp['tax_number'], emp['paye_ref'], f"{taxable_regular_income:.2f}", f"{ot:.2f}", f"{travel:.2f}", f"{total_gross:.2f}", f"{paye:.2f}", f"{uif:.2f}"])
            
    conn.close()
    
    log_action('HR & Payroll', 'Exported Document', f"Exported Ledger-backed EMP501 summary for {year}")
    
    output = Response(si.getvalue(), mimetype='text/csv')
    output.headers["Content-Disposition"] = f"attachment; filename=EMP501_Summary_{year}.csv"
    return output



@app.route('/export_payroll_bank_file')
def export_payroll_bank_file():
    if not session.get('can_payroll') and not session.get('is_superadmin'):
        return "Forbidden", 403
    month_str = (request.args.get('month') or '').strip()
    template_key = (request.args.get('template') or 'generic_csv').strip()
    templates = get_bank_export_templates()
    if template_key not in templates:
        return jsonify({"message": "Unknown bank export template."}), 400
    if not month_str or len(month_str) != 7:
        return jsonify({"message": "Payroll month must be in YYYY-MM format."}), 400

    conn = get_db_connection()
    rows = payroll_payment_rows(conn, session['company_id'], month_str)
    conn.close()

    if not rows:
        return jsonify({"message": "No saved payslips found for the selected month. Save payslips to the ledger first."}), 400

    missing = validate_payroll_bank_rows(rows)
    if missing:
        return jsonify({"message": "Bank details missing or incomplete.", "missing": missing}), 400

    csv_data, label = build_payroll_bank_csv(rows, template_key, month_str, session.get('company_name', 'Company'))
    log_action('HR & Payroll', 'Exported Payroll Bank File', f"Exported {label} for {month_str}")
    output = Response(csv_data, mimetype='text/csv')
    safe_label = template_key.replace('/', '_')
    output.headers['Content-Disposition'] = f'attachment; filename=Payroll_Bank_Export_{safe_label}_{month_str}.csv'
    return output

@app.route('/generate_report', methods=['POST'])
def generate_report():
    data = request.get_json()
    emp_id, emp_name, s_date, e_date = data.get('employee_id'), data.get('employee_name'), data.get('start_date'), data.get('end_date')
    conn = get_db_connection()
    cid = session['company_id']
    bookings = conn.execute("SELECT start, title, booking_notes FROM bookings WHERE company_id=? AND employee LIKE ? AND substr(start, 1, 10) BETWEEN ? AND ? ORDER BY start ASC", (cid, f"%{emp_name}%", s_date, e_date)).fetchall()
    leave = conn.execute("SELECT * FROM leave_records WHERE company_id=? AND employee_id=? AND date_taken BETWEEN ? AND ? ORDER BY date_taken ASC", (cid, emp_id, s_date, e_date)).fetchall()
    emp = conn.execute("SELECT * FROM employees WHERE id=? AND company_id=?", (emp_id, cid)).fetchone()
    workday_hours = get_employee_workday_hours(emp) if emp else 7.0
    conn.close()
    
    dates_worked = []
    for b in bookings:
        work_date = b['start'][:10]
        try:
            weekday = datetime.strptime(work_date, '%Y-%m-%d').strftime('%A')
        except Exception:
            weekday = ''
        dates_worked.append({"date": work_date, "day": weekday, "client": b['title'], "notes": b['booking_notes'] or ""})
    leave_records = [{"id": l['id'], "date": l['date_taken'], "days": l['days'], "type": l['leave_type'], "doc": l['document_file']} for l in leave]
    return jsonify({"total_hours": round(len(dates_worked) * workday_hours, 2), "dates_worked": dates_worked, "total_leave": sum(l['days'] for l in leave_records), "leave_records": leave_records})

# ==========================================================
# 4. INVOICING & QUOTES ROUTES
# ==========================================================
@app.route('/invoicing')
def invoicing_index():
    conn = get_db_connection()
    cid = session['company_id']
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    clients = prepare_client_options(conn.execute("SELECT * FROM clients WHERE company_id=? ORDER BY name ASC, surname ASC, id ASC", (cid,)).fetchall())
    services = conn.execute("SELECT * FROM services WHERE company_id=? ORDER BY name ASC", (cid,)).fetchall()
    projects = conn.execute('''SELECT p.*, c.name AS client_name, c.surname AS client_surname, c.company_name AS client_company_name, c.id AS linked_client_id
                               FROM projects p
                               LEFT JOIN clients c ON c.id=p.client_id AND c.company_id=p.company_id
                               WHERE p.company_id=? AND COALESCE(p.status, '') <> 'Cancelled'
                               ORDER BY p.project_name ASC''', (cid,)).fetchall()
    invoice_page, invoice_per_page, invoice_offset = get_page_args(request.args, prefix='invoice_', default_per_page=50, max_per_page=100)
    quote_page, quote_per_page, quote_offset = get_page_args(request.args, prefix='quote_', default_per_page=50, max_per_page=100)
    invoice_q = (request.args.get('invoice_q') or '').strip()
    quote_q = (request.args.get('quote_q') or '').strip()
    inv_where = 'company_id=?'
    inv_params = [cid]
    if invoice_q:
        like = f'%{invoice_q}%'
        inv_where += " AND (COALESCE(client_name,'') LIKE ? OR COALESCE(status,'') LIKE ? OR COALESCE(date,'') LIKE ? OR COALESCE(due_date,'') LIKE ?)"
        inv_params.extend([like, like, like, like])
    invoice_total = conn.execute(f'SELECT COUNT(*) FROM invoices WHERE {inv_where}', inv_params).fetchone()[0]
    invoices = conn.execute(f'SELECT * FROM invoices WHERE {inv_where} ORDER BY id DESC LIMIT ? OFFSET ?', inv_params + [invoice_per_page, invoice_offset]).fetchall()
    invoice_pagination = pagination_meta(invoice_total, invoice_page, invoice_per_page)
    refresh_expired_quotes(conn, cid)
    conn.commit()
    quote_where = 'company_id=?'
    quote_params = [cid]
    if quote_q:
        like = f'%{quote_q}%'
        quote_where += " AND (COALESCE(client_name,'') LIKE ? OR COALESCE(status,'') LIKE ? OR COALESCE(date,'') LIKE ? OR COALESCE(valid_until,'') LIKE ?)"
        quote_params.extend([like, like, like, like])
    quote_total = conn.execute(f'SELECT COUNT(*) FROM quotes WHERE {quote_where}', quote_params).fetchone()[0]
    quotes = conn.execute(f'SELECT * FROM quotes WHERE {quote_where} ORDER BY id DESC LIMIT ? OFFSET ?', quote_params + [quote_per_page, quote_offset]).fetchall()
    quote_pagination = pagination_meta(quote_total, quote_page, quote_per_page)
    
    settings_rows = conn.execute("SELECT key, value FROM settings WHERE company_id=?", (cid,)).fetchall()
    s_dict = {s['key']: s['value'] for s in settings_rows}
    
    inv_info = s_dict.get('invoice_additional_info', '')
    inv_prefix = s_dict.get('invoice_prefix', 'INV-')
    inv_start = s_dict.get('invoice_start', '1')
    quote_prefix = s_dict.get('quote_prefix', 'QT-')
    quote_start = s_dict.get('quote_start', '1')

    # Assign calculated invoice numbers directly for the dashboard table
    # and calculate outstanding amount from actual recorded payments.
    formatted_invoices = []
    for inv in invoices:
        d = dict(inv)
        try:
            d['formatted_num'] = f"{inv_prefix}{int(inv_start) + d['id'] - 1:04d}"
        except:
            d['formatted_num'] = f"{inv_prefix}{d['id']:04d}"
        project = None
        if d.get('project_id'):
            project = conn.execute('SELECT project_name, project_code FROM projects WHERE id=? AND company_id=?', (d.get('project_id'), cid)).fetchone()
        d['project_name'] = project['project_name'] if project else ''
        d['project_code'] = project['project_code'] if project else ''
        payments = [dict(p) for p in conn.execute(
            "SELECT id, payment_date, amount, payment_method, reference FROM invoice_payments WHERE company_id=? AND invoice_id=? ORDER BY payment_date ASC, id ASC",
            (cid, d['id'])
        ).fetchall()]
        credit_notes = [dict(cn) for cn in conn.execute(
            "SELECT id, credit_date, amount, reason FROM invoice_credit_notes WHERE company_id=? AND invoice_id=? ORDER BY credit_date ASC, id ASC",
            (cid, d['id'])
        ).fetchall()]
        totals = get_invoice_financial_totals(conn, cid, d['id'])
        d['payments'] = payments
        d['payment_count'] = len(payments)
        d['credit_notes'] = credit_notes
        d['credit_count'] = len(credit_notes)
        d['total_paid'] = totals['paid']
        d['total_credited'] = totals['credited']
        d['outstanding_amount'] = totals['outstanding']
        formatted_invoices.append(d)

    formatted_quotes = []
    for q in quotes:
        d = dict(q)
        try:
            d['formatted_num'] = f"{quote_prefix}{int(quote_start) + d['id'] - 1:04d}"
        except:
            d['formatted_num'] = f"{quote_prefix}{d['id']:04d}"
        formatted_quotes.append(d)
    
    conn.close()
    return render_template('invoicing_index.html', company=dict(comp), clients=clients, services=[dict(s) for s in services], projects=[dict(p) for p in projects], invoices=formatted_invoices, quotes=formatted_quotes, invoice_pagination=invoice_pagination, quote_pagination=quote_pagination, invoice_q=invoice_q, quote_q=quote_q, inv_info=inv_info, inv_prefix=inv_prefix, inv_start=inv_start, quote_prefix=quote_prefix, quote_start=quote_start, session=session)

@app.route('/api/save_invoice_settings', methods=['POST'])
def save_invoice_settings():
    if not session.get('can_invoicing') and not session.get('is_superadmin'): return "Forbidden", 403
    data = request.json
    cid = session['company_id']
    conn = get_db_connection()
    
    settings_to_save = {
        'invoice_additional_info': data.get('invoice_additional_info', ''),
        'invoice_prefix': data.get('invoice_prefix', 'INV-'),
        'invoice_start': data.get('invoice_start', '1'),
        'quote_prefix': data.get('quote_prefix', 'QT-'),
        'quote_start': data.get('quote_start', '1')
    }
    
    for key, value in settings_to_save.items():
        exists = conn.execute("SELECT 1 FROM settings WHERE key=? AND company_id=?", (key, cid)).fetchone()
        if exists:
            conn.execute("UPDATE settings SET value=? WHERE key=? AND company_id=?", (value, key, cid))
        else:
            conn.execute("INSERT INTO settings (company_id, key, value) VALUES (?, ?, ?)", (cid, key, value))
            
    conn.commit()
    conn.close()
    log_action('Invoicing', 'Updated Settings', 'Updated invoice layout and sequence settings.')
    return jsonify({"status": "success"})

def get_employee_first_names_for_invoice(employee_names):
    """Return first names from the booking employee field for invoice descriptions."""
    names = []
    seen = set()
    for raw_name in (employee_names or '').split(','):
        full_name = raw_name.strip()
        if not full_name:
            continue
        first_name = full_name.split()[0].strip()
        if first_name and first_name.lower() not in seen:
            names.append(first_name)
            seen.add(first_name.lower())
    return ', '.join(names)


def build_invoice_service_description(service_name, employee_names):
    service_name = (service_name or 'Service').strip() or 'Service'
    employee_first_names = get_employee_first_names_for_invoice(employee_names)
    if employee_first_names:
        return f"{service_name} - {employee_first_names}"
    return service_name


@app.route('/api/uninvoiced', methods=['POST'])
def get_uninvoiced():
    data = request.json
    conn = get_db_connection()
    cid = session['company_id']
    
    try:
        client_id = int(data.get('client_id') or 0)
    except Exception:
        client_id = 0
    if client_id:
        bookings = conn.execute("SELECT * FROM bookings WHERE company_id=? AND client_id=? AND is_invoiced=0 AND (project_id IS NULL OR project_id=0) ORDER BY start ASC", (cid, client_id)).fetchall()
    else:
        bookings = conn.execute("SELECT * FROM bookings WHERE company_id=? AND title=? AND is_invoiced=0 AND (project_id IS NULL OR project_id=0) ORDER BY start ASC", (cid, data.get('client_name'))).fetchall()
    services = conn.execute("SELECT * FROM services WHERE company_id=?", (cid,)).fetchall()
    s_dict = {s['name'].strip(): float(s['client_price'] or 0) for s in services}
    
    res = []
    for b in bookings:
        booking_type = b['booking_type'] or 'Service'
        price = sum(s_dict.get(t.strip(), 0) for t in booking_type.split(',') if t.strip())
        res.append({
            "id": b['id'],
            "date": b['start'][:10],
            "type": booking_type,
            "employee_first_names": get_employee_first_names_for_invoice(b['employee']),
            "description": build_invoice_service_description(booking_type, b['employee']),
            "price": price
        })
    conn.close()
    return jsonify(res)

@app.route('/api/save_invoice', methods=['POST'])
def save_invoice():
    data = request.json
    conn = get_db_connection()
    cid = session['company_id']

    invoice_type = (data.get('invoice_type') or 'standard').strip().lower()
    project_id = None

    # Use the invoice-specific discount submitted from the invoice screen.
    # The client profile discount is only a default that pre-populates the form;
    # users may override it for this individual invoice.
    if invoice_type == 'project':
        try:
            project_id = int(data.get('project_id') or 0)
        except Exception:
            project_id = 0
        project = conn.execute('''SELECT p.*, c.id AS linked_client_id, c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name
                                  FROM projects p
                                  LEFT JOIN clients c ON c.id=p.client_id AND c.company_id=p.company_id
                                  WHERE p.id=? AND p.company_id=?''', (project_id, cid)).fetchone()
        if not project:
            conn.close()
            return jsonify({"status": "error", "message": "Project not found for this company."}), 400
        if (project['status'] or '').lower() == 'cancelled':
            conn.close()
            return jsonify({"status": "error", "message": "Cancelled projects cannot be invoiced."}), 400
        if not project['linked_client_id']:
            conn.close()
            return jsonify({"status": "error", "message": "This project is not linked to a valid client."}), 400
        project_client = get_client_by_id(conn, cid, project['linked_client_id'])
        client_id = project_client['id']
        client_name = client_display_name(project_client)
        project_label = project['project_name'] or 'Project'
        if project['project_code']:
            project_label = f"{project_label} ({project['project_code']})"
        project_fixed_price = float(project['fixed_price'] or 0)
        submitted_items = data.get('items') or []
        if submitted_items:
            items = submitted_items
        else:
            items = [{
                'booking_id': None,
                'service_date': project['start_date'] or data.get('date'),
                'description': f"Project: {project_label}",
                'quantity': 1,
                'unit_price': project_fixed_price,
                'amount': project_fixed_price
            }]
    else:
        try:
            client = resolve_client_from_payload(conn, cid, data, id_key='client_id', name_key='client_name')
        except ValueError as e:
            conn.close()
            return jsonify({"status": "error", "message": str(e)}), 400
        client_id = client['id']
        client_name = client_display_name(client)
        items = data.get('items', [])

    items = [normalise_billing_item(item) for item in (items or [])]
    items = [item for item in items if item.get('description') and float(item.get('amount') or 0) > 0]

    discount_percent = sanitize_percent(data.get('discount_percent', 0))
    original_subtotal = round(sum(float(item.get('amount') or 0) for item in items), 2)
    discount_percent, discount_amount, discounted_subtotal = calculate_invoice_discount(original_subtotal, discount_percent)
    apply_vat = bool(data.get('apply_vat', False))
    vat_amount = round(discounted_subtotal * 0.15, 2) if apply_vat else 0.0
    total = round(discounted_subtotal + vat_amount, 2)
    amount_due_now, balance_remaining = calculate_invoice_due_now(total, data.get('amount_due_now'))
    
    cursor = conn.cursor()
    cursor.execute("""INSERT INTO invoices (company_id, client_id, client_name, date, due_date, subtotal, vat_amount, total, status, discount_percent, discount_amount, amount_due_now, balance_remaining, invoice_type, project_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                   (cid, client_id, client_name, data['date'], data['due_date'], original_subtotal, vat_amount, total, 'Unpaid', discount_percent, discount_amount, amount_due_now, balance_remaining, invoice_type, project_id))
    inv_id = cursor.lastrowid
    
    for item in items:
        cursor.execute("INSERT INTO invoice_items (invoice_id, booking_id, service_date, description, quantity, unit_price, amount) VALUES (?, ?, ?, ?, ?, ?, ?)",
                       (inv_id, item.get('booking_id'), item.get('service_date'), item['description'], item.get('quantity', 1), item.get('unit_price', item['amount']), item['amount']))
        
        if item.get('booking_id'):
            cursor.execute("UPDATE bookings SET is_invoiced=1 WHERE id=? AND company_id=?", (item['booking_id'], cid))

    if invoice_type == 'project' and project_id:
        cursor.execute("UPDATE projects SET status='Invoiced', updated_at=CURRENT_TIMESTAMP WHERE id=? AND company_id=?", (project_id, cid))
            
    conn.commit()
    conn.close()
    log_action('Invoicing', 'Created Invoice', f"Generated Invoice #{inv_id} for {client_name}")
    return jsonify({"status": "success", "invoice_id": inv_id})

@app.route('/api/invoice/<int:inv_id>')
def get_invoice(inv_id):
    conn = get_db_connection()
    cid = session['company_id']
    inv = conn.execute("SELECT * FROM invoices WHERE id=? AND company_id=?", (inv_id, cid)).fetchone()
    if not inv:
        conn.close()
        return jsonify({"status": "error", "message": "Invoice not found."}), 404
    items = conn.execute("SELECT * FROM invoice_items WHERE invoice_id=?", (inv_id,)).fetchall()
    client = get_document_client(conn, cid, inv)
    
    settings_rows = conn.execute("SELECT key, value FROM settings WHERE company_id=?", (cid,)).fetchall()
    s_dict = {s['key']: s['value'] for s in settings_rows}
    inv_prefix = s_dict.get('invoice_prefix', 'INV-')
    inv_start = s_dict.get('invoice_start', '1')
    
    d = dict(inv)
    try:
        d['formatted_num'] = f"{inv_prefix}{int(inv_start) + d['id'] - 1:04d}"
    except:
        d['formatted_num'] = f"{inv_prefix}{d['id']:04d}"

    client_full = inv['client_name']
    client_email = ''
    if client:
        client_dict = dict(client)
        if 'surname' in client_dict and client_dict['surname']:
            client_full = f"{client_dict['name']} {client_dict['surname']}"
        if 'email' in client_dict and client_dict['email']:
            client_email = client_dict['email']

    project = None
    if d.get('project_id'):
        project = conn.execute('SELECT * FROM projects WHERE id=? AND company_id=?', (d.get('project_id'), cid)).fetchone()
    d['project_name'] = project['project_name'] if project else ''
    d['project_code'] = project['project_code'] if project else ''
    d['invoice_type'] = d.get('invoice_type') or 'standard'

    totals = get_invoice_financial_totals(conn, cid, inv_id)
    d['amount_due_now'] = d.get('amount_due_now') if d.get('amount_due_now') is not None else d.get('total', 0)
    d['balance_remaining'] = d.get('balance_remaining') if d.get('balance_remaining') is not None else 0
    d['total_paid'] = totals['paid']
    d['total_credited'] = totals['credited']
    d['outstanding_balance'] = totals['outstanding']
    payments = [dict(p) for p in conn.execute('SELECT * FROM invoice_payments WHERE invoice_id=? AND company_id=? ORDER BY payment_date DESC, id DESC', (inv_id, cid)).fetchall()]
    credit_notes = [dict(cn) for cn in conn.execute('SELECT * FROM invoice_credit_notes WHERE invoice_id=? AND company_id=? ORDER BY credit_date DESC, id DESC', (inv_id, cid)).fetchall()]

    conn.close()
    return jsonify({
        "invoice": d,
        "items": [dict(i) for i in items],
        "client_full_name": client_full,
        "client_email": client_email,
        "client_address": client['address'] if client and 'address' in dict(client) and client['address'] else '',
        "client_company": client['company_name'] if client and 'company_name' in dict(client) and client['company_name'] else '',
        "client_reg": client['registration_number'] if client and 'registration_number' in dict(client) and client['registration_number'] else '',
        "client_vat": client['vat_number'] if client and 'vat_number' in dict(client) and client['vat_number'] else '',
        "payments": payments,
        "credit_notes": credit_notes
    })

@app.route('/api/invoice/<int:inv_id>/status', methods=['POST'])
def update_inv_status(inv_id):
    status = request.json['status']
    conn = get_db_connection()
    conn.execute("UPDATE invoices SET status=? WHERE id=? AND company_id=?", (status, inv_id, session['company_id']))
    conn.commit()
    conn.close()
    log_action('Invoicing', 'Updated Invoice', f"Marked Invoice #{inv_id} as {status}")
    return jsonify({"status": "success"})

@app.route('/api/invoice/<int:inv_id>/payment', methods=['POST'])
def record_invoice_payment(inv_id):
    data = request.json or {}
    cid = session['company_id']
    conn = get_db_connection()
    inv = conn.execute('SELECT * FROM invoices WHERE id=? AND company_id=?', (inv_id, cid)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Invoice not found'}), 404
    totals = get_invoice_financial_totals(conn, cid, inv_id)
    if totals['outstanding'] <= 0:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Cannot record payment because this invoice has no outstanding amount.'}), 400

    amount = sanitize_money(data.get('amount'))
    if amount <= 0:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Payment amount must be greater than zero.'}), 400
    if amount > totals['outstanding']:
        conn.close()
        return jsonify({'status': 'error', 'message': f"Payment amount cannot exceed the outstanding amount of R{totals['outstanding']:.2f}."}), 400

    payment_date = data.get('payment_date') or datetime.now().strftime('%Y-%m-%d')
    cur = conn.cursor()
    cur.execute('''INSERT INTO invoice_payments (company_id, invoice_id, payment_date, amount, payment_method, reference, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?)''',
                (cid, inv_id, payment_date, amount, data.get('payment_method', ''), data.get('reference', ''), data.get('notes', '')))
    payment_id = cur.lastrowid
    update_invoice_payment_status(conn, cid, inv_id)
    conn.commit()
    conn.close()
    log_action('Invoicing', 'Recorded Invoice Payment', f"Recorded payment R{amount:.2f} for Invoice #{inv_id}")
    return jsonify({'status': 'success', 'payment_id': payment_id})

@app.route('/api/receipt/<int:payment_id>')
def get_receipt(payment_id):
    cid = session['company_id']
    conn = get_db_connection()
    payment = conn.execute('SELECT * FROM invoice_payments WHERE id=? AND company_id=?', (payment_id, cid)).fetchone()
    if not payment:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Receipt not found'}), 404
    inv = conn.execute('SELECT * FROM invoices WHERE id=? AND company_id=?', (payment['invoice_id'], cid)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Invoice not found'}), 404
    client = get_document_client(conn, cid, inv)
    settings_rows = conn.execute('SELECT key, value FROM settings WHERE company_id=?', (cid,)).fetchall()
    s_dict = {s['key']: s['value'] for s in settings_rows}
    inv_prefix = s_dict.get('invoice_prefix', 'INV-')
    inv_start = s_dict.get('invoice_start', '1')
    inv_d = dict(inv)
    try:
        inv_d['formatted_num'] = f"{inv_prefix}{int(inv_start) + inv_d['id'] - 1:04d}"
    except Exception:
        inv_d['formatted_num'] = f"{inv_prefix}{inv_d['id']:04d}"

    client_full = inv['client_name']
    client_email = ''
    if client:
        cdict = dict(client)
        if cdict.get('surname'):
            client_full = f"{cdict.get('name', '')} {cdict.get('surname', '')}".strip()
        client_email = cdict.get('email') or ''

    receipt_no = f"RCP-{payment_id:04d}"
    totals = get_invoice_financial_totals(conn, cid, inv['id'])
    conn.close()
    return jsonify({
        'receipt': dict(payment),
        'receipt_no': receipt_no,
        'invoice': inv_d,
        'total_paid': totals['paid'],
        'total_credited': totals['credited'],
        'outstanding_balance': totals['outstanding'],
        'client_full_name': client_full,
        'client_email': client_email,
        'client_address': client['address'] if client and 'address' in dict(client) and client['address'] else '',
        'client_company': client['company_name'] if client and 'company_name' in dict(client) and client['company_name'] else '',
        'client_reg': client['registration_number'] if client and 'registration_number' in dict(client) and client['registration_number'] else '',
        'client_vat': client['vat_number'] if client and 'vat_number' in dict(client) and client['vat_number'] else ''
    })

@app.route('/api/invoice/<int:inv_id>/credit', methods=['POST'])
def credit_invoice(inv_id):
    if not session.get('can_invoicing') and not session.get('is_superadmin'):
        return jsonify({'status': 'error', 'message': 'Forbidden'}), 403
    data = request.json or {}
    cid = session['company_id']
    conn = get_db_connection()

    inv = conn.execute('SELECT * FROM invoices WHERE id=? AND company_id=?', (inv_id, cid)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Invoice not found.'}), 404

    totals = get_invoice_financial_totals(conn, cid, inv_id)
    if totals['outstanding'] <= 0:
        conn.close()
        return jsonify({'status': 'error', 'message': 'This invoice has no outstanding amount to credit.'}), 400

    amount = sanitize_money(data.get('amount'))
    if amount <= 0:
        amount = totals['outstanding']
    if amount > totals['outstanding']:
        conn.close()
        return jsonify({'status': 'error', 'message': f"Credit amount cannot exceed the outstanding amount of R{totals['outstanding']:.2f}."}), 400

    credit_date = data.get('credit_date') or datetime.now().strftime('%Y-%m-%d')
    reason = (data.get('reason') or '').strip()
    cur = conn.cursor()
    cur.execute('''INSERT INTO invoice_credit_notes (company_id, invoice_id, credit_date, amount, reason)
                   VALUES (?, ?, ?, ?, ?)''', (cid, inv_id, credit_date, amount, reason))
    credit_id = cur.lastrowid

    update_invoice_payment_status(conn, cid, inv_id)
    updated_totals = get_invoice_financial_totals(conn, cid, inv_id)

    # Only a full credit note releases associated bookings. Partial credits keep the invoice linked.
    if updated_totals['credited'] >= updated_totals['total'] and updated_totals['total'] > 0:
        items = conn.execute('SELECT booking_id FROM invoice_items WHERE invoice_id=? AND booking_id IS NOT NULL', (inv_id,)).fetchall()
        for item in items:
            conn.execute('UPDATE bookings SET is_invoiced=0 WHERE id=? AND company_id=?', (item['booking_id'], cid))

    conn.commit()
    conn.close()
    log_action('Invoicing', 'Issued Credit Note', f"Credited R{amount:.2f} on Invoice #{inv_id}.")
    return jsonify({'status': 'success', 'credit_id': credit_id})



@app.route('/api/invoice/<int:inv_id>/post_accounting', methods=['POST'])
def post_invoice_accounting_api(inv_id):
    if not (session.get('can_accounting') or session.get('is_superadmin')):
        return jsonify({'status': 'error', 'message': 'Accounting permission is required to post invoices to Accounting.'}), 403
    cid = session['company_id']
    conn = get_db_connection()
    try:
        journal_id = post_invoice_to_accounting_record(conn, cid, inv_id)
        conn.commit()
    except ValueError as exc:
        conn.close()
        return jsonify({'status': 'error', 'message': str(exc)}), 400
    except Exception as exc:
        conn.close()
        return jsonify({'status': 'error', 'message': f'Could not post invoice to Accounting: {exc}'}), 500
    conn.close()
    log_action('Accounting', 'Posted Invoice', f'Posted Invoice #{inv_id} to Accounting Journal #{journal_id}.')
    return jsonify({'status': 'success', 'journal_id': journal_id, 'message': f'Invoice posted to Accounting Journal #{journal_id}.'})


@app.route('/api/credit_note/<int:credit_id>/post_accounting', methods=['POST'])
def post_credit_note_accounting_api(credit_id):
    if not (session.get('can_accounting') or session.get('is_superadmin')):
        return jsonify({'status': 'error', 'message': 'Accounting permission is required to post credit notes to Accounting.'}), 403
    cid = session['company_id']
    conn = get_db_connection()
    try:
        journal_id = post_credit_note_to_accounting_record(conn, cid, credit_id)
        conn.commit()
    except ValueError as exc:
        conn.close()
        return jsonify({'status': 'error', 'message': str(exc)}), 400
    except Exception as exc:
        conn.close()
        return jsonify({'status': 'error', 'message': f'Could not post credit note to Accounting: {exc}'}), 500
    conn.close()
    log_action('Accounting', 'Posted Credit Note', f'Posted Credit Note #{credit_id} to Accounting Journal #{journal_id}.')
    return jsonify({'status': 'success', 'journal_id': journal_id, 'message': f'Credit Note posted to Accounting Journal #{journal_id}.'})


@app.route('/api/credit_note/<int:credit_id>')
def get_credit_note(credit_id):
    cid = session['company_id']
    conn = get_db_connection()
    credit = conn.execute('SELECT * FROM invoice_credit_notes WHERE id=? AND company_id=?', (credit_id, cid)).fetchone()
    if not credit:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Credit note not found'}), 404
    inv = conn.execute('SELECT * FROM invoices WHERE id=? AND company_id=?', (credit['invoice_id'], cid)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Invoice not found'}), 404
    client = get_document_client(conn, cid, inv)
    settings_rows = conn.execute('SELECT key, value FROM settings WHERE company_id=?', (cid,)).fetchall()
    s_dict = {s['key']: s['value'] for s in settings_rows}
    inv_prefix = s_dict.get('invoice_prefix', 'INV-')
    inv_start = s_dict.get('invoice_start', '1')
    inv_d = dict(inv)
    try:
        inv_d['formatted_num'] = f"{inv_prefix}{int(inv_start) + inv_d['id'] - 1:04d}"
    except Exception:
        inv_d['formatted_num'] = f"{inv_prefix}{inv_d['id']:04d}"

    client_full = inv['client_name']
    client_email = ''
    if client:
        cdict = dict(client)
        if cdict.get('surname'):
            client_full = f"{cdict.get('name', '')} {cdict.get('surname', '')}".strip()
        client_email = cdict.get('email') or ''

    totals = get_invoice_financial_totals(conn, cid, inv['id'])
    conn.close()
    return jsonify({
        'credit_note': dict(credit),
        'credit_no': f"CN-{credit_id:04d}",
        'invoice': inv_d,
        'total_paid': totals['paid'],
        'total_credited': totals['credited'],
        'outstanding_balance': totals['outstanding'],
        'client_full_name': client_full,
        'client_email': client_email,
        'client_address': client['address'] if client and 'address' in dict(client) and client['address'] else '',
        'client_company': client['company_name'] if client and 'company_name' in dict(client) and client['company_name'] else '',
        'client_reg': client['registration_number'] if client and 'registration_number' in dict(client) and client['registration_number'] else '',
        'client_vat': client['vat_number'] if client and 'vat_number' in dict(client) and client['vat_number'] else ''
    })



def today_iso():
    return datetime.now().strftime('%Y-%m-%d')


def refresh_expired_quotes(conn, company_id=None):
    """Mark pending quotes as expired when their Valid Until date has passed."""
    today = today_iso()
    if company_id:
        conn.execute("""
            UPDATE quotes
            SET status='Expired'
            WHERE company_id=?
              AND valid_until IS NOT NULL
              AND TRIM(valid_until) <> ''
              AND date(valid_until) < date(?)
              AND COALESCE(status, 'Pending') IN ('Pending', 'Sent')
        """, (company_id, today))
    else:
        conn.execute("""
            UPDATE quotes
            SET status='Expired'
            WHERE valid_until IS NOT NULL
              AND TRIM(valid_until) <> ''
              AND date(valid_until) < date(?)
              AND COALESCE(status, 'Pending') IN ('Pending', 'Sent')
        """, (today,))


def quote_is_expired(quote_row):
    valid_until = quote_row['valid_until'] if quote_row and 'valid_until' in quote_row.keys() else None
    if not valid_until:
        return False
    try:
        return datetime.strptime(valid_until, '%Y-%m-%d').date() < datetime.now().date()
    except Exception:
        return False


@app.route('/api/save_quote', methods=['POST'])
def save_quote():
    data = request.json
    conn = get_db_connection()
    cid = session['company_id']

    try:
        client = resolve_client_from_payload(conn, cid, data, id_key='client_id', name_key='client_name')
    except ValueError as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)}), 400
    client_id = client['id']
    client_name = client_display_name(client)

    items = [normalise_billing_item(item) for item in (data.get('items') or [])]
    items = [item for item in items if (item.get('description') or '').strip()]
    if not items:
        conn.close()
        return jsonify({"status": "error", "message": "Add at least one quote line item with a description."}), 400

    zero_items = [item for item in items if float(item.get('amount') or 0) <= 0]
    if zero_items:
        conn.close()
        return jsonify({"status": "error", "message": "Each quote line with a description must have a quantity and unit price greater than zero."}), 400

    subtotal = round(sum(float(item.get('amount') or 0) for item in items), 2)
    vat_amount = round(subtotal * 0.15, 2) if bool(data.get('apply_vat', False)) else 0.0
    total = round(subtotal + vat_amount, 2)
    
    cursor = conn.cursor()
    cursor.execute("INSERT INTO quotes (company_id, client_id, client_name, date, valid_until, subtotal, vat_amount, total, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                   (cid, client_id, client_name, data['date'], data['valid_until'], subtotal, vat_amount, total, 'Pending'))
    q_id = cursor.lastrowid
    
    for item in items:
        cursor.execute("INSERT INTO quote_items (quote_id, service_date, description, quantity, unit_price, amount) VALUES (?, ?, ?, ?, ?, ?)",
                       (q_id, item.get('service_date'), item['description'], item.get('quantity', 1), item.get('unit_price', item['amount']), item['amount']))
            
    conn.commit()
    conn.close()
    log_action('Invoicing', 'Created Quote', f"Generated Quote #{q_id} for {client_name}")
    return jsonify({"status": "success", "quote_id": q_id})

@app.route('/api/quote/<int:q_id>')
def get_quote(q_id):
    conn = get_db_connection()
    cid = session['company_id']
    refresh_expired_quotes(conn, cid)
    conn.commit()
    q = conn.execute("SELECT * FROM quotes WHERE id=? AND company_id=?", (q_id, cid)).fetchone()
    if not q:
        conn.close()
        return jsonify({"status": "error", "message": "Quote not found."}), 404
    items = conn.execute("SELECT * FROM quote_items WHERE quote_id=?", (q_id,)).fetchall()
    client = get_document_client(conn, cid, q)
    
    settings_rows = conn.execute("SELECT key, value FROM settings WHERE company_id=?", (cid,)).fetchall()
    s_dict = {s['key']: s['value'] for s in settings_rows}
    quote_prefix = s_dict.get('quote_prefix', 'QT-')
    quote_start = s_dict.get('quote_start', '1')
    
    d = dict(q)
    try:
        d['formatted_num'] = f"{quote_prefix}{int(quote_start) + d['id'] - 1:04d}"
    except:
        d['formatted_num'] = f"{quote_prefix}{d['id']:04d}"

    client_full = q['client_name']
    client_email = ''
    if client:
        client_dict = dict(client)
        if 'surname' in client_dict and client_dict['surname']:
            client_full = f"{client_dict['name']} {client_dict['surname']}"
        if 'email' in client_dict and client_dict['email']:
            client_email = client_dict['email']

    conn.close()
    
    return jsonify({
        "quote": d,
        "items": [dict(i) for i in items],
        "client_full_name": client_full,
        "client_email": client_email,
        "client_address": client['address'] if client and 'address' in dict(client) and client['address'] else '',
        "client_company": client['company_name'] if client and 'company_name' in dict(client) and client['company_name'] else '',
        "client_reg": client['registration_number'] if client and 'registration_number' in dict(client) and client['registration_number'] else '',
        "client_vat": client['vat_number'] if client and 'vat_number' in dict(client) and client['vat_number'] else ''
    })

@app.route('/api/quote/<int:q_id>/status', methods=['POST'])
def update_quote_status(q_id):
    requested_status = request.json.get('status')
    allowed = {'Pending', 'Accepted', 'Rejected'}
    if requested_status not in allowed:
        return jsonify({"status": "error", "message": "Invalid quote status."}), 400

    conn = get_db_connection()
    cid = session['company_id']
    refresh_expired_quotes(conn, cid)
    q = conn.execute("SELECT * FROM quotes WHERE id=? AND company_id=?", (q_id, cid)).fetchone()
    if not q:
        conn.close()
        return jsonify({"status": "error", "message": "Quote not found."}), 404

    current_status = q['status'] or 'Pending'
    if current_status == 'Expired':
        conn.close()
        return jsonify({"status": "error", "message": "This quote has expired and cannot be accepted. Create a new quote or update the valid-until date."}), 400
    if current_status == 'Converted':
        conn.close()
        return jsonify({"status": "error", "message": "This quote has already been converted to an invoice."}), 400
    if requested_status == 'Accepted' and quote_is_expired(q):
        conn.execute("UPDATE quotes SET status='Expired' WHERE id=? AND company_id=?", (q_id, cid))
        conn.commit()
        conn.close()
        return jsonify({"status": "error", "message": "This quote has expired and cannot be accepted."}), 400

    conn.execute("UPDATE quotes SET status=? WHERE id=? AND company_id=?", (requested_status, q_id, cid))
    conn.commit()
    conn.close()
    log_action('Invoicing', 'Updated Quote', f"Marked Quote #{q_id} as {requested_status}")
    return jsonify({"status": "success"})


@app.route('/api/quote/<int:q_id>/convert_to_invoice', methods=['POST'])
def convert_quote_to_invoice(q_id):
    conn = get_db_connection()
    cid = session['company_id']
    refresh_expired_quotes(conn, cid)
    q = conn.execute("SELECT * FROM quotes WHERE id=? AND company_id=?", (q_id, cid)).fetchone()
    if not q:
        conn.close()
        return jsonify({"status": "error", "message": "Quote not found."}), 404

    if (q['status'] or 'Pending') == 'Expired' or quote_is_expired(q):
        conn.execute("UPDATE quotes SET status='Expired' WHERE id=? AND company_id=?", (q_id, cid))
        conn.commit()
        conn.close()
        return jsonify({"status": "error", "message": "Expired quotes cannot be converted. Create a new quote first."}), 400

    if (q['status'] or 'Pending') != 'Accepted':
        conn.close()
        return jsonify({"status": "error", "message": "Only accepted quotes can be converted to invoices."}), 400

    q_dict = dict(q)
    existing_invoice_id = q_dict.get('converted_invoice_id')
    if existing_invoice_id:
        conn.close()
        return jsonify({"status": "error", "message": f"This quote has already been converted to invoice #{existing_invoice_id}."}), 400

    items = conn.execute("SELECT * FROM quote_items WHERE quote_id=?", (q_id,)).fetchall()
    if not items:
        conn.close()
        return jsonify({"status": "error", "message": "This quote has no line items to invoice."}), 400

    invoice_date = today_iso()
    due_date = invoice_date
    subtotal = float(q['subtotal'] or 0)
    vat_amount = float(q['vat_amount'] or 0)
    total = float(q['total'] or 0)
    amount_due_now, balance_remaining = calculate_invoice_due_now(total, total)

    cur = conn.cursor()
    cur.execute("""
        INSERT INTO invoices
        (company_id, client_id, client_name, date, due_date, subtotal, vat_amount, total, status, discount_percent, discount_amount, amount_due_now, balance_remaining)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (cid, dict(q).get('client_id'), q['client_name'], invoice_date, due_date, subtotal, vat_amount, total, 'Unpaid', 0, 0, amount_due_now, balance_remaining))
    inv_id = cur.lastrowid

    for item in items:
        quantity = item['quantity'] if 'quantity' in item.keys() and item['quantity'] else 1
        unit_price = item['unit_price'] if 'unit_price' in item.keys() and item['unit_price'] is not None else item['amount']
        cur.execute("INSERT INTO invoice_items (invoice_id, booking_id, service_date, description, quantity, unit_price, amount) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (inv_id, None, item['service_date'], item['description'], quantity, unit_price, item['amount']))

    conn.execute("UPDATE quotes SET status='Converted', converted_invoice_id=?, converted_date=? WHERE id=? AND company_id=?",
                 (inv_id, invoice_date, q_id, cid))
    conn.commit()
    conn.close()
    log_action('Invoicing', 'Converted Quote to Invoice', f"Converted Quote #{q_id} to Invoice #{inv_id}")
    return jsonify({"status": "success", "invoice_id": inv_id})



# --- REAL DOCUMENT PDF GENERATOR FOR INVOICES & QUOTES ---
def _pdf_text_escape(value):
    text = str(value or '')
    text = text.replace('\u00a0', ' ')
    text = text.replace('\r', ' ').replace('\n', ' ')
    text = text.encode('latin-1', 'replace').decode('latin-1')
    return text.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _money(value):
    # Raw PDF text uses base fonts that can render non-breaking spaces
    # inconsistently in some viewers. Keep the Trial Balance style, but use a
    # normal ASCII space as the thousands separator in downloaded PDFs.
    return format_display_money(value).replace('\u00a0', ' ')


def _qty_text(value):
    try:
        qty = float(value if value not in (None, '') else 1)
    except Exception:
        qty = 1.0
    if abs(qty - round(qty)) < 0.0001:
        return str(int(round(qty)))
    return f"{qty:.2f}".rstrip('0').rstrip('.')


def _wrap_pdf_text(text, max_chars):
    words = str(text or '').replace('\r', ' ').replace('\n', ' ').split()
    if not words:
        return ['']
    lines = []
    current = ''
    for word in words:
        if len(word) > max_chars:
            if current:
                lines.append(current)
                current = ''
            while len(word) > max_chars:
                lines.append(word[:max_chars])
                word = word[max_chars:]
            current = word
        elif not current:
            current = word
        elif len(current) + 1 + len(word) <= max_chars:
            current += ' ' + word
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines or ['']


def _send_pdf_email(pdf_bytes, filename, email, doc_name, client_name):
    conn = get_db_connection()
    settings_rows = conn.execute('SELECT key, value FROM settings WHERE company_id=?', (session['company_id'],)).fetchall()
    conn.close()
    s_dict = {r['key']: r['value'] for r in settings_rows}
    if not s_dict.get('smtp_server') or not s_dict.get('smtp_user') or not s_dict.get('smtp_pass'):
        return {"message": "SMTP settings are incomplete. Please configure your Email & Calendar Settings in the Hub."}, 400
    try:
        msg = EmailMessage()
        msg['Subject'] = f"{doc_name.replace('_', ' ')} from {session.get('company_name')}"
        msg['From'] = s_dict.get('sender_email', s_dict.get('smtp_user'))
        msg['To'] = email
        msg.set_content(f"Dear {client_name},\n\nPlease find attached your {doc_name.replace('_', ' ')}.\n\nThank you for your business.\n\nKind regards,\n{session.get('company_name')}")
        msg.add_attachment(pdf_bytes, maintype='application', subtype='pdf', filename=filename)
        port = int(s_dict.get('smtp_port', 465))
        if port == 587:
            with smtplib.SMTP(s_dict['smtp_server'], port, timeout=10) as server:
                server.starttls()
                server.login(s_dict['smtp_user'], s_dict['smtp_pass'])
                server.send_message(msg)
        else:
            with smtplib.SMTP_SSL(s_dict['smtp_server'], port, timeout=10) as server:
                server.login(s_dict['smtp_user'], s_dict['smtp_pass'])
                server.send_message(msg)
        log_action('Invoicing', 'Emailed Document', f"Sent {doc_name} to {email}")
        return {"message": "Email sent successfully!"}, 200
    except Exception as e:
        return {"message": f"Error sending email: {str(e)}"}, 500


def _build_billing_pdf_payload(kind, doc_id):
    cid = session['company_id']
    conn = get_db_connection()
    company_row = conn.execute('SELECT * FROM companies WHERE id=?', (cid,)).fetchone()
    settings_rows = conn.execute('SELECT key, value FROM settings WHERE company_id=?', (cid,)).fetchall()
    s_dict = {s['key']: s['value'] for s in settings_rows}
    if kind == 'invoice':
        doc = conn.execute('SELECT * FROM invoices WHERE id=? AND company_id=?', (doc_id, cid)).fetchone()
        if not doc:
            conn.close()
            return None
        items = conn.execute('SELECT * FROM invoice_items WHERE invoice_id=?', (doc_id,)).fetchall()
        prefix = s_dict.get('invoice_prefix', 'INV-')
        start = s_dict.get('invoice_start', '1')
        try:
            formatted_num = f"{prefix}{int(start) + doc['id'] - 1:04d}"
        except Exception:
            formatted_num = f"{prefix}{doc['id']:04d}"
        totals = get_invoice_financial_totals(conn, cid, doc_id)
        title = 'TAX INVOICE'
        meta_label = 'Due Date'
        meta_value = doc['due_date'] or 'On Receipt'
        totals_rows = [('Subtotal', doc['subtotal'])]
        if float(doc['discount_amount'] or 0) > 0:
            label = 'Discount'
            try:
                if float(doc['discount_percent'] or 0) > 0:
                    label += f" ({float(doc['discount_percent']):.2f}%".replace('.00%', '%') + ')'
            except Exception:
                pass
            totals_rows.append((label, -float(doc['discount_amount'] or 0)))
        totals_rows += [('VAT', doc['vat_amount']), ('TOTAL INVOICE', doc['total'])]
        if doc['amount_due_now'] is not None:
            totals_rows.append(('Amount Due Now', doc['amount_due_now']))
        if doc['balance_remaining'] is not None:
            totals_rows.append(('Balance Remaining', doc['balance_remaining']))
        if totals.get('paid', 0) > 0 or totals.get('credited', 0) > 0:
            totals_rows.append(('Total Paid', totals.get('paid', 0)))
            if totals.get('credited', 0) > 0:
                totals_rows.append(('Total Credited', totals.get('credited', 0)))
            totals_rows.append(('Outstanding Balance', totals.get('outstanding', 0)))
        project_note = ''
        if dict(doc).get('project_id'):
            project = conn.execute('SELECT project_name, project_code FROM projects WHERE id=? AND company_id=?', (doc['project_id'], cid)).fetchone()
            if project:
                project_note = f"Project Invoice: {project['project_name'] or ''}{' (' + project['project_code'] + ')' if project['project_code'] else ''}"
    else:
        doc = conn.execute('SELECT * FROM quotes WHERE id=? AND company_id=?', (doc_id, cid)).fetchone()
        if not doc:
            conn.close()
            return None
        items = conn.execute('SELECT * FROM quote_items WHERE quote_id=?', (doc_id,)).fetchall()
        prefix = s_dict.get('quote_prefix', 'QT-')
        start = s_dict.get('quote_start', '1')
        try:
            formatted_num = f"{prefix}{int(start) + doc['id'] - 1:04d}"
        except Exception:
            formatted_num = f"{prefix}{doc['id']:04d}"
        title = 'OFFICIAL QUOTE'
        meta_label = 'Valid Until'
        meta_value = doc['valid_until'] or 'N/A'
        totals_rows = [('Subtotal', doc['subtotal']), ('VAT', doc['vat_amount']), ('TOTAL', doc['total'])]
        project_note = ''
    client = get_document_client(conn, cid, doc)
    company = dict(company_row) if company_row else {}
    client_full = doc['client_name']
    client_email = ''
    client_lines = []
    if client:
        cdict = dict(client)
        client_full = f"{cdict.get('name', '')} {cdict.get('surname', '')}".strip() if cdict.get('surname') else cdict.get('name', doc['client_name'])
        client_email = cdict.get('email') or ''
        if cdict.get('company_name'):
            client_lines.append(cdict.get('company_name'))
        if cdict.get('registration_number'):
            client_lines.append(f"Reg No: {cdict.get('registration_number')}")
        if cdict.get('vat_number'):
            client_lines.append(f"VAT No: {cdict.get('vat_number')}")
        if cdict.get('address'):
            client_lines.extend([line for line in str(cdict.get('address')).splitlines() if line.strip()])
    additional_info = s_dict.get('invoice_additional_info', '') or ''
    if project_note:
        additional_info = project_note + ('\n\n' + additional_info if additional_info.strip() else '')
    payload = {
        'kind': kind,
        'title': title,
        'number': formatted_num,
        'date': doc['date'],
        'meta_label': meta_label,
        'meta_value': meta_value,
        'company': company,
        'client_name': client_full,
        'client_email': client_email,
        'client_lines': client_lines,
        'items': [dict(i) for i in items],
        'totals_rows': totals_rows,
        'additional_info': additional_info,
        'filename': f"{title.replace(' ', '_')}_{formatted_num}.pdf"
    }
    conn.close()
    return payload




def _jpeg_dimensions(data):
    try:
        i = 2
        while i < len(data):
            if data[i] != 0xFF:
                i += 1
                continue
            marker = data[i + 1]
            i += 2
            if marker in (0xD8, 0xD9):
                continue
            if i + 2 > len(data):
                break
            length = int.from_bytes(data[i:i+2], 'big')
            if marker in (0xC0, 0xC1, 0xC2, 0xC3):
                precision = data[i+2]
                height = int.from_bytes(data[i+3:i+5], 'big')
                width = int.from_bytes(data[i+5:i+7], 'big')
                components = data[i+7]
                return width, height, components
            i += length
    except Exception:
        return None
    return None


def _png_unfilter(raw, width, height, bpp):
    row_len = width * bpp
    rows = []
    pos = 0
    prev = bytearray(row_len)
    for _ in range(height):
        filter_type = raw[pos]
        pos += 1
        scan = bytearray(raw[pos:pos + row_len])
        pos += row_len
        recon = bytearray(row_len)
        for i, val in enumerate(scan):
            left = recon[i - bpp] if i >= bpp else 0
            up = prev[i]
            up_left = prev[i - bpp] if i >= bpp else 0
            if filter_type == 0:
                recon[i] = val
            elif filter_type == 1:
                recon[i] = (val + left) & 0xff
            elif filter_type == 2:
                recon[i] = (val + up) & 0xff
            elif filter_type == 3:
                recon[i] = (val + ((left + up) // 2)) & 0xff
            elif filter_type == 4:
                p = left + up - up_left
                pa = abs(p - left)
                pb = abs(p - up)
                pc = abs(p - up_left)
                pr = left if pa <= pb and pa <= pc else (up if pb <= pc else up_left)
                recon[i] = (val + pr) & 0xff
            else:
                recon[i] = val
        rows.append(bytes(recon))
        prev = recon
    return b''.join(rows)


def _load_pdf_image(path):
    try:
        import zlib, struct
        with open(path, 'rb') as f:
            data = f.read()
        if data[:2] == b'\xff\xd8':
            dims = _jpeg_dimensions(data)
            if not dims:
                return None
            w, h, comps = dims
            colorspace = '/DeviceGray' if comps == 1 else '/DeviceRGB'
            return {'width': w, 'height': h, 'data': data, 'filter': '/DCTDecode', 'colorspace': colorspace, 'bits': 8}
        if data[:8] != b'\x89PNG\r\n\x1a\n':
            return None
        pos = 8
        width = height = bit_depth = color_type = None
        idat = bytearray()
        while pos + 8 <= len(data):
            length = struct.unpack('>I', data[pos:pos+4])[0]
            ctype = data[pos+4:pos+8]
            chunk = data[pos+8:pos+8+length]
            pos += 12 + length
            if ctype == b'IHDR':
                width, height, bit_depth, color_type, comp, filt, interlace = struct.unpack('>IIBBBBB', chunk)
                if comp != 0 or filt != 0 or interlace != 0 or bit_depth != 8:
                    return None
            elif ctype == b'IDAT':
                idat.extend(chunk)
            elif ctype == b'IEND':
                break
        if not width or not height or not idat:
            return None
        if color_type == 6:
            bpp = 4
            decoded = _png_unfilter(zlib.decompress(bytes(idat)), width, height, bpp)
            rgb = bytearray()
            alpha = bytearray()
            for i in range(0, len(decoded), 4):
                rgb.extend(decoded[i:i+3])
                alpha.append(decoded[i+3])
            image = {'width': width, 'height': height, 'data': zlib.compress(bytes(rgb)), 'filter': '/FlateDecode', 'colorspace': '/DeviceRGB', 'bits': 8}
            if any(a != 255 for a in alpha):
                image['smask'] = {'width': width, 'height': height, 'data': zlib.compress(bytes(alpha)), 'filter': '/FlateDecode', 'colorspace': '/DeviceGray', 'bits': 8}
            return image
        if color_type == 2:
            decoded = _png_unfilter(zlib.decompress(bytes(idat)), width, height, 3)
            return {'width': width, 'height': height, 'data': zlib.compress(decoded), 'filter': '/FlateDecode', 'colorspace': '/DeviceRGB', 'bits': 8}
        if color_type == 0:
            decoded = _png_unfilter(zlib.decompress(bytes(idat)), width, height, 1)
            return {'width': width, 'height': height, 'data': zlib.compress(decoded), 'filter': '/FlateDecode', 'colorspace': '/DeviceGray', 'bits': 8}
    except Exception:
        return None
    return None


# Cache decoded logo images in memory so each PDF request does not repeatedly
# read and decode the same logo file. The cache automatically refreshes when
# the file changes on disk.
_PDF_IMAGE_CACHE = {}


def _get_cached_pdf_image(path):
    try:
        stat = os.stat(path)
        key = (os.path.abspath(path), stat.st_mtime_ns, stat.st_size)
        cached = _PDF_IMAGE_CACHE.get(os.path.abspath(path))
        if cached and cached.get('key') == key:
            return cached.get('image')
        image = _load_pdf_image(path)
        if image:
            if len(_PDF_IMAGE_CACHE) > 24:
                _PDF_IMAGE_CACHE.clear()
            _PDF_IMAGE_CACHE[os.path.abspath(path)] = {'key': key, 'image': image}
        return image
    except Exception:
        return None


def _draw_pdf_document(payload):
    page_w, page_h = 595.28, 841.89
    margin = 36.0
    blue = (0.05, 0.62, 0.75)
    dark = (0.12, 0.12, 0.12)
    grey = (0.45, 0.45, 0.45)
    light = (0.95, 0.96, 0.97)
    cmds = []
    pages = []
    page_no = 0

    def cmd(line):
        cmds.append(line)

    def color(c, op='rg'):
        return f"{c[0]:.3f} {c[1]:.3f} {c[2]:.3f} {op}"

    def _text_width(value, size=9, bold=False):
        value = str(value or '')
        # Approximate Helvetica/Helvetica-Bold widths closely enough for reliable
        # right-alignment of totals, dates, document numbers and titles.
        total = 0.0
        for ch in value:
            if ch.isdigit():
                total += 0.556
            elif ch in ',.':
                total += 0.278 if not bold else 0.333
            elif ch in ' -/:':
                total += 0.278 if not bold else 0.333
            elif ch in 'ilI':
                total += 0.240 if not bold else 0.300
            elif ch in 'mwMW':
                total += 0.800 if not bold else 0.900
            else:
                total += 0.520 if not bold else 0.600
        return total * size

    def text(x, y, value, size=9, bold=False, c=dark, align='left'):
        value = str(value or '').replace('\u00a0', ' ')
        text_w = _text_width(value, size, bold)
        if align == 'right':
            x = x - text_w
        elif align == 'center':
            x = x - text_w / 2
        font = 'F2' if bold else 'F1'
        cmd(f"{color(c)} BT /{font} {size:.2f} Tf {x:.2f} {y:.2f} Td ({_pdf_text_escape(value)}) Tj ET")

    def rect(x, y, w, h, stroke=(0,0,0), fill=None):
        if fill is not None:
            cmd(f"{color(fill)} {x:.2f} {y:.2f} {w:.2f} {h:.2f} re f")
        if stroke is not None:
            cmd(f"{color(stroke, 'RG')} {x:.2f} {y:.2f} {w:.2f} {h:.2f} re S")

    def line(x1, y1, x2, y2, stroke=(0,0,0), width=0.5):
        cmd(f"{width:.2f} w {color(stroke, 'RG')} {x1:.2f} {y1:.2f} m {x2:.2f} {y2:.2f} l S")

    table_x = margin
    cols = [70, 245, 40, 75, 82]
    col_x = [table_x]
    for w in cols[:-1]:
        col_x.append(col_x[-1] + w)
    table_w = sum(cols)

    image_resources = {}
    logo_resource_name = None
    logo_info = None
    try:
        logo_file = (payload.get('company') or {}).get('logo_file') or ''
        if logo_file:
            logo_path = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), 'logos', os.path.basename(str(logo_file)))
            logo_info = _get_cached_pdf_image(logo_path) if os.path.exists(logo_path) else None
            if logo_info:
                logo_resource_name = 'ImLogo'
                image_resources[logo_resource_name] = logo_info
    except Exception:
        logo_resource_name = None
        logo_info = None

    def draw_image(name, x, y, w, h):
        cmd(f"q {w:.2f} 0 0 {h:.2f} {x:.2f} {y:.2f} cm /{name} Do Q")

    def table_header(y):
        rect(table_x, y - 18, table_w, 18, stroke=(0.75,0.75,0.75), fill=blue)
        text(col_x[0] + 4, y - 12, 'Date', 8, True, (1,1,1))
        text(col_x[1] + 4, y - 12, 'Description', 8, True, (1,1,1))
        text(col_x[2] + cols[2] - 4, y - 12, 'Qty', 8, True, (1,1,1), 'right')
        text(col_x[3] + cols[3] - 4, y - 12, 'Unit Price', 8, True, (1,1,1), 'right')
        text(col_x[4] + cols[4] - 4, y - 12, 'Line Total', 8, True, (1,1,1), 'right')
        for x in col_x[1:]:
            line(x, y - 18, x, y, stroke=(0.85,0.85,0.85), width=0.4)
        return y - 18

    def footer():
        nonlocal page_no
        line(margin, 30, page_w - margin, 30, stroke=(0.82,0.82,0.82), width=0.4)
        text(margin, 18, payload['company'].get('name', ''), 7, False, grey)
        text(page_w - margin, 18, f"Page {page_no}", 7, False, grey, 'right')

    def new_page(first=False):
        nonlocal cmds, page_no
        if cmds:
            footer()
            pages.append('\n'.join(cmds))
            cmds = []
        page_no += 1
        if first:
            return draw_first_header()
        text(page_w - margin - 8, page_h - 50, payload['title'] + ' (continued)', 14, True, blue, 'right')
        text(page_w - margin - 8, page_h - 66, payload['number'], 9, True, dark, 'right')
        return table_header(page_h - 92)

    def draw_first_header():
        company = payload['company']
        y_top = page_h - margin
        # Move the left branding block roughly two text lines higher while keeping
        # the right document title safely inside the page.
        left_y_top = min(page_h - 18, y_top + 18)
        company_x = margin
        y = left_y_top - 10
        if logo_resource_name and logo_info:
            max_logo_w, max_logo_h = 128.8, 67.2
            ratio = min(max_logo_w / max(float(logo_info.get('width') or 1), 1), max_logo_h / max(float(logo_info.get('height') or 1), 1))
            logo_w = max(1.0, float(logo_info.get('width') or 1) * ratio)
            logo_h = max(1.0, float(logo_info.get('height') or 1) * ratio)
            logo_y = left_y_top - logo_h - 2
            draw_image(logo_resource_name, margin, logo_y, logo_w, logo_h)
            y = logo_y - 12
        text(company_x, y, company.get('name', 'Company Name'), 14, True, dark)
        y -= 16
        for line_text in [
            f"Reg No: {company.get('registration_number')}" if company.get('registration_number') else '',
            company.get('address') or '',
            f"VAT No: {company.get('vat_number')}" if company.get('vat_number') else ''
        ]:
            if not line_text:
                continue
            for part in str(line_text).splitlines():
                text(company_x, y, part, 8, False, grey)
                y -= 10
        title_x = page_w - margin - 18
        title_size = 19 if len(str(payload.get('title') or '')) > 11 else 20
        text(title_x, y_top - 8, payload['title'], title_size, True, blue, 'right')
        text(title_x, y_top - 30, 'No: ' + payload['number'], 9, True, dark, 'right')
        text(title_x, y_top - 44, 'Date: ' + str(payload['date'] or ''), 9, False, dark, 'right')
        text(title_x, y_top - 58, payload['meta_label'] + ': ' + str(payload['meta_value'] or ''), 9, False, dark, 'right')
        header_line_y = min(y_top - 76, y - 8)
        line(margin, header_line_y, page_w - margin, header_line_y, stroke=blue, width=1.2)

        # Build a dynamic Billed To / Quoted To box. Some clients have company,
        # registration, VAT and address lines, so the box must expand and push
        # the table down instead of allowing text to overlap the border.
        client_label = 'QUOTED TO:' if payload.get('kind') == 'quote' else 'BILLED TO:'
        client_name_lines = _wrap_pdf_text(payload.get('client_name') or '', 62)
        client_extra_lines = []
        for cl in payload.get('client_lines', []) or []:
            client_extra_lines.extend(_wrap_pdf_text(cl, 72) if cl else [''])
        client_extra_lines = [line_text for line_text in client_extra_lines if str(line_text).strip()]

        label_h = 13
        name_line_h = 11
        extra_line_h = 9
        box_top_pad = 12
        box_bottom_pad = 10
        box_h = max(52, box_top_pad + label_h + (len(client_name_lines) * name_line_h) + (len(client_extra_lines) * extra_line_h) + box_bottom_pad)
        box_top = header_line_y - 22
        box_y = box_top - box_h
        rect(margin, box_y, table_w, box_h, stroke=(0.85,0.85,0.85), fill=light)

        cy = box_top - 15
        text(margin + 8, cy, client_label, 7, True, grey)
        cy -= 14
        for name_line in client_name_lines:
            text(margin + 8, cy, name_line, 10, True, dark)
            cy -= name_line_h
        for cl in client_extra_lines:
            text(margin + 8, cy, cl, 8, False, dark)
            cy -= extra_line_h
        return table_header(box_y - 18)

    y = new_page(first=True)
    bottom = 78
    for item in payload.get('items') or []:
        desc_lines = _wrap_pdf_text(item.get('description', ''), 46)
        row_h = max(22, 10 + len(desc_lines) * 10)
        if y - row_h < bottom:
            y = new_page(first=False)
        rect(table_x, y - row_h, table_w, row_h, stroke=(0.80,0.80,0.80), fill=None)
        for x in col_x[1:]:
            line(x, y - row_h, x, y, stroke=(0.85,0.85,0.85), width=0.4)
        text(col_x[0] + 4, y - 14, item.get('service_date', '') or '', 8, False, dark)
        dy = y - 14
        for dl in desc_lines:
            text(col_x[1] + 4, dy, dl, 8, False, dark)
            dy -= 10
        qty = _qty_text(item.get('quantity', 1))
        amount = float(item.get('amount') or 0)
        try:
            qty_float = float(item.get('quantity') if item.get('quantity') not in (None, '') else 1)
        except Exception:
            qty_float = 1.0
        try:
            unit_price = float(item.get('unit_price') if item.get('unit_price') not in (None, '') else (amount / qty_float if qty_float else amount))
        except Exception:
            unit_price = amount
        text(col_x[2] + cols[2] - 4, y - 14, qty, 8, False, dark, 'right')
        text(col_x[3] + cols[3] - 4, y - 14, _money(unit_price), 8, False, dark, 'right')
        text(col_x[4] + cols[4] - 4, y - 14, _money(amount), 8, False, dark, 'right')
        y -= row_h

    totals_h = 18 * len(payload.get('totals_rows') or []) + 14
    if y - totals_h < bottom:
        y = new_page(first=False)
    totals_w = 210
    # Keep the totals block aligned with the line-item table but nudge it slightly right
    # so Subtotal / VAT / TOTAL sit neatly inside the table area.
    tx = page_w - margin - totals_w - 12
    y -= 12
    for label, value in payload.get('totals_rows') or []:
        is_total = 'TOTAL' in label.upper() or 'BALANCE' in label.upper()
        if is_total:
            line(tx, y, tx + totals_w, y, stroke=(0.70,0.70,0.70), width=0.8)
        text(tx + 4, y - 12, label, 8.5, is_total, dark)
        try:
            val = float(value or 0)
            val_text = ('- ' if val < 0 else '') + _money(abs(val))
        except Exception:
            val_text = _money(value)
        text(tx + totals_w - 4, y - 12, val_text, 8.5, is_total, dark, 'right')
        y -= 18

    info = (payload.get('additional_info') or '').strip()
    if info:
        lines = []
        for part in info.splitlines():
            lines.extend(_wrap_pdf_text(part, 90) if part else [''])
        info_h = min(160, 14 + len(lines) * 9)
        if y - info_h < bottom:
            y = new_page(first=False)
        rect(margin, y - info_h, table_w, info_h, stroke=(0.85,0.85,0.85), fill=light)
        iy = y - 14
        for line_text in lines[:16]:
            text(margin + 8, iy, line_text, 7.5, False, grey)
            iy -= 9
        y -= info_h

    if y - 45 < bottom:
        y = new_page(first=False)
    line(margin, y - 15, page_w - margin, y - 15, stroke=(0.85,0.85,0.85), width=0.5)
    text(page_w / 2, y - 33, 'Thank you for your business!', 8, False, grey, 'center')
    text(page_w / 2, y - 45, payload['company'].get('name', ''), 8, False, grey, 'center')
    footer()
    pages.append('\n'.join(cmds))
    return _build_raw_pdf(pages, page_w, page_h, image_resources)


def _build_raw_pdf(page_streams, page_w=595.28, page_h=841.89, image_resources=None):
    objects = []
    def add(obj):
        objects.append(obj)
        return len(objects)
    font_regular = add(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    font_bold = add(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")
    image_ids = {}
    image_resources = image_resources or {}
    for name, img in image_resources.items():
        smask_ref = ''
        smask = img.get('smask') if isinstance(img, dict) else None
        if smask:
            smask_dict = f"<< /Type /XObject /Subtype /Image /Width {int(smask['width'])} /Height {int(smask['height'])} /ColorSpace {smask['colorspace']} /BitsPerComponent {int(smask.get('bits', 8))} /Filter {smask['filter']} /Length {len(smask['data'])} >>".encode('ascii')
            smask_obj = smask_dict + b"\nstream\n" + smask['data'] + b"\nendstream"
            smask_id = add(smask_obj)
            smask_ref = f" /SMask {smask_id} 0 R"
        img_dict = f"<< /Type /XObject /Subtype /Image /Width {int(img['width'])} /Height {int(img['height'])} /ColorSpace {img['colorspace']} /BitsPerComponent {int(img.get('bits', 8))} /Filter {img['filter']}{smask_ref} /Length {len(img['data'])} >>".encode('ascii')
        img_obj = img_dict + b"\nstream\n" + img['data'] + b"\nendstream"
        image_ids[name] = add(img_obj)
    page_ids = []
    xobject_resource = ''
    if image_ids:
        xobject_resource = ' /XObject << ' + ' '.join(f'/{name} {obj_id} 0 R' for name, obj_id in image_ids.items()) + ' >>'
    for stream in page_streams:
        data = stream.encode('latin-1', 'replace')
        stream_id = add(b"<< /Length " + str(len(data)).encode('ascii') + b" >>\nstream\n" + data + b"\nendstream")
        page_obj = f"<< /Type /Page /Parent {{PAGES}} 0 R /MediaBox [0 0 {page_w:.2f} {page_h:.2f}] /Resources << /Font << /F1 {font_regular} 0 R /F2 {font_bold} 0 R >>{xobject_resource} >> /Contents {stream_id} 0 R >>".encode('ascii')
        page_ids.append(add(page_obj))
    kids = ' '.join(f'{pid} 0 R' for pid in page_ids)
    pages_id = add(f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>".encode('ascii'))
    for pid in page_ids:
        objects[pid-1] = objects[pid-1].replace(b'{PAGES}', str(pages_id).encode('ascii'))
    catalog_id = add(f"<< /Type /Catalog /Pages {pages_id} 0 R >>".encode('ascii'))
    output = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{idx} 0 obj\n".encode('ascii'))
        output.extend(obj)
        output.extend(b"\nendobj\n")
    xref_pos = len(output)
    output.extend(f"xref\n0 {len(objects)+1}\n".encode('ascii'))
    output.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        output.extend(f"{off:010d} 00000 n \n".encode('ascii'))
    output.extend(f"trailer\n<< /Size {len(objects)+1} /Root {catalog_id} 0 R >>\nstartxref\n{xref_pos}\n%%EOF\n".encode('ascii'))
    return bytes(output)


@app.route('/download/invoice/<int:inv_id>.pdf')
def download_invoice_pdf(inv_id):
    if not session.get('can_invoicing') and not session.get('is_superadmin'):
        return "Forbidden", 403
    payload = _build_billing_pdf_payload('invoice', inv_id)
    if not payload:
        return "Invoice not found", 404
    pdf_bytes = _draw_pdf_document(payload)
    log_action('Invoicing', 'Downloaded Invoice PDF', payload['number'])
    return Response(pdf_bytes, mimetype='application/pdf', headers={'Content-Disposition': f"attachment; filename={payload['filename']}"})


@app.route('/download/quote/<int:q_id>.pdf')
def download_quote_pdf(q_id):
    if not session.get('can_invoicing') and not session.get('is_superadmin'):
        return "Forbidden", 403
    payload = _build_billing_pdf_payload('quote', q_id)
    if not payload:
        return "Quote not found", 404
    pdf_bytes = _draw_pdf_document(payload)
    log_action('Invoicing', 'Downloaded Quote PDF', payload['number'])
    return Response(pdf_bytes, mimetype='application/pdf', headers={'Content-Disposition': f"attachment; filename={payload['filename']}"})


@app.route('/api/email_invoice_pdf/<int:inv_id>', methods=['POST'])
def email_invoice_pdf(inv_id):
    if not session.get('can_invoicing') and not session.get('is_superadmin'):
        return jsonify({'message': 'Forbidden'}), 403
    data = request.get_json(silent=True) or {}
    email = data.get('email')
    if not email:
        return jsonify({'message': 'Missing email address.'}), 400
    payload = _build_billing_pdf_payload('invoice', inv_id)
    if not payload:
        return jsonify({'message': 'Invoice not found.'}), 404
    pdf_bytes = _draw_pdf_document(payload)
    result, status = _send_pdf_email(pdf_bytes, payload['filename'], email, payload['title'] + '_' + payload['number'], payload.get('client_name') or 'Client')
    return jsonify(result), status


@app.route('/api/email_quote_pdf/<int:q_id>', methods=['POST'])
def email_quote_pdf(q_id):
    if not session.get('can_invoicing') and not session.get('is_superadmin'):
        return jsonify({'message': 'Forbidden'}), 403
    data = request.get_json(silent=True) or {}
    email = data.get('email')
    if not email:
        return jsonify({'message': 'Missing email address.'}), 400
    payload = _build_billing_pdf_payload('quote', q_id)
    if not payload:
        return jsonify({'message': 'Quote not found.'}), 404
    pdf_bytes = _draw_pdf_document(payload)
    result, status = _send_pdf_email(pdf_bytes, payload['filename'], email, payload['title'] + '_' + payload['number'], payload.get('client_name') or 'Client')
    return jsonify(result), status


@app.route('/api/email_document', methods=['POST'])
def email_document():
    if not session.get('can_invoicing') and not session.get('is_superadmin'): return jsonify({"message": "Forbidden"}), 403
    
    email = request.form.get('email')
    doc_name = request.form.get('doc_name')
    client_name = request.form.get('client_name')
    pdf_file = request.files.get('pdf')
    
    if not pdf_file or not email: 
        return jsonify({"message": "Missing PDF file or Email address."}), 400
        
    conn = get_db_connection()
    settings_rows = conn.execute('SELECT key, value FROM settings WHERE company_id=?', (session['company_id'],)).fetchall()
    conn.close()
    
    s_dict = {r['key']: r['value'] for r in settings_rows}
    if not s_dict.get('smtp_server') or not s_dict.get('smtp_user') or not s_dict.get('smtp_pass'): 
        return jsonify({"message": "SMTP settings are incomplete. Please configure your Email & Calendar Settings in the Hub."}), 400
        
    try:
        msg = EmailMessage()
        msg['Subject'] = f"{doc_name.replace('_', ' ')} from {session.get('company_name')}"
        msg['From'] = s_dict.get('sender_email', s_dict.get('smtp_user'))
        msg['To'] = email
        msg.set_content(f"Dear {client_name},\n\nPlease find attached your {doc_name.replace('_', ' ')}.\n\nThank you for your business.\n\nKind regards,\n{session.get('company_name')}")
        msg.add_attachment(pdf_file.read(), maintype='application', subtype='pdf', filename=f"{doc_name}.pdf")
        
        port = int(s_dict.get('smtp_port', 465))
        if port == 587:
            with smtplib.SMTP(s_dict['smtp_server'], port, timeout=10) as server:
                server.starttls()
                server.login(s_dict['smtp_user'], s_dict['smtp_pass'])
                server.send_message(msg)
        else:
            with smtplib.SMTP_SSL(s_dict['smtp_server'], port, timeout=10) as server:
                server.login(s_dict['smtp_user'], s_dict['smtp_pass'])
                server.send_message(msg)
            
        log_action('Invoicing', 'Emailed Document', f"Sent {doc_name} to {email}")
        return jsonify({"message": "Email sent successfully!"})
    except Exception as e: 
        return jsonify({"message": f"Error sending email: {str(e)}"}), 500

# --- NEW: CLIENT STATEMENT ROUTE ---
@app.route('/api/client_statement', methods=['POST'])
def client_statement():
    if not session.get('can_invoicing') and not session.get('is_superadmin'): return jsonify({"message": "Forbidden"}), 403
    
    data = request.json
    client_name = data.get('client_name')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    try:
        client_id = int(data.get('client_id') or 0)
    except Exception:
        client_id = 0
    cid = session['company_id']
    
    conn = get_db_connection()
    client = get_client_by_id(conn, cid, client_id) if client_id else find_client_by_display_name(conn, cid, client_name)
    if client:
        client_id = client['id']
        client_name = client_display_name(client)
        query = "SELECT * FROM invoices WHERE client_id=? AND company_id=? AND date >= ? AND date <= ? ORDER BY date ASC"
        invoices = conn.execute(query, (client_id, cid, start_date, end_date)).fetchall()
    else:
        query = "SELECT * FROM invoices WHERE client_name=? AND company_id=? AND date >= ? AND date <= ? ORDER BY date ASC"
        invoices = conn.execute(query, (client_name, cid, start_date, end_date)).fetchall()
    
    settings_rows = conn.execute("SELECT key, value FROM settings WHERE company_id=?", (cid,)).fetchall()
    s_dict = {s['key']: s['value'] for s in settings_rows}
    inv_prefix = s_dict.get('invoice_prefix', 'INV-')
    try:
        inv_start = int(s_dict.get('invoice_start', '1'))
    except:
        inv_start = 1
        
    statement_items = []
    total_invoiced = 0.0
    total_paid = 0.0
    
    for inv in invoices:
        d = dict(inv)
        if d['status'] == 'Credited':
            continue
            
        try:
            formatted_num = f"{inv_prefix}{int(inv_start) + d['id'] - 1:04d}"
        except:
            formatted_num = f"{inv_prefix}{d['id']:04d}"
            
        amt = float(d['total'])
        
        total_invoiced += amt
        if d['status'] == 'Paid':
            total_paid += amt
            
        statement_items.append({
            "date": d['date'],
            "description": f"Tax Invoice #{formatted_num}",
            "amount": amt,
            "status": d['status']
        })
        
    total_due = total_invoiced - total_paid
    
    client_full = client_name
    client_email = ''
    client_dict = {}
    if client:
        client_dict = dict(client)
        if 'surname' in client_dict and client_dict['surname']:
            client_full = f"{client_dict['name']} {client_dict['surname']}"
        if 'email' in client_dict and client_dict['email']:
            client_email = client_dict['email']
            
    conn.close()
    
    return jsonify({
        "client_full_name": client_full,
        "client_email": client_email,
        "client_address": client_dict.get('address') if client_dict.get('address') else '',
        "client_company": client_dict.get('company_name') if client_dict.get('company_name') else '',
        "client_reg": client_dict.get('registration_number') if client_dict.get('registration_number') else '',
        "client_vat": client_dict.get('vat_number') if client_dict.get('vat_number') else '',
        "items": statement_items,
        "total_invoiced": total_invoiced,
        "total_paid": total_paid,
        "total_due": total_due,
        "start_date": start_date,
        "end_date": end_date
    })


# ==========================================================
# ACCOUNTING MODULE - IFRS FOR SMEs FOUNDATION
# ==========================================================
DEFAULT_ACCOUNTING_ACCOUNTS = [
    ('1000', 'Bank', 'asset', 'current_asset', 'debit', 'operating', 1),
    ('1100', 'Accounts Receivable', 'asset', 'current_asset', 'debit', 'operating', 0),
    ('1200', 'Prepaid Expenses', 'asset', 'current_asset', 'debit', 'operating', 0),
    ('1500', 'Property, Plant and Equipment', 'asset', 'non_current_asset', 'debit', 'investing', 0),
    ('1600', 'Accumulated Depreciation', 'asset', 'non_current_asset', 'credit', 'investing', 0),
    ('2000', 'Accounts Payable', 'liability', 'current_liability', 'credit', 'operating', 0),
    ('2100', 'VAT Control', 'liability', 'current_liability', 'credit', 'operating', 0),
    ('2200', 'Payroll Liabilities', 'liability', 'current_liability', 'credit', 'operating', 0),
    ('2500', 'Loan Payable', 'liability', 'non_current_liability', 'credit', 'financing', 0),
    ('3000', 'Owner / Member Equity', 'equity', 'equity', 'credit', 'financing', 0),
    ('3100', 'Retained Earnings', 'equity', 'equity', 'credit', 'financing', 0),
    ('4000', 'Revenue', 'income', 'revenue', 'credit', 'operating', 0),
    ('4100', 'Other Income', 'income', 'other_income', 'credit', 'operating', 0),
    ('5000', 'Cost of Sales', 'cost_of_sales', 'cost_of_sales', 'debit', 'operating', 0),
    ('6000', 'Salaries and Wages', 'expense', 'operating_expense', 'debit', 'operating', 0),
    ('6100', 'Rent Expense', 'expense', 'operating_expense', 'debit', 'operating', 0),
    ('6200', 'Materials and Supplies', 'expense', 'operating_expense', 'debit', 'operating', 0),
    ('6300', 'Transport Expense', 'expense', 'operating_expense', 'debit', 'operating', 0),
    ('6800', 'Finance Costs', 'expense', 'finance_cost', 'debit', 'financing', 0),
    ('6900', 'Income Tax Expense', 'expense', 'income_tax_expense', 'debit', 'operating', 0),
]

IFRS_SECTION_LABELS = {
    'current_asset': 'Current Assets',
    'non_current_asset': 'Non-current Assets',
    'current_liability': 'Current Liabilities',
    'non_current_liability': 'Non-current Liabilities',
    'equity': 'Equity',
    'revenue': 'Revenue',
    'other_income': 'Other Income',
    'cost_of_sales': 'Cost of Sales',
    'operating_expense': 'Operating Expenses',
    'finance_cost': 'Finance Costs',
    'income_tax_expense': 'Income Tax Expense'
}


def _money_float(value):
    try:
        return round(float(value or 0), 2)
    except Exception:
        return 0.0


def _current_company_id():
    return int(session.get('company_id') or 0)


def ensure_default_accounting_accounts(conn, company_id):
    existing = conn.execute('SELECT COUNT(*) FROM accounting_accounts WHERE company_id=?', (company_id,)).fetchone()[0]
    if not existing:
        for code, name, acc_type, section, normal, cf_category, is_cash in DEFAULT_ACCOUNTING_ACCOUNTS:
            conn.execute('''INSERT OR IGNORE INTO accounting_accounts
                            (company_id, account_code, account_name, account_type, report_section, normal_balance, cash_flow_category, is_cash_equivalent, active)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)''',
                         (company_id, code, name, acc_type, section, normal, cf_category, is_cash))
    conn.execute('''INSERT OR IGNORE INTO accounting_settings (company_id, reporting_framework, financial_year_end_month)
                    VALUES (?, 'IFRS for SMEs', 2)''', (company_id,))


def _ensure_account_by_code(conn, company_id, code, name, acc_type, section, normal, cf_category='operating', is_cash=0):
    row = conn.execute('SELECT id FROM accounting_accounts WHERE company_id=? AND account_code=?', (company_id, code)).fetchone()
    if row:
        return int(row['id'])
    cur = conn.execute('''INSERT INTO accounting_accounts
                          (company_id, account_code, account_name, account_type, report_section, normal_balance, cash_flow_category, is_cash_equivalent, active)
                          VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)''',
                       (company_id, code, name, acc_type, section, normal, cf_category, is_cash))
    return int(cur.lastrowid)


def ensure_accounting_posting_defaults(conn, company_id):
    """Ensure invoice/credit note posting accounts exist and are linked in accounting settings."""
    ensure_default_accounting_accounts(conn, company_id)
    ar_id = _ensure_account_by_code(conn, company_id, '1100', 'Accounts Receivable', 'asset', 'current_asset', 'debit', 'operating', 0)
    sales_id = _ensure_account_by_code(conn, company_id, '4000', 'Revenue', 'income', 'revenue', 'credit', 'operating', 0)
    vat_id = _ensure_account_by_code(conn, company_id, '2100', 'VAT Control', 'liability', 'current_liability', 'credit', 'operating', 0)
    credit_id = _ensure_account_by_code(conn, company_id, '4050', 'Credit Notes / Sales Returns', 'income', 'revenue', 'debit', 'operating', 0)
    discount_id = _ensure_account_by_code(conn, company_id, '4060', 'Discount Allowed', 'income', 'revenue', 'debit', 'operating', 0)
    rounding_id = _ensure_account_by_code(conn, company_id, '6990', 'Rounding Differences', 'expense', 'operating_expense', 'debit', 'operating', 0)
    conn.execute('''INSERT OR IGNORE INTO accounting_settings (company_id, reporting_framework, financial_year_end_month)
                    VALUES (?, 'IFRS for SMEs', 2)''', (company_id,))
    current = conn.execute('SELECT * FROM accounting_settings WHERE company_id=?', (company_id,)).fetchone()
    vals = dict(current) if current else {}
    updates = {
        'receivables_account_id': ar_id,
        'sales_revenue_account_id': sales_id,
        'vat_output_account_id': vat_id,
        'vat_control_account_id': vat_id,
        'credit_notes_account_id': credit_id,
        'discount_account_id': discount_id,
        'rounding_account_id': rounding_id,
    }
    for col, val in updates.items():
        if not vals.get(col):
            try:
                conn.execute(f'UPDATE accounting_settings SET {col}=? WHERE company_id=?', (val, company_id))
            except Exception:
                pass
    return {k: int(v) for k, v in updates.items()}


def _accounting_posting_settings(conn, company_id):
    defaults = ensure_accounting_posting_defaults(conn, company_id)
    row = conn.execute('SELECT * FROM accounting_settings WHERE company_id=?', (company_id,)).fetchone()
    vals = dict(row) if row else {}
    out = {}
    for key, default_value in defaults.items():
        try:
            out[key] = int(vals.get(key) or default_value)
        except Exception:
            out[key] = int(default_value)
    return out


def _accounting_account_exists(conn, company_id, account_id):
    if not account_id:
        return False
    row = conn.execute('SELECT id FROM accounting_accounts WHERE id=? AND company_id=? AND active=1', (account_id, company_id)).fetchone()
    return bool(row)


def _invoice_formatted_number(conn, company_id, inv_id):
    rows = conn.execute('SELECT key, value FROM settings WHERE company_id=?', (company_id,)).fetchall()
    s_dict = {r['key']: r['value'] for r in rows}
    prefix = s_dict.get('invoice_prefix', 'INV-')
    start = s_dict.get('invoice_start', '1')
    try:
        return f"{prefix}{int(start) + int(inv_id) - 1:04d}"
    except Exception:
        return f"{prefix}{int(inv_id):04d}"


def _credit_note_formatted_number(credit_id):
    try:
        return f"CN-{int(credit_id):04d}"
    except Exception:
        return f"CN-{credit_id}"


def _accounting_create_posted_journal(conn, company_id, journal_date, reference, description, source_module, source_record_type, source_record_id, lines):
    debit_total = round(sum(_money_float(l.get('debit')) for l in lines), 2)
    credit_total = round(sum(_money_float(l.get('credit')) for l in lines), 2)
    if abs(debit_total - credit_total) > 0.005:
        raise ValueError(f'Accounting entry is out of balance. Debit R{debit_total:.2f}, Credit R{credit_total:.2f}.')
    if debit_total <= 0:
        raise ValueError('Accounting entry amount must be greater than zero.')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    journal_date = _normalise_accounting_date(journal_date)
    if not _is_iso_accounting_date(journal_date):
        journal_date = datetime.now().strftime('%Y-%m-%d')
    cur = conn.execute('''INSERT INTO accounting_journals
                          (company_id, journal_date, reference, description, source_module, status, created_by, posted_by, posted_at, source_record_type, source_record_id)
                          VALUES (?, ?, ?, ?, ?, 'posted', ?, ?, ?, ?, ?)''',
                       (company_id, journal_date, reference, description, source_module, session.get('username'), session.get('username'), now, source_record_type, source_record_id))
    journal_id = cur.lastrowid
    for idx, line in enumerate(lines, start=1):
        account_id = int(line.get('account_id') or 0)
        if not _accounting_account_exists(conn, company_id, account_id):
            raise ValueError('One of the accounting accounts is invalid or inactive.')
        conn.execute('''INSERT INTO accounting_journal_lines
                        (company_id, journal_id, line_no, account_id, description, debit, credit, cash_flow_section)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                     (company_id, journal_id, idx, account_id, line.get('description') or description, _money_float(line.get('debit')), _money_float(line.get('credit')), line.get('cash_flow_section') or 'non_cash'))
    return int(journal_id)


def post_invoice_to_accounting_record(conn, company_id, inv_id):
    ensure_accounting_posting_defaults(conn, company_id)
    inv = conn.execute('SELECT * FROM invoices WHERE id=? AND company_id=?', (inv_id, company_id)).fetchone()
    if not inv:
        raise ValueError('Invoice not found.')
    inv_d = dict(inv)
    if inv_d.get('accounting_journal_id'):
        raise ValueError('This invoice has already been posted to Accounting.')
    total = _money_float(inv_d.get('total'))
    vat_amount = _money_float(inv_d.get('vat_amount'))
    revenue_amount = round(total - vat_amount, 2)
    if total <= 0 or revenue_amount < 0:
        raise ValueError('Invoice total is invalid and cannot be posted to Accounting.')
    settings = _accounting_posting_settings(conn, company_id)
    ref = _invoice_formatted_number(conn, company_id, inv_id)
    desc = f"Invoice {ref} - {inv_d.get('client_name') or 'Client'}"
    lines = [
        {'account_id': settings['receivables_account_id'], 'description': desc, 'debit': total, 'credit': 0, 'cash_flow_section': 'non_cash'},
    ]
    if revenue_amount > 0:
        lines.append({'account_id': settings['sales_revenue_account_id'], 'description': desc, 'debit': 0, 'credit': revenue_amount, 'cash_flow_section': 'non_cash'})
    if vat_amount > 0:
        lines.append({'account_id': settings['vat_output_account_id'], 'description': f"VAT Output - {ref}", 'debit': 0, 'credit': vat_amount, 'cash_flow_section': 'non_cash'})
    journal_id = _accounting_create_posted_journal(conn, company_id, inv_d.get('date'), ref, desc, 'invoice', 'invoice', inv_id, lines)
    conn.execute('''UPDATE invoices
                    SET accounting_status='posted', accounting_journal_id=?, accounting_posted_at=?, accounting_posted_by=?
                    WHERE id=? AND company_id=?''',
                 (journal_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), session.get('username'), inv_id, company_id))
    return journal_id


def post_credit_note_to_accounting_record(conn, company_id, credit_id):
    ensure_accounting_posting_defaults(conn, company_id)
    credit = conn.execute('SELECT * FROM invoice_credit_notes WHERE id=? AND company_id=?', (credit_id, company_id)).fetchone()
    if not credit:
        raise ValueError('Credit note not found.')
    cn = dict(credit)
    if cn.get('accounting_journal_id'):
        raise ValueError('This credit note has already been posted to Accounting.')
    inv = conn.execute('SELECT * FROM invoices WHERE id=? AND company_id=?', (cn.get('invoice_id'), company_id)).fetchone()
    if not inv:
        raise ValueError('Linked invoice not found for this credit note.')
    inv_d = dict(inv)
    amount = _money_float(cn.get('amount'))
    if amount <= 0:
        raise ValueError('Credit note amount must be greater than zero.')
    invoice_total = _money_float(inv_d.get('total'))
    invoice_vat = _money_float(inv_d.get('vat_amount'))
    vat_component = round(amount * (invoice_vat / invoice_total), 2) if invoice_total > 0 and invoice_vat > 0 else 0.0
    net_component = round(amount - vat_component, 2)
    settings = _accounting_posting_settings(conn, company_id)
    cn_ref = _credit_note_formatted_number(credit_id)
    inv_ref = _invoice_formatted_number(conn, company_id, inv_d.get('id'))
    desc = f"Credit Note {cn_ref} against {inv_ref} - {inv_d.get('client_name') or 'Client'}"
    lines = []
    if net_component > 0:
        lines.append({'account_id': settings['credit_notes_account_id'], 'description': desc, 'debit': net_component, 'credit': 0, 'cash_flow_section': 'non_cash'})
    if vat_component > 0:
        lines.append({'account_id': settings['vat_output_account_id'], 'description': f"VAT Output reversal - {cn_ref}", 'debit': vat_component, 'credit': 0, 'cash_flow_section': 'non_cash'})
    lines.append({'account_id': settings['receivables_account_id'], 'description': desc, 'debit': 0, 'credit': amount, 'cash_flow_section': 'non_cash'})
    journal_id = _accounting_create_posted_journal(conn, company_id, cn.get('credit_date'), cn_ref, desc, 'credit_note', 'credit_note', credit_id, lines)
    conn.execute('''UPDATE invoice_credit_notes
                    SET accounting_status='posted', accounting_journal_id=?, accounting_posted_at=?, accounting_posted_by=?
                    WHERE id=? AND company_id=?''',
                 (journal_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), session.get('username'), credit_id, company_id))
    return journal_id


def _accounting_accounts(conn, company_id, active_only=False):
    where = 'WHERE company_id=?'
    params = [company_id]
    if active_only:
        where += ' AND active=1'
    return [dict(a) for a in conn.execute(f'''SELECT * FROM accounting_accounts {where}
                                             ORDER BY account_code ASC, account_name ASC''', params).fetchall()]


def _posted_lines(conn, company_id, start_date=None, end_date=None):
    sql = '''SELECT l.*, a.account_code, a.account_name, a.account_type, a.report_section,
                    a.normal_balance, a.cash_flow_category, a.is_cash_equivalent, j.journal_date, j.reference, j.description AS journal_description
             FROM accounting_journal_lines l
             JOIN accounting_journals j ON j.id = l.journal_id
             JOIN accounting_accounts a ON a.id = l.account_id
             WHERE l.company_id=? AND j.company_id=? AND j.status='posted' '''
    params = [company_id, company_id]
    if start_date:
        sql += ' AND j.journal_date >= ?'
        params.append(start_date)
    if end_date:
        sql += ' AND j.journal_date <= ?'
        params.append(end_date)
    sql += ' ORDER BY j.journal_date, j.id, l.line_no'
    return [dict(r) for r in conn.execute(sql, params).fetchall()]


def _trial_balance_rows(conn, company_id, as_at):
    accounts = _accounting_accounts(conn, company_id, active_only=False)
    totals = {int(a['id']): {'debit': 0.0, 'credit': 0.0} for a in accounts}
    for line in _posted_lines(conn, company_id, None, as_at):
        aid = int(line['account_id'])
        totals.setdefault(aid, {'debit': 0.0, 'credit': 0.0})
        totals[aid]['debit'] += _money_float(line.get('debit'))
        totals[aid]['credit'] += _money_float(line.get('credit'))
    rows = []
    for a in accounts:
        t = totals.get(int(a['id']), {'debit': 0.0, 'credit': 0.0})
        net = round(t['debit'] - t['credit'], 2)
        rows.append({
            'account_id': a['id'],
            'account_code': a['account_code'],
            'account_name': a['account_name'],
            'account_type': a['account_type'],
            'report_section': a['report_section'],
            'debit': round(net, 2) if net >= 0 else 0.0,
            'credit': round(abs(net), 2) if net < 0 else 0.0,
            'raw_debit': round(t['debit'], 2),
            'raw_credit': round(t['credit'], 2)
        })
    return rows


def _income_statement_data(conn, company_id, start_date, end_date):
    sections = {'revenue': [], 'other_income': [], 'cost_of_sales': [], 'operating_expense': [], 'finance_cost': [], 'income_tax_expense': []}
    totals = {k: 0.0 for k in sections}
    balances = {}
    for line in _posted_lines(conn, company_id, start_date, end_date):
        aid = int(line['account_id'])
        rec = balances.setdefault(aid, dict(line, debit_total=0.0, credit_total=0.0))
        rec['debit_total'] += _money_float(line.get('debit'))
        rec['credit_total'] += _money_float(line.get('credit'))
    for rec in balances.values():
        section = rec.get('report_section') or ''
        if section not in sections:
            continue
        if rec.get('account_type') == 'income':
            amount = round(rec['credit_total'] - rec['debit_total'], 2)
        else:
            amount = round(rec['debit_total'] - rec['credit_total'], 2)
        if abs(amount) < 0.005:
            continue
        row = {'account_code': rec['account_code'], 'account_name': rec['account_name'], 'amount': amount}
        sections[section].append(row)
        totals[section] += amount
    revenue_total = round(totals['revenue'] + totals['other_income'], 2)
    gross_profit = round(totals['revenue'] - totals['cost_of_sales'], 2)
    expenses_total = round(totals['operating_expense'] + totals['finance_cost'] + totals['income_tax_expense'], 2)
    profit_loss = round(revenue_total - totals['cost_of_sales'] - expenses_total, 2)
    return {'sections': sections, 'totals': {k: round(v, 2) for k, v in totals.items()}, 'revenue_total': revenue_total, 'gross_profit': gross_profit, 'expenses_total': expenses_total, 'profit_loss': profit_loss}


def _balance_sheet_data(conn, company_id, as_at):
    tb = _trial_balance_rows(conn, company_id, as_at)
    sections = {'current_asset': [], 'non_current_asset': [], 'current_liability': [], 'non_current_liability': [], 'equity': []}
    totals = {k: 0.0 for k in sections}
    for row in tb:
        section = row.get('report_section') or ''
        if section not in sections:
            continue
        if row['account_type'] == 'asset':
            amount = round(float(row['debit']) - float(row['credit']), 2)
        else:
            amount = round(float(row['credit']) - float(row['debit']), 2)
        if abs(amount) < 0.005:
            continue
        out = {'account_code': row['account_code'], 'account_name': row['account_name'], 'amount': amount}
        sections[section].append(out)
        totals[section] += amount
    income = _income_statement_data(conn, company_id, '0001-01-01', as_at)
    current_profit = income.get('profit_loss', 0.0)
    if abs(current_profit) >= 0.005:
        sections['equity'].append({'account_code': '', 'account_name': 'Current period profit / (loss)', 'amount': current_profit})
        totals['equity'] += current_profit
    total_assets = round(totals['current_asset'] + totals['non_current_asset'], 2)
    total_liabilities = round(totals['current_liability'] + totals['non_current_liability'], 2)
    total_equity = round(totals['equity'], 2)
    return {'sections': sections, 'totals': {k: round(v, 2) for k, v in totals.items()}, 'total_assets': total_assets, 'total_liabilities': total_liabilities, 'total_equity': total_equity, 'liabilities_equity_total': round(total_liabilities + total_equity, 2), 'balance_check': round(total_assets - (total_liabilities + total_equity), 2)}




def _general_ledger_data(conn, company_id, start_date=None, end_date=None, account_id=None):
    """Return posted general ledger transactions with running balances per account."""
    ensure_default_accounting_accounts(conn, company_id)
    accounts = _accounting_accounts(conn, company_id, active_only=False)

    where = ['l.company_id=?', 'j.company_id=?', "j.status='posted'"]
    params = [company_id, company_id]
    if start_date:
        where.append('j.journal_date >= ?')
        params.append(start_date)
    if end_date:
        where.append('j.journal_date <= ?')
        params.append(end_date)
    if account_id:
        where.append('l.account_id = ?')
        params.append(int(account_id))

    sql = """SELECT l.*, a.account_code, a.account_name, a.account_type, a.report_section,
                    a.normal_balance, j.journal_date, j.reference, j.description AS journal_description,
                    j.source_module, j.status AS journal_status, j.source_record_type, j.source_record_id
             FROM accounting_journal_lines l
             JOIN accounting_journals j ON j.id = l.journal_id
             JOIN accounting_accounts a ON a.id = l.account_id
             WHERE {where_clause}
             ORDER BY a.account_code ASC, a.account_name ASC, j.journal_date ASC, j.id ASC, l.line_no ASC, l.id ASC""".format(where_clause=' AND '.join(where))
    raw_rows = [dict(r) for r in conn.execute(sql, params).fetchall()]

    # Supporting documents can be attached directly to manual journal lines, or
    # to the original Cash Book line that created one or more General Ledger lines.
    journal_line_file_map = {}
    cashbook_line_file_map = {}
    journal_line_ids = sorted({int(r['id']) for r in raw_rows if r.get('id')})
    cashbook_line_ids = sorted({int(r['source_record_id']) for r in raw_rows if r.get('source_record_type') == 'cash_book_line' and r.get('source_record_id')})

    if journal_line_ids:
        placeholders = ','.join(['?'] * len(journal_line_ids))
        file_rows = conn.execute(f'''SELECT * FROM accounting_transaction_files
                                     WHERE company_id=? AND linked_type='journal_line' AND linked_id IN ({placeholders})
                                     ORDER BY id DESC''', [company_id] + journal_line_ids).fetchall()
        for f in file_rows:
            lid = int(f['linked_id'])
            if lid not in journal_line_file_map:
                journal_line_file_map[lid] = f

    if cashbook_line_ids:
        placeholders = ','.join(['?'] * len(cashbook_line_ids))
        file_rows = conn.execute(f'''SELECT * FROM accounting_transaction_files
                                     WHERE company_id=? AND linked_type='cashbook_line' AND linked_id IN ({placeholders})
                                     ORDER BY id DESC''', [company_id] + cashbook_line_ids).fetchall()
        for f in file_rows:
            lid = int(f['linked_id'])
            if lid not in cashbook_line_file_map:
                cashbook_line_file_map[lid] = f

    account_opening = {int(a['id']): 0.0 for a in accounts}
    if start_date:
        opening_sql = """SELECT l.account_id, COALESCE(SUM(l.debit),0) AS debit_total, COALESCE(SUM(l.credit),0) AS credit_total
                         FROM accounting_journal_lines l
                         JOIN accounting_journals j ON j.id = l.journal_id
                         WHERE l.company_id=? AND j.company_id=? AND j.status='posted' AND j.journal_date < ?"""
        opening_params = [company_id, company_id, start_date]
        if account_id:
            opening_sql += ' AND l.account_id=?'
            opening_params.append(int(account_id))
        opening_sql += ' GROUP BY l.account_id'
        for r in conn.execute(opening_sql, opening_params).fetchall():
            account_opening[int(r['account_id'])] = round(_money_float(r['debit_total']) - _money_float(r['credit_total']), 2)

    rows_by_account = {}
    for row in raw_rows:
        aid = int(row['account_id'])
        rows_by_account.setdefault(aid, []).append(row)

    grouped = []
    selected_accounts = [a for a in accounts if (not account_id or int(a['id']) == int(account_id))]
    selected_accounts.sort(key=lambda a: (str(a.get('account_code') or ''), str(a.get('account_name') or '')))
    for acct in selected_accounts:
        aid = int(acct['id'])
        lines = rows_by_account.get(aid, [])
        opening = round(account_opening.get(aid, 0.0), 2)
        if not lines and abs(opening) < 0.005:
            continue
        running = opening
        out_lines = []
        for line in lines:
            debit = _money_float(line.get('debit'))
            credit = _money_float(line.get('credit'))
            running = round(running + debit - credit, 2)
            file_row = journal_line_file_map.get(int(line.get('id') or 0))
            if not file_row and line.get('source_record_type') == 'cash_book_line' and line.get('source_record_id'):
                file_row = cashbook_line_file_map.get(int(line.get('source_record_id')))
            out_lines.append({
                'id': line.get('id'),
                'journal_id': line.get('journal_id'),
                'journal_date': line.get('journal_date'),
                'reference': line.get('reference') or '',
                'source_module': line.get('source_module') or '',
                'description': line.get('description') or line.get('journal_description') or '',
                'journal_description': line.get('journal_description') or '',
                'debit': debit,
                'credit': credit,
                'running_balance': round(abs(running), 2),
                'balance_side': 'Dr' if running >= 0 else 'Cr',
                'transaction_file': accounting_file_to_dict(file_row) if file_row else None
            })
        closing_raw = running
        grouped.append({
            'account_id': aid,
            'account_code': acct.get('account_code'),
            'account_name': acct.get('account_name'),
            'account_type': acct.get('account_type'),
            'normal_balance': acct.get('normal_balance'),
            'opening_balance': round(abs(opening), 2),
            'opening_side': 'Dr' if opening >= 0 else 'Cr',
            'closing_balance': round(abs(closing_raw), 2),
            'closing_side': 'Dr' if closing_raw >= 0 else 'Cr',
            'debit_total': round(sum(_money_float(l.get('debit')) for l in lines), 2),
            'credit_total': round(sum(_money_float(l.get('credit')) for l in lines), 2),
            'lines': out_lines
        })
    return grouped

def _cash_flow_data(conn, company_id, start_date, end_date):
    opening = 0.0
    closing = 0.0
    sections = {'operating': 0.0, 'investing': 0.0, 'financing': 0.0}
    for line in _posted_lines(conn, company_id, None, start_date):
        if str(line.get('journal_date')) >= str(start_date):
            continue
        if int(line.get('is_cash_equivalent') or 0):
            opening += _money_float(line.get('debit')) - _money_float(line.get('credit'))
    for line in _posted_lines(conn, company_id, None, end_date):
        if int(line.get('is_cash_equivalent') or 0):
            closing += _money_float(line.get('debit')) - _money_float(line.get('credit'))
    for line in _posted_lines(conn, company_id, start_date, end_date):
        if not int(line.get('is_cash_equivalent') or 0):
            continue
        movement = _money_float(line.get('debit')) - _money_float(line.get('credit'))
        section = (line.get('cash_flow_section') or line.get('cash_flow_category') or 'operating').lower()
        if section not in sections:
            section = 'operating'
        sections[section] += movement
    net = round(sum(sections.values()), 2)
    return {'opening_cash': round(opening, 2), 'sections': {k: round(v, 2) for k, v in sections.items()}, 'net_movement': net, 'closing_cash': round(closing, 2), 'calculated_closing_cash': round(opening + net, 2)}


@app.route('/accounting')
def accounting_index():
    conn = get_db_connection()
    ensure_accounting_posting_defaults(conn, _current_company_id())
    conn.commit()
    conn.close()
    return render_template('accounting_index.html', session=session)


@app.route('/api/accounting/bootstrap', methods=['POST'])
def accounting_bootstrap():
    conn = get_db_connection()
    ensure_accounting_posting_defaults(conn, _current_company_id())
    conn.commit()
    conn.close()
    log_action('Accounting', 'Default Chart Created', 'Default IFRS for SMEs chart of accounts and posting defaults ensured.')
    return jsonify({'status': 'success'})


@app.route('/api/accounting/accounts', methods=['GET', 'POST'])
def accounting_accounts_api():
    cid = _current_company_id()
    conn = get_db_connection()
    ensure_default_accounting_accounts(conn, cid)
    if request.method == 'GET':
        accounts = _accounting_accounts(conn, cid, active_only=False)
        conn.close()
        return jsonify({'status': 'success', 'accounts': accounts, 'sections': IFRS_SECTION_LABELS})
    data = request.get_json() or {}
    account_id = data.get('id')
    code = (data.get('account_code') or '').strip()
    name = (data.get('account_name') or '').strip()
    acc_type = (data.get('account_type') or 'expense').strip()
    section = (data.get('report_section') or 'operating_expense').strip()
    normal = (data.get('normal_balance') or ('credit' if acc_type in ['liability', 'equity', 'income'] else 'debit')).strip()
    cf_category = (data.get('cash_flow_category') or 'operating').strip()
    is_cash = 1 if data.get('is_cash_equivalent') else 0
    active = 1 if data.get('active', True) else 0
    if not code or not name:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Account code and name are required.'}), 400
    try:
        if account_id:
            conn.execute('''UPDATE accounting_accounts SET account_code=?, account_name=?, account_type=?, report_section=?, normal_balance=?, cash_flow_category=?, is_cash_equivalent=?, active=?
                            WHERE id=? AND company_id=?''', (code, name, acc_type, section, normal, cf_category, is_cash, active, account_id, cid))
        else:
            conn.execute('''INSERT INTO accounting_accounts (company_id, account_code, account_name, account_type, report_section, normal_balance, cash_flow_category, is_cash_equivalent, active)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', (cid, code, name, acc_type, section, normal, cf_category, is_cash, active))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'status': 'error', 'message': 'An account with this code already exists.'}), 400
    conn.close()
    log_action('Accounting', 'Saved Account', f'Saved chart of accounts item {code} - {name}.')
    return jsonify({'status': 'success'})


@app.route('/api/accounting/settings', methods=['GET', 'POST'])
def accounting_settings_api():
    cid = _current_company_id()
    conn = get_db_connection()
    ensure_accounting_posting_defaults(conn, cid)
    if request.method == 'GET':
        settings = _accounting_posting_settings(conn, cid)
        accounts = _accounting_accounts(conn, cid, active_only=False)
        conn.close()
        return jsonify({'status': 'success', 'settings': settings, 'accounts': accounts})
    data = request.get_json() or {}
    vat_control_account_id = data.get('vat_control_account_id') or None
    if vat_control_account_id:
        acct = conn.execute('SELECT id FROM accounting_accounts WHERE id=? AND company_id=? AND active=1', (vat_control_account_id, cid)).fetchone()
        if not acct:
            conn.close()
            return jsonify({'status': 'error', 'message': 'The selected VAT Control account is not valid or active.'}), 400
    conn.execute('''INSERT OR IGNORE INTO accounting_settings (company_id, reporting_framework, financial_year_end_month)
                    VALUES (?, 'IFRS for SMEs', 2)''', (cid,))
    conn.execute('UPDATE accounting_settings SET vat_control_account_id=? WHERE company_id=?', (vat_control_account_id, cid))
    conn.commit()
    settings = _accounting_posting_settings(conn, cid)
    conn.close()
    log_action('Accounting', 'Updated Accounting Settings', 'Updated Cash Book VAT Control account setting.')
    return jsonify({'status': 'success', 'settings': settings})


@app.route('/api/accounting/journals', methods=['GET', 'POST'])
def accounting_journals_api():
    cid = _current_company_id()
    conn = get_db_connection()
    ensure_default_accounting_accounts(conn, cid)
    if request.method == 'GET':
        journals = [dict(j) for j in conn.execute('''SELECT j.*, COUNT(l.id) AS line_count, COALESCE(SUM(l.debit),0) AS debit_total, COALESCE(SUM(l.credit),0) AS credit_total
                                                   FROM accounting_journals j
                                                   LEFT JOIN accounting_journal_lines l ON l.journal_id=j.id AND l.company_id=j.company_id
                                                   WHERE j.company_id=?
                                                   GROUP BY j.id, j.company_id, j.journal_date, j.reference, j.description, j.source_module, j.status, j.created_by, j.posted_by, j.created_at, j.posted_at
                                                   ORDER BY j.journal_date DESC, j.id DESC LIMIT 100''', (cid,)).fetchall()]
        conn.close()
        return jsonify({'status': 'success', 'journals': journals})
    data = _request_payload_with_optional_files()
    journal_date = data.get('journal_date') or datetime.now().strftime('%Y-%m-%d')
    reference = (data.get('reference') or '').strip()
    description = (data.get('description') or '').strip()
    requested_status = 'posted' if data.get('post_now') else 'draft'
    lines = data.get('lines') or []
    settings = _accounting_posting_settings(conn, cid)
    vat_control_account_id = settings.get('vat_control_account_id')
    vat_control_account = None
    if vat_control_account_id:
        vat_control_account = conn.execute('SELECT * FROM accounting_accounts WHERE id=? AND company_id=? AND active=1', (vat_control_account_id, cid)).fetchone()
    clean_lines = []
    debit_total = 0.0
    credit_total = 0.0
    for idx, line in enumerate(lines, 1):
        account_id = line.get('account_id')
        debit = _money_float(line.get('debit'))
        credit = _money_float(line.get('credit'))
        vat_amount = round(_money_float(line.get('vat_amount')), 2)
        if not account_id or (debit <= 0 and credit <= 0):
            continue
        if debit > 0 and credit > 0:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Line {idx} cannot have both debit and credit amounts.'}), 400
        gross_amount = round(debit if debit > 0 else credit, 2)
        if vat_amount < 0:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Line {idx} has a negative VAT amount.'}), 400
        if vat_amount > gross_amount:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Line {idx} VAT amount cannot be more than the line amount.'}), 400
        if vat_amount > 0 and not vat_control_account:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Select a valid VAT Control account in Accounting Settings before posting manual journal VAT lines.'}), 400
        acct = conn.execute('SELECT id, cash_flow_category FROM accounting_accounts WHERE id=? AND company_id=? AND active=1', (account_id, cid)).fetchone()
        if not acct:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Line {idx} is allocated to an invalid account.'}), 400
        cf_section = (acct['cash_flow_category'] or 'operating').strip()
        if cf_section not in ['operating', 'investing', 'financing', 'non_cash']:
            cf_section = 'operating'
        clean_lines.append({
            'line_no': len(clean_lines) + 1,
            'account_id': int(account_id),
            'description': (line.get('description') or '').strip(),
            'debit': debit,
            'credit': credit,
            'gross_amount': gross_amount,
            'vat_amount': vat_amount,
            'net_amount': round(gross_amount - vat_amount, 2),
            'vat_type': 'Input' if debit > 0 else 'Output',
            'cash_flow_section': cf_section,
            'client_index': line.get('client_index')
        })
        debit_total += debit
        credit_total += credit
    if len(clean_lines) < 2:
        conn.close()
        return jsonify({'status': 'error', 'message': 'A journal entry must have at least two valid lines.'}), 400
    if round(debit_total - credit_total, 2) != 0:
        conn.close()
        return jsonify({'status': 'error', 'message': f'Journal does not balance. Debits {debit_total:.2f} vs Credits {credit_total:.2f}.'}), 400
    cur = conn.execute('''INSERT INTO accounting_journals (company_id, journal_date, reference, description, source_module, status, created_by, posted_by, posted_at)
                          VALUES (?, ?, ?, ?, 'manual', ?, ?, ?, ?)''',
                       (cid, journal_date, reference, description, requested_status, session.get('username'), session.get('username') if requested_status == 'posted' else None, datetime.now().strftime('%Y-%m-%d %H:%M:%S') if requested_status == 'posted' else None))
    journal_id = cur.lastrowid
    stored_line_no = 1
    for line in clean_lines:
        vat_amount = _money_float(line.get('vat_amount'))
        gross_amount = _money_float(line.get('gross_amount'))
        net_amount = round(gross_amount - vat_amount, 2)
        selected_line_id = None
        if line['debit'] > 0:
            if net_amount > 0:
                cur_line = conn.execute('''INSERT INTO accounting_journal_lines (company_id, journal_id, line_no, account_id, description, debit, credit, cash_flow_section, gross_amount, net_amount, vat_amount, vat_type)
                                          VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?, ?, 0, NULL)''',
                                       (cid, journal_id, stored_line_no, line['account_id'], line['description'], net_amount, line['cash_flow_section'], gross_amount if vat_amount > 0 else 0, net_amount if vat_amount > 0 else 0))
                selected_line_id = cur_line.lastrowid
                stored_line_no += 1
            if vat_amount > 0:
                conn.execute('''INSERT INTO accounting_journal_lines (company_id, journal_id, line_no, account_id, description, debit, credit, cash_flow_section, gross_amount, net_amount, vat_amount, vat_type)
                                VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?)''',
                             (cid, journal_id, stored_line_no, vat_control_account_id, f"VAT Control - {line['description'] or reference or journal_id}", vat_amount, line['cash_flow_section'], gross_amount, net_amount, vat_amount, 'Input'))
                stored_line_no += 1
        else:
            if net_amount > 0:
                cur_line = conn.execute('''INSERT INTO accounting_journal_lines (company_id, journal_id, line_no, account_id, description, debit, credit, cash_flow_section, gross_amount, net_amount, vat_amount, vat_type)
                                          VALUES (?, ?, ?, ?, ?, 0, ?, ?, ?, ?, 0, NULL)''',
                                       (cid, journal_id, stored_line_no, line['account_id'], line['description'], net_amount, line['cash_flow_section'], gross_amount if vat_amount > 0 else 0, net_amount if vat_amount > 0 else 0))
                selected_line_id = cur_line.lastrowid
                stored_line_no += 1
            if vat_amount > 0:
                conn.execute('''INSERT INTO accounting_journal_lines (company_id, journal_id, line_no, account_id, description, debit, credit, cash_flow_section, gross_amount, net_amount, vat_amount, vat_type)
                                VALUES (?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?)''',
                             (cid, journal_id, stored_line_no, vat_control_account_id, f"VAT Control - {line['description'] or reference or journal_id}", vat_amount, line['cash_flow_section'], gross_amount, net_amount, vat_amount, 'Output'))
                stored_line_no += 1
        if request.content_type and 'multipart/form-data' in request.content_type and selected_line_id:
            upload = request.files.get(f"journal_file_{line.get('client_index')}")
            if upload and upload.filename:
                try:
                    _save_accounting_transaction_file(conn, cid, 'journal_line', selected_line_id, upload)
                except ValueError as exc:
                    conn.rollback()
                    conn.close()
                    return jsonify({'status': 'error', 'message': str(exc)}), 400
    conn.commit()
    conn.close()
    log_action('Accounting', 'Created Journal', f'Created {requested_status} manual journal {reference or journal_id}.')
    return jsonify({'status': 'success', 'journal_id': journal_id})


@app.route('/api/accounting/journals/<int:journal_id>/post', methods=['POST'])
def accounting_post_journal(journal_id):
    cid = _current_company_id()
    conn = get_db_connection()
    journal = conn.execute('SELECT * FROM accounting_journals WHERE id=? AND company_id=?', (journal_id, cid)).fetchone()
    if not journal:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Journal not found.'}), 404
    lines = [dict(l) for l in conn.execute('SELECT * FROM accounting_journal_lines WHERE journal_id=? AND company_id=?', (journal_id, cid)).fetchall()]
    debit_total = round(sum(_money_float(l.get('debit')) for l in lines), 2)
    credit_total = round(sum(_money_float(l.get('credit')) for l in lines), 2)
    if len(lines) < 2 or debit_total != credit_total:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Cannot post an unbalanced journal.'}), 400
    conn.execute('''UPDATE accounting_journals SET status='posted', posted_by=?, posted_at=? WHERE id=? AND company_id=?''', (session.get('username'), datetime.now().strftime('%Y-%m-%d %H:%M:%S'), journal_id, cid))
    conn.commit()
    conn.close()
    log_action('Accounting', 'Posted Journal', f'Posted manual journal {journal_id}.')
    return jsonify({'status': 'success'})




def _clean_csv_money(value):
    text = str(value or '').strip()
    if not text:
        return 0.0
    negative = False
    if text.startswith('(') and text.endswith(')'):
        negative = True
        text = text[1:-1]
    text = text.replace('R', '').replace('r', '').replace('ZAR', '').replace(',', '').replace(' ', '')
    if text.endswith('-'):
        negative = True
        text = text[:-1]
    try:
        amount = float(text)
    except Exception:
        return 0.0
    if negative:
        amount = -abs(amount)
    return round(amount, 2)


def _normalise_csv_header(value):
    return str(value or '').strip().lower().replace('\ufeff', '').replace('_', ' ').replace('-', ' ')


def _find_csv_column(fieldnames, candidates):
    lookup = {_normalise_csv_header(name): name for name in (fieldnames or [])}
    for candidate in candidates:
        key = _normalise_csv_header(candidate)
        if key in lookup:
            return lookup[key]
    for normalised, original in lookup.items():
        for candidate in candidates:
            if _normalise_csv_header(candidate) in normalised:
                return original
    return None


class CashbookColumnMappingRequired(ValueError):
    def __init__(self, message, fieldnames=None, detected=None):
        super().__init__(message)
        self.fieldnames = fieldnames or []
        self.detected = detected or {}


def _cashbook_bank_formats():
    return {
        'auto': 'Auto-detect',
        'fnb': 'FNB / RMB',
        'standard_bank': 'Standard Bank',
        'absa': 'Absa',
        'nedbank': 'Nedbank',
        'capitec': 'Capitec',
        'investec': 'Investec',
        'tymebank': 'TymeBank / GoTyme',
        'generic': 'Generic CSV',
        'custom': 'Custom Mapping'
    }


def _cashbook_column_candidates(bank_format='auto'):
    bank_format = (bank_format or 'auto').strip().lower()
    common = {
        'date': ['date', 'transaction date', 'trans date', 'posting date', 'value date', 'effective date', 'process date', 'posted date'],
        'description': ['description', 'details', 'narrative', 'transaction description', 'reference', 'memo', 'beneficiary/reference', 'transaction details', 'statement description'],
        'debit': ['debit', 'debits', 'debit amount', 'withdrawal', 'withdrawals', 'money out', 'paid out', 'payments', 'outflow', 'debt'],
        'credit': ['credit', 'credits', 'credit amount', 'deposit', 'deposits', 'money in', 'paid in', 'receipts', 'inflow'],
        'amount': ['amount', 'transaction amount', 'value', 'signed amount', 'movement'],
        'balance': ['balance', 'bank balance', 'running balance', 'available balance', 'closing balance']
    }
    bank_specific = {
        'fnb': {
            'date': ['date', 'transaction date', 'posting date', 'value date'],
            'description': ['description', 'details', 'reference', 'service fee description'],
            'debit': ['debit', 'money out', 'withdrawal', 'withdrawals', 'payments'],
            'credit': ['credit', 'money in', 'deposit', 'deposits', 'receipts'],
            'amount': ['amount', 'transaction amount'],
            'balance': ['balance', 'running balance']
        },
        'standard_bank': {
            'date': ['date', 'transaction date', 'value date', 'posting date'],
            'description': ['description', 'transaction description', 'details'],
            'debit': ['debit', 'debit amount', 'money out', 'withdrawal'],
            'credit': ['credit', 'credit amount', 'money in', 'deposit'],
            'amount': ['amount', 'transaction amount'],
            'balance': ['balance', 'running balance']
        },
        'absa': {
            'date': ['date', 'transaction date', 'posting date', 'value date'],
            'description': ['description', 'transaction description', 'narrative', 'reference'],
            'debit': ['debit', 'debit amount', 'withdrawal', 'money out'],
            'credit': ['credit', 'credit amount', 'deposit', 'money in'],
            'amount': ['amount', 'transaction amount'],
            'balance': ['balance', 'running balance']
        },
        'nedbank': {
            'date': ['date', 'transaction date', 'value date', 'posting date'],
            'description': ['description', 'transaction description', 'details', 'narrative'],
            'debit': ['debit', 'debit amount', 'withdrawals', 'money out'],
            'credit': ['credit', 'credit amount', 'deposits', 'money in'],
            'amount': ['amount', 'transaction amount'],
            'balance': ['balance', 'running balance']
        },
        'capitec': {
            'date': ['date', 'transaction date', 'posting date'],
            'description': ['description', 'beneficiary/reference', 'reference', 'details'],
            'debit': ['debit', 'money out', 'paid out', 'withdrawal'],
            'credit': ['credit', 'money in', 'paid in', 'deposit'],
            'amount': ['amount', 'transaction amount'],
            'balance': ['balance', 'available balance', 'running balance']
        },
        'investec': {
            'date': ['date', 'transaction date', 'value date'],
            'description': ['description', 'details', 'reference', 'narrative'],
            'debit': ['debit', 'debit amount', 'money out'],
            'credit': ['credit', 'credit amount', 'money in'],
            'amount': ['amount', 'transaction amount'],
            'balance': ['balance', 'running balance']
        },
        'tymebank': {
            'date': ['date', 'transaction date', 'posting date'],
            'description': ['description', 'transaction description', 'reference'],
            'debit': ['debit', 'money out', 'paid out', 'withdrawal'],
            'credit': ['credit', 'money in', 'paid in', 'deposit'],
            'amount': ['amount', 'transaction amount'],
            'balance': ['balance', 'running balance']
        },
        'generic': common,
        'auto': common
    }
    selected = bank_specific.get(bank_format, common)
    merged = {}
    for key in common:
        merged[key] = []
        for item in list(selected.get(key, [])) + common[key]:
            if item not in merged[key]:
                merged[key].append(item)
    return merged


def _read_cashbook_csv(file_storage):
    raw = file_storage.read()
    text = None
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            text = raw.decode(encoding)
            break
        except Exception:
            continue
    if text is None:
        raise ValueError('Could not read the CSV file. Please save it as CSV UTF-8 and try again.')
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except Exception:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    fieldnames = [str(f or '').strip().replace('\ufeff', '') for f in (reader.fieldnames or [])]
    if not fieldnames:
        raise ValueError('The CSV file does not contain a header row.')
    rows = list(reader)
    return fieldnames, rows


def _cashbook_mapping_from_request(raw):
    if not raw:
        return {}
    try:
        data = json.loads(raw)
        if not isinstance(data, dict):
            return {}
        return {str(k): str(v).strip() for k, v in data.items() if str(v or '').strip()}
    except Exception:
        return {}


def _detect_cashbook_mapping(fieldnames, bank_format='auto', mapping=None):
    field_set = set(fieldnames or [])
    mapping = mapping or {}
    detected = {}
    key_map = {
        'date': 'transaction_date',
        'description': 'description',
        'debit': 'debit',
        'credit': 'credit',
        'amount': 'amount',
        'balance': 'balance'
    }
    if mapping:
        for short_key, request_key in key_map.items():
            value = mapping.get(request_key) or mapping.get(short_key) or ''
            detected[request_key] = value if value in field_set else ''
        return detected

    candidates = _cashbook_column_candidates(bank_format)
    detected['transaction_date'] = _find_csv_column(fieldnames, candidates.get('date', [])) or ''
    detected['description'] = _find_csv_column(fieldnames, candidates.get('description', [])) or ''
    detected['debit'] = _find_csv_column(fieldnames, candidates.get('debit', [])) or ''
    detected['credit'] = _find_csv_column(fieldnames, candidates.get('credit', [])) or ''
    detected['amount'] = _find_csv_column(fieldnames, candidates.get('amount', [])) or ''
    detected['balance'] = _find_csv_column(fieldnames, candidates.get('balance', [])) or ''
    return detected


def _cashbook_mapping_missing(mapping):
    has_date = bool(mapping.get('transaction_date'))
    has_desc = bool(mapping.get('description'))
    has_money = bool(mapping.get('debit') or mapping.get('credit') or mapping.get('amount'))
    return not (has_date and has_desc and has_money)


def _normalise_accounting_date(value):
    """Convert common bank statement date formats to ISO YYYY-MM-DD for reliable ledger filtering."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip().strip('"').strip("'")
    if not s:
        return ''

    # Remove time portions and common CSV noise.
    s = s.replace('\ufeff', '').strip()
    if 'T' in s and re.match(r'^\d{4}-\d{2}-\d{2}T', s):
        s = s.split('T', 1)[0]
    if ' ' in s and re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}\s+', s):
        s = s.split(' ', 1)[0]

    # Excel serial date support, useful for some exported CSV files.
    try:
        if re.match(r'^\d+(\.0+)?$', s):
            n = int(float(s))
            if 20000 <= n <= 80000:
                return (datetime(1899, 12, 30) + timedelta(days=n)).strftime('%Y-%m-%d')
    except Exception:
        pass

    cleaned = re.sub(r'\s+', ' ', s)
    formats = [
        '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
        '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y',
        '%d/%m/%y', '%d-%m-%y', '%d.%m.%y',
        '%d %b %Y', '%d %B %Y', '%d %b %y', '%d %B %y',
        '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
        '%d/%m/%Y %H:%M:%S', '%d-%m-%Y %H:%M:%S',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(cleaned, fmt).strftime('%Y-%m-%d')
        except Exception:
            continue

    # Last resort for values like 2026-6-3 or 3/6/2026.
    m = re.match(r'^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$', cleaned)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return datetime(y, mo, d).strftime('%Y-%m-%d')
        except Exception:
            pass
    m = re.match(r'^(\d{1,2})[-/.](\d{1,2})[-/.](\d{2,4})$', cleaned)
    if m:
        d, mo, y = m.groups()  # South African bank statements are normally day/month/year.
        y = int(y)
        if y < 100:
            y += 2000
        try:
            return datetime(y, int(mo), int(d)).strftime('%Y-%m-%d')
        except Exception:
            pass
    return s


def _is_iso_accounting_date(value):
    return bool(re.match(r'^\d{4}-\d{2}-\d{2}$', str(value or '').strip()))


def _normalise_existing_cashbook_ledger_dates(conn, company_id):
    """Repair previously imported/posted cash book dates so the General Ledger can find them."""
    updated = 0
    try:
        rows = conn.execute('SELECT id, transaction_date FROM accounting_cashbook_lines WHERE company_id=?', (company_id,)).fetchall()
        for row in rows:
            old = str(row['transaction_date'] or '').strip()
            new = _normalise_accounting_date(old)
            if new and _is_iso_accounting_date(new) and new != old:
                conn.execute('UPDATE accounting_cashbook_lines SET transaction_date=? WHERE id=? AND company_id=?', (new, row['id'], company_id))
                updated += 1
    except Exception:
        pass
    try:
        rows = conn.execute("SELECT id, journal_date FROM accounting_journals WHERE company_id=? AND source_module='cash_book'", (company_id,)).fetchall()
        for row in rows:
            old = str(row['journal_date'] or '').strip()
            new = _normalise_accounting_date(old)
            if new and _is_iso_accounting_date(new) and new != old:
                conn.execute('UPDATE accounting_journals SET journal_date=? WHERE id=? AND company_id=?', (new, row['id'], company_id))
                updated += 1
    except Exception:
        pass
    if updated:
        try:
            conn.commit()
        except Exception:
            pass
    return updated


def _parse_cashbook_csv(file_storage, bank_format='auto', mapping=None):
    fieldnames, rows = _read_cashbook_csv(file_storage)
    bank_format = (bank_format or 'auto').strip().lower()
    mapping = _detect_cashbook_mapping(fieldnames, bank_format, mapping)

    if bank_format == 'custom' and _cashbook_mapping_missing(mapping):
        raise CashbookColumnMappingRequired('Please map the CSV columns before uploading this Custom Mapping cash book.', fieldnames, mapping)
    if _cashbook_mapping_missing(mapping):
        raise CashbookColumnMappingRequired('The app could not confidently detect the date, description and amount/debit/credit columns. Please map the CSV columns and upload again.', fieldnames, mapping)

    date_col = mapping.get('transaction_date')
    desc_col = mapping.get('description')
    debit_col = mapping.get('debit')
    credit_col = mapping.get('credit')
    amount_col = mapping.get('amount')
    balance_col = mapping.get('balance')

    lines = []
    for row_no, row in enumerate(rows, 2):
        if not any(str(v or '').strip() for v in row.values()):
            continue
        raw_date_val = str(row.get(date_col) or '').strip()
        date_val = _normalise_accounting_date(raw_date_val)
        if not _is_iso_accounting_date(date_val):
            raise ValueError(f'CSV row {row_no} has an invalid transaction date: {raw_date_val or "blank"}. Please use or map a date column in YYYY-MM-DD or DD/MM/YYYY format.')
        desc_val = str(row.get(desc_col) or '').strip()
        debit = _clean_csv_money(row.get(debit_col)) if debit_col else 0.0
        credit = _clean_csv_money(row.get(credit_col)) if credit_col else 0.0
        if amount_col and not debit and not credit:
            amount = _clean_csv_money(row.get(amount_col))
            if amount < 0:
                debit = abs(amount)
            elif amount > 0:
                credit = amount
        # Bank statement convention used by Easy Admin Cash Book:
        #   Debit  = money out of the bank account / decrease in bank balance.
        #   Credit = money into the bank account / increase in bank balance.
        # In the general ledger posting step, these are converted into the correct
        # double-entry journal lines for the company's own accounting records.
        debit = abs(debit) if debit else 0.0
        credit = abs(credit) if credit else 0.0
        balance = _clean_csv_money(row.get(balance_col)) if balance_col else 0.0

        # Ignore zero-amount bank statement lines.
        if round(debit, 2) == 0 and round(credit, 2) == 0:
            continue
        if round(debit, 2) > 0 and round(credit, 2) > 0:
            raise ValueError(f'CSV row {row_no} has both a Debit/Money Out and Credit/Money In amount. Please check the bank format or column mapping.')

        lines.append({
            'line_no': len(lines) + 1,
            'transaction_date': date_val,
            'description': desc_val,
            'debit': round(debit, 2),
            'credit': round(credit, 2),
            'balance': round(balance, 2)
        })
    if not lines:
        raise ValueError('No usable bank statement transactions were found in the CSV file.')
    return lines, mapping


def _cashbook_batch_payload(conn, company_id, batch_id):
    batch = conn.execute('''SELECT b.*, a.account_code AS bank_account_code, a.account_name AS bank_account_name
                            FROM accounting_cashbook_batches b
                            LEFT JOIN accounting_accounts a ON a.id=b.bank_account_id AND a.company_id=b.company_id
                            WHERE b.id=? AND b.company_id=?''', (batch_id, company_id)).fetchone()
    if not batch:
        return None
    raw_lines = [dict(r) for r in conn.execute('''SELECT l.*, a.account_code AS allocated_account_code, a.account_name AS allocated_account_name
                                                   FROM accounting_cashbook_lines l
                                                   LEFT JOIN accounting_accounts a ON a.id=l.allocated_account_id AND a.company_id=l.company_id
                                                   WHERE l.batch_id=? AND l.company_id=?
                                                   ORDER BY l.line_no ASC, l.id ASC''', (batch_id, company_id)).fetchall()]
    lines = _attach_accounting_transaction_files(conn, company_id, 'cashbook_line', raw_lines)
    return {'batch': dict(batch), 'lines': lines}


@app.route('/api/accounting/cashbook/batches')
def accounting_cashbook_batches():
    cid = _current_company_id()
    conn = get_db_connection()
    ensure_default_accounting_accounts(conn, cid)
    batches = [dict(r) for r in conn.execute('''SELECT b.*, a.account_code AS bank_account_code, a.account_name AS bank_account_name
                                                 FROM accounting_cashbook_batches b
                                                 LEFT JOIN accounting_accounts a ON a.id=b.bank_account_id AND a.company_id=b.company_id
                                                 WHERE b.company_id=?
                                                 ORDER BY b.created_at DESC, b.id DESC LIMIT 50''', (cid,)).fetchall()]
    conn.close()
    return jsonify({'status': 'success', 'batches': batches})


@app.route('/api/accounting/cashbook/batches/<int:batch_id>')
def accounting_cashbook_batch_get(batch_id):
    cid = _current_company_id()
    conn = get_db_connection()
    payload = _cashbook_batch_payload(conn, cid, batch_id)
    conn.close()
    if not payload:
        return jsonify({'status': 'error', 'message': 'Cash book batch not found.'}), 404
    payload['status'] = 'success'
    return jsonify(payload)


@app.route('/api/accounting/cashbook/inspect', methods=['POST'])
def accounting_cashbook_inspect():
    file_storage = request.files.get('bank_statement')
    bank_format = (request.form.get('bank_format') or 'auto').strip().lower()
    mapping = _cashbook_mapping_from_request(request.form.get('column_mapping'))
    if not file_storage or not file_storage.filename:
        return jsonify({'status': 'error', 'message': 'Select a CSV bank statement file.'}), 400
    try:
        fieldnames, rows = _read_cashbook_csv(file_storage)
        detected = _detect_cashbook_mapping(fieldnames, bank_format, mapping)
        preview = []
        for row in rows[:5]:
            preview.append({h: row.get(h, '') for h in fieldnames})
        return jsonify({
            'status': 'success',
            'bank_format': bank_format,
            'formats': _cashbook_bank_formats(),
            'headers': fieldnames,
            'detected_mapping': detected,
            'mapping_required': bank_format == 'custom' or _cashbook_mapping_missing(detected),
            'preview': preview
        })
    except ValueError as exc:
        return jsonify({'status': 'error', 'message': str(exc)}), 400


@app.route('/api/accounting/cashbook/upload', methods=['POST'])
def accounting_cashbook_upload():
    cid = _current_company_id()
    bank_account_id = request.form.get('bank_account_id')
    file_storage = request.files.get('bank_statement')
    if not bank_account_id:
        return jsonify({'status': 'error', 'message': 'Select the bank account for this statement.'}), 400
    if not file_storage or not file_storage.filename:
        return jsonify({'status': 'error', 'message': 'Select a CSV bank statement file.'}), 400
    conn = get_db_connection()
    ensure_default_accounting_accounts(conn, cid)
    bank = conn.execute('''SELECT * FROM accounting_accounts WHERE id=? AND company_id=? AND active=1 AND is_cash_equivalent=1''', (bank_account_id, cid)).fetchone()
    if not bank:
        conn.close()
        return jsonify({'status': 'error', 'message': 'The selected bank account is not an active Cash / Bank account in the Chart of Accounts.'}), 400
    bank_format = (request.form.get('bank_format') or 'auto').strip().lower()
    mapping = _cashbook_mapping_from_request(request.form.get('column_mapping'))
    try:
        lines, detected_mapping = _parse_cashbook_csv(file_storage, bank_format, mapping)
    except CashbookColumnMappingRequired as exc:
        conn.close()
        return jsonify({'status': 'error', 'message': str(exc), 'mapping_required': True, 'headers': exc.fieldnames, 'detected_mapping': exc.detected, 'formats': _cashbook_bank_formats()}), 400
    except ValueError as exc:
        conn.close()
        return jsonify({'status': 'error', 'message': str(exc)}), 400
    debit_total = round(sum(float(l['debit']) for l in lines), 2)
    credit_total = round(sum(float(l['credit']) for l in lines), 2)
    cur = conn.execute('''INSERT INTO accounting_cashbook_batches
                          (company_id, bank_account_id, original_filename, status, imported_by, line_count, debit_total, credit_total, bank_format, column_mapping_json)
                          VALUES (?, ?, ?, 'draft', ?, ?, ?, ?, ?, ?)''',
                       (cid, int(bank_account_id), secure_filename(file_storage.filename), session.get('username'), len(lines), debit_total, credit_total, bank_format, json.dumps(detected_mapping)))
    batch_id = cur.lastrowid
    for line in lines:
        conn.execute('''INSERT INTO accounting_cashbook_lines
                        (company_id, batch_id, line_no, transaction_date, description, debit, credit, balance, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'draft')''',
                     (cid, batch_id, line['line_no'], line['transaction_date'], line['description'], line['debit'], line['credit'], line['balance']))
    conn.commit()
    payload = _cashbook_batch_payload(conn, cid, batch_id)
    conn.close()
    log_action('Accounting', 'Uploaded Cash Book', f'Uploaded cash book batch {batch_id} with {len(lines)} lines.')
    payload['status'] = 'success'
    return jsonify(payload)


@app.route('/api/accounting/cashbook/batches/<int:batch_id>/save', methods=['POST'])
def accounting_cashbook_save(batch_id):
    cid = _current_company_id()
    data = _request_payload_with_optional_files()
    line_updates = data.get('lines') or []
    conn = get_db_connection()
    batch = conn.execute('SELECT * FROM accounting_cashbook_batches WHERE id=? AND company_id=?', (batch_id, cid)).fetchone()
    if not batch:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Cash book batch not found.'}), 404
    if str(batch['status']) == 'posted':
        conn.close()
        return jsonify({'status': 'error', 'message': 'Posted cash book batches cannot be edited.'}), 400

    for idx, item in enumerate(line_updates, 1):
        raw_line_id = str(item.get('id') or '').strip()
        is_new_line = (not raw_line_id) or raw_line_id.startswith('new_')
        date_raw = (item.get('transaction_date') or '').strip()
        description = (item.get('description') or '').strip()
        debit = round(abs(_money_float(item.get('debit'))), 2)
        credit = round(abs(_money_float(item.get('credit'))), 2)
        account_id = item.get('allocated_account_id') or None
        notes = (item.get('notes') or '').strip()
        vat_amount = round(_money_float(item.get('vat_amount')), 2)
        line_no = int(item.get('line_no') or idx)

        # A completely blank inserted line is ignored so users can insert a row and leave it unused.
        has_file = bool(request.content_type and 'multipart/form-data' in request.content_type and request.files.get(f'cashbook_file_{raw_line_id}'))
        has_any_content = any([date_raw, description, debit, credit, account_id, notes, vat_amount, has_file])
        if is_new_line and not has_any_content:
            continue

        if debit > 0 and credit > 0:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Cash book line {idx} cannot have both Debit/Money Out and Credit/Money In amounts.'}), 400
        if debit <= 0 and credit <= 0:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Cash book line {idx} must have either a Debit/Money Out or Credit/Money In amount.'}), 400
        if vat_amount < 0:
            conn.close()
            return jsonify({'status': 'error', 'message': 'VAT amount cannot be negative.'}), 400
        if vat_amount > max(debit, credit):
            conn.close()
            return jsonify({'status': 'error', 'message': 'VAT amount cannot be more than the bank transaction amount.'}), 400

        transaction_date = _normalise_accounting_date(date_raw)
        if not _is_iso_accounting_date(transaction_date):
            conn.close()
            return jsonify({'status': 'error', 'message': f'Cash book line {idx} has an invalid transaction date. Use YYYY-MM-DD.'}), 400

        cash_flow_section = 'operating'
        if account_id:
            acct = conn.execute('SELECT id, cash_flow_category FROM accounting_accounts WHERE id=? AND company_id=? AND active=1', (account_id, cid)).fetchone()
            if not acct:
                conn.close()
                return jsonify({'status': 'error', 'message': 'One of the selected allocation accounts is not valid.'}), 400
            cash_flow_section = (acct['cash_flow_category'] or 'operating').strip()
            if cash_flow_section not in ['operating', 'investing', 'financing', 'non_cash']:
                cash_flow_section = 'operating'

        line_id = None
        if is_new_line:
            cur = conn.execute('''INSERT INTO accounting_cashbook_lines
                                  (company_id, batch_id, line_no, transaction_date, description, debit, credit, balance, allocated_account_id, vat_amount, notes, cash_flow_section, status)
                                  VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, CASE WHEN ? IS NULL THEN 'draft' ELSE 'allocated' END)''',
                               (cid, batch_id, line_no, transaction_date, description, debit, credit, account_id, vat_amount, notes, cash_flow_section, account_id))
            line_id = cur.lastrowid
        else:
            try:
                line_id = int(raw_line_id)
            except Exception:
                conn.close()
                return jsonify({'status': 'error', 'message': f'Cash book line {idx} has an invalid line ID.'}), 400
            line_row = conn.execute("SELECT id FROM accounting_cashbook_lines WHERE id=? AND batch_id=? AND company_id=? AND status!='posted'", (line_id, batch_id, cid)).fetchone()
            if not line_row:
                conn.close()
                return jsonify({'status': 'error', 'message': f'Cash book line {idx} was not found or has already been posted.'}), 400
            conn.execute('''UPDATE accounting_cashbook_lines
                            SET line_no=?, transaction_date=?, description=?, debit=?, credit=?, allocated_account_id=?, vat_amount=?, notes=?, cash_flow_section=?, status=CASE WHEN ? IS NULL THEN 'draft' ELSE 'allocated' END
                            WHERE id=? AND batch_id=? AND company_id=? AND status!='posted' ''',
                         (line_no, transaction_date, description, debit, credit, account_id, vat_amount, notes, cash_flow_section, account_id, line_id, batch_id, cid))

        if request.content_type and 'multipart/form-data' in request.content_type:
            upload = request.files.get(f'cashbook_file_{raw_line_id}')
            if upload and upload.filename:
                try:
                    _save_accounting_transaction_file(conn, cid, 'cashbook_line', int(line_id), upload)
                except ValueError as exc:
                    conn.rollback()
                    conn.close()
                    return jsonify({'status': 'error', 'message': str(exc)}), 400

    # Re-sequence lines and refresh batch totals after edits/inserts.
    rows = [dict(r) for r in conn.execute('SELECT id FROM accounting_cashbook_lines WHERE batch_id=? AND company_id=? ORDER BY line_no, id', (batch_id, cid)).fetchall()]
    for idx, row in enumerate(rows, 1):
        conn.execute('UPDATE accounting_cashbook_lines SET line_no=? WHERE id=? AND company_id=?', (idx, row['id'], cid))
    totals = conn.execute('''SELECT COUNT(*) AS line_count, COALESCE(SUM(debit),0) AS debit_total, COALESCE(SUM(credit),0) AS credit_total
                             FROM accounting_cashbook_lines WHERE batch_id=? AND company_id=?''', (batch_id, cid)).fetchone()
    conn.execute('''UPDATE accounting_cashbook_batches SET line_count=?, debit_total=?, credit_total=? WHERE id=? AND company_id=?''',
                 (int(totals['line_count'] or 0), round(float(totals['debit_total'] or 0), 2), round(float(totals['credit_total'] or 0), 2), batch_id, cid))
    conn.commit()
    payload = _cashbook_batch_payload(conn, cid, batch_id)
    conn.close()
    payload['status'] = 'success'
    return jsonify(payload)


@app.route('/api/accounting/cashbook/batches/<int:batch_id>/post', methods=['POST'])
def accounting_cashbook_post(batch_id):
    cid = _current_company_id()
    conn = get_db_connection()
    batch = conn.execute('SELECT * FROM accounting_cashbook_batches WHERE id=? AND company_id=?', (batch_id, cid)).fetchone()
    if not batch:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Cash book batch not found.'}), 404
    if str(batch['status']) == 'posted':
        conn.close()
        return jsonify({'status': 'error', 'message': 'Cash book batch already posted.'}), 400
    bank = conn.execute('SELECT * FROM accounting_accounts WHERE id=? AND company_id=? AND active=1 AND is_cash_equivalent=1', (batch['bank_account_id'], cid)).fetchone()
    if not bank:
        conn.close()
        return jsonify({'status': 'error', 'message': 'The linked bank account is not active or is no longer marked as Cash / Bank.'}), 400
    settings = _accounting_posting_settings(conn, cid)
    vat_control_account_id = settings.get('vat_control_account_id') or settings.get('vat_output_account_id')
    vat_control_account = None
    if vat_control_account_id:
        vat_control_account = conn.execute('SELECT * FROM accounting_accounts WHERE id=? AND company_id=? AND active=1', (vat_control_account_id, cid)).fetchone()
    lines = [dict(r) for r in conn.execute('SELECT * FROM accounting_cashbook_lines WHERE batch_id=? AND company_id=? ORDER BY line_no, id', (batch_id, cid)).fetchall()]
    if not lines:
        conn.close()
        return jsonify({'status': 'error', 'message': 'This cash book batch has no lines to post.'}), 400
    for line in lines:
        debit = _money_float(line.get('debit'))
        credit = _money_float(line.get('credit'))
        if debit <= 0 and credit <= 0:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Line {line.get("line_no")} has no debit or credit amount.'}), 400
        if debit > 0 and credit > 0:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Line {line.get("line_no")} cannot have both debit and credit amounts.'}), 400
        vat_amount = _money_float(line.get('vat_amount'))
        amount = round(credit if credit > 0 else debit, 2)
        if vat_amount < 0:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Line {line.get("line_no")} has a negative VAT amount.'}), 400
        if vat_amount > amount:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Line {line.get("line_no")} VAT amount cannot be more than the bank transaction amount.'}), 400
        if vat_amount > 0 and not vat_control_account:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Select a valid VAT Control account in Accounting Settings before posting VAT cash book lines.'}), 400
        if not line.get('allocated_account_id'):
            conn.close()
            return jsonify({'status': 'error', 'message': f'Line {line.get("line_no")} is not allocated to a Chart of Account.'}), 400
        acct = conn.execute('SELECT * FROM accounting_accounts WHERE id=? AND company_id=? AND active=1', (line.get('allocated_account_id'), cid)).fetchone()
        if not acct:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Line {line.get("line_no")} is allocated to an invalid account.'}), 400

    created_journals = []
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for line in lines:
        debit = _money_float(line.get('debit'))
        credit = _money_float(line.get('credit'))
        amount = round(credit if credit > 0 else debit, 2)
        vat_amount = _money_float(line.get('vat_amount'))
        net_amount = round(amount - vat_amount, 2)
        ref = f"CB-{batch_id}-{line.get('line_no')}"
        desc = line.get('description') or f"Cash book line {line.get('line_no')}"
        cf_section = (line.get('cash_flow_section') or 'operating').strip()
        if cf_section not in ['operating', 'investing', 'financing', 'non_cash']:
            cf_section = 'operating'
        journal_date = _normalise_accounting_date(line.get('transaction_date'))
        if not _is_iso_accounting_date(journal_date):
            journal_date = datetime.now().strftime('%Y-%m-%d')
        cur = conn.execute('''INSERT INTO accounting_journals
                              (company_id, journal_date, reference, description, source_module, status, created_by, posted_by, posted_at, source_record_type, source_record_id)
                              VALUES (?, ?, ?, ?, 'cash_book', 'posted', ?, ?, ?, 'cash_book_line', ?)''',
                           (cid, journal_date, ref, desc, session.get('username'), session.get('username'), now, line.get('id')))
        journal_id = cur.lastrowid
        line_no = 1
        # Convert bank statement direction into ledger postings:
        #   Bank statement Credit (money in)  -> Dr Bank full amount, Cr allocated net, Cr VAT control.
        #   Bank statement Debit  (money out) -> Dr allocated net, Dr VAT control, Cr Bank full amount.
        if credit > 0:
            conn.execute('''INSERT INTO accounting_journal_lines (company_id, journal_id, line_no, account_id, description, debit, credit, cash_flow_section)
                            VALUES (?, ?, ?, ?, ?, ?, 0, ?)''', (cid, journal_id, line_no, batch['bank_account_id'], desc, amount, cf_section))
            line_no += 1
            if net_amount > 0:
                conn.execute('''INSERT INTO accounting_journal_lines (company_id, journal_id, line_no, account_id, description, debit, credit, cash_flow_section)
                                VALUES (?, ?, ?, ?, ?, 0, ?, ?)''', (cid, journal_id, line_no, line.get('allocated_account_id'), desc, net_amount, cf_section))
                line_no += 1
            if vat_amount > 0:
                conn.execute('''INSERT INTO accounting_journal_lines (company_id, journal_id, line_no, account_id, description, debit, credit, cash_flow_section)
                                VALUES (?, ?, ?, ?, ?, 0, ?, ?)''', (cid, journal_id, line_no, vat_control_account_id, f"VAT Control - {desc}", vat_amount, cf_section))
        else:
            if net_amount > 0:
                conn.execute('''INSERT INTO accounting_journal_lines (company_id, journal_id, line_no, account_id, description, debit, credit, cash_flow_section)
                                VALUES (?, ?, ?, ?, ?, ?, 0, ?)''', (cid, journal_id, line_no, line.get('allocated_account_id'), desc, net_amount, cf_section))
                line_no += 1
            if vat_amount > 0:
                conn.execute('''INSERT INTO accounting_journal_lines (company_id, journal_id, line_no, account_id, description, debit, credit, cash_flow_section)
                                VALUES (?, ?, ?, ?, ?, ?, 0, ?)''', (cid, journal_id, line_no, vat_control_account_id, f"VAT Control - {desc}", vat_amount, cf_section))
                line_no += 1
            conn.execute('''INSERT INTO accounting_journal_lines (company_id, journal_id, line_no, account_id, description, debit, credit, cash_flow_section)
                            VALUES (?, ?, ?, ?, ?, 0, ?, ?)''', (cid, journal_id, line_no, batch['bank_account_id'], desc, amount, cf_section))
        conn.execute('''UPDATE accounting_cashbook_lines SET status='posted', linked_journal_id=? WHERE id=? AND batch_id=? AND company_id=?''', (journal_id, line.get('id'), batch_id, cid))
        created_journals.append(journal_id)
    conn.execute('''UPDATE accounting_cashbook_batches SET status='posted', posted_by=?, posted_at=? WHERE id=? AND company_id=?''', (session.get('username'), now, batch_id, cid))
    conn.commit()
    payload = _cashbook_batch_payload(conn, cid, batch_id)
    conn.close()
    log_action('Accounting', 'Posted Cash Book', f'Posted cash book batch {batch_id} and created {len(created_journals)} journals.')
    payload['status'] = 'success'
    payload['created_journals'] = created_journals
    return jsonify(payload)


@app.route('/api/accounting/reports/general_ledger')
def accounting_general_ledger():
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    account_id = request.args.get('account_id') or None
    try:
        account_id = int(account_id) if account_id else None
    except Exception:
        account_id = None
    conn = get_db_connection()
    _normalise_existing_cashbook_ledger_dates(conn, cid)
    data = _general_ledger_data(conn, cid, start_date, end_date, account_id)
    conn.close()
    return jsonify({'status': 'success', 'start_date': start_date, 'end_date': end_date, 'account_id': account_id, 'framework': 'IFRS for SMEs', 'accounts': data})


@app.route('/api/accounting/reports/trial_balance')
def accounting_trial_balance():
    cid = _current_company_id()
    as_at = request.args.get('as_at') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    ensure_default_accounting_accounts(conn, cid)
    rows = _trial_balance_rows(conn, cid, as_at)
    conn.close()
    return jsonify({'status': 'success', 'as_at': as_at, 'rows': rows, 'debit_total': round(sum(float(r['debit']) for r in rows), 2), 'credit_total': round(sum(float(r['credit']) for r in rows), 2)})


@app.route('/api/accounting/reports/income_statement')
def accounting_income_statement():
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    data = _income_statement_data(conn, cid, start_date, end_date)
    conn.close()
    data.update({'status': 'success', 'start_date': start_date, 'end_date': end_date, 'framework': 'IFRS for SMEs'})
    return jsonify(data)


@app.route('/api/accounting/reports/balance_sheet')
def accounting_balance_sheet():
    cid = _current_company_id()
    as_at = request.args.get('as_at') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    data = _balance_sheet_data(conn, cid, as_at)
    conn.close()
    data.update({'status': 'success', 'as_at': as_at, 'framework': 'IFRS for SMEs'})
    return jsonify(data)


@app.route('/api/accounting/reports/cash_flow')
def accounting_cash_flow():
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    data = _cash_flow_data(conn, cid, start_date, end_date)
    conn.close()
    data.update({'status': 'success', 'start_date': start_date, 'end_date': end_date, 'framework': 'IFRS for SMEs'})
    return jsonify(data)


# ==========================================================
# ACCOUNTING REPORT PDF EXPORTS
# ==========================================================
def _accounting_company_payload(conn, company_id):
    row = conn.execute('SELECT * FROM companies WHERE id=?', (company_id,)).fetchone()
    company = dict(row) if row else {}
    return {
        'name': company.get('name') or session.get('company_name') or 'Company',
        'address': company.get('address') or '',
        'registration_number': company.get('registration_number') or '',
        'vat_number': company.get('vat_number') or '',
        'logo_file': company.get('logo_file') or ''
    }


def _accounting_pdf_safe_filename(value):
    value = re.sub(r'[^A-Za-z0-9_\-]+', '_', str(value or 'Accounting_Report')).strip('_')
    return (value or 'Accounting_Report') + '.pdf'


def _accounting_pdf_money(value):
    try:
        val = float(value or 0)
        return ('-' if val < 0 else '') + _money(abs(val))
    except Exception:
        return _money(value)


def _accounting_make_row(cells, bold=False, shade=False, dark=False, total=False):
    return {'cells': [str(c if c is not None else '') for c in cells], 'bold': bool(bold), 'shade': bool(shade), 'dark': bool(dark), 'total': bool(total)}


def _draw_accounting_report_pdf(payload):
    orientation = str((payload or {}).get('orientation') or 'portrait').lower()
    if orientation == 'landscape':
        page_w, page_h = 841.89, 595.28
    else:
        page_w, page_h = 595.28, 841.89
    margin = 36.0
    table_w = page_w - (margin * 2)
    dark = (0.10, 0.12, 0.16)
    blue = (0.05, 0.32, 0.56)
    grey = (0.42, 0.45, 0.50)
    light = (0.95, 0.96, 0.97)
    white = (1, 1, 1)
    cmds = []
    pages = []
    page_no = 0
    image_resources = {}
    logo_resource_name = None
    logo_info = None

    def cmd(line):
        cmds.append(line)

    def color(c, op='rg'):
        return f"{c[0]:.3f} {c[1]:.3f} {c[2]:.3f} {op}"

    def _tw(value, size=9, bold=False):
        value = str(value or '')
        total = 0.0
        for ch in value:
            if ch.isdigit(): total += 0.556
            elif ch in ',.': total += 0.278 if not bold else 0.333
            elif ch in ' -/:()': total += 0.278 if not bold else 0.333
            elif ch in 'ilI': total += 0.240 if not bold else 0.300
            elif ch in 'mwMW': total += 0.800 if not bold else 0.900
            else: total += 0.520 if not bold else 0.600
        return total * size

    def text_at(x, y, value, size=8.5, bold=False, c=dark, align='left'):
        value = str(value or '')
        if align == 'right':
            x -= _tw(value, size, bold)
        elif align == 'center':
            x -= _tw(value, size, bold) / 2
        font = 'F2' if bold else 'F1'
        cmd(f"{color(c)} BT /{font} {size:.2f} Tf {x:.2f} {y:.2f} Td ({_pdf_text_escape(value)}) Tj ET")

    def rect(x, y, w, h, stroke=None, fill=None):
        if fill is not None:
            cmd(f"{color(fill)} {x:.2f} {y:.2f} {w:.2f} {h:.2f} re f")
        if stroke is not None:
            cmd(f"{color(stroke, 'RG')} {x:.2f} {y:.2f} {w:.2f} {h:.2f} re S")

    def line(x1, y1, x2, y2, stroke=(0.78,0.78,0.78), width=0.45):
        cmd(f"{width:.2f} w {color(stroke, 'RG')} {x1:.2f} {y1:.2f} m {x2:.2f} {y2:.2f} l S")

    def draw_image(name, x, y, w, h):
        cmd(f"q {w:.2f} 0 0 {h:.2f} {x:.2f} {y:.2f} cm /{name} Do Q")

    company = payload.get('company') or {}
    try:
        logo_file = company.get('logo_file') or ''
        if logo_file:
            logo_path = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), 'logos', os.path.basename(str(logo_file)))
            logo_info = _get_cached_pdf_image(logo_path) if os.path.exists(logo_path) else None
            if logo_info:
                logo_resource_name = 'AcctLogo'
                image_resources[logo_resource_name] = logo_info
    except Exception:
        logo_resource_name = None
        logo_info = None

    def footer():
        line(margin, 28, page_w - margin, 28, stroke=(0.82,0.82,0.82), width=0.4)
        text_at(margin, 16, company.get('name', ''), 7, False, grey)
        text_at(page_w - margin, 16, f"Page {page_no}", 7, False, grey, 'right')

    def start_page(first=False):
        nonlocal cmds, page_no
        if cmds:
            footer()
            pages.append('\n'.join(cmds))
            cmds = []
        page_no += 1
        y_top = page_h - margin
        if first:
            lx = margin
            ly = y_top
            if logo_resource_name and logo_info:
                max_logo_w, max_logo_h = 100.0, 50.0
                ratio = min(max_logo_w / max(float(logo_info.get('width') or 1), 1), max_logo_h / max(float(logo_info.get('height') or 1), 1))
                logo_w = max(1.0, float(logo_info.get('width') or 1) * ratio)
                logo_h = max(1.0, float(logo_info.get('height') or 1) * ratio)
                draw_image(logo_resource_name, lx, ly - logo_h, logo_w, logo_h)
                detail_x = lx + max_logo_w + 16
            else:
                detail_x = lx
            text_at(detail_x, y_top - 10, company.get('name', ''), 12, True, dark)
            cy = y_top - 24
            if company.get('address'):
                for addr_line in _wrap_pdf_text(company.get('address'), 48)[:2]:
                    text_at(detail_x, cy, addr_line, 8, False, grey)
                    cy -= 10
            if company.get('registration_number'):
                text_at(detail_x, cy, 'Reg No: ' + str(company.get('registration_number')), 8, False, grey)
                cy -= 10
            if company.get('vat_number'):
                text_at(detail_x, cy, 'VAT No: ' + str(company.get('vat_number')), 8, False, grey)
            text_at(page_w - margin, y_top - 10, payload.get('title', 'Accounting Report'), 18, True, blue, 'right')
            text_at(page_w - margin, y_top - 28, payload.get('subtitle', ''), 8.5, False, grey, 'right')
            text_at(page_w - margin, y_top - 41, 'IFRS for SMEs', 8, True, dark, 'right')
            line(margin, y_top - 70, page_w - margin, y_top - 70, stroke=(0.75,0.80,0.86), width=0.8)
            return y_top - 92
        text_at(margin, y_top - 12, payload.get('title', 'Accounting Report') + ' (continued)', 12, True, blue)
        text_at(page_w - margin, y_top - 12, payload.get('subtitle', ''), 8, False, grey, 'right')
        line(margin, y_top - 28, page_w - margin, y_top - 28, stroke=(0.75,0.80,0.86), width=0.5)
        return y_top - 48

    columns = payload.get('columns') or []
    if not columns:
        columns = [{'label':'Description','width':table_w*0.75,'align':'left'}, {'label':'Amount','width':table_w*0.25,'align':'right'}]

    # Normalise accounting report column widths before drawing.
    # Some report payloads only provide labels/alignments. Without widths the old PDF
    # renderer calculated zero-width columns, which made all text print on top of itself.
    has_widths = any(float(c.get('width') or 0) > 0 for c in columns)
    if not has_widths:
        preferred = []
        flexible_indexes = []
        for idx, c in enumerate(columns):
            label = str(c.get('label') or '').strip().lower()
            if 'description' in label or 'account' in label or 'client' in label:
                preferred.append(0.0)
                flexible_indexes.append(idx)
            elif 'date' in label:
                preferred.append(68.0)
            elif 'reference' in label or 'journal' in label:
                preferred.append(76.0)
            elif 'type' in label or 'status' in label or 'normal' in label:
                preferred.append(64.0)
            elif any(token in label for token in ['amount', 'vat', 'debit', 'credit', 'total', 'incl', 'excl', 'balance']):
                preferred.append(82.0)
            elif 'code' in label:
                preferred.append(58.0)
            else:
                preferred.append(74.0)
        fixed_total = sum(preferred)
        flexible_count = max(1, len(flexible_indexes))
        flex_width = max(85.0, (table_w - fixed_total) / flexible_count)
        for idx, c in enumerate(columns):
            if idx in flexible_indexes:
                c['width'] = flex_width
            else:
                c['width'] = preferred[idx]

    total_defined_w = sum(float(c.get('width') or 0) for c in columns)
    scale = table_w / total_defined_w if total_defined_w else 1
    for c in columns:
        c['draw_width'] = max(16.0, float(c.get('width') or 0) * scale)
    col_x = [margin]
    for c in columns[:-1]:
        col_x.append(col_x[-1] + c['draw_width'])

    def draw_table_header(y):
        h = 18
        rect(margin, y - h, table_w, h, stroke=(0.72,0.72,0.72), fill=blue)
        for i, c in enumerate(columns):
            x = col_x[i]
            w = c['draw_width']
            align = c.get('align') or 'left'
            tx = x + 4 if align != 'right' else x + w - 4
            text_at(tx, y - 12, c.get('label', ''), 7.6, True, white, align)
            if i > 0:
                line(x, y - h, x, y, stroke=(0.88,0.88,0.88), width=0.3)
        return y - h

    def cell_lines(value, width):
        value = str(value or '')
        max_chars = max(8, int(width / 4.4))
        return _wrap_pdf_text(value, max_chars) or ['']

    def draw_row(y, row):
        cells = row.get('cells') if isinstance(row, dict) else row
        bold = bool(isinstance(row, dict) and row.get('bold'))
        shade = bool(isinstance(row, dict) and row.get('shade'))
        dark_row = bool(isinstance(row, dict) and row.get('dark'))
        total = bool(isinstance(row, dict) and row.get('total'))
        wrapped = []
        for i, c in enumerate(columns):
            wrapped.append(cell_lines(cells[i] if i < len(cells) else '', c['draw_width'] - 8))
        max_lines = max(len(w) for w in wrapped) if wrapped else 1
        row_h = max(18, 8 + max_lines * 9)
        if y - row_h < 52:
            y = start_page(False)
            y = draw_table_header(y)
        fill = dark if dark_row else (light if shade else None)
        stroke = (0.82,0.82,0.82)
        rect(margin, y - row_h, table_w, row_h, stroke=stroke, fill=fill)
        if total:
            line(margin, y, page_w - margin, y, stroke=(0.18,0.18,0.18), width=0.8)
        for i, c in enumerate(columns):
            x = col_x[i]
            w = c['draw_width']
            if i > 0:
                line(x, y - row_h, x, y, stroke=(0.88,0.88,0.88), width=0.3)
            align = c.get('align') or 'left'
            tx_base = x + 4 if align != 'right' else x + w - 4
            ty = y - 13
            for ln in wrapped[i]:
                text_at(tx_base, ty, ln, 7.5, bold, white if dark_row else dark, align)
                ty -= 9
        return y - row_h

    y = start_page(True)
    for group in payload.get('groups') or []:
        title = group.get('title') or ''
        if title:
            if y - 24 < 52:
                y = start_page(False)
            rect(margin, y - 18, table_w, 18, stroke=(0.82,0.82,0.82), fill=light)
            text_at(margin + 6, y - 12, title, 8.5, True, dark)
            y -= 18
        y = draw_table_header(y)
        rows = group.get('rows') or []
        if not rows:
            y = draw_row(y, _accounting_make_row(['No records found'] + [''] * (len(columns)-1)))
        for row in rows:
            y = draw_row(y, row)
        y -= 10

    if not pages and not cmds:
        y = start_page(True)
    footer()
    pages.append('\n'.join(cmds))
    return _build_raw_pdf(pages, page_w, page_h, image_resources)


def _accounting_pdf_response(pdf_bytes, filename):
    return Response(pdf_bytes, mimetype='application/pdf', headers={'Content-Disposition': f'attachment; filename={filename}'})


# ==========================================================
# ACCOUNTING REPORT EXCEL EXPORTS
# ==========================================================
def _accounting_xlsx_safe_filename(value):
    value = re.sub(r'[^A-Za-z0-9_\-]+', '_', str(value or 'Accounting_Report')).strip('_')
    return (value or 'Accounting_Report') + '.xlsx'


def _safe_excel_sheet_name(value):
    value = re.sub(r'[\\/*?:\[\]]+', ' ', str(value or 'Accounting Report')).strip()
    return (value or 'Accounting Report')[:31]


def _excel_num(value):
    if value is None or value == '':
        return ''
    try:
        return round(float(value or 0), 2)
    except Exception:
        return value


def _accounting_excel_response(workbook, filename):
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


def _build_accounting_report_xlsx(payload):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
    except Exception as exc:
        raise RuntimeError('Excel export requires openpyxl and Pillow. Run pip install -r requirements.txt and restart the app.') from exc

    company = payload.get('company') or {}
    title = payload.get('title') or 'Accounting Report'
    subtitle = payload.get('subtitle') or ''
    columns = payload.get('columns') or [{'label': 'Description'}, {'label': 'Amount', 'align': 'right'}]
    groups = payload.get('groups') or [{'rows': []}]

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_excel_sheet_name(title)

    dark_fill = PatternFill('solid', fgColor='1F2937')
    blue_fill = PatternFill('solid', fgColor='0D4F8B')
    light_fill = PatternFill('solid', fgColor='EEF2F7')
    total_fill = PatternFill('solid', fgColor='E7F0FA')
    white_font = Font(color='FFFFFF', bold=True)
    title_font = Font(size=16, bold=True, color='0D4F8B')
    subtitle_font = Font(size=10, color='6B7280')
    header_font = Font(bold=True, color='FFFFFF')
    bold_font = Font(bold=True)
    thin_side = Side(style='thin', color='D9DEE6')
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    money_format = '#,##0.00;[Red]-#,##0.00;-'

    max_col = max(len(columns), 4)
    detail_col = 1
    logo_added = False
    try:
        logo_file = company.get('logo_file') or ''
        if logo_file:
            logo_path = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), 'logos', os.path.basename(str(logo_file)))
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                max_w, max_h = 120, 64
                ratio = min(max_w / max(float(img.width or 1), 1), max_h / max(float(img.height or 1), 1))
                img.width = int(max(1, float(img.width or 1) * ratio))
                img.height = int(max(1, float(img.height or 1) * ratio))
                ws.add_image(img, 'A1')
                ws.column_dimensions['A'].width = 18
                ws.row_dimensions[1].height = 42
                ws.row_dimensions[2].height = 18
                detail_col = 3
                logo_added = True
    except Exception:
        detail_col = 1
        logo_added = False

    ws.cell(1, detail_col, company.get('name') or 'Company').font = Font(size=13, bold=True, color='111827')
    ws.cell(2, detail_col, company.get('address') or '').font = subtitle_font
    ws.cell(3, detail_col, ('Reg No: ' + str(company.get('registration_number'))) if company.get('registration_number') else '').font = subtitle_font
    ws.cell(4, detail_col, ('VAT No: ' + str(company.get('vat_number'))) if company.get('vat_number') else '').font = subtitle_font

    title_row = 6 if logo_added else 6
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=max_col)
    ws.cell(title_row, 1, title).font = title_font
    ws.cell(title_row, 1).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=title_row + 1, start_column=1, end_row=title_row + 1, end_column=max_col)
    ws.cell(title_row + 1, 1, subtitle).font = subtitle_font
    ws.cell(title_row + 1, 1).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=title_row + 2, start_column=1, end_row=title_row + 2, end_column=max_col)
    ws.cell(title_row + 2, 1, 'IFRS for SMEs').font = Font(size=10, bold=True, color='111827')
    ws.cell(title_row + 2, 1).alignment = Alignment(horizontal='center')

    row_no = title_row + 4

    def style_header_row(r):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(r, c)
            cell.fill = blue_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='right' if (columns[c-1].get('align') == 'right') else 'left')

    def write_table_header():
        nonlocal row_no
        for idx, col in enumerate(columns, start=1):
            ws.cell(row_no, idx, col.get('label') or '')
        style_header_row(row_no)
        row_no += 1

    def write_row(row):
        nonlocal row_no
        cells = row.get('cells') if isinstance(row, dict) else row
        bold = bool(isinstance(row, dict) and row.get('bold'))
        shade = bool(isinstance(row, dict) and row.get('shade'))
        dark = bool(isinstance(row, dict) and row.get('dark'))
        total = bool(isinstance(row, dict) and row.get('total'))
        for idx, col in enumerate(columns, start=1):
            value = cells[idx - 1] if idx - 1 < len(cells) else ''
            cell = ws.cell(row_no, idx, value)
            cell.border = border
            if col.get('align') == 'right' or isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='top')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            if isinstance(value, (int, float)):
                cell.number_format = money_format
            if dark:
                cell.fill = dark_fill
                cell.font = white_font
            elif shade:
                cell.fill = light_fill
                cell.font = bold_font if bold else Font()
            elif total:
                cell.fill = total_fill
                cell.font = bold_font
            elif bold:
                cell.font = bold_font
        row_no += 1

    for group in groups:
        group_title = group.get('title') or ''
        if group_title:
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=len(columns))
            cell = ws.cell(row_no, 1, group_title)
            cell.fill = light_fill
            cell.font = bold_font
            cell.border = border
            row_no += 1
        write_table_header()
        rows = group.get('rows') or []
        if not rows:
            rows = [_accounting_make_row(['No records found'] + [''] * (len(columns) - 1))]
        for r in rows:
            write_row(r)
        row_no += 1

    for idx, col in enumerate(columns, start=1):
        label = str(col.get('label') or '').lower()
        if 'description' in label or 'account' in label:
            width = 34
        elif 'date' in label:
            width = 13
        elif 'reference' in label:
            width = 16
        elif 'source' in label:
            width = 14
        else:
            width = 15
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = f'A{title_row + 4}'
    return wb



@app.route('/download/accounting/reports/trial_balance.pdf')
def download_accounting_trial_balance_pdf():
    if not session.get('can_accounting') and not session.get('is_superadmin'):
        return 'Forbidden', 403
    cid = _current_company_id()
    as_at = request.args.get('as_at') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    ensure_default_accounting_accounts(conn, cid)
    company = _accounting_company_payload(conn, cid)
    rows = _trial_balance_rows(conn, cid, as_at)
    display_rows = []
    debit_total = credit_total = 0.0
    for r in rows:
        if abs(float(r.get('debit') or 0)) < 0.005 and abs(float(r.get('credit') or 0)) < 0.005:
            continue
        debit_total += float(r.get('debit') or 0)
        credit_total += float(r.get('credit') or 0)
        display_rows.append(_accounting_make_row([r.get('account_code'), r.get('account_name'), _accounting_pdf_money(r.get('debit')), _accounting_pdf_money(r.get('credit'))]))
    display_rows.append(_accounting_make_row(['', 'Totals', _accounting_pdf_money(debit_total), _accounting_pdf_money(credit_total)], bold=True, shade=True, total=True))
    conn.close()
    payload = {'company': company, 'title': 'Trial Balance', 'subtitle': f'As at {as_at}', 'columns': [
        {'label':'Code','width':70}, {'label':'Account','width':280}, {'label':'Debit','width':85,'align':'right'}, {'label':'Credit','width':85,'align':'right'}
    ], 'groups': [{'rows': display_rows}]}
    return _accounting_pdf_response(_draw_accounting_report_pdf(payload), _accounting_pdf_safe_filename(f'Trial_Balance_{as_at}'))


@app.route('/download/accounting/reports/income_statement.pdf')
def download_accounting_income_statement_pdf():
    if not session.get('can_accounting') and not session.get('is_superadmin'):
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    data = _income_statement_data(conn, cid, start_date, end_date)
    conn.close()
    rows = []
    labels = [('revenue','Revenue'),('other_income','Other Income'),('cost_of_sales','Cost of Sales'),('operating_expense','Operating Expenses'),('finance_cost','Finance Costs'),('income_tax_expense','Income Tax Expense')]
    for key, label in labels:
        rows.append(_accounting_make_row(['', label, ''], bold=True, shade=True))
        section_total = 0.0
        for r in data.get('sections', {}).get(key, []) or []:
            section_total += float(r.get('amount') or 0)
            rows.append(_accounting_make_row([r.get('account_code'), r.get('account_name'), _accounting_pdf_money(r.get('amount'))]))
        rows.append(_accounting_make_row(['', f'Total {label}', _accounting_pdf_money(section_total)], bold=True, total=True))
        if key == 'cost_of_sales':
            rows.append(_accounting_make_row(['', 'Gross Profit', _accounting_pdf_money(data.get('gross_profit'))], bold=True, shade=True, total=True))
    rows.append(_accounting_make_row(['', 'Profit / (Loss)', _accounting_pdf_money(data.get('profit_loss'))], bold=True, dark=True, total=True))
    payload = {'company': company, 'title': 'Income Statement', 'subtitle': f'{start_date} to {end_date}', 'columns': [
        {'label':'Code','width':70}, {'label':'Description','width':330}, {'label':'Amount','width':120,'align':'right'}
    ], 'groups': [{'rows': rows}]}
    return _accounting_pdf_response(_draw_accounting_report_pdf(payload), _accounting_pdf_safe_filename(f'Income_Statement_{start_date}_to_{end_date}'))


@app.route('/download/accounting/reports/balance_sheet.pdf')
def download_accounting_balance_sheet_pdf():
    if not session.get('can_accounting') and not session.get('is_superadmin'):
        return 'Forbidden', 403
    cid = _current_company_id()
    as_at = request.args.get('as_at') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    data = _balance_sheet_data(conn, cid, as_at)
    conn.close()
    rows = []
    labels = [('current_asset','Current Assets'),('non_current_asset','Non-current Assets'),('current_liability','Current Liabilities'),('non_current_liability','Non-current Liabilities'),('equity','Equity')]
    for key, label in labels:
        rows.append(_accounting_make_row(['', label, ''], bold=True, shade=True))
        section_total = 0.0
        for r in data.get('sections', {}).get(key, []) or []:
            section_total += float(r.get('amount') or 0)
            rows.append(_accounting_make_row([r.get('account_code'), r.get('account_name'), _accounting_pdf_money(r.get('amount'))]))
        rows.append(_accounting_make_row(['', f'Total {label}', _accounting_pdf_money(section_total)], bold=True, total=True))
        if key == 'non_current_asset':
            rows.append(_accounting_make_row(['', 'Total Assets', _accounting_pdf_money(data.get('total_assets'))], bold=True, shade=True, total=True))
        if key == 'equity':
            rows.append(_accounting_make_row(['', 'Total Liabilities and Equity', _accounting_pdf_money(data.get('liabilities_equity_total'))], bold=True, shade=True, total=True))
            rows.append(_accounting_make_row(['', 'Balance Check', _accounting_pdf_money(data.get('balance_check'))], bold=True))
    payload = {'company': company, 'title': 'Balance Sheet', 'subtitle': f'As at {as_at}', 'columns': [
        {'label':'Code','width':70}, {'label':'Description','width':330}, {'label':'Amount','width':120,'align':'right'}
    ], 'groups': [{'rows': rows}]}
    return _accounting_pdf_response(_draw_accounting_report_pdf(payload), _accounting_pdf_safe_filename(f'Balance_Sheet_{as_at}'))


@app.route('/download/accounting/reports/cash_flow.pdf')
def download_accounting_cash_flow_pdf():
    if not session.get('can_accounting') and not session.get('is_superadmin'):
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    data = _cash_flow_data(conn, cid, start_date, end_date)
    conn.close()
    sections = data.get('sections') or {}
    rows = [
        _accounting_make_row(['Opening Cash / Bank', _accounting_pdf_money(data.get('opening_cash'))]),
        _accounting_make_row(['Net cash from Operating Activities', _accounting_pdf_money(sections.get('operating'))]),
        _accounting_make_row(['Net cash from Investing Activities', _accounting_pdf_money(sections.get('investing'))]),
        _accounting_make_row(['Net cash from Financing Activities', _accounting_pdf_money(sections.get('financing'))]),
        _accounting_make_row(['Net Cash Movement', _accounting_pdf_money(data.get('net_movement'))], bold=True, shade=True, total=True),
        _accounting_make_row(['Closing Cash / Bank', _accounting_pdf_money(data.get('closing_cash'))], bold=True, dark=True, total=True),
    ]
    payload = {'company': company, 'title': 'Cash Flow Statement', 'subtitle': f'{start_date} to {end_date}', 'columns': [
        {'label':'Description','width':390}, {'label':'Amount','width':130,'align':'right'}
    ], 'groups': [{'rows': rows}]}
    return _accounting_pdf_response(_draw_accounting_report_pdf(payload), _accounting_pdf_safe_filename(f'Cash_Flow_{start_date}_to_{end_date}'))


@app.route('/download/accounting/reports/general_ledger.pdf')
def download_accounting_general_ledger_pdf():
    if not session.get('can_accounting') and not session.get('is_superadmin'):
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    account_id = request.args.get('account_id') or None
    try:
        account_id = int(account_id) if account_id else None
    except Exception:
        account_id = None
    conn = get_db_connection()
    _normalise_existing_cashbook_ledger_dates(conn, cid)
    company = _accounting_company_payload(conn, cid)
    data = _general_ledger_data(conn, cid, start_date, end_date, account_id)
    conn.close()
    groups = []
    for acc in data:
        rows = [_accounting_make_row(['', '', '', 'Opening Balance', '', '', f"{_accounting_pdf_money(acc.get('opening_balance'))} {acc.get('opening_side') or ''}"], bold=True, shade=True)]
        for ln in acc.get('lines') or []:
            rows.append(_accounting_make_row([
                ln.get('journal_date') or '', ln.get('reference') or '', str(ln.get('source_module') or '').replace('_', ' '),
                ln.get('description') or '', _accounting_pdf_money(ln.get('debit')) if float(ln.get('debit') or 0) else '',
                _accounting_pdf_money(ln.get('credit')) if float(ln.get('credit') or 0) else '',
                f"{_accounting_pdf_money(ln.get('running_balance'))} {ln.get('balance_side') or ''}"
            ]))
        rows.append(_accounting_make_row(['', '', '', 'Account Movement', _accounting_pdf_money(acc.get('debit_total')), _accounting_pdf_money(acc.get('credit_total')), f"{_accounting_pdf_money(acc.get('closing_balance'))} {acc.get('closing_side') or ''}"], bold=True, shade=True, total=True))
        groups.append({'title': f"{acc.get('account_code')} - {acc.get('account_name')}", 'rows': rows})
    payload = {'company': company, 'title': 'General Ledger', 'subtitle': f'{start_date} to {end_date}', 'columns': [
        {'label':'Date','width':58}, {'label':'Reference','width':70}, {'label':'Source','width':58}, {'label':'Description','width':150}, {'label':'Debit','width':60,'align':'right'}, {'label':'Credit','width':60,'align':'right'}, {'label':'Balance','width':70,'align':'right'}
    ], 'groups': groups or [{'rows': []}]}
    return _accounting_pdf_response(_draw_accounting_report_pdf(payload), _accounting_pdf_safe_filename(f'General_Ledger_{start_date}_to_{end_date}'))



@app.route('/download/accounting/reports/trial_balance.xlsx')
def download_accounting_trial_balance_xlsx():
    if not session.get('can_accounting') and not session.get('is_superadmin'):
        return 'Forbidden', 403
    cid = _current_company_id()
    as_at = request.args.get('as_at') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    ensure_default_accounting_accounts(conn, cid)
    company = _accounting_company_payload(conn, cid)
    rows = _trial_balance_rows(conn, cid, as_at)
    display_rows = []
    debit_total = credit_total = 0.0
    for r in rows:
        debit = float(r.get('debit') or 0)
        credit = float(r.get('credit') or 0)
        if abs(debit) < 0.005 and abs(credit) < 0.005:
            continue
        debit_total += debit
        credit_total += credit
        display_rows.append(_accounting_make_row([r.get('account_code'), r.get('account_name'), _excel_num(debit), _excel_num(credit)]))
    display_rows.append(_accounting_make_row(['', 'Totals', _excel_num(debit_total), _excel_num(credit_total)], bold=True, shade=True, total=True))
    conn.close()
    payload = {'company': company, 'title': 'Trial Balance', 'subtitle': f'As at {as_at}', 'columns': [
        {'label':'Code'}, {'label':'Account'}, {'label':'Debit','align':'right'}, {'label':'Credit','align':'right'}
    ], 'groups': [{'rows': display_rows}]}
    return _accounting_excel_response(_build_accounting_report_xlsx(payload), _accounting_xlsx_safe_filename(f'Trial_Balance_{as_at}'))


@app.route('/download/accounting/reports/income_statement.xlsx')
def download_accounting_income_statement_xlsx():
    if not session.get('can_accounting') and not session.get('is_superadmin'):
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    data = _income_statement_data(conn, cid, start_date, end_date)
    conn.close()
    rows = []
    labels = [('revenue','Revenue'),('other_income','Other Income'),('cost_of_sales','Cost of Sales'),('operating_expense','Operating Expenses'),('finance_cost','Finance Costs'),('income_tax_expense','Income Tax Expense')]
    for key, label in labels:
        rows.append(_accounting_make_row(['', label, ''], bold=True, shade=True))
        section_total = 0.0
        for r in data.get('sections', {}).get(key, []) or []:
            amount = float(r.get('amount') or 0)
            section_total += amount
            rows.append(_accounting_make_row([r.get('account_code'), r.get('account_name'), _excel_num(amount)]))
        rows.append(_accounting_make_row(['', f'Total {label}', _excel_num(section_total)], bold=True, total=True))
        if key == 'cost_of_sales':
            rows.append(_accounting_make_row(['', 'Gross Profit', _excel_num(data.get('gross_profit'))], bold=True, shade=True, total=True))
    rows.append(_accounting_make_row(['', 'Profit / (Loss)', _excel_num(data.get('profit_loss'))], bold=True, dark=True, total=True))
    payload = {'company': company, 'title': 'Income Statement', 'subtitle': f'{start_date} to {end_date}', 'columns': [
        {'label':'Code'}, {'label':'Description'}, {'label':'Amount','align':'right'}
    ], 'groups': [{'rows': rows}]}
    return _accounting_excel_response(_build_accounting_report_xlsx(payload), _accounting_xlsx_safe_filename(f'Income_Statement_{start_date}_to_{end_date}'))


@app.route('/download/accounting/reports/balance_sheet.xlsx')
def download_accounting_balance_sheet_xlsx():
    if not session.get('can_accounting') and not session.get('is_superadmin'):
        return 'Forbidden', 403
    cid = _current_company_id()
    as_at = request.args.get('as_at') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    data = _balance_sheet_data(conn, cid, as_at)
    conn.close()
    rows = []
    labels = [('current_asset','Current Assets'),('non_current_asset','Non-current Assets'),('current_liability','Current Liabilities'),('non_current_liability','Non-current Liabilities'),('equity','Equity')]
    for key, label in labels:
        rows.append(_accounting_make_row(['', label, ''], bold=True, shade=True))
        section_total = 0.0
        for r in data.get('sections', {}).get(key, []) or []:
            amount = float(r.get('amount') or 0)
            section_total += amount
            rows.append(_accounting_make_row([r.get('account_code'), r.get('account_name'), _excel_num(amount)]))
        rows.append(_accounting_make_row(['', f'Total {label}', _excel_num(section_total)], bold=True, total=True))
        if key == 'non_current_asset':
            rows.append(_accounting_make_row(['', 'Total Assets', _excel_num(data.get('total_assets'))], bold=True, shade=True, total=True))
        if key == 'equity':
            rows.append(_accounting_make_row(['', 'Total Liabilities and Equity', _excel_num(data.get('liabilities_equity_total'))], bold=True, shade=True, total=True))
            rows.append(_accounting_make_row(['', 'Balance Check', _excel_num(data.get('balance_check'))], bold=True))
    payload = {'company': company, 'title': 'Balance Sheet', 'subtitle': f'As at {as_at}', 'columns': [
        {'label':'Code'}, {'label':'Description'}, {'label':'Amount','align':'right'}
    ], 'groups': [{'rows': rows}]}
    return _accounting_excel_response(_build_accounting_report_xlsx(payload), _accounting_xlsx_safe_filename(f'Balance_Sheet_{as_at}'))


@app.route('/download/accounting/reports/cash_flow.xlsx')
def download_accounting_cash_flow_xlsx():
    if not session.get('can_accounting') and not session.get('is_superadmin'):
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    data = _cash_flow_data(conn, cid, start_date, end_date)
    conn.close()
    sections = data.get('sections') or {}
    rows = [
        _accounting_make_row(['Opening Cash / Bank', _excel_num(data.get('opening_cash'))]),
        _accounting_make_row(['Net cash from Operating Activities', _excel_num(sections.get('operating'))]),
        _accounting_make_row(['Net cash from Investing Activities', _excel_num(sections.get('investing'))]),
        _accounting_make_row(['Net cash from Financing Activities', _excel_num(sections.get('financing'))]),
        _accounting_make_row(['Net Cash Movement', _excel_num(data.get('net_movement'))], bold=True, shade=True, total=True),
        _accounting_make_row(['Closing Cash / Bank', _excel_num(data.get('closing_cash'))], bold=True, dark=True, total=True),
    ]
    payload = {'company': company, 'title': 'Cash Flow Statement', 'subtitle': f'{start_date} to {end_date}', 'columns': [
        {'label':'Description'}, {'label':'Amount','align':'right'}
    ], 'groups': [{'rows': rows}]}
    return _accounting_excel_response(_build_accounting_report_xlsx(payload), _accounting_xlsx_safe_filename(f'Cash_Flow_{start_date}_to_{end_date}'))


@app.route('/download/accounting/reports/general_ledger.xlsx')
def download_accounting_general_ledger_xlsx():
    if not session.get('can_accounting') and not session.get('is_superadmin'):
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    account_id = request.args.get('account_id') or None
    try:
        account_id = int(account_id) if account_id else None
    except Exception:
        account_id = None
    conn = get_db_connection()
    _normalise_existing_cashbook_ledger_dates(conn, cid)
    company = _accounting_company_payload(conn, cid)
    data = _general_ledger_data(conn, cid, start_date, end_date, account_id)
    conn.close()
    groups = []
    for acc in data:
        rows = [_accounting_make_row(['', '', '', 'Opening Balance', '', '', _excel_num(acc.get('opening_balance')), acc.get('opening_side') or ''], bold=True, shade=True)]
        for ln in acc.get('lines') or []:
            rows.append(_accounting_make_row([
                ln.get('journal_date') or '', ln.get('reference') or '', str(ln.get('source_module') or '').replace('_', ' '),
                ln.get('description') or '', _excel_num(ln.get('debit')) if float(ln.get('debit') or 0) else '',
                _excel_num(ln.get('credit')) if float(ln.get('credit') or 0) else '',
                _excel_num(ln.get('running_balance')), ln.get('balance_side') or ''
            ]))
        rows.append(_accounting_make_row(['', '', '', 'Account Movement', _excel_num(acc.get('debit_total')), _excel_num(acc.get('credit_total')), _excel_num(acc.get('closing_balance')), acc.get('closing_side') or ''], bold=True, shade=True, total=True))
        groups.append({'title': f"{acc.get('account_code')} - {acc.get('account_name')}", 'rows': rows})
    payload = {'company': company, 'title': 'General Ledger', 'subtitle': f'{start_date} to {end_date}', 'columns': [
        {'label':'Date'}, {'label':'Reference'}, {'label':'Source'}, {'label':'Description'}, {'label':'Debit','align':'right'}, {'label':'Credit','align':'right'}, {'label':'Balance','align':'right'}, {'label':'Side'}
    ], 'groups': groups or [{'rows': []}]}
    return _accounting_excel_response(_build_accounting_report_xlsx(payload), _accounting_xlsx_safe_filename(f'General_Ledger_{start_date}_to_{end_date}'))

# ==========================================================
# ACCOUNTING MANAGEMENT REPORTS (SALES / COA / VAT)
# ==========================================================
def _accounting_report_require_permission():
    return bool(session.get('can_accounting') or session.get('is_superadmin'))


def _accounting_client_display_from_fields(first_name='', surname='', company_name='', fallback=''):
    person = ' '.join([str(first_name or '').strip(), str(surname or '').strip()]).strip()
    return person or str(company_name or fallback or 'Client').strip() or 'Client'


def _sales_report_data(conn, company_id, start_date, end_date):
    rows = []
    invoice_rows = conn.execute('''SELECT i.*, c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name
                                   FROM invoices i
                                   LEFT JOIN clients c ON c.id=i.client_id AND c.company_id=i.company_id
                                   WHERE i.company_id=? AND i.date>=? AND i.date<=? AND COALESCE(i.accounting_status, '')='posted'
                                   ORDER BY i.date ASC, i.id ASC''', (company_id, start_date, end_date)).fetchall()
    for inv in invoice_rows:
        d = dict(inv)
        vat = _money_float(d.get('vat_amount'))
        total = _money_float(d.get('total'))
        excl = round(total - vat, 2)
        rows.append({
            'date': d.get('date') or '',
            'document_type': 'Invoice',
            'reference': _invoice_formatted_number(conn, company_id, d.get('id')),
            'client': _accounting_client_display_from_fields(d.get('client_first_name'), d.get('client_surname'), d.get('client_company_name'), d.get('client_name')),
            'excl_vat': excl,
            'vat_amount': vat,
            'total_incl_vat': total,
            'journal_id': d.get('accounting_journal_id') or '',
            'journal_reference': f"JRN-{int(d.get('accounting_journal_id')):04d}" if d.get('accounting_journal_id') else ''
        })

    credit_rows = conn.execute('''SELECT cn.*, i.total AS invoice_total, i.vat_amount AS invoice_vat, i.client_name,
                                         c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name
                                  FROM invoice_credit_notes cn
                                  JOIN invoices i ON i.id=cn.invoice_id AND i.company_id=cn.company_id
                                  LEFT JOIN clients c ON c.id=i.client_id AND c.company_id=i.company_id
                                  WHERE cn.company_id=? AND cn.credit_date>=? AND cn.credit_date<=? AND COALESCE(cn.accounting_status, '')='posted'
                                  ORDER BY cn.credit_date ASC, cn.id ASC''', (company_id, start_date, end_date)).fetchall()
    for cn in credit_rows:
        d = dict(cn)
        amount = _money_float(d.get('amount'))
        invoice_total = _money_float(d.get('invoice_total'))
        invoice_vat = _money_float(d.get('invoice_vat'))
        vat_component = round(amount * (invoice_vat / invoice_total), 2) if invoice_total > 0 and invoice_vat > 0 else 0.0
        net_component = round(amount - vat_component, 2)
        rows.append({
            'date': d.get('credit_date') or '',
            'document_type': 'Credit Note',
            'reference': _credit_note_formatted_number(d.get('id')),
            'client': _accounting_client_display_from_fields(d.get('client_first_name'), d.get('client_surname'), d.get('client_company_name'), d.get('client_name')),
            'excl_vat': -net_component,
            'vat_amount': -vat_component,
            'total_incl_vat': -amount,
            'journal_id': d.get('accounting_journal_id') or '',
            'journal_reference': f"JRN-{int(d.get('accounting_journal_id')):04d}" if d.get('accounting_journal_id') else ''
        })
    rows.sort(key=lambda r: (r.get('date') or '', r.get('reference') or ''))
    return {
        'rows': rows,
        'sales_excl_vat': round(sum(_money_float(r.get('excl_vat')) for r in rows), 2),
        'vat_amount': round(sum(_money_float(r.get('vat_amount')) for r in rows), 2),
        'total_incl_vat': round(sum(_money_float(r.get('total_incl_vat')) for r in rows), 2)
    }


def _chart_of_accounts_report_data(conn, company_id):
    rows = []
    type_labels = {'asset': 'Assets', 'liability': 'Liabilities', 'equity': 'Equity', 'income': 'Income', 'cost_of_sales': 'Cost of Sales', 'expense': 'Expenses'}
    for a in _accounting_accounts(conn, company_id, active_only=False):
        rows.append({
            'account_code': a.get('account_code') or '',
            'account_name': a.get('account_name') or '',
            'account_type': a.get('account_type') or '',
            'account_type_label': type_labels.get(a.get('account_type'), a.get('account_type') or ''),
            'report_section': a.get('report_section') or '',
            'report_section_label': IFRS_SECTION_LABELS.get(a.get('report_section'), a.get('report_section') or ''),
            'normal_balance': a.get('normal_balance') or '',
            'cash_flow_category': a.get('cash_flow_category') or '',
            'is_cash_equivalent': int(a.get('is_cash_equivalent') or 0),
            'active': int(a.get('active') or 0)
        })
    return {'rows': rows}



def _bank_reconciliation_accounts(conn, company_id):
    ensure_default_accounting_accounts(conn, company_id)
    rows = conn.execute("""SELECT id, account_code, account_name
                           FROM accounting_accounts
                           WHERE company_id=? AND active=1 AND (COALESCE(is_cash_equivalent,0)=1 OR account_type='asset')
                           ORDER BY COALESCE(is_cash_equivalent,0) DESC, account_code, account_name""", (company_id,)).fetchall()
    accounts = []
    seen = set()
    for r in rows:
        d = dict(r)
        if d.get('id') in seen:
            continue
        seen.add(d.get('id'))
        accounts.append(d)
    return accounts


def _bank_reconciliation_recent(conn, company_id, limit=20):
    rows = conn.execute("""SELECT r.*, a.account_code, a.account_name
                           FROM accounting_bank_reconciliations r
                           LEFT JOIN accounting_accounts a ON a.id=r.bank_account_id AND a.company_id=r.company_id
                           WHERE r.company_id=?
                           ORDER BY r.recon_date DESC, r.id DESC
                           LIMIT ?""", (company_id, limit)).fetchall()
    return [dict(r) for r in rows]


def _bank_reconciliation_selected_line_ids(conn, company_id, reconciliation_id):
    if not reconciliation_id:
        return set()
    rows = conn.execute("""SELECT journal_line_id FROM accounting_bank_reconciliation_lines
                           WHERE company_id=? AND reconciliation_id=? AND COALESCE(reconciled,0)=1""", (company_id, reconciliation_id)).fetchall()
    return {int(r['journal_line_id']) for r in rows if r['journal_line_id'] is not None}


def _bank_reconciliation_data(conn, company_id, start_date, end_date, bank_account_id=None, statement_balance=0, reconciliation_id=None, selected_line_ids=None):
    ensure_default_accounting_accounts(conn, company_id)
    accounts = _bank_reconciliation_accounts(conn, company_id)
    if not bank_account_id and accounts:
        bank_account_id = accounts[0].get('id')
    try:
        bank_account_id = int(bank_account_id or 0)
    except Exception:
        bank_account_id = 0
    bank_account = None
    if bank_account_id:
        bank_account = conn.execute("""SELECT * FROM accounting_accounts
                                       WHERE id=? AND company_id=? AND active=1""", (bank_account_id, company_id)).fetchone()
    if not bank_account:
        return {
            'accounts': accounts, 'bank_account_id': None, 'bank_account': '', 'start_date': start_date, 'end_date': end_date,
            'statement_balance': _money_float(statement_balance), 'gl_balance': 0.0,
            'outstanding_deposits': 0.0, 'outstanding_payments': 0.0,
            'adjusted_statement_balance': _money_float(statement_balance), 'difference_amount': 0.0,
            'rows': [], 'recent': _bank_reconciliation_recent(conn, company_id), 'reconciliation_id': reconciliation_id
        }

    if reconciliation_id:
        rec_row = conn.execute("""SELECT * FROM accounting_bank_reconciliations
                                  WHERE id=? AND company_id=?""", (reconciliation_id, company_id)).fetchone()
        if rec_row:
            recd = dict(rec_row)
            bank_account_id = int(recd.get('bank_account_id') or bank_account_id)
            start_date = recd.get('start_date') or start_date
            end_date = recd.get('recon_date') or end_date
            statement_balance = recd.get('statement_balance')
            bank_account = conn.execute("""SELECT * FROM accounting_accounts
                                           WHERE id=? AND company_id=? AND active=1""", (bank_account_id, company_id)).fetchone()

    statement_balance = _money_float(statement_balance)
    bal = conn.execute("""SELECT COALESCE(SUM(COALESCE(l.debit,0) - COALESCE(l.credit,0)), 0) AS balance
                          FROM accounting_journal_lines l
                          JOIN accounting_journals j ON j.id=l.journal_id AND j.company_id=l.company_id
                          WHERE l.company_id=? AND l.account_id=? AND COALESCE(j.status,'')='posted'
                                AND j.journal_date<=?""", (company_id, bank_account_id, end_date)).fetchone()
    gl_balance = _money_float(bal.get('balance') if hasattr(bal, 'get') else (bal['balance'] if bal else 0))

    trans = conn.execute("""SELECT l.id AS journal_line_id, l.journal_id, l.description AS line_description, l.debit, l.credit,
                                  j.journal_date, j.reference, j.description AS journal_description, j.source_module, j.source_record_type, j.source_record_id
                           FROM accounting_journal_lines l
                           JOIN accounting_journals j ON j.id=l.journal_id AND j.company_id=l.company_id
                           WHERE l.company_id=? AND l.account_id=? AND COALESCE(j.status,'')='posted'
                                 AND j.journal_date>=? AND j.journal_date<=?
                           ORDER BY j.journal_date ASC, j.id ASC, l.line_no ASC, l.id ASC""", (company_id, bank_account_id, start_date, end_date)).fetchall()
    rows = []
    # Bank reconciliation now compares the GL bank balance directly to the bank statement
    # closing balance. Outstanding deposits/payments are no longer used in the calculation.
    outstanding_deposits = 0.0
    outstanding_payments = 0.0
    for r in trans:
        d = dict(r)
        debit = _money_float(d.get('debit'))
        credit = _money_float(d.get('credit'))
        movement = round(debit - credit, 2)
        rows.append({
            'journal_line_id': d.get('journal_line_id'),
            'journal_id': d.get('journal_id'),
            'date': d.get('journal_date') or '',
            'reference': d.get('reference') or f"JRN-{int(d.get('journal_id') or 0):04d}",
            'source': str(d.get('source_module') or '').replace('_', ' ').title(),
            'description': d.get('line_description') or d.get('journal_description') or '',
            'debit': debit,
            'credit': credit,
            'movement': movement
        })
    adjusted_statement_balance = round(statement_balance, 2)
    difference_amount = round(gl_balance - statement_balance, 2)
    return {
        'accounts': accounts,
        'bank_account_id': bank_account_id,
        'bank_account': f"{bank_account['account_code']} - {bank_account['account_name']}" if bank_account else '',
        'start_date': start_date,
        'end_date': end_date,
        'statement_balance': statement_balance,
        'gl_balance': gl_balance,
        'outstanding_deposits': outstanding_deposits,
        'outstanding_payments': outstanding_payments,
        'adjusted_statement_balance': adjusted_statement_balance,
        'difference_amount': difference_amount,
        'rows': rows,
        'recent': _bank_reconciliation_recent(conn, company_id),
        'reconciliation_id': int(reconciliation_id or 0) if reconciliation_id else None
    }

def _vat_report_data(conn, company_id, start_date, end_date):
    # SARS-friendly VAT report: show both the VAT amount and the gross VAT-applicable
    # transaction value for the selected period. VAT Control account movements alone
    # do not always contain the gross amount, so source documents are used where possible.
    settings = _accounting_posting_settings(conn, company_id)
    vat_account_id = settings.get('vat_control_account_id') or settings.get('vat_output_account_id')
    vat_account = None
    if vat_account_id:
        vat_account = conn.execute('SELECT * FROM accounting_accounts WHERE id=? AND company_id=?', (vat_account_id, company_id)).fetchone()
    if not vat_account:
        return {
            'vat_account': '', 'vat_account_id': None, 'rows': [],
            'output_gross': 0.0, 'output_net': 0.0, 'output_vat': 0.0,
            'input_gross': 0.0, 'input_net': 0.0, 'input_vat': 0.0,
            'vat_adjustments': 0.0, 'net_vat': 0.0
        }

    rows = []
    source_journal_ids = set()

    invoice_rows = conn.execute('''SELECT i.*, c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name
                                   FROM invoices i
                                   LEFT JOIN clients c ON c.id=i.client_id AND c.company_id=i.company_id
                                   WHERE i.company_id=? AND i.date>=? AND i.date<=? AND COALESCE(i.accounting_status, '')='posted'
                                         AND COALESCE(i.vat_amount, 0) <> 0
                                   ORDER BY i.date ASC, i.id ASC''', (company_id, start_date, end_date)).fetchall()
    for inv in invoice_rows:
        d = dict(inv)
        gross = _money_float(d.get('total'))
        vat = _money_float(d.get('vat_amount'))
        net = round(gross - vat, 2)
        if d.get('accounting_journal_id'):
            source_journal_ids.add(int(d.get('accounting_journal_id')))
        ref = _invoice_formatted_number(conn, company_id, d.get('id'))
        rows.append({
            'date': d.get('date') or '',
            'reference': ref,
            'source': 'Invoice',
            'description': _accounting_client_display_from_fields(d.get('client_first_name'), d.get('client_surname'), d.get('client_company_name'), d.get('client_name')),
            'vat_type': 'Output',
            'gross_amount': gross,
            'net_amount': net,
            'vat_amount': vat,
            'output_gross': gross,
            'output_net': net,
            'output_vat': vat,
            'input_gross': 0.0,
            'input_net': 0.0,
            'input_vat': 0.0,
            'vat_adjustment': 0.0,
            'net_vat': vat,
        })

    credit_rows = conn.execute('''SELECT cn.*, i.total AS invoice_total, i.vat_amount AS invoice_vat, i.client_name,
                                         c.name AS client_first_name, c.surname AS client_surname, c.company_name AS client_company_name
                                  FROM invoice_credit_notes cn
                                  JOIN invoices i ON i.id=cn.invoice_id AND i.company_id=cn.company_id
                                  LEFT JOIN clients c ON c.id=i.client_id AND c.company_id=i.company_id
                                  WHERE cn.company_id=? AND cn.credit_date>=? AND cn.credit_date<=? AND COALESCE(cn.accounting_status, '')='posted'
                                  ORDER BY cn.credit_date ASC, cn.id ASC''', (company_id, start_date, end_date)).fetchall()
    for cn in credit_rows:
        d = dict(cn)
        gross_abs = _money_float(d.get('amount'))
        invoice_total = _money_float(d.get('invoice_total'))
        invoice_vat = _money_float(d.get('invoice_vat'))
        vat_abs = round(gross_abs * (invoice_vat / invoice_total), 2) if invoice_total > 0 and invoice_vat > 0 else 0.0
        if vat_abs <= 0:
            continue
        net_abs = round(gross_abs - vat_abs, 2)
        gross = -gross_abs
        vat = -vat_abs
        net = -net_abs
        if d.get('accounting_journal_id'):
            source_journal_ids.add(int(d.get('accounting_journal_id')))
        rows.append({
            'date': d.get('credit_date') or '',
            'reference': _credit_note_formatted_number(d.get('id')),
            'source': 'Credit Note',
            'description': _accounting_client_display_from_fields(d.get('client_first_name'), d.get('client_surname'), d.get('client_company_name'), d.get('client_name')),
            'vat_type': 'Output',
            'gross_amount': gross,
            'net_amount': net,
            'vat_amount': vat,
            'output_gross': gross,
            'output_net': net,
            'output_vat': vat,
            'input_gross': 0.0,
            'input_net': 0.0,
            'input_vat': 0.0,
            'vat_adjustment': 0.0,
            'net_vat': vat,
        })

    cashbook_rows = conn.execute('''SELECT l.*, b.id AS batch_id, b.bank_account_id, a.account_code AS allocated_account_code, a.account_name AS allocated_account_name
                                    FROM accounting_cashbook_lines l
                                    JOIN accounting_cashbook_batches b ON b.id=l.batch_id AND b.company_id=l.company_id
                                    LEFT JOIN accounting_accounts a ON a.id=l.allocated_account_id AND a.company_id=l.company_id
                                    WHERE l.company_id=? AND COALESCE(l.status, '')='posted' AND COALESCE(l.vat_amount, 0) <> 0
                                          AND l.transaction_date>=? AND l.transaction_date<=?
                                    ORDER BY l.transaction_date ASC, l.batch_id ASC, l.line_no ASC''', (company_id, start_date, end_date)).fetchall()
    for cb in cashbook_rows:
        d = dict(cb)
        debit = _money_float(d.get('debit'))
        credit = _money_float(d.get('credit'))
        gross_abs = round(credit if credit > 0 else debit, 2)
        vat_abs = _money_float(d.get('vat_amount'))
        if gross_abs <= 0 or vat_abs <= 0:
            continue
        net_abs = round(gross_abs - vat_abs, 2)
        if d.get('linked_journal_id'):
            source_journal_ids.add(int(d.get('linked_journal_id')))
        desc = d.get('description') or f"Cash book line {d.get('line_no')}"
        account_label = ''
        if d.get('allocated_account_code') or d.get('allocated_account_name'):
            account_label = f" · {d.get('allocated_account_code') or ''} {d.get('allocated_account_name') or ''}".strip()
        if debit > 0:
            rows.append({
                'date': d.get('transaction_date') or '',
                'reference': f"CB-{d.get('batch_id')}-{d.get('line_no')}",
                'source': 'Cash Book',
                'description': f"{desc}{account_label}",
                'vat_type': 'Input',
                'gross_amount': gross_abs,
                'net_amount': net_abs,
                'vat_amount': vat_abs,
                'output_gross': 0.0,
                'output_net': 0.0,
                'output_vat': 0.0,
                'input_gross': gross_abs,
                'input_net': net_abs,
                'input_vat': vat_abs,
                'vat_adjustment': 0.0,
                'net_vat': -vat_abs,
            })
        else:
            rows.append({
                'date': d.get('transaction_date') or '',
                'reference': f"CB-{d.get('batch_id')}-{d.get('line_no')}",
                'source': 'Cash Book',
                'description': f"{desc}{account_label}",
                'vat_type': 'Output',
                'gross_amount': gross_abs,
                'net_amount': net_abs,
                'vat_amount': vat_abs,
                'output_gross': gross_abs,
                'output_net': net_abs,
                'output_vat': vat_abs,
                'input_gross': 0.0,
                'input_net': 0.0,
                'input_vat': 0.0,
                'vat_adjustment': 0.0,
                'net_vat': vat_abs,
            })

    adj_sql = '''SELECT j.id AS journal_id, j.journal_date, j.reference, j.source_module, j.source_record_type, j.description AS journal_description,
                        l.description AS line_description, l.debit, l.credit, l.gross_amount, l.net_amount, l.vat_amount, l.vat_type
                 FROM accounting_journal_lines l
                 JOIN accounting_journals j ON j.id=l.journal_id AND j.company_id=l.company_id
                 WHERE l.company_id=? AND l.account_id=? AND j.status='posted' AND j.journal_date>=? AND j.journal_date<=?
                       AND COALESCE(j.source_module, '') NOT IN ('invoice', 'credit_note', 'cash_book')
                 ORDER BY j.journal_date ASC, j.id ASC, l.line_no ASC'''
    for r in conn.execute(adj_sql, (company_id, vat_account_id, start_date, end_date)).fetchall():
        d = dict(r)
        if d.get('journal_id') and int(d.get('journal_id')) in source_journal_ids:
            continue
        debit = _money_float(d.get('debit'))
        credit = _money_float(d.get('credit'))
        gross = _money_float(d.get('gross_amount'))
        net = _money_float(d.get('net_amount'))
        vat_meta = _money_float(d.get('vat_amount'))
        vat_type = (d.get('vat_type') or '').strip().title()
        if gross > 0 and vat_meta > 0 and vat_type in ['Input', 'Output']:
            if vat_type == 'Input':
                rows.append({
                    'date': d.get('journal_date') or '',
                    'reference': d.get('reference') or '',
                    'source': str(d.get('source_module') or d.get('source_record_type') or 'manual').replace('_', ' ').title(),
                    'description': d.get('line_description') or d.get('journal_description') or 'Manual journal VAT',
                    'vat_type': 'Input',
                    'gross_amount': gross,
                    'net_amount': net,
                    'vat_amount': vat_meta,
                    'output_gross': 0.0,
                    'output_net': 0.0,
                    'output_vat': 0.0,
                    'input_gross': gross,
                    'input_net': net,
                    'input_vat': vat_meta,
                    'vat_adjustment': 0.0,
                    'net_vat': -vat_meta,
                })
            else:
                rows.append({
                    'date': d.get('journal_date') or '',
                    'reference': d.get('reference') or '',
                    'source': str(d.get('source_module') or d.get('source_record_type') or 'manual').replace('_', ' ').title(),
                    'description': d.get('line_description') or d.get('journal_description') or 'Manual journal VAT',
                    'vat_type': 'Output',
                    'gross_amount': gross,
                    'net_amount': net,
                    'vat_amount': vat_meta,
                    'output_gross': gross,
                    'output_net': net,
                    'output_vat': vat_meta,
                    'input_gross': 0.0,
                    'input_net': 0.0,
                    'input_vat': 0.0,
                    'vat_adjustment': 0.0,
                    'net_vat': vat_meta,
                })
            continue
        effect = round(credit - debit, 2)
        if abs(effect) <= 0.005:
            continue
        rows.append({
            'date': d.get('journal_date') or '',
            'reference': d.get('reference') or '',
            'source': str(d.get('source_module') or d.get('source_record_type') or 'manual').replace('_', ' ').title(),
            'description': d.get('line_description') or d.get('journal_description') or 'VAT adjustment',
            'vat_type': 'Adjustment',
            'gross_amount': 0.0,
            'net_amount': 0.0,
            'vat_amount': effect,
            'output_gross': 0.0,
            'output_net': 0.0,
            'output_vat': 0.0,
            'input_gross': 0.0,
            'input_net': 0.0,
            'input_vat': 0.0,
            'vat_adjustment': effect,
            'net_vat': effect,
        })

    rows.sort(key=lambda r: (r.get('date') or '', r.get('reference') or '', r.get('source') or ''))
    output_gross = round(sum(_money_float(r.get('output_gross')) for r in rows), 2)
    output_net = round(sum(_money_float(r.get('output_net')) for r in rows), 2)
    output_vat = round(sum(_money_float(r.get('output_vat')) for r in rows), 2)
    input_gross = round(sum(_money_float(r.get('input_gross')) for r in rows), 2)
    input_net = round(sum(_money_float(r.get('input_net')) for r in rows), 2)
    input_vat = round(sum(_money_float(r.get('input_vat')) for r in rows), 2)
    vat_adjustments = round(sum(_money_float(r.get('vat_adjustment')) for r in rows), 2)
    return {
        'vat_account': f"{vat_account['account_code']} - {vat_account['account_name']}",
        'vat_account_id': vat_account_id,
        'rows': rows,
        'output_gross': output_gross,
        'output_net': output_net,
        'output_vat': output_vat,
        'input_gross': input_gross,
        'input_net': input_net,
        'input_vat': input_vat,
        'vat_adjustments': vat_adjustments,
        'net_vat': round(output_vat - input_vat + vat_adjustments, 2)
    }

@app.route('/api/accounting/reports/sales')
def accounting_sales_report():
    if not _accounting_report_require_permission():
        return jsonify({'status': 'error', 'message': 'Accounting permission required.'}), 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    data = _sales_report_data(conn, cid, start_date, end_date)
    conn.close()
    data.update({'status': 'success', 'start_date': start_date, 'end_date': end_date, 'framework': 'IFRS for SMEs'})
    return jsonify(data)


@app.route('/api/accounting/reports/chart_of_accounts')
def accounting_chart_of_accounts_report():
    if not _accounting_report_require_permission():
        return jsonify({'status': 'error', 'message': 'Accounting permission required.'}), 403
    cid = _current_company_id()
    conn = get_db_connection()
    ensure_default_accounting_accounts(conn, cid)
    data = _chart_of_accounts_report_data(conn, cid)
    conn.close()
    data.update({'status': 'success', 'framework': 'IFRS for SMEs'})
    return jsonify(data)


@app.route('/api/accounting/reports/vat')
def accounting_vat_report():
    if not _accounting_report_require_permission():
        return jsonify({'status': 'error', 'message': 'Accounting permission required.'}), 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    data = _vat_report_data(conn, cid, start_date, end_date)
    conn.close()
    data.update({'status': 'success', 'start_date': start_date, 'end_date': end_date, 'framework': 'IFRS for SMEs'})
    return jsonify(data)


@app.route('/download/accounting/reports/sales.pdf')
def download_accounting_sales_pdf():
    if not _accounting_report_require_permission():
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    data = _sales_report_data(conn, cid, start_date, end_date)
    conn.close()
    rows = [_accounting_make_row([r.get('date'), r.get('document_type'), r.get('reference'), r.get('client'), _accounting_pdf_money(r.get('excl_vat')), _accounting_pdf_money(r.get('vat_amount')), _accounting_pdf_money(r.get('total_incl_vat'))]) for r in data.get('rows') or []]
    rows.append(_accounting_make_row(['', '', '', 'Totals', _accounting_pdf_money(data.get('sales_excl_vat')), _accounting_pdf_money(data.get('vat_amount')), _accounting_pdf_money(data.get('total_incl_vat'))], bold=True, shade=True, total=True))
    payload = {'company': company, 'title': 'Sales Report', 'subtitle': f'{start_date} to {end_date}', 'columns': [
        {'label':'Date'}, {'label':'Type'}, {'label':'Reference'}, {'label':'Client'}, {'label':'Excl. VAT','align':'right'}, {'label':'VAT','align':'right'}, {'label':'Incl. VAT','align':'right'}
    ], 'groups': [{'rows': rows}]}
    return _accounting_pdf_response(_draw_accounting_report_pdf(payload), _accounting_pdf_safe_filename(f'Sales_Report_{start_date}_to_{end_date}'))


@app.route('/download/accounting/reports/sales.xlsx')
def download_accounting_sales_xlsx():
    if not _accounting_report_require_permission():
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    data = _sales_report_data(conn, cid, start_date, end_date)
    conn.close()
    rows = [_accounting_make_row([r.get('date'), r.get('document_type'), r.get('reference'), r.get('client'), _excel_num(r.get('excl_vat')), _excel_num(r.get('vat_amount')), _excel_num(r.get('total_incl_vat')), r.get('journal_reference') or '']) for r in data.get('rows') or []]
    rows.append(_accounting_make_row(['', '', '', 'Totals', _excel_num(data.get('sales_excl_vat')), _excel_num(data.get('vat_amount')), _excel_num(data.get('total_incl_vat')), ''], bold=True, shade=True, total=True))
    payload = {'company': company, 'title': 'Sales Report', 'subtitle': f'{start_date} to {end_date}', 'columns': [
        {'label':'Date'}, {'label':'Type'}, {'label':'Reference'}, {'label':'Client'}, {'label':'Excl. VAT','align':'right'}, {'label':'VAT','align':'right'}, {'label':'Incl. VAT','align':'right'}, {'label':'Journal'}
    ], 'groups': [{'rows': rows}]}
    return _accounting_excel_response(_build_accounting_report_xlsx(payload), _accounting_xlsx_safe_filename(f'Sales_Report_{start_date}_to_{end_date}'))


@app.route('/download/accounting/reports/chart_of_accounts.pdf')
def download_accounting_chart_of_accounts_pdf():
    if not _accounting_report_require_permission():
        return 'Forbidden', 403
    cid = _current_company_id()
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    ensure_default_accounting_accounts(conn, cid)
    data = _chart_of_accounts_report_data(conn, cid)
    conn.close()
    rows = [_accounting_make_row([r.get('account_code'), r.get('account_name'), r.get('account_type_label'), r.get('report_section_label'), r.get('normal_balance'), r.get('cash_flow_category'), 'Yes' if r.get('is_cash_equivalent') else 'No', 'Active' if r.get('active') else 'Inactive']) for r in data.get('rows') or []]
    payload = {'company': company, 'title': 'Chart of Accounts Report', 'subtitle': 'IFRS for SMEs account structure', 'columns': [
        {'label':'Code'}, {'label':'Account'}, {'label':'Type'}, {'label':'Report Section'}, {'label':'Normal'}, {'label':'Cash Flow'}, {'label':'Cash/Bank'}, {'label':'Status'}
    ], 'groups': [{'rows': rows}]}
    return _accounting_pdf_response(_draw_accounting_report_pdf(payload), _accounting_pdf_safe_filename('Chart_of_Accounts_Report'))


@app.route('/download/accounting/reports/chart_of_accounts.xlsx')
def download_accounting_chart_of_accounts_xlsx():
    if not _accounting_report_require_permission():
        return 'Forbidden', 403
    cid = _current_company_id()
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    ensure_default_accounting_accounts(conn, cid)
    data = _chart_of_accounts_report_data(conn, cid)
    conn.close()
    rows = [_accounting_make_row([r.get('account_code'), r.get('account_name'), r.get('account_type_label'), r.get('report_section_label'), r.get('normal_balance'), r.get('cash_flow_category'), 'Yes' if r.get('is_cash_equivalent') else 'No', 'Active' if r.get('active') else 'Inactive']) for r in data.get('rows') or []]
    payload = {'company': company, 'title': 'Chart of Accounts Report', 'subtitle': 'IFRS for SMEs account structure', 'columns': [
        {'label':'Code'}, {'label':'Account'}, {'label':'Type'}, {'label':'Report Section'}, {'label':'Normal'}, {'label':'Cash Flow'}, {'label':'Cash/Bank'}, {'label':'Status'}
    ], 'groups': [{'rows': rows}]}
    return _accounting_excel_response(_build_accounting_report_xlsx(payload), _accounting_xlsx_safe_filename('Chart_of_Accounts_Report'))


@app.route('/download/accounting/reports/vat.pdf')
def download_accounting_vat_pdf():
    if not _accounting_report_require_permission():
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    data = _vat_report_data(conn, cid, start_date, end_date)
    conn.close()
    net_label = 'VAT Payable' if _money_float(data.get('net_vat')) >= 0 else 'VAT Refundable'
    summary_rows = [
        _accounting_make_row(['', '', '', 'Output VAT-applicable sales/receipts', _accounting_pdf_money(data.get('output_gross')), _accounting_pdf_money(data.get('output_net')), _accounting_pdf_money(data.get('output_vat'))], bold=True, shade=True),
        _accounting_make_row(['', '', '', 'Input VAT-applicable purchases/expenses', _accounting_pdf_money(data.get('input_gross')), _accounting_pdf_money(data.get('input_net')), _accounting_pdf_money(data.get('input_vat'))], bold=True, shade=True),
    ]
    if abs(_money_float(data.get('vat_adjustments'))) > 0.005:
        summary_rows.append(_accounting_make_row(['', '', '', 'VAT adjustments without gross value', '', '', _accounting_pdf_money(data.get('vat_adjustments'))], bold=True, shade=True))
    summary_rows.append(_accounting_make_row(['', '', '', net_label, '', '', _accounting_pdf_money(abs(_money_float(data.get('net_vat'))))], bold=True, shade=True, total=True))

    detail_rows = []
    for r in data.get('rows') or []:
        source_desc = ((r.get('source') or '') + ': ' + (r.get('description') or '')).strip(': ')
        detail_rows.append(_accounting_make_row([
            r.get('date'), r.get('reference'), r.get('vat_type'), source_desc,
            _accounting_pdf_money(r.get('gross_amount')) if _money_float(r.get('gross_amount')) else '',
            _accounting_pdf_money(r.get('net_amount')) if _money_float(r.get('net_amount')) else '',
            _accounting_pdf_money(r.get('vat_amount')) if _money_float(r.get('vat_amount')) else ''
        ]))
    detail_rows.append(_accounting_make_row(['', '', '', 'Totals', _accounting_pdf_money(_money_float(data.get('output_gross')) + _money_float(data.get('input_gross'))), _accounting_pdf_money(_money_float(data.get('output_net')) + _money_float(data.get('input_net'))), _accounting_pdf_money(data.get('net_vat'))], bold=True, shade=True, total=True))
    payload = {'company': company, 'title': 'VAT Report', 'subtitle': f'{start_date} to {end_date} · {data.get("vat_account") or "VAT Control Account not configured"}', 'orientation': 'landscape', 'columns': [
        {'label':'Date','width':70}, {'label':'Reference','width':78}, {'label':'VAT Type','width':62}, {'label':'Description','width':330}, {'label':'Gross Amount','width':86,'align':'right'}, {'label':'Net Amount','width':86,'align':'right'}, {'label':'VAT Amount','width':86,'align':'right'}
    ], 'groups': [{'title': 'SARS VAT Summary', 'rows': summary_rows}, {'title': 'VAT Transaction Detail', 'rows': detail_rows}]}
    return _accounting_pdf_response(_draw_accounting_report_pdf(payload), _accounting_pdf_safe_filename(f'VAT_Report_{start_date}_to_{end_date}'))


@app.route('/download/accounting/reports/vat.xlsx')
def download_accounting_vat_xlsx():
    if not _accounting_report_require_permission():
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    company = _accounting_company_payload(conn, cid)
    data = _vat_report_data(conn, cid, start_date, end_date)
    conn.close()
    net_label = 'VAT Payable' if _money_float(data.get('net_vat')) >= 0 else 'VAT Refundable'
    summary_rows = [
        _accounting_make_row(['', '', '', 'Output VAT-applicable sales/receipts', _excel_num(data.get('output_gross')), _excel_num(data.get('output_net')), _excel_num(data.get('output_vat'))], bold=True, shade=True),
        _accounting_make_row(['', '', '', 'Input VAT-applicable purchases/expenses', _excel_num(data.get('input_gross')), _excel_num(data.get('input_net')), _excel_num(data.get('input_vat'))], bold=True, shade=True),
    ]
    if abs(_money_float(data.get('vat_adjustments'))) > 0.005:
        summary_rows.append(_accounting_make_row(['', '', '', 'VAT adjustments without gross value', '', '', _excel_num(data.get('vat_adjustments'))], bold=True, shade=True))
    summary_rows.append(_accounting_make_row(['', '', '', net_label, '', '', _excel_num(abs(_money_float(data.get('net_vat'))))], bold=True, shade=True, total=True))

    detail_rows = []
    for r in data.get('rows') or []:
        source_desc = ((r.get('source') or '') + ': ' + (r.get('description') or '')).strip(': ')
        detail_rows.append(_accounting_make_row([
            r.get('date'), r.get('reference'), r.get('vat_type'), source_desc,
            _excel_num(r.get('gross_amount')) if _money_float(r.get('gross_amount')) else '',
            _excel_num(r.get('net_amount')) if _money_float(r.get('net_amount')) else '',
            _excel_num(r.get('vat_amount')) if _money_float(r.get('vat_amount')) else ''
        ]))
    detail_rows.append(_accounting_make_row(['', '', '', 'Totals', _excel_num(_money_float(data.get('output_gross')) + _money_float(data.get('input_gross'))), _excel_num(_money_float(data.get('output_net')) + _money_float(data.get('input_net'))), _excel_num(data.get('net_vat'))], bold=True, shade=True, total=True))
    payload = {'company': company, 'title': 'VAT Report', 'subtitle': f'{start_date} to {end_date} · {data.get("vat_account") or "VAT Control Account not configured"}', 'columns': [
        {'label':'Date'}, {'label':'Reference'}, {'label':'VAT Type'}, {'label':'Description'}, {'label':'Gross Amount','align':'right'}, {'label':'Net Amount','align':'right'}, {'label':'VAT Amount','align':'right'}
    ], 'groups': [{'title': 'SARS VAT Summary', 'rows': summary_rows}, {'title': 'VAT Transaction Detail', 'rows': detail_rows}]}
    return _accounting_excel_response(_build_accounting_report_xlsx(payload), _accounting_xlsx_safe_filename(f'VAT_Report_{start_date}_to_{end_date}'))

@app.route('/api/accounting/reports/bank_reconciliation')
def accounting_bank_reconciliation_report():
    if not _accounting_report_require_permission():
        return jsonify({'status': 'error', 'message': 'Forbidden'}), 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    bank_account_id = request.args.get('bank_account_id') or None
    statement_balance = request.args.get('statement_balance') or 0
    reconciliation_id = request.args.get('reconciliation_id') or None
    try:
        reconciliation_id = int(reconciliation_id) if reconciliation_id else None
    except Exception:
        reconciliation_id = None
    conn = get_db_connection()
    try:
        data = _bank_reconciliation_data(conn, cid, start_date, end_date, bank_account_id, statement_balance, reconciliation_id)
        data['status'] = 'success'
        return jsonify(data)
    finally:
        conn.close()


@app.route('/api/accounting/reports/bank_reconciliation/save', methods=['POST'])
def accounting_bank_reconciliation_save():
    if not _accounting_report_require_permission():
        return jsonify({'status': 'error', 'message': 'Forbidden'}), 403
    cid = _current_company_id()
    data = request.get_json() or {}
    start_date = data.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = data.get('end_date') or data.get('recon_date') or datetime.now().strftime('%Y-%m-%d')
    bank_account_id = data.get('bank_account_id')
    statement_balance = _money_float(data.get('statement_balance'))
    notes = (data.get('notes') or '').strip()
    try:
        bank_account_id = int(bank_account_id or 0)
    except Exception:
        bank_account_id = 0
    conn = get_db_connection()
    try:
        acct = conn.execute("""SELECT id FROM accounting_accounts WHERE id=? AND company_id=? AND active=1""", (bank_account_id, cid)).fetchone()
        if not acct:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Please select a valid bank account.'}), 400
        calc = _bank_reconciliation_data(conn, cid, start_date, end_date, bank_account_id, statement_balance)
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cur = conn.execute("""INSERT INTO accounting_bank_reconciliations
                              (company_id, bank_account_id, start_date, recon_date, statement_balance, gl_balance,
                               outstanding_deposits, outstanding_payments, adjusted_statement_balance, difference_amount,
                               notes, created_by, updated_at)
                              VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                           (cid, bank_account_id, start_date, end_date, calc.get('statement_balance'), calc.get('gl_balance'),
                            calc.get('outstanding_deposits'), calc.get('outstanding_payments'), calc.get('adjusted_statement_balance'),
                            calc.get('difference_amount'), notes, session.get('username'), now))
        reconciliation_id = cur.lastrowid
        conn.commit()
        saved = _bank_reconciliation_data(conn, cid, start_date, end_date, bank_account_id, statement_balance, reconciliation_id)
        saved['status'] = 'success'
        saved['message'] = 'Bank reconciliation saved.'
        log_action('Accounting', 'Saved Bank Reconciliation', f'Saved bank reconciliation {reconciliation_id} for account {bank_account_id}.')
        return jsonify(saved)
    except Exception as exc:
        try: conn.rollback()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(exc)}), 500
    finally:
        try: conn.close()
        except Exception: pass


def _bank_reconciliation_report_payload(conn, company_id, start_date, end_date, bank_account_id=None, statement_balance=0, reconciliation_id=None):
    company = _accounting_company_payload(conn, company_id)
    data = _bank_reconciliation_data(conn, company_id, start_date, end_date, bank_account_id, statement_balance, reconciliation_id)
    summary_rows = [
        _accounting_make_row(['Bank Account', data.get('bank_account') or '', '', ''], bold=True, shade=True),
        _accounting_make_row(['Statement Closing Balance', _accounting_pdf_money(data.get('statement_balance')), '', '']),
        _accounting_make_row(['General Ledger Bank Balance', _accounting_pdf_money(data.get('gl_balance')), '', '']),
        _accounting_make_row(['Difference', _accounting_pdf_money(data.get('difference_amount')), '', ''], bold=True, shade=True, total=True),
    ]
    detail_rows = []
    for r in data.get('rows') or []:
        detail_rows.append(_accounting_make_row([
            r.get('date') or '', r.get('reference') or '', r.get('source') or '', r.get('description') or '',
            _accounting_pdf_money(r.get('debit')) if _money_float(r.get('debit')) else '',
            _accounting_pdf_money(r.get('credit')) if _money_float(r.get('credit')) else '',
            _accounting_pdf_money(r.get('movement')) if _money_float(r.get('movement')) else ''
        ]))
    return {
        'data': data,
        'payload': {
            'company': company,
            'title': 'Bank Reconciliation',
            'subtitle': f'{start_date} to {end_date}',
            'orientation': 'landscape',
            'columns': [
                {'label': 'Date', 'width': 68}, {'label': 'Reference', 'width': 80}, {'label': 'Source', 'width': 70},
                {'label': 'Description', 'width': 350}, {'label': 'Debit', 'width': 85, 'align': 'right'},
                {'label': 'Credit', 'width': 85, 'align': 'right'}, {'label': 'Movement', 'width': 85, 'align': 'right'}
            ],
            'groups': [
                {'title': 'Reconciliation Summary', 'columns': [{'label': 'Item'}, {'label': 'Amount'}, {'label': ''}, {'label': ''}], 'rows': summary_rows},
                {'title': 'Bank Ledger Transactions', 'rows': detail_rows}
            ]
        }
    }


@app.route('/download/accounting/reports/bank_reconciliation.pdf')
def download_accounting_bank_reconciliation_pdf():
    if not _accounting_report_require_permission():
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    bank_account_id = request.args.get('bank_account_id') or None
    statement_balance = request.args.get('statement_balance') or 0
    reconciliation_id = request.args.get('reconciliation_id') or None
    try:
        reconciliation_id = int(reconciliation_id) if reconciliation_id else None
    except Exception:
        reconciliation_id = None
    conn = get_db_connection()
    try:
        result = _bank_reconciliation_report_payload(conn, cid, start_date, end_date, bank_account_id, statement_balance, reconciliation_id)
        return _accounting_pdf_response(_draw_accounting_report_pdf(result['payload']), _accounting_pdf_safe_filename(f'Bank_Reconciliation_{result["data"].get("end_date") or end_date}'))
    finally:
        conn.close()


@app.route('/download/accounting/reports/bank_reconciliation.xlsx')
def download_accounting_bank_reconciliation_xlsx():
    if not _accounting_report_require_permission():
        return 'Forbidden', 403
    cid = _current_company_id()
    start_date = request.args.get('start_date') or datetime.now().strftime('%Y-01-01')
    end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
    bank_account_id = request.args.get('bank_account_id') or None
    statement_balance = request.args.get('statement_balance') or 0
    reconciliation_id = request.args.get('reconciliation_id') or None
    try:
        reconciliation_id = int(reconciliation_id) if reconciliation_id else None
    except Exception:
        reconciliation_id = None
    conn = get_db_connection()
    try:
        result = _bank_reconciliation_report_payload(conn, cid, start_date, end_date, bank_account_id, statement_balance, reconciliation_id)
        data = result['data']
        summary_rows = [
            _accounting_make_row(['Bank Account', data.get('bank_account') or '', '', ''], bold=True, shade=True),
            _accounting_make_row(['Statement Closing Balance', _excel_num(data.get('statement_balance')), '', '']),
            _accounting_make_row(['General Ledger Bank Balance', _excel_num(data.get('gl_balance')), '', '']),
            _accounting_make_row(['Difference', _excel_num(data.get('difference_amount')), '', ''], bold=True, shade=True, total=True),
        ]
        detail_rows = []
        for r in data.get('rows') or []:
            detail_rows.append(_accounting_make_row([
                r.get('date') or '', r.get('reference') or '', r.get('source') or '', r.get('description') or '',
                _excel_num(r.get('debit')) if _money_float(r.get('debit')) else '',
                _excel_num(r.get('credit')) if _money_float(r.get('credit')) else '',
                _excel_num(r.get('movement')) if _money_float(r.get('movement')) else ''
            ]))
        payload = result['payload']
        payload['groups'] = [
            {'title': 'Reconciliation Summary', 'columns': [{'label': 'Item'}, {'label': 'Amount'}, {'label': ''}, {'label': ''}], 'rows': summary_rows},
            {'title': 'Bank Ledger Transactions', 'rows': detail_rows}
        ]
        return _accounting_excel_response(_build_accounting_report_xlsx(payload), _accounting_xlsx_safe_filename(f'Bank_Reconciliation_{data.get("end_date") or end_date}'))
    finally:
        conn.close()

@app.route('/health')
def health_check():
    return 'OK', 200

init_db()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
